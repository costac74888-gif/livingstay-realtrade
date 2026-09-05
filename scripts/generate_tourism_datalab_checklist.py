"""Generate the operational (not official publication) Data Lab checklist."""
from datetime import date
from pathlib import Path
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

P={"visitor_sgg":"202507-202606","foreign_sgg":"202507-202606","visitor_sido":"202507-202606","foreign_sido":"202507-202606","visitor_trend":"202507-202606","foreign_trend":"202507-202606","foreign_country":"202507-202606","surge_domestic_dong":"202608-202608","surge_foreign_dong":"202608-202608","consumption_region":"202508-202607","consumption_trend":"202508-202607","consumption_sector":"202508-202607","search_sgg":"202508-202607","search_trend":"202508-202607","search_ranking":"202508-202607","lodging_search_rank":"기간 미상","lodging_sector":"202601-202607","camping_sector":"202601-202607","camping_site_type":"202602-202608"}
names={"visitor_sgg":"지역별 방문자 수(기초지자체별)","foreign_sgg":"외국인 지역별 방문자 수(기초지자체별)","visitor_sido":"지역별 방문자 수(광역별)","foreign_sido":"외국인 지역별 방문자 수(광역별)","visitor_trend":"방문자 수 추이","foreign_trend":"외국인 방문자 수 추이","foreign_country":"외국인 방문자 거주지","surge_domestic_dong":"방문자 급등동네(내국인)","surge_foreign_dong":"방문자 급등동네(외국인)","consumption_region":"지역별 지출액","consumption_trend":"관광소비 추이","consumption_sector":"업종별 지출액","search_sgg":"지역별 검색건수","search_trend":"검색건수 추이","search_ranking":"지역별 관광지 검색순위","lodging_search_rank":"관광숙박 검색순위","lodging_sector":"업종별 분포","camping_sector":"캠핑장 업종별 분포","camping_site_type":"캠핑사이트 유형별 현황"}
wb=Workbook(); g=wb.active; g.title="안내"
g.append(["관광 데이터랩 정기 다운로드·반영 체크리스트"]); g.append(["기준일: 2026-09-05 / 포털 로그인·약관·메뉴·기간 선택으로 다운로드는 수동입니다."])
g.append(["https://datalab.visitkorea.or.kr/datalab/portal/main/getMainForm.do"]); g.append(["점검 일정은 운영 권장사항이며 공식 발행일 또는 발행 보장이 아닙니다."])
for r in range(1,5): g.merge_cells(start_row=r,start_column=1,end_row=r,end_column=12)
g["A1"].font=Font(size=16,bold=True,color="FFFFFF");g["A1"].fill=PatternFill("solid",fgColor="164E63")
w=wb.create_sheet("정기다운로드목록"); heads=["stat_type","한국어 데이터명","다운로드 메뉴 묶음","예상 내부 파일명","최신 보유 기간","실제 보유 원본 예시","권장 주기","확인 기준일","다음 권장 점검일","포털 선택/필터","우선순위","운영 메모"];w.append(heads)
zips=sorted(Path("attached_assets").glob("*데이터랩*.zip"))
for t,period in P.items():
 annual=t in {"lodging_sector","camping_sector","camping_site_type"}; group="관광지 검색" if "search" in t or t=="lodging_search_rank" else ("관광소비" if "consumption" in t else "방문자/외국인")
 w.append([t,names[t],group,f"{names[t]}_YYYYMM-YYYYMM.csv",period,zips[0].name if zips else "-", "연간" if annual else "월간",date(2026,9,5),date(2027,1,10) if annual else date(2026,10,5),f"전국 · 기간 · stat_type={t}"+(" · 중분류=숙박" if t=="lodging_search_rank" else ""),"높음" if t in {"lodging_search_rank","visitor_sgg","foreign_sgg"} else "보통","권장 일정(공식 발행 보장 아님)"])
i=wb.create_sheet("기존보유원본");i.append(["구분","파일명","source period","내부 CSV 요약","확인 기준일","비고"])
for p in zips:
 with zipfile.ZipFile(p) as z: summary=", ".join(Path(x.filename).name for x in z.infolist() if x.filename.lower().endswith(".csv"))
 i.append(["Data Lab ZIP",p.name,p.name.split("_")[2] if len(p.name.split("_"))>2 else "기간 미상",summary,date(2026,9,5),"실제 첨부 자산"])
for p in sorted(Path("attached_assets").glob("*관광지_검색순위*.csv")): i.append(["단독 CSV",p.name,"기간 미상",p.name,date(2026,9,5),"실제 첨부 자산"])
for s in wb:
 s.freeze_panes="A2";s.sheet_view.showGridLines=False;s.sheet_properties.pageSetUpPr.fitToPage=True;s.page_setup.orientation="landscape";s.page_setup.fitToWidth=1;s.print_title_rows="1:1";s.oddFooter.center.text="운영 권장 일정 · Page &P of &N"
 for row in s:
  for c in row:c.alignment=Alignment(wrap_text=True,vertical="top")
 if s in (w,i):
  s.auto_filter.ref=s.dimensions; tab=Table(displayName="T"+s.title.encode().hex()[:12],ref=s.dimensions);tab.tableStyleInfo=TableStyleInfo(name="TableStyleMedium2",showRowStripes=True);s.add_table(tab)
 for idx,col in enumerate(s.columns,1):
  from openpyxl.utils import get_column_letter
  s.column_dimensions[get_column_letter(idx)].width=min(55,max(13,max(len(str(c.value or "")) for c in col)+2))
for s in (w,i):
 for row in range(2,s.max_row+1):
  for col in ([8,9] if s==w else [5]):s.cell(row,col).number_format="yyyy-mm-dd"
Path("exports").mkdir(exist_ok=True);wb.save("exports/관광데이터랩_정기다운로드_목록_20260905.xlsx")