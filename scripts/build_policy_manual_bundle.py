#!/usr/bin/env python3
"""Build and validate the comprehensive policy manual release bundle."""
from __future__ import annotations

import csv
import hashlib
import html
import re
import shutil
import subprocess
import textwrap
import zipfile
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
EXPORTS = ROOT / "exports"
DATE = "2026-09-04"
BASE = f"HOME_AND_STAY_COMPREHENSIVE_OPERATIONS_MANUAL_V01_0_{DATE}"
SOURCE = DOCS / f"{BASE}.md"
FONT = Path("/nix/store/1wd0fbh9pwn9cna5vkj762b72yw974qp-nanum-20200506/share/fonts/NanumBarunGothic.ttf")
BOLD = FONT.with_name("NanumBarunGothicBold.ttf")


def inline(value: str) -> str:
    value = html.escape(value)
    value = re.sub(r"`([^`]+)`", r"<code>\1</code>", value)
    value = re.sub(r"\*\*([^*]+)\*\*", r"<strong>\1</strong>", value)
    return value


def slug(title: str) -> str:
    value = re.sub(r"[^\w가-힣]+", "-", title.lower()).strip("-")
    return value or "section"


def markdown_to_html(md: str) -> str:
    lines = md.splitlines()
    out, toc, in_code, in_ul, in_ol, in_table = [], [], False, False, False, False
    table_rows = []

    def close_lists():
        nonlocal in_ul, in_ol
        if in_ul:
            out.append("</ul>"); in_ul = False
        if in_ol:
            out.append("</ol>"); in_ol = False

    def close_table():
        nonlocal in_table, table_rows
        if not in_table:
            return
        out.append("<table>")
        for index, cells in enumerate(table_rows):
            tag = "th" if index == 0 else "td"
            out.append("<tr>" + "".join(f"<{tag}>{inline(c)}</{tag}>" for c in cells) + "</tr>")
        out.append("</table>")
        in_table, table_rows = False, []

    for raw in lines:
        line = raw.rstrip()
        if line.startswith("```"):
            close_lists(); close_table()
            out.append("</code></pre>" if in_code else "<pre><code>")
            in_code = not in_code
            continue
        if in_code:
            out.append(html.escape(line) + "\n")
            continue
        if re.match(r"^\|.*\|$", line):
            close_lists()
            cells = [c.strip() for c in line.strip("|").split("|")]
            if all(re.fullmatch(r":?-{3,}:?", c) for c in cells):
                continue
            in_table = True; table_rows.append(cells); continue
        close_table()
        match = re.match(r"^(#{1,3})\s+(.+)$", line)
        if match:
            close_lists()
            level, title = len(match.group(1)), match.group(2)
            anchor = slug(title)
            out.append(f"<h{level} id=\"{anchor}\">{inline(title)}</h{level}>")
            if level == 2:
                toc.append((title, anchor))
            continue
        if line == "---":
            close_lists(); out.append("<hr>"); continue
        if line.startswith("- "):
            if not in_ul: close_lists(); out.append("<ul>"); in_ul = True
            out.append(f"<li>{inline(line[2:])}</li>"); continue
        ordered = re.match(r"^\d+\.\s+(.+)$", line)
        if ordered:
            if not in_ol: close_lists(); out.append("<ol>"); in_ol = True
            out.append(f"<li>{inline(ordered.group(1))}</li>"); continue
        close_lists()
        if line:
            out.append(f"<p>{inline(line)}</p>")
    close_lists(); close_table()
    toc_html = '<nav class="toc"><h2>문서 목차</h2><ol>' + "".join(
        f'<li><a href="#{anchor}">{inline(title)}</a></li>' for title, anchor in toc
    ) + "</ol></nav>"
    body = "\n".join(out)
    first_h1_end = body.find("</h1>") + 5
    return body[:first_h1_end] + toc_html + body[first_h1_end:]


def build_html(md: str, path: Path) -> None:
    css = """
    @font-face{font-family:PolicyKR;src:url('file://%s')}@font-face{font-family:PolicyKR;src:url('file://%s');font-weight:700}
    @page{size:A4;margin:16mm 15mm 18mm;@bottom-center{content:counter(page) " / " counter(pages);font-size:8pt;color:#667}}
    *{box-sizing:border-box}body{font-family:PolicyKR,'Malgun Gothic',sans-serif;color:#18212f;font-size:9.4pt;line-height:1.55;max-width:190mm;margin:auto}
    h1{font-size:25pt;color:#17324d;border-bottom:3px solid #c49745;padding-bottom:10px}h2{font-size:16pt;color:#17324d;border-bottom:1px solid #ccd5df;padding-bottom:5px;break-after:avoid}
    h3{font-size:12.5pt;color:#8a5b16;break-after:avoid}table{width:100%%;border-collapse:collapse;margin:8px 0 15px;font-size:8.2pt}tr{break-inside:avoid}
    th{background:#17324d;color:#fff}th,td{border:1px solid #b9c4cf;padding:4px 5px;vertical-align:top}pre{white-space:pre-wrap;background:#17202b;color:white;padding:9px}
    code{background:#eef1f4;padding:1px 3px}pre code{background:transparent}.toc{break-after:page}.toc a{color:#17324d;text-decoration:none}a{color:#165a92}
    """ % (FONT, BOLD)
    document = f'<!doctype html><html lang="ko"><head><meta charset="utf-8"><title>홈앤스테이 종합매뉴얼 V.01.0</title><style>{css}</style></head><body>{markdown_to_html(md)}</body></html>'
    path.write_text(document, encoding="utf-8")


def build_pdf(md: str, path: Path) -> None:
    doc = fitz.open()
    font = str(FONT)
    bold = str(BOLD)
    page = doc.new_page(width=595, height=842)
    page.insert_font(fontname="KR", fontfile=font)
    page.insert_font(fontname="KRB", fontfile=bold)
    y, page_no = 55, 1

    def new_page():
        nonlocal page, y, page_no
        page.insert_text((280, 820), str(page_no), fontname="KR", fontsize=8, color=(.35, .4, .45))
        page_no += 1
        page = doc.new_page(width=595, height=842)
        page.insert_font(fontname="KR", fontfile=font)
        page.insert_font(fontname="KRB", fontfile=bold)
        y = 55

    for raw in md.splitlines():
        line = re.sub(r"[`*_]", "", raw.strip())
        if not line or line == "---" or line.startswith("|---"):
            y += 5; continue
        if line.startswith("# "): size, lead, face = 22, 31, "KRB"; line = line[2:]
        elif line.startswith("## "): size, lead, face = 15, 24, "KRB"; line = line[3:]
        elif line.startswith("### "): size, lead, face = 11.5, 18, "KRB"; line = line[4:]
        else: size, lead, face = 8.8, 13, "KR"
        if line.startswith("|"):
            line = "  |  ".join(cell.strip() for cell in line.strip("|").split("|"))
            size = 7.2
        width = max(28, int(92 * 8.8 / size))
        wrapped = textwrap.wrap(line, width=width, break_long_words=False) or [""]
        needed = lead + (len(wrapped) - 1) * (size + 3)
        if y + needed > 790:
            new_page()
        for index, part in enumerate(wrapped):
            page.insert_text((48, y), part, fontname=face, fontsize=size, color=(.08, .15, .22))
            y += lead if index == 0 else size + 3
    page.insert_text((280, 820), str(page_no), fontname="KR", fontsize=8, color=(.35, .4, .45))
    doc.set_metadata({"title": "홈앤스테이 종합 운영·통계·공개정책 매뉴얼", "author": "홈앤스테이", "subject": "V.01.0 / 2026-09-04"})
    doc.save(path, deflate=True)


def build_docx(md: str, path: Path) -> None:
    def p(style, text):
        escaped = html.escape(text)
        return f'<w:p><w:pPr><w:pStyle w:val="{style}"/></w:pPr><w:r><w:t xml:space="preserve">{escaped}</w:t></w:r></w:p>'
    def table(rows):
        xml = ['<w:tbl><w:tblPr><w:tblW w:w="0" w:type="auto"/><w:tblBorders><w:top w:val="single" w:sz="4"/><w:left w:val="single" w:sz="4"/><w:bottom w:val="single" w:sz="4"/><w:right w:val="single" w:sz="4"/><w:insideH w:val="single" w:sz="4"/><w:insideV w:val="single" w:sz="4"/></w:tblBorders></w:tblPr>']
        for row_index, row in enumerate(rows):
            xml.append("<w:tr>")
            for cell in row:
                shade = '<w:shd w:fill="17324D"/>' if row_index == 0 else ""
                bold = "<w:b/>" if row_index == 0 else ""
                color = '<w:color w:val="FFFFFF"/>' if row_index == 0 else ""
                xml.append(f'<w:tc><w:tcPr>{shade}</w:tcPr><w:p><w:r><w:rPr>{bold}{color}</w:rPr><w:t>{html.escape(cell)}</w:t></w:r></w:p></w:tc>')
            xml.append("</w:tr>")
        xml.append("</w:tbl>")
        return "".join(xml)
    blocks, lines, index = [], md.splitlines(), 0
    while index < len(lines):
        line = re.sub(r"[`*_]", "", lines[index].strip())
        if re.match(r"^\|.*\|$", line):
            rows = []
            while index < len(lines) and re.match(r"^\|.*\|$", lines[index].strip()):
                cells = [re.sub(r"[`*_]", "", c.strip()) for c in lines[index].strip().strip("|").split("|")]
                if not all(re.fullmatch(r":?-{3,}:?", c) for c in cells):
                    rows.append(cells)
                index += 1
            if rows: blocks.append(table(rows))
            continue
        index += 1
        if not line or line == "---": continue
        style = "Normal"
        for prefix, candidate in (("# ", "Title"), ("## ", "Heading1"), ("### ", "Heading2")):
            if line.startswith(prefix): line, style = line[len(prefix):], candidate; break
        blocks.append(p(style, line))
    document = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:body>' + "".join(blocks) + '<w:sectPr><w:pgSz w:w="11906" w:h="16838"/><w:pgMar w:top="900" w:right="850" w:bottom="1000" w:left="850"/><w:footerReference w:type="default" r:id="rId1" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/></w:sectPr></w:body></w:document>'
    styles = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><w:styles xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:docDefaults><w:rPrDefault><w:rPr><w:rFonts w:ascii="NanumBarunGothic" w:eastAsia="NanumBarunGothic"/><w:sz w:val="19"/></w:rPr></w:rPrDefault></w:docDefaults><w:style w:type="paragraph" w:default="1" w:styleId="Normal"><w:name w:val="Normal"/></w:style><w:style w:type="paragraph" w:styleId="Title"><w:name w:val="Title"/><w:rPr><w:b/><w:sz w:val="44"/></w:rPr></w:style><w:style w:type="paragraph" w:styleId="Heading1"><w:name w:val="heading 1"/><w:rPr><w:b/><w:sz w:val="30"/></w:rPr></w:style><w:style w:type="paragraph" w:styleId="Heading2"><w:name w:val="heading 2"/><w:rPr><w:b/><w:sz w:val="24"/></w:rPr></w:style></w:styles>'
    footer = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><w:ftr xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"><w:p><w:pPr><w:jc w:val="center"/></w:pPr><w:r><w:fldChar w:fldCharType="begin"/></w:r><w:r><w:instrText>PAGE</w:instrText></w:r><w:r><w:fldChar w:fldCharType="end"/></w:r></w:p></w:ftr>'
    content_types = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/><Override PartName="/word/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.styles+xml"/><Override PartName="/word/footer1.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.footer+xml"/></Types>'
    root_rels = '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/></Relationships>'
    doc_rels = '<?xml version="1.0" encoding="UTF-8"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/footer" Target="footer1.xml"/><Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/></Relationships>'
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", content_types); z.writestr("_rels/.rels", root_rels)
        z.writestr("word/document.xml", document); z.writestr("word/styles.xml", styles)
        z.writestr("word/footer1.xml", footer); z.writestr("word/_rels/document.xml.rels", doc_rels)


def build_register_xlsx(csv_path: Path, output: Path) -> None:
    with csv_path.open(encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.reader(fh))
    wb = Workbook(); ws = wb.active; ws.title = "문서대장"
    for row in rows: ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="17324D")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    widths = [24, 42, 14, 16, 20, 12, 16, 70, 26]
    for index, width in enumerate(widths, 1): ws.column_dimensions[chr(64 + index)].width = width
    for row in ws.iter_rows():
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    readme = wb.create_sheet("상태·대체관계")
    readme.append(["상태", "의미"]); readme.append(["current", "현재 운영 판단 기준"])
    readme.append(["superseded-preserved", "종합매뉴얼이 탐색 기준을 대체하나 상세 근거로 보존"])
    readme.append(["historical", "당시 증거로 보존, 현재 절차 지시 아님"])
    wb.save(output)


def validate(outputs: list[Path], zip_path: Path) -> str:
    md = SOURCE.read_text(encoding="utf-8")
    required = ["## 1.", "## 11.", "### 4.8", "### 6.7", "### 10.3", "일반야영", "자동차야영", "글램핑", "카라반", "부록 A. 공통 용어집", "상태별 처리표", "기준선·dry-run·승인·장애 기록 양식"]
    assert all(token in md for token in required)
    headings = [int(x) for x in re.findall(r"^## (\d+)\.", md, re.M)]
    assert headings == list(range(1, 13)), headings
    pdf = fitz.open(outputs[2]); pdf_text = "".join(page.get_text() for page in pdf)
    assert "홈앤스테이" in pdf_text and "카라반" in pdf_text
    with zipfile.ZipFile(outputs[1]) as z:
        assert z.testzip() is None and "word/document.xml" in z.namelist()
        docx_xml = z.read("word/document.xml").decode("utf-8")
        assert "홈앤스테이" in docx_xml and docx_xml.count("<w:tbl>") >= 10
    html_text = outputs[0].read_text(encoding="utf-8")
    assert html_text.count('href="#') >= 12 and "PolicyKR" in html_text
    with zipfile.ZipFile(zip_path) as z:
        assert z.testzip() is None
        names = set(z.namelist())
        for path in outputs + [SOURCE, DOCS / "POLICY_DOCUMENT_REGISTER_V01_0_2026-09-04.md", DOCS / "POLICY_DOCUMENT_REGISTER_V01_0_2026-09-04.csv"]:
            assert path.name in names
    toc_link_count = html_text.count('href="#')
    report = [
        "HOME & STAY POLICY MANUAL RELEASE VALIDATION",
        f"source_sha256={hashlib.sha256(SOURCE.read_bytes()).hexdigest()}",
        f"chapters={headings}",
        f"html_toc_links={toc_link_count}",
        f"pdf_pages={len(pdf)}",
        "pdf_korean_text=PASS",
        "docx_package_and_korean_text=PASS",
        f"docx_tables={docx_xml.count('<w:tbl>')}",
        "zip_crc_and_required_files=PASS",
        "camping_facility_and_four_site_categories=PASS",
    ]
    return "\n".join(report) + "\n"


def main() -> None:
    EXPORTS.mkdir(exist_ok=True)
    md = SOURCE.read_text(encoding="utf-8")
    html_path, docx_path, pdf_path = [EXPORTS / f"{BASE}.{ext}" for ext in ("html", "docx", "pdf")]
    build_html(md, html_path); build_docx(md, docx_path); build_pdf(md, pdf_path)
    register_xlsx = EXPORTS / f"HOME_AND_STAY_POLICY_BOOK_REGISTER_V01_0_{DATE}.xlsx"
    build_register_xlsx(DOCS / f"POLICY_DOCUMENT_REGISTER_V01_0_{DATE}.csv", register_xlsx)
    zip_path = EXPORTS / f"HOME_AND_STAY_POLICY_MANUALS_{DATE}.zip"
    bundle = [
        SOURCE,
        DOCS / f"POLICY_DOCUMENT_REGISTER_V01_0_{DATE}.md",
        DOCS / f"POLICY_DOCUMENT_REGISTER_V01_0_{DATE}.csv",
        html_path, docx_path, pdf_path, register_xlsx,
    ]
    detailed_sources = [
        DOCS / "DATA_COLLECTION_GOVERNANCE_OPERATIONS_POLICY_V03_0_2026-09-04.md",
        DOCS / "DATA_COLLECTION_SYNC_OPERATIONS_MANUAL_V01_0_2026-09-04.md",
        DOCS / "LODGING_STATISTICS_PUBLIC_POLICY_MANUAL_V01_0_2026-09-04.md",
        DOCS / "PUBLIC_DATA_PRIVACY_PROTECTION_POLICY_V01_0_2026-09-04.md",
        DOCS / "PHOTO_RETRIEVAL_OPERATIONS_MANUAL_V01_0_2026-09-04.md",
        DOCS / "COMPREHENSIVE_MANUAL_BASELINE_AND_OUTLINE_V01_0_2026-09-04.md",
    ]
    readme = EXPORTS / "POLICY_BUNDLE_README.md"
    readme.write_text("# 홈앤스테이 종합매뉴얼 배포 묶음\n\n권위 원본은 종합매뉴얼 Markdown이며 HTML·DOCX·PDF는 동일 원본의 배포본입니다. 개별 정책·매뉴얼은 상세 근거와 역사 보존을 위해 프로젝트 docs에 유지됩니다.\n", encoding="utf-8")
    report_path = EXPORTS / "VALIDATION_REPORT.txt"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for path in bundle + detailed_sources + [readme]: z.write(path, path.name)
    report_path.write_text(validate([html_path, docx_path, pdf_path, register_xlsx], zip_path), encoding="utf-8")
    with zipfile.ZipFile(zip_path, "a", zipfile.ZIP_DEFLATED) as z:
        z.write(report_path, report_path.name)
    print(report_path.read_text(encoding="utf-8"), end="")


if __name__ == "__main__":
    main()