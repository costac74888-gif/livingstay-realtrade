# -*- coding: utf-8 -*-
"""
app.py — 검색 API + 정적 페이지 서빙 (Replit에서 바로 실행)

엔드포인트
------------------------------------------------------------
GET /                          → static/index.html 서빙
GET /api/transactions          → 게시판(전체 최신순) or 검색 결과
    쿼리파라미터:
      q       : 건물명 또는 주소 검색어 (부분일치, 보조 수단)
      si_do   : 시/도 (예: '경기도') — 행정접미사 떼고 코어 이름으로 비교
                ('서울'과 '서울특별시'가 동일하게 취급됨)
      sgg_nm  : 시/군/구 (예: '수원시') — 정확히 일치
      umd_nm  : 읍/면/동 (예: '매산로1가') — 정확히 일치
      year    : 계약연도 (예: '2026', 'all'이면 전체)
      page/size
GET /api/regions               → 계층형 지역 트리 (시도 > 시군구 > 읍면동, 각 count)
GET /api/health                → 배치 마지막 실행 시각/건수 확인용
"""

import os
import re
import io
import sys
import json
import time
import hashlib
import threading
import subprocess
import secrets as _secrets
from functools import wraps
from urllib.parse import quote, urlencode
import requests
from werkzeug.security import generate_password_hash, check_password_hash

from sms_util import send_sms
from email_util import send_email
import storage_util
import addr_norm
from psycopg2 import errors as psycopg2_errors
from psycopg2.extras import execute_values
from flask import Flask, request, jsonify, send_from_directory, Response, abort, session, redirect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from datetime import datetime
from db import get_conn, init_db
from address_utils import (
    normalize_umd_nm, sido_core, sido_match_clause,
    build_authority_index, match_authority_contact,
    BjdongMap, parse_jibun,
)
from store_info_util import build_pnu, get_stores_by_pnu

# 서버 기동 시각 — 정적 SDK URL 캐시 무효화용 (기동할 때만 바뀜)
SERVER_BOOT_V = str(int(time.time()))

# 파트너(중개사/운영업체) 1곳이 무료로 담당 등록할 수 있는 건물 수 상한.
# 가격정책 확정 전 임시 무료 캡 — 정책 확정 시 이 상수만 조정하면 됨.
MAX_FREE_BUILDINGS = 5

# 정적 JS/CSS 자산에 배포마다 바뀌는 버전 쿼리스트링(?v=SERVER_BOOT_V)을 붙여
# 새 배포 때 브라우저가 무조건 새 파일을 받도록 한다(캐시버스팅). 버전 값은
# 하드코딩하지 않고 서버 기동 시각(=배포마다 갱신)을 재사용한다.
_ASSET_VER_RE = re.compile(r'(src|href)="(/static/(?:js|css)/[^"?]+\.(?:js|css))"')


def _inject_asset_version(html):
    return _ASSET_VER_RE.sub(
        lambda m: f'{m.group(1)}="{m.group(2)}?v={SERVER_BOOT_V}"', html
    )

app = Flask(__name__, static_folder="static")

# 대출상담사 '상담 가능 상품' 허용 목록 — 프로필 체크박스/B화면 태그 공통 기준.
# 순서 = 노출 순서 (생활숙박시설 담보대출이 항상 최상단).
LOAN_CONSULTANT_PRODUCTS = [
    "생활숙박시설 담보대출",
    "주택담보대출",
    "전세대출",
    "사업자대출",
    "정책자금·채무조정",
    "차량담보대출",
]
# 관리자 세션(서명된 쿠키) 서명 키. FLASK_SECRET_KEY가 없으면 세션이 유지되지
# 않아 관리자 로그인이 동작하지 않으므로 Secrets에 반드시 등록되어 있어야 한다.
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "")
if not app.secret_key:
    # 세션 쿠키는 이 키로 서명된다. 빈 키면 누구나 세션(회원/관리자)을 위조할 수
    # 있으므로 로그인 기능의 신뢰가 붕괴된다. 없으면 기동 자체를 중단한다.
    raise RuntimeError(
        "FLASK_SECRET_KEY가 설정되지 않았습니다. Secrets에 등록 후 다시 실행하세요."
    )

# 세션 쿠키 보안 플래그.
#  - HTTPONLY: JS(document.cookie)에서 세션 쿠키를 읽지 못하게 하여 XSS 탈취 방지.
#  - SECURE: HTTPS로만 전송(리플릿은 dev/prod 모두 HTTPS).
#  - SAMESITE=Lax: 카카오 OAuth 콜백처럼 외부에서 되돌아오는 top-level GET 이동에는
#    쿠키가 실려야 state(CSRF) 검증이 가능하므로 Strict가 아닌 Lax를 사용한다.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_SAMESITE="Lax",
)


def get_client_ip():
    """
    rate limiter가 IP별로 카운터를 나누는 기준 IP를 돌려준다.

    리플릿은 앱 앞에 여러 프록시 홉이 있고(확인된 체인 예:
    '<client>, 10.x, 10.x, 127.0.0.1'), 엣지 프록시가 클라이언트가 보낸
    X-Forwarded-For를 무시하고 새로 세팅한다(위조 XFF가 제거되는 것 확인함).
    따라서 XFF 맨 앞(최초) 항목이 실제 방문자 IP다.
    remote_addr(=마지막 홉 127.0.0.1)을 키로 쓰면 모든 사용자가 한 카운터를
    공유하게 되므로 절대 쓰면 안 된다.
    """
    xff = request.headers.get("X-Forwarded-For")
    if xff:
        return xff.split(",")[0].strip()
    return get_remote_address()


# 메모리 기반 rate limiter (별도 인프라 불필요).
# 기본값은 걸지 않고, 쓰기성 API에만 데코레이터로 개별 제한을 건다.
limiter = Limiter(
    key_func=get_client_ip,
    app=app,
    storage_uri="memory://",
)


@app.errorhandler(429)
def ratelimit_handler(e):
    return (
        jsonify({"message": "너무 많은 요청입니다. 잠시 후 다시 시도해주세요."}),
        429,
    )

# 방문 기록(page_views)용 고정 salt — 원본 IP를 그대로 저장하지 않고 해시할 때 쓴다.
# 코드에 평문 salt를 박지 않으려고 FLASK_SECRET_KEY를 재사용(이미 시크릿). 없으면 폴백.
_PAGE_VIEW_SALT = os.environ.get("FLASK_SECRET_KEY", "") or "livingstay_pageview_salt_v1"


def _record_page_view(path, resp_status):
    """실제 사용자 페이지(GET·200)만 page_views에 1건 기록한다.
    개인정보 최소수집: 방문자 IP는 원본 저장 없이 sha256(IP+salt)로만 남긴다.
    통계 기록 실패가 절대 실서비스 요청을 죽이지 않도록 전 구간을 try/except로 감싼다."""
    conn = None
    cur = None
    try:
        ip = get_client_ip() or ""
        ip_hash = hashlib.sha256((ip + _PAGE_VIEW_SALT).encode("utf-8")).hexdigest()
        ua = (request.headers.get("User-Agent") or "")[:500]
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO page_views (path, ip_hash, user_agent) VALUES (%s, %s, %s)",
            [path[:500], ip_hash, ua],
        )
        conn.commit()
    except Exception:
        # 통계 때문에 사용자 요청이 실패하면 안 된다 → 조용히 무시.
        pass
    finally:
        # 예외가 나도 커서/커넥션은 반드시 정리해 연결 누수를 막는다.
        if cur is not None:
            try:
                cur.close()
            except Exception:
                pass
        if conn is not None:
            try:
                conn.close()
            except Exception:
                pass


@app.after_request
def _log_page_view(resp):
    """사용자 페이지 조회만 집계: GET·200이고 /api·/admin·/static 이 아닌 경로.
    (실제 집계 대상: /, /building/<id>, /apply/agent, /apply/operator 등)"""
    try:
        if request.method == "GET" and resp.status_code == 200:
            path = request.path or "/"
            excluded = any(
                path == p or path.startswith(p + "/")
                for p in ("/api", "/admin", "/static")
            )
            # 정적 자산(favicon.ico, .js, .css 등)도 페이지 조회가 아니므로 제외
            is_asset = "." in path.rsplit("/", 1)[-1]
            if not excluded and not is_asset:
                _record_page_view(path, resp.status_code)
    except Exception:
        pass
    return resp


# 앱 부팅 시 스키마를 보장한다 (building_requests 정정 컬럼 등).
# init_db는 CREATE/ALTER ... IF NOT EXISTS라 여러 번 호출해도 안전(멱등).
# 이렇게 해야 배포 직후(아직 sync 스크립트가 안 돈 시점)에도 요청 API가 500 없이 동작한다.
init_db()


def _serve_app_shell():
    # 정적 index.html을 읽어 카카오맵 JS 키만 서버에서 주입해 서빙한다.
    # (프론트 소스에 키를 직접 박지 않고, 환경변수/시크릿에서 안전하게 넣는다.)
    kakao_js_key = os.environ.get("KAKAO_JS_KEY", "")
    html_path = os.path.join(app.static_folder, "index.html")
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    html = html.replace("{{KAKAO_JS_KEY}}", quote(kakao_js_key, safe=""))
    html = html.replace("{{KAKAO_SDK_V}}", SERVER_BOOT_V)
    html = _inject_asset_version(html)
    resp = Response(html, mimetype="text/html")
    # 진입 HTML은 캐시하지 않아 항상 최신 SDK URL(_v)을 받도록 한다.
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


@app.route("/")
def index():
    return _serve_app_shell()


@app.route("/manifest.json")
def pwa_manifest():
    """PWA 매니페스트 — 루트 경로로 서빙 (모든 페이지의 <link rel="manifest">가 참조)."""
    return send_from_directory(app.static_folder, "manifest.json", mimetype="application/manifest+json")


@app.route("/favicon.ico")
def favicon_ico():
    """브라우저가 관성적으로 루트 /favicon.ico를 요청하는 경우 대응."""
    return send_from_directory(os.path.join(app.static_folder, "img"), "favicon.ico")


@app.route("/building/<int:building_id>")
def building_page(building_id):
    """건물 상세 — 별도 페이지가 아니라 홈화면(index.html)을 그대로 서빙한다.

    상세 내용은 프런트(main.js renderBuildingPanel)에서 좌측 패널 안에 그린다.
    새로고침/공유 링크로 /building/<id>에 직접 들어와도 index.html이 로드되며,
    main.js가 URL을 확인해 자동으로 해당 건물 상세를 좌측 패널에 표시한다.
    (static/building.html은 롤백 대비 남겨두되 더 이상 서빙하지 않는다.)
    """
    return _serve_app_shell()


_AUTHORITY_INDEX = None


def get_authority_index():
    """lodging_authority_contacts(135행)로 담당부서 매칭 인덱스를 1회 만들어 캐시한다.

    데이터가 정적(엑셀 적재본)이고 gunicorn 자동 리로드가 없어 프로세스 단위 캐시로 충분.
    적재를 다시 하면(load_authority_contacts.py) 앱을 재시작해야 새 값이 반영된다.
    """
    global _AUTHORITY_INDEX
    if _AUTHORITY_INDEX is None:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT region_name_raw, dept, phone FROM lodging_authority_contacts")
        rows = [dict(r) for r in cur.fetchall()]
        cur.close()
        conn.close()
        _AUTHORITY_INDEX = build_authority_index(rows)
    return _AUTHORITY_INDEX


@app.route("/api/building/<int:building_id>")
def get_building(building_id):
    """건물 상세페이지용 단건 조회 — master_buildings 기준."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT mb.building_name, mb.name_pending, mb.road_address, mb.jibun_address,
               mb.lodging_type, mb.lodging_type_detail,
               mb.sgg_cd, mb.umd_nm, mb.jibun,
               mb.units, mb.biz_units, mb.lat, mb.lng, mb.sgg_text,
               mb.use_apr_day, mb.tot_pkng_cnt, mb.grnd_flr_cnt, mb.ugrnd_flr_cnt,
               mb.tot_area, mb.plat_area, mb.hhld_cnt, mb.strct_nm,
               lt.address AS address
        FROM master_buildings mb
        LEFT JOIN LATERAL (
            SELECT t.address
            FROM transactions t
            WHERE (
                    mb.sgg_cd IS NOT NULL AND mb.umd_nm IS NOT NULL AND mb.jibun IS NOT NULL
                    AND t.sgg_cd = mb.sgg_cd
                    AND t.umd_nm = mb.umd_nm
                    AND t.jibun  = mb.jibun
                  )
               OR (
                    (mb.sgg_cd IS NULL OR mb.umd_nm IS NULL OR mb.jibun IS NULL)
                    AND mb.building_name <> '-'
                    AND t.building_name = mb.building_name
                  )
            ORDER BY (t.building_name = mb.building_name) DESC NULLS LAST, t.deal_date DESC
            LIMIT 1
        ) lt ON TRUE
        WHERE mb.id = %s
    """, [building_id])
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({"error": "not found"}), 404

    # 전속중개사: agent_buildings에 이 건물이 등록된 approved 중개사 — 최대 3명.
    # 정렬: priority_score DESC(유료 우선노출 자리, 현재 전부 0) → 동점자는 RANDOM().
    # 단일 쿼리 한 줄로 처리해야 전원 0점일 때도 매번 완전 랜덤이 된다 (2단계 분리 금지).
    cur.execute("""
        SELECT a.id, a.office_name, a.owner_name, a.phone, a.subdomain_slug, a.photo_url
        FROM agent_buildings ab
        JOIN agents a ON a.id = ab.agent_id
        WHERE ab.master_building_id = %s
          AND a.status = 'approved'
          AND COALESCE(a.is_visible, TRUE)
        ORDER BY COALESCE(a.priority_score, 0) DESC, RANDOM()
        LIMIT 3
    """, [building_id])
    agent_rows = cur.fetchall()

    # 담당 운영업체: operator_buildings에 이 건물이 등록된 approved 운영업체 목록.
    # 화면(B화면 위탁운영/하우스키핑 카드)에서 category별로 골라 최대 3곳씩 표시한다.
    # 정렬은 중개사와 동일하게 priority_score DESC, RANDOM() (카드별 LIMIT은 화면에서 적용).
    cur.execute("""
        SELECT o.company_name, o.category, o.subdomain_slug
        FROM operator_buildings ob
        JOIN operators o ON o.id = ob.operator_id
        WHERE ob.master_building_id = %s
          AND o.status = 'approved'
          AND COALESCE(o.is_visible, TRUE)
        ORDER BY COALESCE(o.priority_score, 0) DESC, RANDOM()
    """, [building_id])
    operator_rows = [dict(r) for r in cur.fetchall()]

    # 숙박업 영업신고(행안부) — 이 건물 주소(도로명+건물번호 정규화)로 매칭된 '영업/정상' 사업장.
    # 정렬: 등록 운영업체(operators approved+노출중)와 상호가 매칭되면 priority_score DESC 우선 →
    #       그 안/미등록끼리는 매번 랜덤. 신고율 = 영업/정상 객실수 합 / 총 세대수(units).
    lodgings = []
    lodging_room_total = 0
    try:
        lr_rows = []
        road_norm = addr_norm.normalize_road_prefix(row["road_address"])
        if road_norm:
            cur.execute("""
                SELECT lr.biz_name, lr.permit_date, lr.room_count, lr.biz_name_norm
                FROM lodging_registry lr
                WHERE lr.road_norm = %s
                  AND lr.biz_status_name = '영업/정상'
            """, [road_norm])
            lr_rows = cur.fetchall()
        if not lr_rows:
            # 2차: 도로명 매칭 실패(도로명 없는 건물 등) 시 지번 정규화 키로 매칭
            jibun_key = addr_norm.get_building_jibun_key(row)
            if jibun_key:
                cur.execute("""
                    SELECT lr.biz_name, lr.permit_date, lr.room_count, lr.biz_name_norm
                    FROM lodging_registry lr
                    WHERE lr.jibun_norm = %s
                      AND lr.biz_status_name = '영업/정상'
                """, [jibun_key])
                lr_rows = cur.fetchall()
        if lr_rows:
            op_map = {}
            if lr_rows:
                # 상호 정규화명 → 등록 운영업체(노출중, priority 최고) 매핑
                cur.execute("""
                    SELECT o.company_name, o.subdomain_slug,
                           COALESCE(o.priority_score, 0) AS priority_score
                    FROM operators o
                    WHERE o.status = 'approved' AND COALESCE(o.is_visible, TRUE)
                """)
                for o in cur.fetchall():
                    norm = addr_norm.normalize_name(o["company_name"])
                    if norm and (norm not in op_map
                                 or o["priority_score"] > op_map[norm]["priority_score"]):
                        op_map[norm] = dict(o)
            import random as _random
            for r in lr_rows:
                d = dict(r)
                norm = d.pop("biz_name_norm", None) or addr_norm.normalize_name(d["biz_name"])
                op = op_map.get(norm) if norm else None
                d["registered"] = op is not None
                d["operator_slug"] = op["subdomain_slug"] if op else None
                d["_prio"] = op["priority_score"] if op else -1
                d["_rand"] = _random.random()
                lodging_room_total += int(d.get("room_count") or 0)
                lodgings.append(d)
            # 등록 운영업체(priority DESC) 최상단 → 나머지는 매번 랜덤
            lodgings.sort(key=lambda d: (-int(d["registered"]), -d["_prio"], d["_rand"]))
            for d in lodgings:
                d.pop("_prio", None)
                d.pop("_rand", None)
    except Exception:
        app.logger.exception("영업신고 매칭 실패 (building_id=%s)", building_id)
        lodgings = []
        lodging_room_total = 0
    cur.close()
    conn.close()

    result = dict(row)
    agents_list = []
    for r in agent_rows:
        agent_d = dict(r)
        # 스토리지 키는 노출하지 않고, 사진이 있으면 공개 프록시 URL만 내려준다.
        agent_d["photo_src"] = f"/api/partners/agent-photo/{agent_d['id']}" if agent_d.pop("photo_url", None) else None
        agents_list.append(agent_d)
    result["agents"] = agents_list
    # 하위호환: 단일 agent를 쓰던 기존 코드용 (첫 번째 = 최우선 노출)
    result["agent"] = agents_list[0] if agents_list else None
    result["operators"] = operator_rows
    # B화면 행정 카드용: 영업/정상 영업신고 목록 + 신고율(영업 객실수 합 / 총 세대수)
    result["lodgings"] = lodgings
    result["lodging_room_total"] = lodging_room_total
    units = result.get("units")
    result["lodging_report_rate"] = (
        round(lodging_room_total * 100.0 / units, 1) if units and lodging_room_total else None
    )

    # 담당부처/연락처: sgg_text를 지자체 담당부서 인덱스와 매칭.
    #   source='exact'(시군구 전용) | 'fallback'(시도 대표) → 화면에서 "(시/도 대표)" 꼬리표 판단용.
    #   매칭 실패면 dept=phone=None → 화면은 "확인중" 유지.
    #   매칭/인덱스 조회가 어떤 이유로 실패해도 건물 상세 전체가 죽지 않도록 방어(담당부처만 확인중).
    try:
        dept, phone, source = match_authority_contact(result.get("sgg_text"), get_authority_index())
    except Exception:
        app.logger.exception("담당부서 매칭 실패 (building_id=%s)", building_id)
        dept, phone, source = None, None, None
    result["authority_dept"] = dept
    result["authority_phone"] = phone
    result["authority_source"] = source if dept is not None else None
    return jsonify(result)


# ── 상거래정보(주변 상가업소) ───────────────────────────────────────────
# 법정동코드 매핑은 zip 로딩이 무거워서 앱 프로세스당 1회만 lazy 로딩한다.
_bjdong_map = None
_bjdong_lock = threading.Lock()


def _get_bjdong_map():
    global _bjdong_map
    if _bjdong_map is None:
        with _bjdong_lock:
            if _bjdong_map is None:
                _bjdong_map = BjdongMap(os.environ.get("BJDONG_CODE_CSV", "법정동코드_전체자료.zip"))
    return _bjdong_map


# 건물별 상가업소 결과 1시간 메모리 캐시 — 같은 건물 반복 조회 시 외부 API 호출 절약.
# (gunicorn 워커별 캐시. 정확성보다 호출량 절감 목적이라 워커 간 공유 안 해도 충분)
_STORES_CACHE_TTL = 3600
_STORES_CACHE_TTL_EMPTY = 300  # 실패/0건은 일시 장애일 수 있어 5분만 캐시(곧 재시도)
_stores_cache = {}
_stores_cache_lock = threading.Lock()


@app.route("/api/building/<int:building_id>/nearby-stores")
def get_building_nearby_stores(building_id):
    """이 건물(지번, PNU 기준)의 상가업소 목록 — 업종별 개수 + 층별 목록.

    키가 없거나(pnu 산출 불가) 조회 실패/0건이면 {"available": False} —
    화면은 기존 "준비 중" 카드를 유지한다.
    """
    now = time.time()
    with _stores_cache_lock:
        hit = _stores_cache.get(building_id)
        if hit:
            ttl = _STORES_CACHE_TTL if hit[1].get("available") else _STORES_CACHE_TTL_EMPTY
            if now - hit[0] < ttl:
                return jsonify(hit[1])

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT sgg_cd, umd_nm, jibun FROM master_buildings WHERE id = %s",
        [building_id],
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return jsonify({"error": "not found"}), 404

    payload = {"available": False, "total": 0, "categories": [], "stores": []}
    try:
        if row["sgg_cd"] and row["umd_nm"] and row["jibun"]:
            bjd = _get_bjdong_map().find_bjdong_cd(row["sgg_cd"], row["umd_nm"])
            if bjd:
                plat_gb, bun, ji = parse_jibun(row["jibun"])
                pnu = build_pnu(row["sgg_cd"], bjd, plat_gb, bun, ji)
                stores = get_stores_by_pnu(pnu) if pnu else []
                if stores:
                    counts = {}
                    for s in stores:
                        cat = s["category"] or "기타"
                        counts[cat] = counts.get(cat, 0) + 1
                    categories = sorted(counts.items(), key=lambda kv: -kv[1])

                    def _floor_sort_key(s):
                        f = s["floor"]
                        try:
                            return (0, int(f))
                        except (ValueError, TypeError):
                            return (1, 0)  # 층 정보 없는 업소는 뒤로

                    stores_sorted = sorted(stores, key=_floor_sort_key)
                    payload = {
                        "available": True,
                        "total": len(stores),
                        "categories": [{"category": c, "count": n} for c, n in categories],
                        "stores": stores_sorted,
                    }
    except Exception:
        # 상거래정보는 부가 정보 — 실패해도 500 내지 말고 "준비 중" 유지
        app.logger.exception("상가업소 조회 실패 (building_id=%s)", building_id)
        payload = {"available": False, "total": 0, "categories": [], "stores": []}

    with _stores_cache_lock:
        _stores_cache[building_id] = (now, payload)
    return jsonify(payload)


@app.route("/api/transactions")
def get_transactions():
    q = request.args.get("q", "").strip()
    si_do = request.args.get("si_do", "").strip()
    sgg_nm = request.args.get("sgg_nm", "").strip()
    umd_nm = request.args.get("umd_nm", "").strip()
    year = request.args.get("year", "").strip()
    lodging_type = request.args.get("lodging_type", "").strip()
    page = max(int(request.args.get("page", 1)), 1)
    size = min(int(request.args.get("size", 20)), 200)
    offset = (page - 1) * size

    where = ["1=1"]
    params = []

    if q:
        where.append("(building_name ILIKE %s OR address ILIKE %s)")
        params += [f"%{q}%", f"%{q}%"]
    if si_do:
        # '서울' vs '서울특별시' 표기 편차 흡수 — 지도와 동일한 코어 이름 비교 규칙 사용
        where.append(sido_match_clause("si_do"))
        params.append(sido_core(si_do))
    if sgg_nm:
        where.append("sgg_nm = %s")
        params.append(sgg_nm)
    if umd_nm:
        where.append("umd_nm = %s")
        params.append(umd_nm)
    if year and year != "all":
        where.append("deal_date LIKE %s")
        params.append(f"{year}-%")
    if lodging_type == "복합":
        # '호텔·콘도'처럼 여러 용도가 병기된 건물만 (백엔드가 LIKE '%·%'로 처리)
        where.append("lodging_type LIKE %s")
        params.append("%·%")
    elif lodging_type:
        where.append("lodging_type = %s")
        params.append(lodging_type)

    # 선택적 building_id → 해당 건물의 실거래만(get_buildings_geo/get_monthly_trend와 동일 전략).
    #   - 지번키(sgg_cd+umd_nm+jibun)가 모두 있으면 지번 정확 매칭(동명 건물 오염 방지)
    #   - 셋 중 하나라도 NULL이면 building_name 폴백('-'/미존재/빈값은 0매칭 처리)
    #   - 정수가 아니거나 없으면 기존 동작(q/지역/연도 등) 그대로 유지(하위호환)
    building_id = request.args.get("building_id", "").strip()
    if building_id.isdigit():
        mconn = get_conn()
        mcur = mconn.cursor()
        mcur.execute("""
            SELECT building_name, sgg_cd, umd_nm, jibun
            FROM master_buildings WHERE id = %s
        """, [int(building_id)])
        b = mcur.fetchone()
        mcur.close()
        mconn.close()
        if b and b["sgg_cd"] and b["umd_nm"] and b["jibun"]:
            where.append("sgg_cd = %s AND umd_nm = %s AND jibun = %s")
            params += [b["sgg_cd"], b["umd_nm"], b["jibun"]]
        else:
            name = (b["building_name"] if b else None) or ""
            where.append("building_name = %s")
            params.append(name if name and name != "-" else "\x00")

    where_sql = " AND ".join(where)

    conn = get_conn()
    cur = conn.cursor()

    cur.execute(f"SELECT COUNT(*) c FROM transactions WHERE {where_sql}", params)
    total = cur.fetchone()["c"]

    # master_building_id: 각 거래의 지번키(sgg_cd+umd_nm+jibun)로 master_buildings를
    # 역매칭(get_buildings_geo/get_monthly_trend와 동일한 지번 정확 매칭 전략).
    #   - 세 키가 모두 있고 대응 건물이 있으면 그 건물 id, 없으면 NULL.
    #   - 필터/정렬/페이지네이션(where_sql/ORDER BY/LIMIT)에는 영향 없음.
    cur.execute(f"""
        SELECT building_name, address, si_do, sgg_nm, umd_nm, jibun, sgg_cd,
               area, price, deal_date, deal_type, floor,
               lodging_type, lodging_type_detail, match_source,
               (SELECT mb.id FROM master_buildings mb
                 WHERE mb.sgg_cd = transactions.sgg_cd
                   AND mb.umd_nm = transactions.umd_nm
                   AND mb.jibun = transactions.jibun
                 ORDER BY mb.id LIMIT 1) AS master_building_id
        FROM transactions
        WHERE {where_sql}
        ORDER BY deal_date DESC, id DESC
        LIMIT %s OFFSET %s
    """, params + [size, offset])
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()

    return jsonify({"total": total, "page": page, "size": size, "items": rows})


@app.route("/api/buildings-geo")
def get_buildings_geo():
    """지도 마커용 — 좌표(lat/lng)가 있는 마스터 건물.

    선택적 필터(지역/건물명/용도)를 /api/transactions와 동일한 파라미터
    이름으로 지원한다. 단, 기간(year)은 건물 위치와 무관하므로 지도에는
    적용하지 않는다(게시판 전용).

    master_buildings에는 si_do/sgg_nm 컬럼이 없고 sgg_text('서울특별시 서초구')
    와 umd_nm만 있으므로:
      - si_do  : 표기 편차('서울' vs '서울특별시')로 인한 누락을 막기 위해
                 양쪽 모두 행정접미사를 떼어낸 코어 이름으로 정확 비교
      - sgg_nm : sgg_text 포함 매칭
      - umd_nm : 공백 유무 차이('손양면 동호리' vs '손양면동호리')를 흡수하기
                 위해 공백 제거 후 포함 매칭
    """
    q = request.args.get("q", "").strip()
    si_do = request.args.get("si_do", "").strip()
    sgg_nm = request.args.get("sgg_nm", "").strip()
    umd_nm = request.args.get("umd_nm", "").strip()
    lodging_type = request.args.get("lodging_type", "").strip()

    where = ["lat IS NOT NULL", "lng IS NOT NULL"]
    params = []

    if q:
        where.append("(building_name ILIKE %s OR road_address ILIKE %s OR jibun_address ILIKE %s)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if si_do:
        # '서울' vs '서울특별시' 표기 편차 흡수 — 게시판과 동일한 코어 이름 비교 규칙 사용.
        # sgg_text('서울특별시 서초구')의 첫 토큰을 정규화해 비교한다.
        where.append(sido_match_clause("split_part(sgg_text, ' ', 1)"))
        params.append(sido_core(si_do))
    if sgg_nm:
        where.append("sgg_text LIKE %s")
        params.append(f"%{sgg_nm}%")
    if umd_nm:
        where.append("REPLACE(umd_nm, ' ', '') ILIKE %s")
        params.append(f"%{umd_nm.replace(' ', '')}%")
    if lodging_type == "복합":
        where.append("lodging_type LIKE %s")
        params.append("%·%")
    elif lodging_type:
        where.append("lodging_type = %s")
        params.append(lodging_type)

    where_sql = " AND ".join(where)

    conn = get_conn()
    cur = conn.cursor()
    # 각 건물의 '가장 최근 실거래가'를 지번(sgg_cd+umd_nm+jibun) 기준으로 1건만 붙인다.
    # 건물명 매칭은 마스터에 건물명이 "-"처럼 여러 건물이 공유하는 플레이스홀더로
    # 채워진 경우, "-"인 실거래 1건이 "-" 이름의 모든 건물에 잘못 붙는 버그가 있어
    # 지번 튜플 매칭으로 대체한다(sync_batch.py가 transactions에 적재할 때 쓰는
    # (sgg_cd, 정규화 umd_nm, jibun)과 동일한 키).
    #   - 지번 3개 컬럼이 모두 있으면 지번으로 정확 매칭.
    #   - 셋 중 하나라도 NULL이라 지번 매칭이 불가능하면, 예외적으로 건물명으로
    #     한 번 더 시도하되 "-" 같은 플레이스홀더 이름은 제외한다.
    # 같은 지번에 여러 건물(동/호로 구분되는 단지)이 있으면 서로 다른 건물의 거래가
    # 섞일 수 있으므로, 지번이 같은 후보 안에서 건물명(t.building_name = mb.building_name)
    # 까지 정확히 일치하는 거래를 최우선으로 고르고(name_exact DESC), 그런 거래가
    # 없으면 그 지번의 최신 거래를 대체값으로 쓴다(그다음 deal_date DESC).
    #   - latest_price_exact=TRUE  : 건물명까지 정확히 일치한 확정 거래
    #   - latest_price_exact=FALSE : 같은 필지의 대체(참고) 거래
    # N+1 방지를 위해 LEFT JOIN LATERAL로 건물당 최신 1행만 조회하고,
    # 실거래 이력이 없으면 latest_price/latest_deal_date가 NULL로 반환된다.
    cur.execute(f"""
        SELECT mb.id, mb.building_name, mb.road_address, mb.lat, mb.lng, mb.lodging_type,
               lt.price AS latest_price, lt.deal_date AS latest_deal_date,
               lt.floor AS latest_floor, lt.area AS latest_area,
               lt.deal_type AS latest_deal_type,
               COALESCE(lt.name_exact, FALSE) AS latest_price_exact,
               lt.address AS address
        FROM master_buildings mb
        LEFT JOIN LATERAL (
            SELECT t.price, t.deal_date, t.floor, t.area, t.deal_type, t.address,
                   (t.building_name = mb.building_name) AS name_exact
            FROM transactions t
            WHERE (
                    mb.sgg_cd IS NOT NULL AND mb.umd_nm IS NOT NULL AND mb.jibun IS NOT NULL
                    AND t.sgg_cd = mb.sgg_cd
                    AND t.umd_nm = mb.umd_nm
                    AND t.jibun  = mb.jibun
                  )
               OR (
                    (mb.sgg_cd IS NULL OR mb.umd_nm IS NULL OR mb.jibun IS NULL)
                    AND mb.building_name <> '-'
                    AND t.building_name = mb.building_name
                  )
            ORDER BY (t.building_name = mb.building_name) DESC NULLS LAST, t.deal_date DESC
            LIMIT 1
        ) lt ON TRUE
        WHERE {where_sql}
        ORDER BY mb.id
    """, params)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"total": len(rows), "items": rows})


# 실거래 추이 집계의 공통 상수/로직 — /api/monthly-trend(A/B화면)와
# /api/admin/stats(관리자 대시보드)가 동일하게 공유한다(중복 방지).
TREND_FLOOR_YM = "2020-01"  # 백필 목표 하한 — 이 이전 계약월은 집계에서 제외


def _trend_bucket_items(agg, now):
    """월별 집계 dict를 추이 차트용 버킷 목록으로 변환한다.

    agg: {"YYYY-MM": {"cnt": int, "sum_price": int}} (TREND_FLOOR_YM 이전은 이미 제외된 상태)
    반환: (items, granularity)
      - items: [{"ym": "YYYY-MM"|"YYYY-Qn", "count": int, "sum_price": int}, ...] (과거→최근, 빈 버킷 0)
      - granularity: "month" | "quarter" (기간이 24개월 초과면 분기로 자동 전환)
    데이터가 전혀 없으면 최근 12개월(전부 0) month 버킷을 반환한다(하위호환).
    """
    now_ym = f"{now.year:04d}-{now.month:02d}"

    def _month_range(start_ym_, end_ym_):
        """start~end(포함) YYYY-MM 목록 (과거 → 최근)."""
        sy, sm = int(start_ym_[:4]), int(start_ym_[5:7])
        ey, em = int(end_ym_[:4]), int(end_ym_[5:7])
        out = []
        while (sy, sm) <= (ey, em):
            out.append(f"{sy:04d}-{sm:02d}")
            sm += 1
            if sm == 13:
                sm = 1
                sy += 1
        return out

    if agg:
        start_ym = min(agg.keys())
        if start_ym > now_ym:  # 미래 계약일만 있는 극단 케이스 방어
            start_ym = now_ym
    else:
        # 데이터 없음 → 기존과 동일하게 최근 12개월(전부 0)
        y, m = now.year, now.month
        m -= 11
        while m <= 0:
            m += 12
            y -= 1
        start_ym = f"{y:04d}-{m:02d}"

    months = _month_range(start_ym, now_ym)

    if len(months) > 24:
        # 24개월 초과 → 분기별(3개월) 버킷으로 자동 전환. 라벨은 "2025-Q1" 형식.
        def _q(ym):
            return f"{ym[:4]}-Q{(int(ym[5:7]) - 1) // 3 + 1}"
        qagg = {}
        for ym, v in agg.items():
            k = _q(ym)
            cur_v = qagg.setdefault(k, {"cnt": 0, "sum_price": 0})
            cur_v["cnt"] += v["cnt"]
            cur_v["sum_price"] += v["sum_price"]
        quarters = []
        for ym in months:
            k = _q(ym)
            if not quarters or quarters[-1] != k:
                quarters.append(k)
        items = [{
            "ym": q,
            "count": qagg.get(q, {}).get("cnt", 0),
            "sum_price": qagg.get(q, {}).get("sum_price", 0),
        } for q in quarters]
        return items, "quarter"

    items = [{
        "ym": ym,
        "count": agg.get(ym, {}).get("cnt", 0),
        "sum_price": agg.get(ym, {}).get("sum_price", 0),
    } for ym in months]
    return items, "month"


@app.route("/api/monthly-trend")
def get_monthly_trend():
    """
    실거래 추세 집계 (좌측 패널 '실거래추세' 콤보차트용).
    - count     : 버킷별 거래건수 (막대)
    - sum_price : 버킷별 거래금액 합계, 만원 단위 (선)
    시작월은 실제 데이터의 최소 계약월(하한 2020-01)이며, 기간이 24개월을
    초과하면 분기("YYYY-Qn") 버킷으로 집계한다(응답 granularity: month|quarter).
    빈 버킷은 0으로 채우고, 데이터가 전혀 없으면 최근 12개월 0버킷을 반환한다.

    선택적 building_id가 있으면 해당 건물(master_buildings.building_name)의
    실거래만 집계하고, 없으면 기존처럼 전체를 집계한다(하위호환).
    """
    now = datetime.now()
    # 집계 시작월은 '실제 데이터의 최소 계약월'로 하되, 백필 목표 범위(TREND_FLOOR_YM) 이전은 잘라낸다.
    # 버킷 계산(월/분기 전환 포함)은 _trend_bucket_items 공용 헬퍼 사용.
    where = ["deal_date IS NOT NULL", "substring(deal_date, 1, 7) >= %s"]
    params = [TREND_FLOOR_YM]

    # 선택적 building_id → 해당 건물의 실거래만 집계(하위호환: 없거나 정수 아니면 전체 집계).
    # 정확도: A화면 마커(get_buildings_geo)와 동일한 키 전략을 쓴다.
    #   - 지번키(sgg_cd+umd_nm+jibun)가 모두 있으면 지번으로 정확 매칭
    #     (건물명은 유니크 키가 아니라 동명 건물 거래가 섞일 수 있어 지번을 우선).
    #   - 셋 중 하나라도 NULL이면 예외적으로 건물명 매칭('-' 플레이스홀더는 제외).
    # 정수가 아닌 값은 무시하고 전체 집계로 폴백해 500(정수 캐스팅 오류)을 막는다.
    building_id = request.args.get("building_id", "").strip()
    if building_id.isdigit():
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT building_name, sgg_cd, umd_nm, jibun
            FROM master_buildings WHERE id = %s
        """, [int(building_id)])
        b = cur.fetchone()
        cur.close()
        conn.close()
        if b and b["sgg_cd"] and b["umd_nm"] and b["jibun"]:
            where.append("sgg_cd = %s AND umd_nm = %s AND jibun = %s")
            params += [b["sgg_cd"], b["umd_nm"], b["jibun"]]
        else:
            # 지번키 불완전 → 건물명 폴백. 건물 미존재/이름 없음/'-'는 매칭 0으로 처리.
            # (기존 "\x00" 플레이스홀더는 psycopg2가 NUL 거부로 500을 내던 버그라 FALSE 조건으로 교체)
            name = (b["building_name"] if b else None) or ""
            if name and name != "-":
                where.append("building_name = %s")
                params.append(name)
            else:
                where.append("FALSE")

    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT substring(deal_date, 1, 7) AS ym,
               COUNT(*) AS cnt,
               COALESCE(SUM(price), 0) AS sum_price
        FROM transactions
        WHERE {" AND ".join(where)}
        GROUP BY ym
    """, params)
    agg = {r["ym"]: {"cnt": r["cnt"], "sum_price": int(r["sum_price"] or 0)} for r in cur.fetchall()}
    cur.close()
    conn.close()

    items, granularity = _trend_bucket_items(agg, now)
    return jsonify({"items": items, "granularity": granularity})


@app.route("/api/tx-count")
def get_tx_count():
    """전체 실거래 건수 (실시간 COUNT — 관리자 대시보드 '누적 거래' KPI와 동일 기준).
    메인 지도 상단 '실거래 N건' 표시에 사용. 백필/동기화로 늘어나므로 매 로드마다 조회."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) AS c FROM transactions")
    count = int(cur.fetchone()["c"])
    cur.close()
    conn.close()
    return jsonify({"count": count})


@app.route("/api/regions")
def get_regions():
    """시도 > 시군구 > 읍면동 계층 트리 (계층 검색 드롭다운용)"""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT si_do, sgg_nm, umd_nm, COUNT(*) c
        FROM transactions
        WHERE si_do IS NOT NULL
        GROUP BY si_do, sgg_nm, umd_nm
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    tree = {}
    for r in rows:
        sd, sg, um, c = r["si_do"], r["sgg_nm"], r["umd_nm"], r["c"]
        tree.setdefault(sd, {"count": 0, "sgg": {}})
        tree[sd]["count"] += c
        tree[sd]["sgg"].setdefault(sg, {"count": 0, "umd": {}})
        tree[sd]["sgg"][sg]["count"] += c
        tree[sd]["sgg"][sg]["umd"][um] = tree[sd]["sgg"][sg]["umd"].get(um, 0) + c

    return jsonify(tree)


@app.route("/api/years")
def get_years():
    """
    실거래 연도 목록 (기간 필터 드롭다운용).
    - 실제 데이터에 존재하는 연도만 노출 (데이터 없는 연도는 드롭다운에서 제외)
    - 데이터가 하나도 없으면 현재 연도만 표시
    - 새 연도 데이터가 들어오면 자동으로 목록에 추가됨 (하드코딩 아님)
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT DISTINCT LEFT(deal_date, 4) y FROM transactions WHERE deal_date IS NOT NULL")
    data_years = {r["y"] for r in cur.fetchall() if r["y"]}
    cur.close()
    conn.close()

    current_year = datetime.now().year
    years = sorted(data_years, reverse=True) if data_years else [str(current_year)]

    return jsonify({"years": years, "current_year": str(current_year)})


@app.route("/api/favorites")
def get_favorites():
    """
    관심단지 전용 조회 — /api/transactions의 size 상한(200)과 무관하게
    저장된 관심단지 키(building_name|address) 전체를 한 번에 정확히 조회한다.
    쿼리파라미터: keys = "건물명|주소" 쌍을 쉼표(,)로 연결
    """
    raw_keys = request.args.get("keys", "").strip()
    if not raw_keys:
        return jsonify({"items": [], "total": 0})

    pairs = []
    for token in raw_keys.split(","):
        if "|" not in token:
            continue
        name, addr = token.split("|", 1)
        pairs.append((name, addr))

    if not pairs:
        return jsonify({"items": [], "total": 0})

    conn = get_conn()
    cur = conn.cursor()
    # 미매칭 거래는 building_name이 NULL이고, 프론트 favKey는 이를 문자열 "null"로 저장한다.
    # SQL에서 = 'null'은 실제 NULL과 매칭되지 않으므로 그런 항목은 IS NULL로 조회한다.
    conditions = []
    params = []
    for name, addr in pairs:
        if name in ("null", "undefined", ""):
            conditions.append("(building_name IS NULL AND address = %s)")
            params.append(addr)
        else:
            conditions.append("(building_name = %s AND address = %s)")
            params.extend([name, addr])
    conditions = " OR ".join(conditions)
    cur.execute(f"""
        SELECT building_name, address, si_do, sgg_nm, umd_nm, jibun, sgg_cd,
               area, price, deal_date, deal_type, floor,
               lodging_type, lodging_type_detail, match_source,
               (SELECT mb.id FROM master_buildings mb
                 WHERE mb.sgg_cd = transactions.sgg_cd
                   AND mb.umd_nm = transactions.umd_nm
                   AND mb.jibun = transactions.jibun
                 ORDER BY mb.id LIMIT 1) AS master_building_id
        FROM transactions
        WHERE {conditions}
        ORDER BY deal_date DESC, id DESC
    """, params)
    rows = [dict(r) for r in cur.fetchall()]

    # 실거래가 없는 관심단지 fallback — transactions에서 못 찾은 키는 master_buildings를
    # 직접 매칭(도로명주소 일치 또는 "읍면동+지번" 공백제거 비교)해 거래 없이도
    # 이름/지역/상세링크(master_building_id)를 돌려준다. (홈 좌측 관심단지 위젯용)
    covered = {(r["building_name"], r["address"]) for r in rows}
    for name, addr in pairs:
        norm_name = None if name in ("null", "undefined", "") else name
        if (norm_name, addr) in covered:
            continue
        cur.execute("""
            SELECT id, building_name, sgg_text, umd_nm, jibun
            FROM master_buildings
            WHERE road_address = %s
               OR REPLACE(umd_nm || jibun, ' ', '') = REPLACE(%s, ' ', '')
            ORDER BY (building_name = %s) DESC NULLS LAST, id
            LIMIT 1
        """, (addr, addr, norm_name))
        mb = cur.fetchone()
        if not mb:
            continue
        rows.append({
            "building_name": norm_name, "address": addr,
            "si_do": None, "sgg_nm": mb["sgg_text"], "umd_nm": mb["umd_nm"],
            "jibun": mb["jibun"], "sgg_cd": None,
            "area": None, "price": None, "deal_date": None, "deal_type": None,
            "floor": None, "lodging_type": None, "lodging_type_detail": None,
            "match_source": "master_fallback",
            "master_building_id": mb["id"],
        })
    cur.close()
    conn.close()
    return jsonify({"items": rows, "total": len(rows)})


def _fill_master_coords(cur, master_id, road_address):
    """신규 master_buildings 행에 카카오 지오코딩(geocode_buildings.geocode_address)으로
    lat/lng을 채운다. 이미 열려 있는 커서를 받아 UPDATE만 수행하고 커밋은 호출측에 맡긴다.

    주소 미인식/카카오 API 오류/키 미설정 등으로 실패해도 예외를 밖으로 던지지 않아
    건물 등록 자체는 막지 않는다(좌표만 NULL로 남기고 넘어감)."""
    if not road_address:
        return
    try:
        from geocode_buildings import geocode_address
        result = geocode_address(road_address)
        if not result:
            return
        lat, lng = result
        # 보조 UPDATE는 SAVEPOINT로 감싼다. UPDATE 자체가 DB 오류를 내도
        # 메인 INSERT 트랜잭션이 aborted 상태로 오염되지 않아, 호출측 commit이
        # 정상 진행되고 건물 등록은 좌표 NULL로라도 유지된다.
        cur.execute("SAVEPOINT geocode_sp")
        try:
            cur.execute(
                "UPDATE master_buildings SET lat=%s, lng=%s WHERE id=%s",
                (lat, lng, master_id),
            )
            cur.execute("RELEASE SAVEPOINT geocode_sp")
        except Exception:
            cur.execute("ROLLBACK TO SAVEPOINT geocode_sp")
            raise
    except Exception as e:
        app.logger.warning(
            "신규 건물 지오코딩 실패(등록은 계속) id=%s: %s / %s",
            master_id, road_address, e,
        )


@app.route("/api/submit-building", methods=["POST"])
@limiter.limit("3 per minute; 10 per hour")
def submit_building():
    """
    사용자가 "내 건물이 목록에 없다"며 도로명주소를 제출하면:
      1) building_requests에 요청 기록
      2) 도로명→지번 변환(JUSO) → building_registry.classify_lodging_type()로 실시간 재검증
      3) 사용자가 고른 용도(suggested)는 참고용으로만 기록하고, 실제 반영은 검증 결과만 사용.
         검증 통과 시에만 master_buildings(신마스터)에 편입(source='user_submitted').
         판정 불가 시 사유와 함께 거절 (요청 기록에 남김)
    """
    from address_utils import road_to_jibun, BjdongMap, parse_jibun
    from building_registry import classify_lodging_type, resolve_api_building_name
    import os as _os

    data = request.get_json(force=True) or {}
    road_address = (data.get("road_address") or "").strip()
    building_name_hint = (data.get("building_name_hint") or "").strip()
    suggested_lodging_type = (data.get("suggested_lodging_type") or "").strip()  # 참고용, 신뢰 안 함
    requester_note = (data.get("requester_note") or "").strip()

    if not road_address:
        return jsonify({"status": "error", "message": "주소를 입력해주세요."}), 400

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO building_requests (request_type, road_address, building_name_hint, suggested_lodging_type, requester_note)
        VALUES ('new', %s, %s, %s, %s) RETURNING id
    """, (road_address, building_name_hint, suggested_lodging_type, requester_note))
    request_id = cur.fetchone()["id"]
    conn.commit()

    def fail(reason, http_code=200):
        cur.execute("""
            UPDATE building_requests SET status='rejected', reject_reason=%s, processed_at=NOW()
            WHERE id=%s
        """, (reason, request_id))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "rejected", "message": reason}), http_code

    try:
        juso = road_to_jibun(road_address)
    except Exception as e:
        return fail(f"주소 변환 중 오류: {e}")

    if not juso:
        return fail("입력하신 주소로 지번 정보를 찾지 못했습니다. 도로명주소를 다시 확인해주세요.")

    si_do = juso.get("siNm", "")
    sgg_nm = juso.get("sggNm", "")
    # umd_nm은 마스터/실거래 매칭키다. sync_batch가 emdNm+liNm을 공백제거해 저장하므로
    # 여기서도 동일한 표준 함수로 정규화해야 이후 실거래 sync에서 이 건물이 누락되지 않는다.
    umd_nm = normalize_umd_nm(juso.get("emdNm", "") + juso.get("liNm", ""))
    bun = juso.get("lnbrMnnm", "0")
    ji = juso.get("lnbrSlno", "0")
    jibun_str = f"{bun}-{ji}" if ji not in ("0", "", None) else bun

    bjdong_csv = _os.environ.get("BJDONG_CODE_CSV", "법정동코드 전체자료.csv")
    bjdong = BjdongMap(bjdong_csv)
    sgg_cd = bjdong.find_sgg_cd(si_do, sgg_nm)
    if not sgg_cd:
        return fail("법정동코드 매칭에 실패했습니다. 주소 표기를 확인해주세요.")

    bjdong_cd = bjdong.find_bjdong_cd(sgg_cd, umd_nm)
    if not bjdong_cd:
        return fail("읍/면/동 코드 매칭에 실패했습니다.")

    plat_gb, bun2, ji2 = parse_jibun(jibun_str)

    try:
        label, detail, title, reason = classify_lodging_type(sgg_cd, bjdong_cd, plat_gb, bun2, ji2)
    except Exception as e:
        return fail(f"건축물대장 조회 중 오류: {e}")

    if label is None:
        return fail(f"건축물대장으로 확인한 결과 판정이 어렵습니다 ({reason}). "
                     f"집합건축물(생숙/호텔/콘도)이 맞는지 다시 확인해주세요.")

    # 검증 통과 → 사용자가 뭐라고 골랐든 상관없이, 여기서 확정된 label만 반영
    # 건물명도 동일 원칙: API(건축물대장) 명칭이 있으면 무조건 그것으로 확정하고
    # 사용자 입력(building_name_hint)은 building_requests에 참고용으로만 남긴다.
    # API 명칭이 없으면 "읍면동 지번" 임시명으로 등록하고 name_pending=TRUE 표시.
    api_bld_nm = resolve_api_building_name(title)
    if api_bld_nm:
        building_name = api_bld_nm
        name_pending = False
    else:
        building_name = f"{umd_nm} {jibun_str}"
        name_pending = True
    sgg_text = f"{si_do} {sgg_nm}".strip()
    road_addr_final = title["new_plat_plc"] or title["plat_plc"] or road_address

    # 같은 지번의 건물이 이미 신마스터에 있으면 중복 INSERT 대신 검증값으로 갱신한다
    # (같은 주소를 여러 번 요청해도 마스터 키가 중복되지 않도록).
    cur.execute(
        "SELECT id, building_name FROM master_buildings WHERE sgg_cd=%s AND umd_nm=%s AND jibun=%s",
        (sgg_cd, umd_nm, jibun_str),
    )
    existing = cur.fetchone()
    if existing:
        master_id = existing["id"]
        cur.execute("""
            UPDATE master_buildings
            SET lodging_type=%s, lodging_type_detail=%s, verified_at=NOW()
            WHERE id=%s
        """, (label, detail, master_id))
        # 임시명(name_pending) 상태였던 건물이 재제출로 API 명칭이 확인되면 그때 확정한다.
        # (관리자가 손질한 기존 확정 명칭은 덮어쓰지 않도록 name_pending=TRUE인 경우에만.)
        if api_bld_nm:
            cur.execute("""
                UPDATE master_buildings
                SET building_name=%s, name_pending=FALSE
                WHERE id=%s AND name_pending IS TRUE
            """, (api_bld_nm, master_id))
        else:
            # 안전장치: API 명칭이 여전히 없더라도, 기존 이름이 "(이름 미상)"/"-"/빈값처럼
            # 무의미하면 최소한 "읍면동 지번" 임시명으로 바꾸고 name_pending=TRUE 유지.
            existing_name = (existing["building_name"] or "").strip()
            if existing_name in ("", "-", "(이름 미상)"):
                cur.execute("""
                    UPDATE master_buildings
                    SET building_name=%s, name_pending=TRUE
                    WHERE id=%s
                """, (f"{umd_nm} {jibun_str}", master_id))
    else:
        cur.execute("""
            INSERT INTO master_buildings
                (building_name, road_address, sgg_text, sgg_cd, umd_nm, jibun, units, source,
                 lodging_type, lodging_type_detail, verified_at, name_pending)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 'user_submitted', %s, %s, NOW(), %s)
            RETURNING id
        """, (building_name, road_addr_final, sgg_text, sgg_cd, umd_nm, jibun_str, title["ho_cnt"], label, detail, name_pending))
        master_id = cur.fetchone()["id"]
        # 신규 편입 건물의 좌표를 도로명주소로 즉시 채운다(실패해도 등록은 계속).
        _fill_master_coords(cur, master_id, road_addr_final)

    mismatch_note = ""
    if suggested_lodging_type and suggested_lodging_type != label:
        mismatch_note = f" (제출하신 예상 용도 '{suggested_lodging_type}'와 다르게, 건축물대장 확인 결과는 '{label}'입니다.)"

    cur.execute("""
        UPDATE building_requests
        SET status='verified', verified_lodging_type=%s, master_building_id=%s, processed_at=NOW()
        WHERE id=%s
    """, (label, master_id, request_id))
    # 응답 문구/필드는 추정값이 아니라 DB에 실제 저장된 최종 상태를 기준으로 한다.
    # (기존 건물이 이미 확정명인데 이번 API 응답만 비어 있는 경우 등 분기 오류 방지.)
    cur.execute("SELECT building_name, name_pending FROM master_buildings WHERE id=%s", (master_id,))
    final = cur.fetchone()
    building_name = final["building_name"]
    name_pending = bool(final["name_pending"])
    conn.commit()
    cur.close()
    conn.close()

    # 명칭 미확정(name_pending)이면 이름까지 확정된 것처럼 오해하지 않도록 문구를 분리한다.
    if name_pending:
        result_message = (
            f"용도는 '{label}'(으)로 확인되어 등록되었습니다.{mismatch_note} "
            f"건물명은 아직 공식 확인 전이라 임시로 '{building_name}'으로 표시됩니다. "
            f"정확한 명칭을 아신다면 건물상세 페이지에서 '건물명 제안하기'로 알려주세요. "
            f"다음 실거래 갱신부터 이 건물의 거래가 표시됩니다."
        )
    else:
        result_message = (
            f"'{building_name}'이(가) '{label}'(으)로 확인되어 등록되었습니다.{mismatch_note} "
            f"다음 실거래 갱신부터 이 건물의 거래가 표시됩니다."
        )
    return jsonify({
        "status": "verified",
        "message": result_message,
        "name_pending": name_pending,
        "building_name": building_name,
        "lodging_type": label,
        "units": title["ho_cnt"],
    })


@app.route("/apply/agent")
def apply_agent_page():
    """중개사 회원신청(C화면) 정적 폼 HTML 서빙.

    카카오맵이 필요 없는 단순 정적 폼이므로 키 주입 없이 그대로 서빙한다.
    """
    html_path = os.path.join(app.static_folder, "apply_agent.html")
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    html = _inject_asset_version(html)
    resp = Response(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


def _handle_apply_upload(applicant_type, allowed_doc_types):
    """C/D 신청서 서류 업로드 공통 처리.

    - multipart/form-data: file(파일) + doc_type(문서 종류)
    - 확장자 pdf/jpg/jpeg/png, 파일당 5MB, 매직 바이트로 실제 내용 검증
    - 저장 후 내부 참조 키(applications/{type}/{uuid}/{doc_type}.{ext})를 반환.
      URL이 아니므로 서명 없이는 외부에서 접근할 수 없다.
    """
    f = request.files.get("file")
    doc_type = (request.form.get("doc_type") or "").strip()
    if not f or not f.filename:
        return jsonify({"ok": False, "message": "파일을 선택해주세요."}), 400
    if doc_type not in allowed_doc_types:
        return jsonify({"ok": False, "message": "허용되지 않은 문서 종류입니다."}), 400

    filename = f.filename
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in storage_util.ALLOWED_EXTENSIONS:
        return jsonify({"ok": False, "message": "PDF, JPG, PNG 파일만 업로드할 수 있습니다."}), 400

    data = f.read(storage_util.MAX_FILE_BYTES + 1)
    if len(data) > storage_util.MAX_FILE_BYTES:
        return jsonify({"ok": False, "message": "파일 크기는 5MB 이하여야 합니다."}), 400
    if len(data) < 16:
        return jsonify({"ok": False, "message": "파일이 비어 있거나 손상되었습니다."}), 400
    if not storage_util.check_magic_bytes(data, ext):
        return jsonify({"ok": False, "message": "파일 내용이 확장자와 일치하지 않습니다. 실제 PDF/JPG/PNG 파일만 업로드해주세요."}), 400

    key = storage_util.build_doc_key(applicant_type, doc_type, ext)
    try:
        storage_util.upload_doc(key, data)
    except Exception:
        app.logger.exception("서류 업로드 실패 (%s/%s)", applicant_type, doc_type)
        return jsonify({"ok": False, "message": "파일 저장 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요."}), 500

    return jsonify({"ok": True, "doc_ref": key})


@app.route("/api/apply/agent/upload", methods=["POST"])
@limiter.limit("10 per hour")
def apply_agent_upload():
    """중개사 신청 서류 업로드 (비로그인, IP 기준 rate limit)."""
    # 여권용 사진은 화면에 <img>로 노출되므로 이미지 파일만 허용 (PDF 차단 — 운영업체 로고와 동일)
    if (request.form.get("doc_type") or "").strip() == "photo":
        f = request.files.get("file")
        ext = f.filename.rsplit(".", 1)[-1].lower() if f and f.filename and "." in f.filename else ""
        if ext not in storage_util.LOGO_EXTENSIONS:
            return jsonify({"ok": False, "message": "사진은 JPG 또는 PNG 이미지 파일만 업로드할 수 있습니다."}), 400
    return _handle_apply_upload("agent", storage_util.AGENT_DOC_TYPES)


@app.route("/api/apply/operator/upload", methods=["POST"])
@limiter.limit("10 per hour")
def apply_operator_upload():
    """운영업체 신청 서류 업로드 (비로그인, IP 기준 rate limit)."""
    # 로고는 화면에 <img>로 노출되므로 이미지 파일만 허용 (PDF 차단)
    if (request.form.get("doc_type") or "").strip() == "logo":
        f = request.files.get("file")
        ext = f.filename.rsplit(".", 1)[-1].lower() if f and f.filename and "." in f.filename else ""
        if ext not in storage_util.LOGO_EXTENSIONS:
            return jsonify({"ok": False, "message": "로고는 JPG 또는 PNG 이미지 파일만 업로드할 수 있습니다."}), 400
    return _handle_apply_upload("operator", storage_util.OPERATOR_DOC_TYPES)


def _clean_doc_ref(value, applicant_type, doc_type):
    """신청서 제출 시 넘어온 서류 참조 키 검증.

    - 빈 값이면 None (서류는 선택 항목)
    - 우리가 발급한 키 형식이 아니거나, 실제 스토리지에 없으면 에러 문자열 반환
    """
    ref = (value or "").strip()
    if not ref:
        return None, None
    if not storage_util.is_valid_doc_ref(ref, applicant_type, {doc_type}):
        return None, "서류 참조가 올바르지 않습니다. 파일을 다시 업로드해주세요."
    if not storage_util.doc_exists(ref):
        return None, "업로드된 서류를 찾을 수 없습니다. 파일을 다시 업로드해주세요."
    return ref, None


# ---- 번호 정규화/표시 헬퍼 (전화번호·사업자등록번호·중개사무소등록번호 공통 규칙) ----
# 저장: 숫자만 남긴다. 표시: format_phone()/format_biz_reg_number()로 하이픈 포함 재포맷.

def _digits_only(s):
    """숫자 이외 문자(하이픈·공백 등)를 전부 제거한다."""
    return re.sub(r"\D", "", s or "")


def format_phone(p):
    """숫자만 저장된 전화번호를 하이픈 포함 표시용으로 포맷한다 (실패 시 원문 반환)."""
    d = _digits_only(p)
    if not d:
        return p or ""
    if d.startswith("02"):
        if len(d) == 9:
            return f"{d[:2]}-{d[2:5]}-{d[5:]}"
        if len(d) == 10:
            return f"{d[:2]}-{d[2:6]}-{d[6:]}"
    if len(d) == 10:
        return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    if len(d) == 11:
        return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    return p or ""


def format_biz_reg_number(b):
    """숫자만 저장된 사업자등록번호(10자리)를 000-00-00000 형태로 포맷한다."""
    d = _digits_only(b)
    if len(d) == 10:
        return f"{d[:3]}-{d[3:5]}-{d[5:]}"
    return b or ""


def _validate_phone_digits(d):
    """숫자만 남긴 전화번호가 10~11자리인지 검사."""
    return 10 <= len(d) <= 11


def _validate_biz_reg_digits(d):
    """숫자만 남긴 사업자등록번호가 10자리인지 검사."""
    return len(d) == 10


_APPLICANT_TYPE_KR = {"agent": "중개사", "operator": "운영지원업체", "loan_consultant": "대출상담사"}


def _send_application_received_email(applicant_type, company_name, to_email):
    """신청 접수 직후 안내 이메일 발송 — 실패해도 접수 처리에는 영향 없음(예외 삼킴)."""
    try:
        type_kr = _APPLICANT_TYPE_KR.get(applicant_type, applicant_type)
        received_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        html = f"""
        <div style="font-family:'Apple SD Gothic Neo','Malgun Gothic',sans-serif; max-width:520px; margin:0 auto; color:#16202E;">
          <h2 style="font-size:18px; border-bottom:2px solid #B4863F; padding-bottom:8px;">홈앤스테이 (HOME &amp; STAY)</h2>
          <p style="font-size:15px; font-weight:700;">신청이 접수되었습니다.</p>
          <table style="font-size:14px; border-collapse:collapse; margin:12px 0;">
            <tr><td style="padding:4px 12px 4px 0; color:#6b7280;">신청유형</td><td>{type_kr}</td></tr>
            <tr><td style="padding:4px 12px 4px 0; color:#6b7280;">업체명</td><td>{company_name}</td></tr>
            <tr><td style="padding:4px 12px 4px 0; color:#6b7280;">접수일시</td><td>{received_at}</td></tr>
          </table>
          <p style="font-size:13px; color:#6b7280;">검토 후 담당자가 연락드리겠습니다. 승인 결과는 문자와 이메일로 안내됩니다.</p>
        </div>
        """
        ok, msg = send_email(to_email, "[홈앤스테이] 신청이 접수되었습니다", html)
        if not ok:
            app.logger.warning("신청 접수 이메일 발송 실패 (%s, %s): %s", applicant_type, to_email, msg)
    except Exception:
        app.logger.exception("신청 접수 이메일 발송 중 오류 (%s, %s)", applicant_type, to_email)


def _send_approval_email(applicant_type, to_email, login_id, temp_pw, login_url):
    """승인 완료 이메일(SMS와 동일 내용) — 실패해도 승인 처리에는 영향 없음. (ok, msg) 반환."""
    try:
        type_kr = _APPLICANT_TYPE_KR.get(applicant_type, applicant_type)
        html = f"""
        <div style="font-family:'Apple SD Gothic Neo','Malgun Gothic',sans-serif; max-width:520px; margin:0 auto; color:#16202E;">
          <h2 style="font-size:18px; border-bottom:2px solid #B4863F; padding-bottom:8px;">홈앤스테이 (HOME &amp; STAY)</h2>
          <p style="font-size:15px; font-weight:700;">{type_kr} 승인이 완료되었습니다.</p>
          <table style="font-size:14px; border-collapse:collapse; margin:12px 0;">
            <tr><td style="padding:4px 12px 4px 0; color:#6b7280;">로그인 ID(이메일)</td><td>{login_id}</td></tr>
            <tr><td style="padding:4px 12px 4px 0; color:#6b7280;">임시비밀번호</td><td><b>{temp_pw}</b></td></tr>
            <tr><td style="padding:4px 12px 4px 0; color:#6b7280;">로그인</td><td><a href="{login_url}">{login_url}</a></td></tr>
          </table>
          <p style="font-size:13px; color:#B00020; font-weight:600;">최초 로그인 후 반드시 비밀번호를 변경해주세요.</p>
        </div>
        """
        ok, msg = send_email(to_email, "[홈앤스테이] 승인되었습니다 — 로그인 안내", html)
        if not ok:
            app.logger.warning("승인 이메일 발송 실패 (%s, %s): %s", applicant_type, to_email, msg)
        return ok, msg
    except Exception as e:
        app.logger.exception("승인 이메일 발송 중 오류 (%s, %s)", applicant_type, to_email)
        return False, str(e)


@app.route("/api/apply/agent", methods=["POST"])
@limiter.limit("3 per minute; 10 per hour")
def apply_agent():
    """중개사 회원신청 접수 API.

    텍스트 항목만 받아 applications 테이블에 applicant_type='agent',
    status='submitted'로 INSERT한다. 서류(자격증/등록증 등)는 이번엔 미사용이라
    doc_* 및 intro_text는 NULL로 둔다.
    """
    data = request.get_json(force=True) or {}

    office_or_company_name = (data.get("office_or_company_name") or "").strip()
    owner_name = (data.get("owner_name") or "").strip()
    # 번호류는 하이픈·공백 입력 여부와 무관하게 숫자만 남겨 저장 (표시할 때 재포맷)
    reg_number = _digits_only(data.get("reg_number"))
    biz_reg_number = _digits_only(data.get("biz_reg_number"))
    phone = _digits_only(data.get("phone"))
    email = (data.get("email") or "").strip()
    # 희망지역 → 희망건물로 변경. 구버전 호환을 위해 preferred_region도 함께 받아둔다.
    preferred_building = (data.get("preferred_building") or "").strip()
    preferred_region = (data.get("preferred_region") or "").strip()
    # 건물 상세(B화면)에서 진입 시 함께 오는 희망건물 id — 숫자가 아니면 무시
    raw_bid = str(data.get("preferred_building_id") or "").strip()
    preferred_building_id = int(raw_bid) if raw_bid.isdigit() else None

    # 필수값 검증
    missing = []
    if not office_or_company_name:
        missing.append("중개사무소명")
    if not owner_name:
        missing.append("대표자")
    if not reg_number:
        missing.append("등록번호")
    if not biz_reg_number:
        missing.append("사업자등록번호")
    if not phone:
        missing.append("연락처")
    if not email:
        missing.append("이메일")
    if missing:
        return jsonify({"ok": False, "message": "필수 항목을 입력해주세요: " + ", ".join(missing)}), 400

    # 번호 형식 검증 (숫자만 남긴 기준)
    if not _validate_biz_reg_digits(biz_reg_number):
        return jsonify({"ok": False, "message": "사업자등록번호 형식이 올바르지 않습니다. (숫자 10자리)"}), 400
    if not _validate_phone_digits(phone):
        return jsonify({"ok": False, "message": "전화번호 형식이 올바르지 않습니다. (숫자 10~11자리)"}), 400

    # 법적 동의 서버측 재검증 — 클라이언트 우회 방지 (둘 다 명시적 true여야 함)
    if data.get("terms") is not True or data.get("privacy") is not True:
        return jsonify({"ok": False, "message": "필수 약관(이용약관, 개인정보 수집·이용)에 모두 동의해주세요."}), 400

    # 간단한 이메일 형식 체크 (@ 앞뒤로 내용, . 포함)
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return jsonify({"ok": False, "message": "이메일 형식이 올바르지 않습니다."}), 400

    # 서류 참조 키 검증 — 사업자등록증·중개사무소등록증은 필수, 나머지는 선택
    doc_refs = {}
    for field, doc_type in (("doc_license_url", "license"),
                            ("doc_office_reg_url", "office_reg"),
                            ("doc_biz_reg_url", "biz_reg"),
                            ("doc_photo_url", "photo")):
        ref, err = _clean_doc_ref(data.get(field), "agent", doc_type)
        if err:
            return jsonify({"ok": False, "message": err}), 400
        doc_refs[field] = ref
    if not doc_refs["doc_biz_reg_url"]:
        return jsonify({"ok": False, "message": "사업자등록증 사본 첨부는 필수입니다."}), 400
    if not doc_refs["doc_office_reg_url"]:
        return jsonify({"ok": False, "message": "중개사무소등록증 사본 첨부는 필수입니다."}), 400

    conn = get_conn()
    cur = conn.cursor()

    # 희망건물 id는 실제 건물마스터에 있는 것만 저장 (없으면 조용히 NULL — 이름은 별도 보존)
    if preferred_building_id is not None:
        cur.execute("SELECT 1 FROM master_buildings WHERE id=%s", [preferred_building_id])
        if not cur.fetchone():
            preferred_building_id = None

    cur.execute("""
        INSERT INTO applications
            (applicant_type, office_or_company_name, owner_name, reg_number,
             biz_reg_number, phone, email, preferred_region, preferred_building, status,
             intro_text, doc_license_url, doc_office_reg_url, doc_biz_reg_url,
             doc_photo_url, preferred_building_id,
             terms_agreed_at, privacy_agreed_at)
        VALUES ('agent', %s, %s, %s, %s, %s, %s, %s, %s, 'submitted',
                NULL, %s, %s, %s, %s, %s, NOW(), NOW())
        RETURNING id
    """, (office_or_company_name, owner_name, reg_number,
          biz_reg_number or None, phone, email, preferred_region or None,
          preferred_building or None,
          doc_refs["doc_license_url"], doc_refs["doc_office_reg_url"],
          doc_refs["doc_biz_reg_url"], doc_refs["doc_photo_url"],
          preferred_building_id))
    new_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()

    # 접수 안내 이메일 — 실패해도 접수는 이미 확정
    _send_application_received_email("agent", office_or_company_name, email)

    return jsonify({"ok": True, "id": new_id})


@app.route("/apply/operator")
def apply_operator_page():
    """운영업체 등록신청(D화면) 정적 폼 HTML 서빙.

    apply_agent_page()과 동일하게, 카카오맵이 필요 없는 단순 정적 폼이므로
    키 주입 없이 그대로 서빙한다.
    """
    html_path = os.path.join(app.static_folder, "apply_operator.html")
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    html = _inject_asset_version(html)
    resp = Response(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


# 운영지원업체 업종: 이 5개만 허용한다(operators.category와 동일 기준).
# '대출상담사'는 별도 엔티티(loan_consultants)로 분리되어 신규 신청은 /apply/loan 으로 받는다.
# (기존 operators 테이블의 대출상담사 행은 데이터 보존 차원에서 그대로 둔다)
OPERATOR_CATEGORIES = {"위탁운영", "청소", "세탁", "용품", "인테리어"}


@app.route("/api/apply/operator", methods=["POST"])
@limiter.limit("3 per minute; 10 per hour")
def apply_operator():
    """운영업체 등록신청 접수 API.

    apply/agent와 동일한 구조로 텍스트 항목만 받아 applications 테이블에
    applicant_type='operator', status='submitted'로 INSERT한다. 서류(명함/영업
    허가증 등)는 이번엔 미사용이라 doc_* 및 reg_number/intro_text는 NULL로 둔다.
    """
    data = request.get_json(force=True) or {}

    office_or_company_name = (data.get("office_or_company_name") or "").strip()
    owner_name = (data.get("owner_name") or "").strip()
    category = (data.get("category") or "").strip()
    biz_reg_number = _digits_only(data.get("biz_reg_number"))
    phone = _digits_only(data.get("phone"))
    email = (data.get("email") or "").strip()
    website_url = (data.get("website_url") or "").strip()
    preferred_region = (data.get("preferred_region") or "").strip()

    # 필수값 검증
    missing = []
    if not office_or_company_name:
        missing.append("업체명")
    if not owner_name:
        missing.append("대표자")
    if not category:
        missing.append("업종")
    if not biz_reg_number:
        missing.append("사업자등록번호")
    if not phone:
        missing.append("연락처")
    if not email:
        missing.append("이메일")
    if missing:
        return jsonify({"ok": False, "message": "필수 항목을 입력해주세요: " + ", ".join(missing)}), 400

    # 번호 형식 검증 (숫자만 남긴 기준)
    if not _validate_biz_reg_digits(biz_reg_number):
        return jsonify({"ok": False, "message": "사업자등록번호 형식이 올바르지 않습니다. (숫자 10자리)"}), 400
    if not _validate_phone_digits(phone):
        return jsonify({"ok": False, "message": "전화번호 형식이 올바르지 않습니다. (숫자 10~11자리)"}), 400

    # 법적 동의 서버측 재검증 — 클라이언트 우회 방지 (둘 다 명시적 true여야 함)
    if data.get("terms") is not True or data.get("privacy") is not True:
        return jsonify({"ok": False, "message": "필수 약관(이용약관, 개인정보 수집·이용)에 모두 동의해주세요."}), 400

    # 업종은 허용된 6개 중 하나만
    if category not in OPERATOR_CATEGORIES:
        return jsonify({"ok": False, "message": "업종은 다음 중 하나여야 합니다: " + ", ".join(sorted(OPERATOR_CATEGORIES))}), 400

    # 간단한 이메일 형식 체크 (apply/agent와 동일 정규식)
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return jsonify({"ok": False, "message": "이메일 형식이 올바르지 않습니다."}), 400

    # 서류 참조 키 검증 — 사업자등록증은 필수, 나머지(명함/영업허가증/로고)는 선택
    doc_refs = {}
    for field, doc_type in (("doc_biz_reg_url", "biz_reg"),
                            ("doc_business_card_url", "business_card"),
                            ("doc_biz_license_url", "biz_license"),
                            ("doc_logo_url", "logo")):
        ref, err = _clean_doc_ref(data.get(field), "operator", doc_type)
        if err:
            return jsonify({"ok": False, "message": err}), 400
        doc_refs[field] = ref
    if not doc_refs["doc_biz_reg_url"]:
        return jsonify({"ok": False, "message": "사업자등록증 사본 첨부는 필수입니다."}), 400

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO applications
            (applicant_type, office_or_company_name, owner_name, category,
             biz_reg_number, phone, email, website_url, preferred_region, status,
             reg_number, intro_text, doc_biz_reg_url, doc_business_card_url,
             doc_biz_license_url, doc_logo_url, terms_agreed_at, privacy_agreed_at)
        VALUES ('operator', %s, %s, %s, %s, %s, %s, %s, %s, 'submitted',
                NULL, NULL, %s, %s, %s, %s, NOW(), NOW())
        RETURNING id
    """, (office_or_company_name, owner_name, category,
          biz_reg_number or None, phone, email,
          website_url or None, preferred_region or None,
          doc_refs["doc_biz_reg_url"], doc_refs["doc_business_card_url"],
          doc_refs["doc_biz_license_url"], doc_refs["doc_logo_url"]))
    new_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()

    # 접수 안내 이메일 — 실패해도 접수는 이미 확정
    _send_application_received_email("operator", office_or_company_name, email)

    return jsonify({"ok": True, "id": new_id})


@app.route("/apply/loan")
def apply_loan_page():
    """대출상담사 등록신청 정적 폼 HTML 서빙 (apply_agent_page()와 동일 패턴)."""
    html_path = os.path.join(app.static_folder, "apply_loan_consultant.html")
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    html = _inject_asset_version(html)
    resp = Response(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


@app.route("/api/apply/loan", methods=["POST"])
@limiter.limit("3 per minute; 10 per hour")
def apply_loan():
    """대출상담사 등록신청 접수 API.

    apply/agent와 동일 구조로 applications 테이블에 applicant_type='loan_consultant',
    status='submitted'로 INSERT한다. 대출모집인 등록번호는 applications.reg_number
    컬럼을 재사용해 저장한다 (관리자가 loanconsultant.or.kr에서 조회 후 승인).
    """
    data = request.get_json(force=True) or {}

    office_or_company_name = (data.get("office_or_company_name") or "").strip()
    owner_name = (data.get("owner_name") or "").strip()
    license_number = (data.get("license_number") or "").strip()
    biz_reg_number = _digits_only(data.get("biz_reg_number"))
    phone = _digits_only(data.get("phone"))
    email = (data.get("email") or "").strip()

    missing = []
    if not office_or_company_name:
        missing.append("소속 회사명")
    if not owner_name:
        missing.append("성명")
    if not license_number:
        missing.append("대출모집인 등록번호")
    if not phone:
        missing.append("연락처")
    if not email:
        missing.append("이메일")
    if missing:
        return jsonify({"ok": False, "message": "필수 항목을 입력해주세요: " + ", ".join(missing)}), 400

    # 번호 형식 검증 — 전화번호는 필수, 사업자등록번호는 선택(입력 시에만 검사)
    if not _validate_phone_digits(phone):
        return jsonify({"ok": False, "message": "전화번호 형식이 올바르지 않습니다. (숫자 10~11자리)"}), 400
    if biz_reg_number and not _validate_biz_reg_digits(biz_reg_number):
        return jsonify({"ok": False, "message": "사업자등록번호 형식이 올바르지 않습니다. (숫자 10자리)"}), 400

    # 법적 동의 서버측 재검증 — 클라이언트 우회 방지 (둘 다 명시적 true여야 함)
    if data.get("terms") is not True or data.get("privacy") is not True:
        return jsonify({"ok": False, "message": "필수 약관(이용약관, 개인정보 수집·이용)에 모두 동의해주세요."}), 400

    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return jsonify({"ok": False, "message": "이메일 형식이 올바르지 않습니다."}), 400

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO applications
            (applicant_type, office_or_company_name, owner_name, reg_number,
             biz_reg_number, phone, email, status, terms_agreed_at, privacy_agreed_at)
        VALUES ('loan_consultant', %s, %s, %s, %s, %s, %s, 'submitted', NOW(), NOW())
        RETURNING id
    """, (office_or_company_name, owner_name, license_number,
          biz_reg_number or None, phone, email))
    new_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()

    # 접수 안내 이메일 — 실패해도 접수는 이미 확정
    _send_application_received_email("loan_consultant", office_or_company_name, email)

    return jsonify({"ok": True, "id": new_id})


@app.route("/api/loan-consultants")
def loan_consultants_list():
    """승인된 대출상담사 공개 목록 — B화면 '금융' 카드에서 사용.

    연락처(phone)는 상담 연결 목적상 노출한다 (중개사 카드와 동일 정책).
    license_number는 '금융감독원 등록 대출모집인' 뱃지 표기용으로 포함.
    정렬: priority_score DESC(유료 우선노출 자리, 현재 전부 0), 나머지는 RANDOM() — 3명만 반환.
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT office_name, owner_name, phone, subdomain_slug,
               license_number, consultant_products, kakao_chat_url, intro_text
        FROM loan_consultants
        WHERE status = 'approved'
          AND COALESCE(is_visible, TRUE)
        ORDER BY COALESCE(priority_score, 0) DESC, RANDOM()
        LIMIT 3
    """)
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"ok": True, "items": items})


@app.route("/api/loan-consultants/all")
def loan_consultants_list_all():
    """승인된 대출상담사 전체 공개 목록 — /loan-consultants 전체 목록 페이지에서 사용."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT office_name, owner_name, phone, subdomain_slug,
               license_number, consultant_products, kakao_chat_url, intro_text
        FROM loan_consultants
        WHERE status = 'approved'
          AND COALESCE(is_visible, TRUE)
        ORDER BY COALESCE(priority_score, 0) DESC, approved_at DESC NULLS LAST, id DESC
    """)
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"ok": True, "items": items})


@app.route("/loan-consultants")
def loan_consultants_list_page():
    """전체 대출상담사 목록 페이지 (공개)."""
    return _serve_static_html("loan_consultants_list.html")


# ---- 파트너 소개 (공개, /partner 페이지) ----

@app.route("/partners-directory")
def partners_directory_page():
    """파트너 소개 게시판 (공개) — 로고 없이 텍스트/표 기반 (상표권 이슈 회피)."""
    return _serve_static_html("partners_directory.html")


@app.route("/api/partners/directory")
def partners_directory_api():
    """승인 + 노출중(approved & is_visible) 파트너 전체 목록 — /partners-directory 게시판용.

    타입별 정렬(요구사항):
      - 중개사/운영지원업체: priority_score DESC → 담당건물수 DESC → 승인일 DESC
      - 대출상담사:          priority_score DESC → 승인일 DESC
    서비스지역:
      - 중개사/운영지원업체: 담당 건물들의 시도(sgg_text 첫 토큰) 목록, 없으면 신청서 희망지역
      - 대출상담사: 신청서 희망지역, 없으면 '전국'
    """
    def _safe_url(u):
        # javascript:/data: 등 위험 스킴 차단 — http(s)만 링크로 노출
        u = (u or "").strip()
        return u if u.lower().startswith(("http://", "https://")) else None

    conn = get_conn()
    cur = conn.cursor()
    items = []
    try:
        cur.execute("""
            SELECT a.office_name AS name, a.phone, a.subdomain_slug AS slug,
                   COALESCE(a.priority_score, 0) AS priority_score, a.approved_at,
                   (SELECT COUNT(*) FROM agent_buildings ab WHERE ab.agent_id = a.id) AS building_count,
                   (SELECT string_agg(DISTINCT split_part(mb.sgg_text, ' ', 1), ', ')
                      FROM agent_buildings ab JOIN master_buildings mb ON mb.id = ab.master_building_id
                     WHERE ab.agent_id = a.id AND mb.sgg_text IS NOT NULL) AS region_bld,
                   (SELECT ap.preferred_region FROM applications ap
                     WHERE ap.linked_agent_id = a.id ORDER BY ap.id DESC LIMIT 1) AS region_app
            FROM agents a
            WHERE a.status = 'approved' AND COALESCE(a.is_visible, TRUE)
            ORDER BY COALESCE(a.priority_score, 0) DESC, building_count DESC,
                     a.approved_at DESC NULLS LAST, a.id DESC
        """)
        for r in cur.fetchall():
            items.append({
                "type": "agent",
                "name": r["name"],
                "region": r["region_bld"] or r["region_app"] or "-",
                "phone": format_phone(r["phone"]) if r["phone"] else None,
                "website": None,
                "tags": [],
                "building_count": r["building_count"],
                "license_number": None,
                "link": f"/agent/{r['slug']}" if r["slug"] else None,
            })

        cur.execute("""
            SELECT o.company_name AS name, o.phone, o.subdomain_slug AS slug, o.category,
                   o.website_url,
                   COALESCE(o.priority_score, 0) AS priority_score, o.approved_at,
                   (SELECT COUNT(*) FROM operator_buildings ob WHERE ob.operator_id = o.id) AS building_count,
                   (SELECT string_agg(DISTINCT split_part(mb.sgg_text, ' ', 1), ', ')
                      FROM operator_buildings ob JOIN master_buildings mb ON mb.id = ob.master_building_id
                     WHERE ob.operator_id = o.id AND mb.sgg_text IS NOT NULL) AS region_bld,
                   (SELECT ap.preferred_region FROM applications ap
                     WHERE ap.linked_operator_id = o.id ORDER BY ap.id DESC LIMIT 1) AS region_app
            FROM operators o
            WHERE o.status = 'approved' AND COALESCE(o.is_visible, TRUE)
            ORDER BY COALESCE(o.priority_score, 0) DESC, building_count DESC,
                     o.approved_at DESC NULLS LAST, o.id DESC
        """)
        for r in cur.fetchall():
            items.append({
                "type": "operator",
                "name": r["name"],
                "region": r["region_bld"] or r["region_app"] or "-",
                "phone": format_phone(r["phone"]) if r["phone"] else None,
                "website": _safe_url(r["website_url"]),
                "tags": [t for t in [(r["category"] or "").strip()] if t],
                "building_count": r["building_count"],
                "license_number": None,
                "link": f"/operator/{r['slug']}" if r["slug"] else None,
            })

        cur.execute("""
            SELECT lc.office_name AS name, lc.owner_name, lc.phone, lc.subdomain_slug AS slug,
                   lc.license_number, lc.consultant_products,
                   COALESCE(lc.priority_score, 0) AS priority_score, lc.approved_at,
                   (SELECT ap.preferred_region FROM applications ap
                     WHERE ap.linked_loan_consultant_id = lc.id ORDER BY ap.id DESC LIMIT 1) AS region_app
            FROM loan_consultants lc
            WHERE lc.status = 'approved' AND COALESCE(lc.is_visible, TRUE)
            ORDER BY COALESCE(lc.priority_score, 0) DESC, lc.approved_at DESC NULLS LAST, lc.id DESC
        """)
        for r in cur.fetchall():
            products = [p.strip() for p in (r["consultant_products"] or "").split(",") if p.strip()]
            items.append({
                "type": "loan",
                "name": r["name"],
                "region": r["region_app"] or "전국",
                "phone": format_phone(r["phone"]) if r["phone"] else None,
                "website": None,
                "tags": products,
                "building_count": None,  # 대출상담사는 담당건물 대신 등록번호 뱃지
                "license_number": r["license_number"],
                # 대출상담사는 개별 공개 프로필 페이지가 없어 전체 목록 페이지로 연결
                "link": "/loan-consultants",
            })
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "items": items})


@app.route("/api/partners/operators")
def partners_operators_list():
    """로고가 등록된 승인 운영지원업체 공개 목록 — /partner '등록된 파트너' 섹션용.

    로고 없는 승인 업체는 시각 노출 대상이 아니므로 제외한다.
    로고 이미지는 스토리지 키를 직접 노출하지 않고 프록시(/api/partners/operator-logo/<id>)로 서빙.
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, company_name, subdomain_slug
        FROM operators
        WHERE status = 'approved' AND logo_url IS NOT NULL AND logo_url <> ''
        ORDER BY approved_at DESC NULLS LAST, id DESC
        LIMIT 24
    """)
    items = [{
        "company_name": r["company_name"],
        "subdomain_slug": r["subdomain_slug"],
        "logo_src": f"/api/partners/operator-logo/{r['id']}",
    } for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"ok": True, "items": items})


@app.route("/api/partners/operator-logo/<int:operator_id>")
def partners_operator_logo(operator_id):
    """승인 업체 로고 이미지 공개 프록시 — 승인 + 로고 보유 업체만 서빙 (팝업 이미지 프록시와 동일 패턴)."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT logo_url FROM operators WHERE id=%s AND status='approved'", [operator_id])
    row = cur.fetchone()
    cur.close()
    conn.close()
    key = (row or {}).get("logo_url")
    if not key or not storage_util.is_valid_doc_ref(key, "operator", {"logo"}):
        abort(404)
    try:
        data = storage_util.download_bytes(key)
    except Exception:
        abort(404)
    ext = key.rsplit(".", 1)[-1].lower()
    mime = "image/png" if ext == "png" else "image/jpeg"
    resp = Response(data, mimetype=mime)
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp


@app.route("/api/partners/agent-photo/<int:agent_id>")
def partners_agent_photo(agent_id):
    """승인 중개사 프로필 사진 공개 프록시 — 운영업체 로고 프록시와 동일 패턴."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT photo_url FROM agents WHERE id=%s AND status='approved'", [agent_id])
    row = cur.fetchone()
    cur.close()
    conn.close()
    key = (row or {}).get("photo_url")
    if not key or not storage_util.is_valid_doc_ref(key, "agent", {"photo"}):
        abort(404)
    try:
        data = storage_util.download_bytes(key)
    except Exception:
        abort(404)
    ext = key.rsplit(".", 1)[-1].lower()
    mime = "image/png" if ext == "png" else "image/jpeg"
    resp = Response(data, mimetype=mime)
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp


@app.route("/api/request-correction", methods=["POST"])
@limiter.limit("3 per minute; 10 per hour")
def request_correction():
    """
    이미 목록에 있는 건물의 용도 라벨이 잘못됐다고 생각될 때 정정을 요청하는 API.

    핵심 원칙: 사용자가 "이거 아니고 저거예요"라고 제안해도, 그 제안을 그대로 반영하지
    않는다. 반드시 building_registry.classify_lodging_type()으로 그 자리에서 다시
    조회해서, 실제로 확인된 결과만 반영한다. 사용자 제안과 재검증 결과가 같으면
    "확인되어 반영됨", 다르면 "확인해봤지만 제안하신 내용과는 다릅니다"로 응답한다.
    """
    from address_utils import BjdongMap, parse_jibun
    from building_registry import classify_lodging_type, resolve_api_building_name
    import os as _os

    data = request.get_json(force=True) or {}
    sgg_cd = (data.get("sgg_cd") or "").strip()
    umd_nm = (data.get("umd_nm") or "").strip()
    jibun = (data.get("jibun") or "").strip()
    suggested_lodging_type = (data.get("suggested_lodging_type") or "").strip()
    suggested_building_name = (data.get("suggested_building_name") or "").strip()
    requester_note = (data.get("requester_note") or "").strip()

    if not (sgg_cd and umd_nm and jibun):
        return jsonify({"status": "error", "message": "대상 건물 정보가 올바르지 않습니다."}), 400

    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO building_requests
            (request_type, target_sgg_cd, target_umd_nm, target_jibun, suggested_lodging_type,
             suggested_building_name, requester_note)
        VALUES ('correction', %s, %s, %s, %s, %s, %s) RETURNING id
    """, (sgg_cd, umd_nm, jibun, suggested_lodging_type, suggested_building_name, requester_note))
    request_id = cur.fetchone()["id"]
    conn.commit()

    def fail(reason):
        cur.execute("""
            UPDATE building_requests SET status='rejected', reject_reason=%s, processed_at=NOW()
            WHERE id=%s
        """, (reason, request_id))
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"status": "rejected", "message": reason})

    # 매칭키 정규화: 마스터는 umd_nm을 공백 없이("봉평면면온리"), 실거래는 공백 포함으로
    # ("봉평면 면온리") 저장하므로, 양쪽 컬럼에서 공백을 제거해 비교해야 면/리 지역도 매칭된다.
    umd_key = normalize_umd_nm(umd_nm)

    cur.execute("""
        SELECT id, building_name, lodging_type, name_pending FROM master_buildings
        WHERE sgg_cd=%s AND REPLACE(umd_nm, ' ', '')=%s AND jibun=%s
    """, (sgg_cd, umd_key, jibun))
    building = cur.fetchone()
    if not building:
        return fail("해당 건물을 마스터 목록에서 찾지 못했습니다.")

    bjdong_csv = _os.environ.get("BJDONG_CODE_CSV", "법정동코드 전체자료.csv")
    bjdong = BjdongMap(bjdong_csv)
    bjdong_cd = bjdong.find_bjdong_cd(sgg_cd, umd_nm)
    if not bjdong_cd:
        return fail("읍/면/동 코드 매칭에 실패했습니다.")

    plat_gb, bun, ji = parse_jibun(jibun)

    try:
        label, detail, title, reason = classify_lodging_type(sgg_cd, bjdong_cd, plat_gb, bun, ji)
    except Exception as e:
        return fail(f"건축물대장 재조회 중 오류: {e}")

    if label is None:
        return fail(f"재검증했지만 판정이 어렵습니다 ({reason}). 기존 값을 그대로 유지합니다.")

    old_label = building["lodging_type"]
    changed = (label != old_label)

    # 건물명 수정요청 처리 — 사용자 제안명을 그대로 반영하지 않고,
    # 이번 재조회에서 함께 받아온 건축물대장 명칭(title["bld_nm"])으로만 확정한다.
    api_bld_nm = resolve_api_building_name(title)
    name_changed = False
    name_review = False
    name_message = ""
    if suggested_building_name:
        if api_bld_nm:
            if api_bld_nm != (building["building_name"] or ""):
                cur.execute("""
                    UPDATE master_buildings SET building_name=%s, name_pending=FALSE
                    WHERE id=%s
                """, (api_bld_nm, building["id"]))
                name_changed = True
                name_message = f" 건물명은 건축물대장에서 '{api_bld_nm}'(으)로 확인되어 반영했습니다."
            else:
                if building.get("name_pending"):
                    cur.execute("UPDATE master_buildings SET name_pending=FALSE WHERE id=%s",
                                (building["id"],))
                name_message = f" 건물명은 건축물대장 확인 결과 기존 명칭 '{api_bld_nm}'이 맞습니다."
        else:
            # API에 명칭이 없음 → 제안명은 기록만 하고 마스터는 그대로(지번 임시명 유지).
            # 관리자 수정요청 이력에 '명칭 확인 필요'로 노출된다.
            name_review = True
            name_message = " 제안하신 건물명은 건축물대장에서 확인되지 않아, 관리자 확인 후 반영 예정입니다."

    if changed:
        cur.execute("""
            UPDATE master_buildings SET lodging_type=%s, lodging_type_detail=%s, verified_at=NOW()
            WHERE id=%s
        """, (label, detail, building["id"]))
        cur.execute("""
            UPDATE transactions SET lodging_type=%s, lodging_type_detail=%s
            WHERE sgg_cd=%s AND REPLACE(umd_nm, ' ', '')=%s AND jibun=%s
        """, (label, detail, sgg_cd, umd_key, jibun))

    # 명칭 미확인 건은 'name_review' 상태로 남겨 관리자 이력에서 "명칭 확인 필요"로 노출한다.
    final_status = "name_review" if name_review else "verified"
    cur.execute("""
        UPDATE building_requests
        SET status=%s, verified_lodging_type=%s, changed=%s, master_building_id=%s, processed_at=NOW()
        WHERE id=%s
    """, (final_status, label, changed or name_changed, building["id"], request_id))
    conn.commit()
    cur.close()
    conn.close()

    if changed:
        message = f"재검증 결과 '{old_label or '미확인'}' → '{label}'(으)로 확인되어 반영했습니다."
    else:
        message = f"건축물대장을 다시 확인했지만, 기존 라벨 '{old_label or '미확인'}'이 맞는 것으로 확인됐습니다."
        if suggested_lodging_type and suggested_lodging_type != label:
            message += f" (제안하신 '{suggested_lodging_type}'과는 다릅니다.)"
    message += name_message

    return jsonify({
        "status": final_status,
        "changed": changed,
        "name_changed": name_changed,
        "name_review": name_review,
        "lodging_type": label,
        "message": message,
    })


# ============================================================
# 관리자(E화면) — admin_users 기반 이메일/비밀번호 로그인 + 건물마스터 CRUD
# 로그인 성공 시 서명된 세션 쿠키에 admin=True, admin_user_id=행 id를 저장한다.
# require_admin 및 /api/admin/* 나머지 API는 session["admin"] 여부만 확인한다.
# ============================================================

def require_admin(f):
    """세션에 admin=True가 없으면 차단한다.
    /api/admin/* 요청은 401 JSON, 그 외(/admin/*)는 /admin/login으로 리다이렉트."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("admin"):
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
            return redirect("/admin/login")
        return f(*args, **kwargs)
    return wrapper


def _serve_static_html(filename):
    """정적 HTML을 no-cache 헤더와 함께 서빙 (apply 페이지들과 동일 방식)."""
    html_path = os.path.join(app.static_folder, filename)
    with open(html_path, encoding="utf-8") as fp:
        html = fp.read()
    html = _inject_asset_version(html)
    resp = Response(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


@app.route("/notices")
def notices_page():
    """공지사항 페이지 (현재는 정적 안내만)."""
    return _serve_static_html("notices.html")


@app.route("/menu")
def menu_page():
    """모바일 전용 전체 메뉴 페이지 (햄버거 버튼에서 진입)."""
    return _serve_static_html("menu.html")


@app.route("/mypage")
def mypage_page():
    """마이페이지 — 로그인 여부는 /api/auth/me로 클라이언트에서 판단한다."""
    return _serve_static_html("mypage.html")


@app.route("/transactions")
def transactions_page():
    """실거래목록 전용 페이지 (검색 필터 + 게시판, /api/transactions 재사용)."""
    return _serve_static_html("transactions.html")


@app.route("/terms")
def terms_page():
    """이용약관 페이지 (정적)."""
    return _serve_static_html("terms.html")


@app.route("/partner")
def partner_page():
    """파트너(중개사·운영업체) 등록 안내 페이지."""
    return _serve_static_html("partner.html")


@app.route("/agents")
def agents_landing_page():
    """담당중개사 아웃바운드/소개용 랜딩 페이지 (?company=업체명 개인화)."""
    return _serve_static_html("agents.html")


@app.route("/operators")
def operators_landing_page():
    """위탁운영업체 이메일 아웃바운드용 랜딩 페이지 (?company=업체명 개인화)."""
    return _serve_static_html("operators.html")


@app.route("/loan-partners")
def loan_partners_landing_page():
    """대출상담사(금융 파트너) 아웃바운드용 랜딩 페이지 (?company=업체명 개인화)."""
    return _serve_static_html("loan_partners.html")


@app.route("/privacy")
def privacy_page():
    """개인정보처리방침 페이지 — 뼈대는 정적, 본문은 /api/legal/privacy에서 로드."""
    return _serve_static_html("privacy.html")


# ---- 약관/개인정보처리방침 (legal_documents) ----
# doc_type은 'terms' 또는 'privacy' 두 값만 허용한다.
_LEGAL_DOC_TYPES = ("terms", "privacy")


@app.route("/api/legal/<doc_type>")
def public_legal_get(doc_type):
    """공개 조회 — 인증 불필요. /terms, /privacy 페이지가 본문을 채울 때 사용."""
    if doc_type not in _LEGAL_DOC_TYPES:
        return jsonify({"ok": False, "message": "존재하지 않는 문서입니다."}), 404
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT content, to_char(updated_at, 'YYYY-MM-DD') AS updated_at "
            "FROM legal_documents WHERE doc_type = %s",
            [doc_type],
        )
        row = cur.fetchone()
    finally:
        cur.close()
        conn.close()
    if not row:
        return jsonify({"ok": False, "message": "존재하지 않는 문서입니다."}), 404
    return jsonify({"ok": True, "doc_type": doc_type,
                    "content": row["content"], "updated_at": row["updated_at"]})


@app.route("/api/admin/legal/<doc_type>")
@require_admin
def admin_legal_get(doc_type):
    """관리자 조회 — 현재 저장된 본문 반환."""
    if doc_type not in _LEGAL_DOC_TYPES:
        return jsonify({"ok": False, "message": "존재하지 않는 문서입니다."}), 404
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT content, to_char(updated_at, 'YYYY-MM-DD HH24:MI') AS updated_at "
            "FROM legal_documents WHERE doc_type = %s",
            [doc_type],
        )
        row = cur.fetchone()
    finally:
        cur.close()
        conn.close()
    if not row:
        return jsonify({"ok": False, "message": "존재하지 않는 문서입니다."}), 404
    return jsonify({"ok": True, "doc_type": doc_type,
                    "content": row["content"], "updated_at": row["updated_at"]})


@app.route("/api/admin/legal/<doc_type>", methods=["PUT"])
@require_admin
def admin_legal_update(doc_type):
    """관리자 저장 — content 통째로 교체. 없으면 새로 만든다(upsert)."""
    if doc_type not in _LEGAL_DOC_TYPES:
        return jsonify({"ok": False, "message": "존재하지 않는 문서입니다."}), 404
    data = request.get_json(force=True, silent=True) or {}
    content = (data.get("content") or "").strip()
    if not content:
        return jsonify({"ok": False, "message": "본문은 비울 수 없습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """INSERT INTO legal_documents (doc_type, content, updated_at)
               VALUES (%s, %s, NOW())
               ON CONFLICT (doc_type)
               DO UPDATE SET content = EXCLUDED.content, updated_at = NOW()""",
            [doc_type, content],
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


# =====================================================================
# 일반 회원(users) 인증 — 이메일/비밀번호 + 카카오 소셜 로그인
# 관리자(admin_users / session["admin"])와는 완전히 분리된 세션 키를 쓴다.
#   - 일반 회원 세션 키: session["user_id"]
#   - 관리자 세션 키   : session["admin"]
# =====================================================================

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _valid_email(email):
    return bool(email) and len(email) <= 254 and _EMAIL_RE.match(email) is not None


def current_user():
    """세션의 user_id로 현재 로그인한 일반 회원 행을 돌려준다. 없으면 None."""
    uid = session.get("user_id")
    if not uid:
        return None
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, email, name, provider, COALESCE(email_alert_enabled, TRUE) AS email_alert_enabled"
            " FROM users WHERE id = %s AND status <> 'withdrawn'",
            (uid,),
        )
        return cur.fetchone()
    finally:
        cur.close()
        conn.close()


@app.route("/api/auth/signup", methods=["POST"])
@limiter.limit("5 per minute; 20 per hour")
def auth_signup():
    """이메일 회원가입 → 성공 시 자동 로그인."""
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    name = (data.get("name") or "").strip()

    if not _valid_email(email):
        return jsonify({"ok": False, "message": "올바른 이메일 형식이 아닙니다."}), 400
    if len(password) < 8:
        return jsonify({"ok": False, "message": "비밀번호는 8자 이상이어야 합니다."}), 400
    if not name:
        return jsonify({"ok": False, "message": "이름을 입력해주세요."}), 400

    # 법적 동의 재검증 — 클라이언트 체크박스와 무관하게 서버에서도 필수 3개를 강제.
    # (만 14세 이상 / 이용약관 / 개인정보 수집·이용. 하나라도 빠지면 400)
    if not (data.get("age14") is True and data.get("terms") is True and data.get("privacy") is True):
        return jsonify({"ok": False, "message": "필수 약관(만 14세 이상, 이용약관, 개인정보 수집·이용)에 모두 동의해주세요."}), 400
    marketing = data.get("marketing") is True  # 선택 동의 — 체크한 경우에만 시각 기록

    pw_hash = generate_password_hash(password)
    conn = get_conn()
    cur = conn.cursor()
    try:
        # 이메일 중복 확인 (대소문자 무시). DB UNIQUE 제약도 있지만 친절한 메시지를 위해 먼저 확인.
        cur.execute("SELECT id FROM users WHERE LOWER(email) = %s", (email,))
        if cur.fetchone():
            return jsonify({"ok": False, "message": "이미 가입된 이메일입니다."}), 400
        cur.execute(
            """INSERT INTO users (email, password_hash, name, provider, last_login_at,
                                  terms_agreed_at, privacy_agreed_at, marketing_agreed_at)
               VALUES (%s, %s, %s, 'email', NOW(),
                       NOW(), NOW(), CASE WHEN %s THEN NOW() ELSE NULL END)
               RETURNING id""",
            (email, pw_hash, name, marketing),
        )
        new_id = cur.fetchone()["id"]
        conn.commit()
    except Exception:
        conn.rollback()
        # DB UNIQUE 제약 위반 등 → 중복으로 간주(경합 상황 대비)
        return jsonify({"ok": False, "message": "이미 가입된 이메일입니다."}), 400
    finally:
        cur.close()
        conn.close()

    session["user_id"] = new_id
    session.permanent = True
    return jsonify({"ok": True})


@app.route("/api/auth/login", methods=["POST"])
@limiter.limit("5 per minute; 20 per hour")
def auth_login():
    """이메일/비밀번호 로그인. 실패 시 계정 존재 여부를 노출하지 않도록 메시지를 통일한다."""
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    fail_msg = "이메일 또는 비밀번호가 올바르지 않습니다."

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, password_hash, status FROM users WHERE LOWER(email) = %s",
            (email,),
        )
        row = cur.fetchone()
        # 카카오 전용 가입자(password_hash NULL)·탈퇴 계정은 이메일 로그인 불가 → 동일 메시지로 통일.
        if (not row or not row["password_hash"] or row.get("status") == "withdrawn"
                or not check_password_hash(row["password_hash"], password)):
            return jsonify({"ok": False, "message": fail_msg}), 401
        cur.execute("UPDATE users SET last_login_at = NOW() WHERE id = %s", (row["id"],))
        conn.commit()
        uid = row["id"]
    finally:
        cur.close()
        conn.close()

    session["user_id"] = uid
    # "로그인 상태 유지" 체크 시에만 31일 유지(permanent), 아니면 브라우저 세션 쿠키(닫으면 만료).
    session.permanent = bool(data.get("remember"))
    return jsonify({"ok": True})


@app.route("/api/auth/logout", methods=["POST"])
def auth_logout():
    """일반 회원 로그아웃 — 관리자 세션은 건드리지 않고 회원 키만 제거한다."""
    session.pop("user_id", None)
    session.pop("kakao_oauth_state", None)
    return jsonify({"ok": True})


@app.route("/api/auth/me")
def auth_me():
    """로그인 상태 조회. 프런트 헤더가 로그인/로그아웃 표시를 결정하는 데 쓴다."""
    u = current_user()
    if not u:
        return jsonify({"logged_in": False})
    return jsonify({
        "logged_in": True,
        "name": u.get("name"),
        "email": u.get("email"),
        "provider": u.get("provider"),
        "email_alert_enabled": bool(u.get("email_alert_enabled", True)),
    })


@app.route("/api/auth/me", methods=["PUT"])
@limiter.limit("5 per minute; 20 per hour")
def auth_update_name():
    """이름 변경 — 로그인 필요. 이메일/카카오 공통. name만 바꾼다."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    data = request.get_json(force=True, silent=True) or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "message": "이름을 입력해주세요."}), 400
    if len(name) > 50:
        return jsonify({"ok": False, "message": "이름은 50자 이하로 입력해주세요."}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE users SET name = %s WHERE id = %s", (name, u["id"]))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "name": name, "email": u.get("email"), "provider": u.get("provider")})


@app.route("/api/auth/email-alert", methods=["PUT"])
@limiter.limit("10 per minute")
def auth_update_email_alert():
    """실거래 이메일 알림 수신 여부 변경 — 로그인 필요. 인앱 알림에는 영향 없음."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    data = request.get_json(force=True, silent=True) or {}
    raw = data.get("enabled")
    if not isinstance(raw, bool):
        return jsonify({"ok": False, "message": "enabled 값은 true/false여야 합니다."}), 400
    enabled = raw
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE users SET email_alert_enabled = %s WHERE id = %s", (enabled, u["id"]))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "email_alert_enabled": enabled})


@app.route("/api/auth/password", methods=["PUT"])
@limiter.limit("5 per minute; 20 per hour")
def auth_change_password():
    """비밀번호 변경 — 로그인 필요, 이메일 계정만. 현재 비밀번호 확인 후 교체."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    if u.get("provider") != "email":
        return jsonify({"ok": False, "message": "카카오 로그인 계정은 비밀번호 변경이 필요 없습니다."}), 400
    data = request.get_json(force=True, silent=True) or {}
    current_pw = data.get("current_password") or ""
    new_pw = data.get("new_password") or ""
    if len(new_pw) < 8:
        return jsonify({"ok": False, "message": "비밀번호는 8자 이상이어야 합니다."}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT password_hash FROM users WHERE id = %s", (u["id"],))
        row = cur.fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], current_pw):
            return jsonify({"ok": False, "message": "현재 비밀번호가 올바르지 않습니다."}), 401
        cur.execute(
            "UPDATE users SET password_hash = %s WHERE id = %s",
            (generate_password_hash(new_pw), u["id"]),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/auth/me", methods=["DELETE"])
@limiter.limit("5 per minute; 20 per hour")
def auth_withdraw():
    """회원탈퇴 — 로그인 필요. 완전삭제 대신 status='withdrawn' 소프트삭제 후 세션 초기화."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE users SET status = 'withdrawn' WHERE id = %s", (u["id"],))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    session.pop("user_id", None)
    session.pop("kakao_oauth_state", None)
    return jsonify({"ok": True})


# ---- 카카오 소셜 로그인 (OAuth 2.0 Authorization Code) ----

_KAKAO_AUTHORIZE_URL = "https://kauth.kakao.com/oauth/authorize"
_KAKAO_TOKEN_URL = "https://kauth.kakao.com/oauth/token"
_KAKAO_USERME_URL = "https://kapi.kakao.com/v2/user/me"


def _kakao_redirect_uri():
    """콜백 URL을 현재 요청 호스트 기준으로 만든다(https 강제).
    카카오 개발자센터의 'Redirect URI'에 이 값이 정확히 등록돼 있어야 한다.
    (개발 도메인·배포 도메인 두 개를 모두 등록하면 양쪽에서 동작)"""
    host = request.host
    return f"https://{host}/auth/kakao/callback"


@app.route("/auth/kakao/start")
def kakao_start():
    """카카오 인증 페이지로 리다이렉트. CSRF 방지용 state를 세션에 저장한다."""
    client_id = os.environ.get("KAKAO_REST_API_KEY", "")
    if not client_id:
        return redirect("/?login_error=1")
    state = _secrets.token_urlsafe(24)
    session["kakao_oauth_state"] = state
    params = {
        "client_id": client_id,
        "redirect_uri": _kakao_redirect_uri(),
        "response_type": "code",
        "state": state,
    }
    return redirect(f"{_KAKAO_AUTHORIZE_URL}?{urlencode(params)}")


@app.route("/auth/kakao/callback")
def kakao_callback():
    """카카오 콜백: code→token→사용자정보→users upsert→세션 저장→홈으로.
    어떤 단계든 실패/거부 시 에러 페이지 대신 홈으로 리다이렉트(?login_error=1)."""
    # 사용자가 동의 취소하면 error 파라미터가 붙어 돌아온다.
    if request.args.get("error"):
        return redirect("/?login_error=1")

    # CSRF: 세션에 저장한 state와 콜백 state가 일치해야 한다. (1회용 → 즉시 제거)
    expected_state = session.pop("kakao_oauth_state", None)
    if not expected_state or request.args.get("state") != expected_state:
        return redirect("/?login_error=1")

    code = request.args.get("code")
    if not code:
        return redirect("/?login_error=1")

    client_id = os.environ.get("KAKAO_REST_API_KEY", "")
    client_secret = os.environ.get("KAKAO_CLIENT_SECRET", "")  # 선택: 카카오 콘솔에서 'client_secret' 사용 설정 시에만 필요
    if not client_id:
        return redirect("/?login_error=1")

    try:
        token_data = {
            "grant_type": "authorization_code",
            "client_id": client_id,
            "redirect_uri": _kakao_redirect_uri(),
            "code": code,
        }
        if client_secret:
            token_data["client_secret"] = client_secret
        tok = requests.post(
            _KAKAO_TOKEN_URL,
            data=token_data,
            headers={"Content-Type": "application/x-www-form-urlencoded;charset=utf-8"},
            timeout=10,
        )
        tok.raise_for_status()
        access_token = tok.json().get("access_token")
        if not access_token:
            return redirect("/?login_error=1")

        me = requests.get(
            _KAKAO_USERME_URL,
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=10,
        )
        me.raise_for_status()
        me_json = me.json()
    except Exception:
        return redirect("/?login_error=1")

    kakao_id = str(me_json.get("id") or "").strip()
    if not kakao_id:
        return redirect("/?login_error=1")
    account = me_json.get("kakao_account") or {}
    profile = account.get("profile") or {}
    email = (account.get("email") or "").strip().lower() or None
    nickname = (profile.get("nickname") or "").strip() or "카카오회원"

    conn = get_conn()
    cur = conn.cursor()
    try:
        # kakao_id 기준 upsert: 있으면 로그인(마지막 로그인 갱신), 없으면 신규 생성.
        cur.execute("SELECT id, status FROM users WHERE kakao_id = %s", (kakao_id,))
        row = cur.fetchone()
        if row and row.get("status") == "withdrawn":
            # 탈퇴한 카카오 계정은 자동 부활시키지 않는다(로그인 차단).
            conn.rollback()
            return redirect("/?login_error=1")
        if row:
            uid = row["id"]
            cur.execute("UPDATE users SET last_login_at = NOW() WHERE id = %s", (uid,))
        else:
            # 이메일이 이미 이메일가입으로 존재하면 충돌 방지를 위해 email은 비우고 카카오 계정으로 신규 생성.
            email_to_store = email
            if email_to_store:
                cur.execute("SELECT id FROM users WHERE LOWER(email) = %s", (email_to_store,))
                if cur.fetchone():
                    email_to_store = None
            cur.execute(
                """INSERT INTO users (email, password_hash, name, provider, kakao_id, last_login_at)
                   VALUES (%s, NULL, %s, 'kakao', %s, NOW()) RETURNING id""",
                (email_to_store, nickname, kakao_id),
            )
            uid = cur.fetchone()["id"]
        conn.commit()
    except Exception:
        conn.rollback()
        return redirect("/?login_error=1")
    finally:
        cur.close()
        conn.close()

    session["user_id"] = uid
    session.permanent = True
    return redirect("/")


# =====================================================================
# 로그인 회원 관심단지(user_favorites) — 비로그인은 기존 localStorage만 사용.
#   프론트 favKey = "building_name|address" 규칙과 동일하게 (building_name,address)로 저장한다.
#   미매칭(건물명 NULL) 거래는 프론트에서 "null" 문자열로 표현하므로 저장 시 NULL로 정규화한다.
# =====================================================================

def _norm_fav_name(name):
    """프론트 favKey의 건물명 부분을 서버 저장용으로 정규화. "null"/"undefined"/빈값 → None."""
    if name is None:
        return None
    s = str(name).strip()
    if s in ("", "null", "undefined"):
        return None
    return s


@app.route("/api/favorites/mine")
def favorites_mine():
    """로그인 회원의 관심단지 목록 + 각 단지의 최신 실거래가 + 건물상세 링크용 building_id."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    conn = get_conn()
    cur = conn.cursor()
    try:
        # lt: 관심키(건물명+주소)에 해당하는 최신 실거래 1건.
        # building_id 결정 순서(실거래 비의존):
        #   1) uf.master_building_id — 저장 시점에 프론트가 넘긴 건물 id (가장 확실)
        #   2) bid: 거래 경유 역매칭 (기존 저장분 하위호환)
        #   3) bid2: master_buildings 직접 매칭 — 도로명주소 일치 또는
        #      "읍면동+지번" 조합이 uf.address와 일치(공백 제거 비교) → 실거래 없어도 링크됨
        cur.execute("""
            SELECT uf.building_name, uf.address, uf.created_at,
                   lt.price, lt.deal_date, lt.area, lt.floor, lt.deal_type,
                   lt.lodging_type, lt.lodging_type_detail,
                   COALESCE(uf.master_building_id, bid.id, bid2.id) AS building_id
            FROM user_favorites uf
            LEFT JOIN LATERAL (
                SELECT t.price, t.deal_date, t.area, t.floor, t.deal_type,
                       t.lodging_type, t.lodging_type_detail
                FROM transactions t
                WHERE ((uf.building_name IS NULL AND t.building_name IS NULL)
                       OR t.building_name = uf.building_name)
                  AND t.address = uf.address
                ORDER BY t.deal_date DESC, t.id DESC
                LIMIT 1
            ) lt ON TRUE
            LEFT JOIN LATERAL (
                SELECT mb.id
                FROM transactions t2
                JOIN master_buildings mb
                  ON mb.sgg_cd = t2.sgg_cd AND mb.umd_nm = t2.umd_nm AND mb.jibun = t2.jibun
                WHERE ((uf.building_name IS NULL AND t2.building_name IS NULL)
                       OR t2.building_name = uf.building_name)
                  AND t2.address = uf.address
                ORDER BY (mb.building_name = uf.building_name) DESC NULLS LAST, mb.id
                LIMIT 1
            ) bid ON TRUE
            LEFT JOIN LATERAL (
                SELECT mb.id
                FROM master_buildings mb
                WHERE mb.road_address = uf.address
                   OR REPLACE(mb.umd_nm || mb.jibun, ' ', '') = REPLACE(uf.address, ' ', '')
                ORDER BY (mb.building_name = uf.building_name) DESC NULLS LAST, mb.id
                LIMIT 1
            ) bid2 ON TRUE
            WHERE uf.user_id = %s
            ORDER BY uf.created_at DESC, uf.id DESC
        """, (u["id"],))
        rows = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "items": rows, "total": len(rows)})


@app.route("/api/favorites/mine", methods=["POST"])
def favorites_mine_add():
    """관심단지 1건 저장 — 이미 있으면 무시(중복 스킵)."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    data = request.get_json(force=True, silent=True) or {}
    name = _norm_fav_name(data.get("building_name"))
    addr = (data.get("address") or "").strip()
    if not addr:
        return jsonify({"ok": False, "message": "주소가 필요합니다."}), 400
    # 프론트가 알고 있는 master_buildings.id — 실거래 없는 건물도 상세 링크 유지용(선택값)
    bid = data.get("building_id")
    try:
        bid = int(bid) if bid is not None else None
        if bid is not None and bid <= 0:
            bid = None
    except (TypeError, ValueError):
        bid = None
    conn = get_conn()
    cur = conn.cursor()
    try:
        if bid is not None:
            # 존재하는 건물 id만 저장(임의 값 방지)
            cur.execute("SELECT 1 FROM master_buildings WHERE id = %s", (bid,))
            if not cur.fetchone():
                bid = None
        # 표현식 UNIQUE 인덱스(uq_user_favorites) 기준으로 원자적 dedup — 동시요청 안전.
        # 이미 저장된 관심단지에 building_id 없이 남아 있으면 이번 값으로 채워준다(백필).
        cur.execute(
            "INSERT INTO user_favorites (user_id, building_name, address, master_building_id) "
            "VALUES (%s, %s, %s, %s) "
            "ON CONFLICT (user_id, COALESCE(building_name, ''), address) DO UPDATE "
            "SET master_building_id = COALESCE(user_favorites.master_building_id, EXCLUDED.master_building_id)",
            (u["id"], name, addr, bid),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/favorites/mine", methods=["DELETE"])
def favorites_mine_remove():
    """관심단지 1건 삭제."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    data = request.get_json(force=True, silent=True) or {}
    name = _norm_fav_name(data.get("building_name"))
    addr = (data.get("address") or "").strip()
    if not addr:
        return jsonify({"ok": False, "message": "주소가 필요합니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "DELETE FROM user_favorites "
            "WHERE user_id = %s AND COALESCE(building_name,'') = COALESCE(%s,'') AND address = %s",
            (u["id"], name, addr),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/favorites/migrate", methods=["POST"])
def favorites_migrate():
    """로그인 직후 1회 호출용. localStorage favKey 배열을 받아 없는 것만 채우고,
    합쳐진 최종 관심키 목록을 돌려준다(프론트가 localStorage를 이 값으로 동기화)."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    data = request.get_json(force=True, silent=True) or {}
    keys = data.get("keys") or []
    pairs = []
    if isinstance(keys, list):
        for token in keys:
            if not isinstance(token, str) or "|" not in token:
                continue
            name, addr = token.split("|", 1)
            addr = addr.strip()
            if not addr:
                continue
            pairs.append((_norm_fav_name(name), addr))
    conn = get_conn()
    cur = conn.cursor()
    try:
        for name, addr in pairs:
            # 표현식 UNIQUE 인덱스 기준 원자적 dedup — 동시요청/중복키 안전
            cur.execute(
                "INSERT INTO user_favorites (user_id, building_name, address) VALUES (%s, %s, %s) "
                "ON CONFLICT (user_id, COALESCE(building_name, ''), address) DO NOTHING",
                (u["id"], name, addr),
            )
        conn.commit()
        cur.execute(
            "SELECT building_name, address FROM user_favorites "
            "WHERE user_id = %s ORDER BY created_at ASC, id ASC",
            (u["id"],),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()
    merged = []
    for r in rows:
        nm = r["building_name"]
        merged.append(f"{nm if nm is not None else 'null'}|{r['address']}")
    return jsonify({"ok": True, "keys": merged})


# =====================================================================
# 실거래 알림 구독(user_alert_subscriptions) — user_favorites 와 동일 패턴.
#   관심저장과 독립적으로 켜고 끌 수 있는 별도 테이블. 새 실거래가 들어오면
#   sync_batch.py 가 이 구독을 조회해 notifications 를 만든다.
# =====================================================================

@app.route("/api/alerts/mine")
def alerts_mine():
    """로그인 회원의 실거래 알림 구독 목록 + 각 단지 최신 실거래가 + building_id."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT us.building_name, us.address, us.created_at,
                   lt.price, lt.deal_date, lt.area, lt.floor, lt.deal_type,
                   lt.lodging_type, lt.lodging_type_detail,
                   bid.id AS building_id
            FROM user_alert_subscriptions us
            LEFT JOIN LATERAL (
                SELECT t.price, t.deal_date, t.area, t.floor, t.deal_type,
                       t.lodging_type, t.lodging_type_detail
                FROM transactions t
                WHERE ((us.building_name IS NULL AND t.building_name IS NULL)
                       OR t.building_name = us.building_name)
                  AND t.address = us.address
                ORDER BY t.deal_date DESC, t.id DESC
                LIMIT 1
            ) lt ON TRUE
            LEFT JOIN LATERAL (
                SELECT mb.id
                FROM transactions t2
                JOIN master_buildings mb
                  ON mb.sgg_cd = t2.sgg_cd AND mb.umd_nm = t2.umd_nm AND mb.jibun = t2.jibun
                WHERE ((us.building_name IS NULL AND t2.building_name IS NULL)
                       OR t2.building_name = us.building_name)
                  AND t2.address = us.address
                ORDER BY (mb.building_name = us.building_name) DESC NULLS LAST, mb.id
                LIMIT 1
            ) bid ON TRUE
            WHERE us.user_id = %s
            ORDER BY us.created_at DESC, us.id DESC
        """, (u["id"],))
        rows = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    # 프론트가 구독 여부 판정에 쓰는 favKey 목록도 함께 내려준다.
    keys = [f"{(r['building_name'] if r['building_name'] is not None else 'null')}|{r['address']}" for r in rows]
    return jsonify({"ok": True, "items": rows, "keys": keys, "total": len(rows)})


@app.route("/api/alerts/mine", methods=["POST"])
def alerts_mine_add():
    """실거래 알림 구독 1건 추가 — 이미 있으면 무시(중복 스킵)."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    data = request.get_json(force=True, silent=True) or {}
    name = _norm_fav_name(data.get("building_name"))
    addr = (data.get("address") or "").strip()
    if not addr:
        return jsonify({"ok": False, "message": "주소가 필요합니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO user_alert_subscriptions (user_id, building_name, address) VALUES (%s, %s, %s) "
            "ON CONFLICT (user_id, COALESCE(building_name, ''), address) DO NOTHING",
            (u["id"], name, addr),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/alerts/mine", methods=["DELETE"])
def alerts_mine_remove():
    """실거래 알림 구독 1건 삭제."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    data = request.get_json(force=True, silent=True) or {}
    name = _norm_fav_name(data.get("building_name"))
    addr = (data.get("address") or "").strip()
    if not addr:
        return jsonify({"ok": False, "message": "주소가 필요합니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "DELETE FROM user_alert_subscriptions "
            "WHERE user_id = %s AND COALESCE(building_name,'') = COALESCE(%s,'') AND address = %s",
            (u["id"], name, addr),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/alerts/migrate", methods=["POST"])
def alerts_migrate():
    """로그인 직후 1회 호출용. localStorage 알림구독(favKey 배열)을 서버로 이관하고,
    합쳐진 최종 구독키 목록을 돌려준다(프론트가 localStorage를 이 값으로 동기화)."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    data = request.get_json(force=True, silent=True) or {}
    keys = data.get("keys") or []
    pairs = []
    if isinstance(keys, list):
        for token in keys:
            if not isinstance(token, str) or "|" not in token:
                continue
            name, addr = token.split("|", 1)
            addr = addr.strip()
            if not addr:
                continue
            pairs.append((_norm_fav_name(name), addr))
    conn = get_conn()
    cur = conn.cursor()
    try:
        for name, addr in pairs:
            cur.execute(
                "INSERT INTO user_alert_subscriptions (user_id, building_name, address) VALUES (%s, %s, %s) "
                "ON CONFLICT (user_id, COALESCE(building_name, ''), address) DO NOTHING",
                (u["id"], name, addr),
            )
        conn.commit()
        cur.execute(
            "SELECT building_name, address FROM user_alert_subscriptions "
            "WHERE user_id = %s ORDER BY created_at ASC, id ASC",
            (u["id"],),
        )
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()
    merged = [f"{(r['building_name'] if r['building_name'] is not None else 'null')}|{r['address']}" for r in rows]
    return jsonify({"ok": True, "keys": merged})


# =====================================================================
# 알림함(notifications) — 헤더 벨 아이콘. 새 실거래 발생 시 sync_batch 가 쌓아둔다.
# =====================================================================

@app.route("/api/notifications/mine")
def notifications_mine():
    """최근 알림 목록 — 안읽음 우선, 최신순, 최대 30개. 클릭 이동용 building_id 포함."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT n.id, n.title, n.body, n.building_name, n.address,
                   n.is_read, n.created_at,
                   bid.id AS building_id
            FROM notifications n
            LEFT JOIN LATERAL (
                SELECT mb.id
                FROM transactions t2
                JOIN master_buildings mb
                  ON mb.sgg_cd = t2.sgg_cd AND mb.umd_nm = t2.umd_nm AND mb.jibun = t2.jibun
                WHERE ((n.building_name IS NULL AND t2.building_name IS NULL)
                       OR t2.building_name = n.building_name)
                  AND t2.address = n.address
                ORDER BY (mb.building_name = n.building_name) DESC NULLS LAST, mb.id
                LIMIT 1
            ) bid ON TRUE
            WHERE n.user_id = %s
            ORDER BY n.is_read ASC, n.created_at DESC, n.id DESC
            LIMIT 30
        """, (u["id"],))
        rows = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "items": rows, "total": len(rows)})


@app.route("/api/notifications/unread-count")
def notifications_unread_count():
    """헤더 벨 뱃지용 안 읽은 알림 개수."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT COUNT(*) AS c FROM notifications WHERE user_id = %s AND is_read = FALSE",
            (u["id"],),
        )
        c = cur.fetchone()["c"]
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "count": c})


@app.route("/api/notifications/mine/read-all", methods=["POST"])
def notifications_read_all():
    """전체 읽음 처리."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE notifications SET is_read = TRUE WHERE user_id = %s AND is_read = FALSE",
            (u["id"],),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/notifications/mine/read", methods=["POST"])
def notifications_read_one():
    """알림 1건 읽음 처리 — 항목 클릭 시 사용."""
    u = current_user()
    if not u:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    data = request.get_json(force=True, silent=True) or {}
    nid = data.get("id")
    if not isinstance(nid, int):
        try:
            nid = int(nid)
        except (TypeError, ValueError):
            return jsonify({"ok": False, "message": "알림 id가 필요합니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "UPDATE notifications SET is_read = TRUE WHERE id = %s AND user_id = %s",
            (nid, u["id"]),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/admin/login")
def admin_login_page():
    return _serve_static_html("admin_login.html")


@app.route("/api/admin/login", methods=["POST"])
@limiter.limit("5 per minute; 30 per hour")
def admin_login():
    """admin_users 테이블 기반 이메일/비밀번호 로그인.
    실패 시 이메일 존재 여부를 드러내지 않도록 통일된 메시지로 401을 반환한다."""
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip()
    password = data.get("password") or ""
    fail = jsonify({"ok": False, "message": "이메일 또는 비밀번호가 올바르지 않습니다."}), 401
    if not email or not password:
        return fail

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, password_hash FROM admin_users WHERE email = %s",
            (email,),
        )
        row = cur.fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], password):
            return fail
        cur.execute("UPDATE admin_users SET last_login_at = NOW() WHERE id = %s", (row["id"],))
        conn.commit()
    finally:
        cur.close()
        conn.close()

    session["admin"] = True
    session["admin_user_id"] = row["id"]
    session.permanent = True
    return jsonify({"ok": True})


@app.route("/api/admin/logout", methods=["POST"])
def admin_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/admin/password", methods=["PUT"])
@require_admin
@limiter.limit("5 per minute; 20 per hour")
def admin_change_password():
    """관리자 비밀번호 변경 — 로그인 필요. 현재 비밀번호 확인 후 새 비밀번호로 교체."""
    admin_id = session.get("admin_user_id")
    if not admin_id:
        return jsonify({"ok": False, "message": "다시 로그인해주세요."}), 401
    data = request.get_json(force=True, silent=True) or {}
    current_pw = data.get("current_password") or ""
    new_pw = data.get("new_password") or ""
    if len(new_pw) < 8:
        return jsonify({"ok": False, "message": "새 비밀번호는 8자 이상이어야 합니다."}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT password_hash FROM admin_users WHERE id = %s", (admin_id,))
        row = cur.fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], current_pw):
            return jsonify({"ok": False, "message": "현재 비밀번호가 올바르지 않습니다."}), 401
        cur.execute(
            "UPDATE admin_users SET password_hash = %s WHERE id = %s",
            (generate_password_hash(new_pw), admin_id),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


# ============================================================
# 중개사(agents) 로그인 — 승인된 중개사만. admin 로그인과 같은 패턴.
# 세션에 agent_id 저장. require_agent 로 보호.
# ============================================================

def require_agent(f):
    """세션에 agent_id가 없으면 차단한다.
    /api/* 요청은 401 JSON, 그 외는 /agent/login으로 리다이렉트."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("agent_id"):
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
            return redirect("/agent/login")
        return f(*args, **kwargs)
    return wrapper


@app.route("/agent/login")
def agent_login_page():
    return _serve_static_html("agent_login.html")


@app.route("/agent/<slug>")
def agent_profile_page(slug):
    """중개사 공개 프로필 페이지. Flask는 정적 룰(/agent/login)을 우선 매칭하므로 충돌 없음."""
    return _serve_static_html("agent_profile.html")


@app.route("/api/agent/profile/<slug>")
def agent_public_profile(slug):
    """중개사 공개 프로필 API — 인증 불필요. approved 상태만 노출."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, office_name, owner_name, phone, photo_url, intro_text
        FROM agents
        WHERE subdomain_slug = %s AND status = 'approved'
    """, [slug])
    agent = cur.fetchone()
    if not agent:
        cur.close()
        conn.close()
        return jsonify({"error": "not found"}), 404
    cur.execute("""
        SELECT ab.master_building_id, mb.building_name, mb.lodging_type,
               COALESCE(ab.sale_count, 0)      AS sale_count,
               COALESCE(ab.jeonse_count, 0)    AS jeonse_count,
               COALESCE(ab.wolse_count, 0)     AS wolse_count,
               COALESCE(ab.shortterm_count, 0) AS shortterm_count
        FROM agent_buildings ab
        JOIN master_buildings mb ON mb.id = ab.master_building_id
        WHERE ab.agent_id = %s
        ORDER BY mb.building_name
    """, [agent["id"]])
    buildings = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    total_listings = sum(
        b["sale_count"] + b["jeonse_count"] + b["wolse_count"] + b["shortterm_count"]
        for b in buildings
    )
    return jsonify({
        "office_name": agent["office_name"],
        "owner_name": agent["owner_name"],
        "phone": agent["phone"],
        "photo_url": agent["photo_url"],
        "intro_text": agent["intro_text"],
        "buildings": buildings,
        "building_count": len(buildings),
        "total_listings": total_listings,
    })


@app.route("/api/agent/login", methods=["POST"])
@limiter.limit("5 per minute; 20 per hour")
def agent_login():
    """agents 테이블 기반 이메일/비밀번호 로그인. status='approved'만 허용.
    실패 시 이메일 존재 여부를 드러내지 않도록 통일된 메시지로 401을 반환한다."""
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip()
    password = data.get("password") or ""
    fail = jsonify({"ok": False, "message": "이메일 또는 비밀번호가 올바르지 않습니다."}), 401
    if not email or not password:
        return fail

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, password_hash, status FROM agents WHERE LOWER(email) = LOWER(%s)",
            (email,),
        )
        row = cur.fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], password):
            return fail
        if row["status"] != "approved":
            return jsonify({"ok": False, "message": "승인된 중개사 계정이 아닙니다."}), 403
    finally:
        cur.close()
        conn.close()

    session["agent_id"] = row["id"]
    session.permanent = True
    return jsonify({"ok": True})


@app.route("/api/agent/logout", methods=["POST"])
def agent_logout():
    session.pop("agent_id", None)
    return jsonify({"ok": True})


@app.route("/api/agent/password", methods=["PUT"])
@require_agent
@limiter.limit("5 per minute; 20 per hour")
def agent_change_password():
    """중개사 비밀번호 변경 — 현재 비밀번호 확인 후 교체 (admin/mypage와 같은 패턴)."""
    agent_id = session.get("agent_id")
    data = request.get_json(force=True, silent=True) or {}
    current_pw = data.get("current_password") or ""
    new_pw = data.get("new_password") or ""
    if len(new_pw) < 8:
        return jsonify({"ok": False, "message": "새 비밀번호는 8자 이상이어야 합니다."}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT password_hash FROM agents WHERE id = %s", (agent_id,))
        row = cur.fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], current_pw):
            return jsonify({"ok": False, "message": "현재 비밀번호가 올바르지 않습니다."}), 401
        cur.execute(
            "UPDATE agents SET password_hash = %s WHERE id = %s",
            (generate_password_hash(new_pw), agent_id),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


# ---- 중개사 대시보드(본인 데이터 관리) — 전부 require_agent, session["agent_id"] 스코프 ----

@app.route("/agent/dashboard")
@require_agent
def agent_dashboard_page():
    return _serve_static_html("agent_dashboard.html")


@app.route("/api/agent/me")
@require_agent
def agent_me():
    agent_id = session["agent_id"]
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT office_name, owner_name, phone, photo_url, intro_text, subdomain_slug,
                   COALESCE(is_visible, TRUE) AS is_visible
            FROM agents WHERE id = %s
        """, [agent_id])
        me = cur.fetchone()
        if not me:
            return jsonify({"ok": False, "message": "계정을 찾을 수 없습니다."}), 404
        cur.execute("""
            SELECT ab.master_building_id, mb.building_name, mb.lodging_type,
                   COALESCE(ab.sale_count, 0)      AS sale_count,
                   COALESCE(ab.jeonse_count, 0)    AS jeonse_count,
                   COALESCE(ab.wolse_count, 0)     AS wolse_count,
                   COALESCE(ab.shortterm_count, 0) AS shortterm_count
            FROM agent_buildings ab
            JOIN master_buildings mb ON mb.id = ab.master_building_id
            WHERE ab.agent_id = %s
            ORDER BY mb.building_name
        """, [agent_id])
        buildings = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    out = dict(me)
    out["buildings"] = buildings
    return jsonify(out)


@app.route("/api/agent/me", methods=["PUT"])
@require_agent
def agent_me_update():
    """phone / photo_url / intro_text 부분 업데이트 — 전달된 키만 수정."""
    agent_id = session["agent_id"]
    data = request.get_json(force=True, silent=True) or {}
    allowed = ["phone", "photo_url", "intro_text"]
    sets, params = [], []
    for k in allowed:
        if k in data:
            v = data.get(k)
            if v is not None and not isinstance(v, str):
                return jsonify({"ok": False, "message": f"{k} 값이 올바르지 않습니다."}), 400
            v = (v or "").strip() or None
            if k == "phone":
                # 하이픈 유무와 무관하게 숫자만 저장 + 자릿수 검증 (표시할 때 재포맷)
                v = _digits_only(v) or None
                if v and not _validate_phone_digits(v):
                    return jsonify({"ok": False, "message": "전화번호 형식이 올바르지 않습니다. (숫자 10~11자리)"}), 400
            if k == "photo_url" and v and not (v.startswith("http://") or v.startswith("https://")):
                return jsonify({"ok": False, "message": "사진 URL은 http(s)://로 시작해야 합니다."}), 400
            sets.append(f"{k} = %s")
            params.append(v)
    if not sets:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다."}), 400
    params.append(agent_id)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"UPDATE agents SET {', '.join(sets)} WHERE id = %s", params)
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/agent/visibility", methods=["PUT"])
@require_agent
def agent_visibility_update():
    """노출 여부 토글 — 본인 세션 기준. is_visible만 갱신 (status와 무관)."""
    data = request.get_json(force=True, silent=True) or {}
    v = data.get("is_visible")
    if not isinstance(v, bool):
        return jsonify({"ok": False, "message": "is_visible 값은 true/false여야 합니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE agents SET is_visible = %s WHERE id = %s", (v, session["agent_id"]))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "is_visible": v})


@app.route("/api/agent/buildings", methods=["POST"])
@require_agent
def agent_building_add():
    agent_id = session["agent_id"]
    data = request.get_json(force=True, silent=True) or {}
    try:
        mbid = int(data.get("master_building_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "message": "건물 ID가 올바르지 않습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT 1 FROM master_buildings WHERE id = %s", [mbid])
        if not cur.fetchone():
            return jsonify({"ok": False, "message": "존재하지 않는 건물입니다."}), 404
        # 무료 캡: 담당 건물 수가 MAX_FREE_BUILDINGS 이상이면 추가 등록 차단(안내만, 결제 미도입)
        cur.execute("SELECT COUNT(*) c FROM agent_buildings WHERE agent_id = %s", [agent_id])
        if cur.fetchone()["c"] >= MAX_FREE_BUILDINGS:
            return jsonify({
                "ok": False,
                "message": f"무료 등록 가능 건물 수({MAX_FREE_BUILDINGS}개)를 초과했습니다. 추가 등록은 준비 중입니다.",
            }), 400
        try:
            cur.execute(
                "INSERT INTO agent_buildings (agent_id, master_building_id) VALUES (%s, %s)",
                [agent_id, mbid],
            )
            conn.commit()
        except psycopg2_errors.UniqueViolation:
            conn.rollback()
            return jsonify({"ok": False, "message": "이미 등록된 단지입니다."}), 400
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/agent/buildings/<int:mbid>", methods=["DELETE"])
@require_agent
def agent_building_delete(mbid):
    agent_id = session["agent_id"]
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "DELETE FROM agent_buildings WHERE agent_id = %s AND master_building_id = %s",
            [agent_id, mbid],
        )
        deleted = cur.rowcount
        conn.commit()
    finally:
        cur.close()
        conn.close()
    if not deleted:
        return jsonify({"ok": False, "message": "등록되지 않은 단지입니다."}), 404
    return jsonify({"ok": True})


@app.route("/api/agent/buildings/<int:mbid>/counts", methods=["PUT"])
@require_agent
def agent_building_counts(mbid):
    agent_id = session["agent_id"]
    data = request.get_json(force=True, silent=True) or {}
    fields = ["sale_count", "jeonse_count", "wolse_count", "shortterm_count"]
    values = []
    for k in fields:
        v = data.get(k)
        if not isinstance(v, int) or isinstance(v, bool) or v < 0:
            return jsonify({"ok": False, "message": "매물 수는 0 이상의 정수여야 합니다."}), 400
        values.append(v)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE agent_buildings
            SET sale_count=%s, jeonse_count=%s, wolse_count=%s, shortterm_count=%s, updated_at=NOW()
            WHERE agent_id = %s AND master_building_id = %s
        """, values + [agent_id, mbid])
        updated = cur.rowcount
        conn.commit()
    finally:
        cur.close()
        conn.close()
    if not updated:
        return jsonify({"ok": False, "message": "등록되지 않은 단지입니다."}), 404
    return jsonify({"ok": True})


@app.route("/api/agent/buildings/search")
@require_agent
def agent_building_search():
    """단지관리 모달용 건물명 검색 — 이미 등록된 건물은 already_added 표시."""
    agent_id = session["agent_id"]
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify({"items": []})
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT mb.id, mb.building_name, mb.lodging_type, mb.sgg_text, mb.umd_nm,
                   (ab.id IS NOT NULL) AS already_added
            FROM master_buildings mb
            LEFT JOIN agent_buildings ab
              ON ab.master_building_id = mb.id AND ab.agent_id = %s
            WHERE mb.building_name ILIKE %s AND mb.building_name <> '-'
            ORDER BY mb.building_name
            LIMIT 20
        """, [agent_id, f"%{q}%"])
        items = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return jsonify({"items": items})


@app.route("/api/agent/leads")
@require_agent
def agent_leads():
    """나에게 배정된 매물의뢰 목록 — routed_agent_id = 내 agent_id."""
    agent_id = session["agent_id"]
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT lr.id, lr.deal_type, lr.desired_price, lr.contact_phone,
                   lr.routed_reason, lr.status,
                   mb.id AS master_building_id, mb.building_name,
                   to_char(lr.created_at, 'YYYY-MM-DD HH24:MI') AS created_at
            FROM listing_requests lr
            JOIN master_buildings mb ON mb.id = lr.master_building_id
            WHERE lr.routed_agent_id = %s
            ORDER BY lr.created_at DESC, lr.id DESC
            LIMIT 200
        """, [agent_id])
        items = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return jsonify({"items": items})


# 매물의뢰 상태 진행 순서 — 순방향만 허용(건너뛰기 가능, 역방향 금지). 관리자 API로는 변경 불가.
_LEAD_STATUS_ORDER = {"submitted": 0, "in_progress": 1, "done": 2}


@app.route("/api/agent/leads/<int:lead_id>/status", methods=["PUT"])
@require_agent
def agent_lead_update_status(lead_id):
    """내게 배정된 매물의뢰의 상태 변경 — submitted → in_progress → done 순방향만."""
    agent_id = session["agent_id"]
    data = request.get_json(force=True, silent=True) or {}
    new_status = (data.get("status") or "").strip()
    if new_status not in _LEAD_STATUS_ORDER:
        return jsonify({"ok": False, "message": "잘못된 상태값입니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        # 행 잠금(FOR UPDATE)으로 동시 요청을 직렬화 — 순방향-only 규칙이 경쟁 상황에서도 깨지지 않게 한다.
        cur.execute("SELECT routed_agent_id, status FROM listing_requests WHERE id = %s FOR UPDATE", [lead_id])
        row = cur.fetchone()
        if not row:
            conn.rollback()
            return jsonify({"ok": False, "message": "존재하지 않는 의뢰입니다."}), 404
        if row["routed_agent_id"] != agent_id:
            # 다른 중개사에게 배정된 건은 수정 불가.
            conn.rollback()
            return jsonify({"ok": False, "message": "권한이 없습니다."}), 403
        cur_rank = _LEAD_STATUS_ORDER.get(row["status"], 0)
        if _LEAD_STATUS_ORDER[new_status] <= cur_rank:
            conn.rollback()
            return jsonify({"ok": False, "message": "상태는 순방향(신규→처리중→완료)으로만 변경할 수 있습니다."}), 400
        cur.execute("UPDATE listing_requests SET status = %s WHERE id = %s", [new_status, lead_id])
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "status": new_status})


# ============================================================
# 매물의뢰 접수 + 중개사 라우팅
#   ① exclusive: 그 건물을 agent_buildings에 등록한 approved 중개사 (최근 갱신순 1명)
#   ② region   : 같은 sgg_text 지역에 건물을 등록한 approved 중개사들 (전원 SMS, 대표 1명 저장)
#   ③ house    : 하우스 계정(홈스퀘어부동산중개법인)
# ============================================================

_LISTING_DEAL_TYPES = {"매매", "전세", "월세", "단기임대"}
_PHONE_RE = re.compile(r"^0\d{1,2}-?\d{3,4}-?\d{4}$")
_HOUSE_OFFICE_NAME = "홈스퀘어부동산중개법인"


@app.route("/api/listing-requests", methods=["POST"])
@limiter.limit("5 per hour")
def create_listing_request():
    user = current_user()
    if not user:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401

    data = request.get_json(force=True, silent=True) or {}
    try:
        mb_id = int(data.get("master_building_id") or 0)
    except (TypeError, ValueError):
        mb_id = 0
    deal_type = (data.get("deal_type") or "").strip()
    desired_price = (data.get("desired_price") or "").strip()[:100]
    contact_phone = (data.get("contact_phone") or "").strip()

    # 거래유형별 구조화 희망가(만원 단위 정수, 선택) — 단기임대는 자유텍스트만 사용.
    # 값이 전달됐는데 정수가 아니거나 범위(1~1억 만원)를 벗어나면 400 (조용한 유실 방지).
    def _parse_krw(field, allowed):
        v = data.get(field)
        if v is None or v == "":
            return None, None
        if not allowed:
            return None, f"{field}는 이 거래유형에서 사용할 수 없습니다."
        try:
            n = int(v)
        except (TypeError, ValueError):
            return None, "희망가는 만원 단위 숫자로 입력해주세요."
        if not (0 < n <= 100_000_000):
            return None, "희망가 숫자 범위가 올바르지 않습니다. (1~1억 만원)"
        return n, None
    price_krw, err1 = _parse_krw("price_krw", deal_type in ("매매", "전세", "월세"))
    monthly_rent_krw, err2 = _parse_krw("monthly_rent_krw", deal_type == "월세")
    if err1 or err2:
        return jsonify({"ok": False, "message": err1 or err2}), 400

    if not mb_id:
        return jsonify({"ok": False, "message": "건물 정보가 없습니다."}), 400
    if deal_type not in _LISTING_DEAL_TYPES:
        return jsonify({"ok": False, "message": "거래유형은 매매/전세/월세/단기임대 중 하나여야 합니다."}), 400
    if not _PHONE_RE.match(contact_phone):
        return jsonify({"ok": False, "message": "연락처 형식이 올바르지 않습니다. 예) 010-1234-5678"}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, building_name, sgg_text FROM master_buildings WHERE id = %s", [mb_id])
        bld = cur.fetchone()
        if not bld:
            return jsonify({"ok": False, "message": "등록되지 않은 건물입니다."}), 404

        routed_agent_id = None
        routed_reason = None
        notify_agents = []  # [{id, phone, office_name}] — SMS 수신 대상 (첫 번째가 배정자)

        # ① 전속(exclusive): 그 건물 담당 approved 중개사들.
        #    선정 기준을 건물카드 노출 로직과 동일하게 priority_score DESC → RANDOM() 으로 통일.
        #    첫 번째 1명만 routed_agent_id(상태변경 권한), 나머지는 참고용 알림만 발송.
        cur.execute("""
            SELECT a.id, a.phone, a.office_name
            FROM agent_buildings ab
            JOIN agents a ON a.id = ab.agent_id AND a.status = 'approved'
            WHERE ab.master_building_id = %s
              AND COALESCE(a.is_visible, TRUE)
            ORDER BY COALESCE(a.priority_score, 0) DESC, RANDOM()
        """, [mb_id])
        rows = cur.fetchall()
        if rows:
            routed_agent_id = rows[0]["id"]
            routed_reason = "exclusive"
            notify_agents = [dict(r) for r in rows]
        else:
            # ② 지역(region): 같은 시군구에 건물을 등록한 approved 중개사들 (하우스 계정 제외)
            sgg = (bld["sgg_text"] or "").strip()
            if sgg:
                cur.execute("""
                    SELECT a.id, a.phone, a.office_name, MAX(ab.updated_at) AS last_active
                    FROM agent_buildings ab
                    JOIN agents a ON a.id = ab.agent_id AND a.status = 'approved'
                    JOIN master_buildings mb ON mb.id = ab.master_building_id
                    WHERE mb.sgg_text = %s AND a.office_name <> %s
                      AND COALESCE(a.is_visible, TRUE)
                    GROUP BY a.id, a.phone, a.office_name
                    ORDER BY last_active DESC NULLS LAST
                """, [sgg, _HOUSE_OFFICE_NAME])
                rows = cur.fetchall()
                if rows:
                    routed_agent_id = rows[0]["id"]
                    routed_reason = "region"
                    notify_agents = [dict(r) for r in rows]
            if routed_agent_id is None:
                # ③ 하우스 계정 — 없으면 routed_agent_id NULL로라도 접수는 저장
                routed_reason = "house"
                cur.execute("""
                    SELECT id, phone, office_name FROM agents
                    WHERE office_name = %s AND status = 'approved'
                    ORDER BY id LIMIT 1
                """, [_HOUSE_OFFICE_NAME])
                row = cur.fetchone()
                if row:
                    routed_agent_id = row["id"]
                    notify_agents = [dict(row)]

        cur.execute("""
            INSERT INTO listing_requests
                (user_id, master_building_id, deal_type, desired_price, contact_phone,
                 routed_agent_id, routed_reason, price_krw, monthly_rent_krw)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, [user["id"], mb_id, deal_type, desired_price or None, contact_phone,
              routed_agent_id, routed_reason, price_krw, monthly_rent_krw])
        req_id = cur.fetchone()["id"]
        conn.commit()
    finally:
        cur.close()
        conn.close()

    # SMS 알림 — 실패해도 접수 자체는 성공 처리 (send_sms는 예외를 던지지 않음)
    sms_body = (
        f"[홈앤스테이] 매물의뢰 접수 — {bld['building_name']} / {deal_type}"
        + (f" / 희망가 {desired_price}" if desired_price else "")
        + f" / 연락처 {contact_phone}"
    )
    sms_results = []
    for ag in notify_agents:
        # 배정자(routed_agent_id)에게는 배정 문자, 나머지 담당중개사에게는 참고용 문자만.
        # (상태변경 권한은 배정자 1명에게만 있음)
        if ag["id"] == routed_agent_id:
            body = sms_body + " / 대시보드에서 확인해주세요"
        else:
            body = "[참고용] " + sms_body + " / 다른 담당중개사에게 배정되었습니다"
        if ag.get("phone"):
            sent, msg = send_sms(ag["phone"], body)
        else:
            sent, msg = False, "중개사 전화번호 없음"
        sms_results.append({"agent_id": ag["id"], "sent": sent, "message": msg})

    return jsonify({
        "ok": True, "id": req_id,
        "routed_reason": routed_reason, "routed_agent_id": routed_agent_id,
        "notified": len([r for r in sms_results if r["sent"]]),
    })


@app.route("/api/listing-requests/mine")
def my_listing_requests():
    """내가 접수한 매물의뢰 목록 — 마이페이지 '매물의뢰 현황'용 (건물명/거래유형/상태)."""
    user = current_user()
    if not user:
        return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT lr.id, lr.deal_type, lr.desired_price, lr.status,
                   to_char(lr.created_at, 'YYYY-MM-DD') AS created_date,
                   mb.id AS building_id, mb.building_name,
                   a.office_name AS agent_office_name, a.subdomain_slug AS agent_slug
            FROM listing_requests lr
            JOIN master_buildings mb ON mb.id = lr.master_building_id
            LEFT JOIN agents a ON a.id = lr.routed_agent_id
            WHERE lr.user_id = %s
            ORDER BY lr.created_at DESC
        """, [user["id"]])
        items = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "items": items})


# ============================================================
# 운영업체(operators) 로그인 — 승인된 운영업체만. agent 로그인과 같은 패턴.
# 세션에 operator_id 저장. require_operator 로 보호.
# ============================================================

def require_operator(f):
    """세션에 operator_id가 없으면 차단한다.
    /api/* 요청은 401 JSON, 그 외는 /operator/login으로 리다이렉트."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("operator_id"):
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
            return redirect("/operator/login")
        return f(*args, **kwargs)
    return wrapper


@app.route("/operator/login")
def operator_login_page():
    return _serve_static_html("operator_login.html")


@app.route("/api/operator/login", methods=["POST"])
@limiter.limit("5 per minute; 20 per hour")
def operator_login():
    """operators 테이블 기반 이메일/비밀번호 로그인. status='approved'만 허용.
    실패 시 이메일 존재 여부를 드러내지 않도록 통일된 메시지로 401을 반환한다."""
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip()
    password = data.get("password") or ""
    fail = jsonify({"ok": False, "message": "이메일 또는 비밀번호가 올바르지 않습니다."}), 401
    if not email or not password:
        return fail

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, password_hash, status FROM operators WHERE LOWER(email) = LOWER(%s)",
            (email,),
        )
        row = cur.fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], password):
            return fail
        if row["status"] != "approved":
            return jsonify({"ok": False, "message": "승인된 운영지원업체 계정이 아닙니다."}), 403
    finally:
        cur.close()
        conn.close()

    session["operator_id"] = row["id"]
    session.permanent = True
    return jsonify({"ok": True})


@app.route("/api/operator/logout", methods=["POST"])
def operator_logout():
    session.pop("operator_id", None)
    return jsonify({"ok": True})


@app.route("/api/operator/password", methods=["PUT"])
@require_operator
@limiter.limit("5 per minute; 20 per hour")
def operator_change_password():
    """운영업체 비밀번호 변경 — 현재 비밀번호 확인 후 교체 (agent와 같은 패턴)."""
    operator_id = session.get("operator_id")
    data = request.get_json(force=True, silent=True) or {}
    current_pw = data.get("current_password") or ""
    new_pw = data.get("new_password") or ""
    if len(new_pw) < 8:
        return jsonify({"ok": False, "message": "새 비밀번호는 8자 이상이어야 합니다."}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT password_hash FROM operators WHERE id = %s", (operator_id,))
        row = cur.fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], current_pw):
            return jsonify({"ok": False, "message": "현재 비밀번호가 올바르지 않습니다."}), 401
        cur.execute(
            "UPDATE operators SET password_hash = %s WHERE id = %s",
            (generate_password_hash(new_pw), operator_id),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


# ---- 운영업체 공개 프로필 + 대시보드(본인 데이터 관리) — agent와 동일 패턴 ----
# 차이점: 매물 개수(매매/전세/월세/단기임대) 개념 없음. "담당 단지 + 메모(note)"만 관리.

@app.route("/operator/<slug>")
def operator_profile_page(slug):
    """운영업체 공개 프로필 페이지. Flask는 정적 룰(/operator/login, /operator/dashboard)을 우선 매칭하므로 충돌 없음."""
    return _serve_static_html("operator_profile.html")


@app.route("/api/operator/profile/<slug>")
def operator_public_profile(slug):
    """운영업체 공개 프로필 API — 인증 불필요. approved 상태만 노출."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, company_name, owner_name, category, phone, photo_url, intro_text
            FROM operators
            WHERE subdomain_slug = %s AND status = 'approved'
        """, [slug])
        op = cur.fetchone()
        if not op:
            return jsonify({"error": "not found"}), 404
        cur.execute("""
            SELECT ob.master_building_id, mb.building_name, mb.lodging_type, ob.note
            FROM operator_buildings ob
            JOIN master_buildings mb ON mb.id = ob.master_building_id
            WHERE ob.operator_id = %s
            ORDER BY mb.building_name
        """, [op["id"]])
        buildings = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return jsonify({
        "company_name": op["company_name"],
        "owner_name": op["owner_name"],
        "category": op["category"],
        "phone": op["phone"],
        "photo_url": op["photo_url"],
        "intro_text": op["intro_text"],
        "buildings": buildings,
        "building_count": len(buildings),
    })


@app.route("/operator/dashboard")
@require_operator
def operator_dashboard_page():
    return _serve_static_html("operator_dashboard.html")


@app.route("/api/operator/me")
@require_operator
def operator_me():
    operator_id = session["operator_id"]
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT company_name, owner_name, category, phone, photo_url, intro_text, subdomain_slug,
                   COALESCE(is_visible, TRUE) AS is_visible
            FROM operators WHERE id = %s
        """, [operator_id])
        me = cur.fetchone()
        if not me:
            return jsonify({"ok": False, "message": "계정을 찾을 수 없습니다."}), 404
        cur.execute("""
            SELECT ob.master_building_id, mb.building_name, mb.lodging_type, ob.note
            FROM operator_buildings ob
            JOIN master_buildings mb ON mb.id = ob.master_building_id
            WHERE ob.operator_id = %s
            ORDER BY mb.building_name
        """, [operator_id])
        buildings = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    out = dict(me)
    out["buildings"] = buildings
    return jsonify(out)


@app.route("/api/operator/me", methods=["PUT"])
@require_operator
def operator_me_update():
    """phone / photo_url / intro_text 부분 업데이트 — 전달된 키만 수정 (agent와 동일 패턴)."""
    operator_id = session["operator_id"]
    data = request.get_json(force=True, silent=True) or {}
    allowed = ["phone", "photo_url", "intro_text"]
    sets, params = [], []
    for k in allowed:
        if k in data:
            v = data.get(k)
            if v is not None and not isinstance(v, str):
                return jsonify({"ok": False, "message": f"{k} 값이 올바르지 않습니다."}), 400
            v = (v or "").strip() or None
            if k == "phone":
                # 하이픈 유무와 무관하게 숫자만 저장 + 자릿수 검증 (표시할 때 재포맷)
                v = _digits_only(v) or None
                if v and not _validate_phone_digits(v):
                    return jsonify({"ok": False, "message": "전화번호 형식이 올바르지 않습니다. (숫자 10~11자리)"}), 400
            if k == "photo_url" and v and not (v.startswith("http://") or v.startswith("https://")):
                return jsonify({"ok": False, "message": "사진 URL은 http(s)://로 시작해야 합니다."}), 400
            sets.append(f"{k} = %s")
            params.append(v)
    if not sets:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다."}), 400
    params.append(operator_id)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"UPDATE operators SET {', '.join(sets)} WHERE id = %s", params)
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/operator/visibility", methods=["PUT"])
@require_operator
def operator_visibility_update():
    """노출 여부 토글 — 본인 세션 기준. is_visible만 갱신 (agent와 동일 패턴)."""
    data = request.get_json(force=True, silent=True) or {}
    v = data.get("is_visible")
    if not isinstance(v, bool):
        return jsonify({"ok": False, "message": "is_visible 값은 true/false여야 합니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE operators SET is_visible = %s WHERE id = %s", (v, session["operator_id"]))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "is_visible": v})


# ============================================================
# 대출상담사 로그인/대시보드 — agent/operator와 동일 패턴.
# 세션에 loan_consultant_id 저장. require_loan_consultant 로 보호.
# ============================================================

def require_loan_consultant(f):
    """세션에 loan_consultant_id가 없으면 차단한다.
    /api/* 요청은 401 JSON, 그 외는 /loan-consultant/login으로 리다이렉트."""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("loan_consultant_id"):
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "message": "로그인이 필요합니다."}), 401
            return redirect("/loan-consultant/login")
        return f(*args, **kwargs)
    return wrapper


@app.route("/loan-consultant/login")
def loan_consultant_login_page():
    return _serve_static_html("loan_consultant_login.html")


@app.route("/api/loan-consultant/login", methods=["POST"])
@limiter.limit("5 per minute; 20 per hour")
def loan_consultant_login():
    """loan_consultants 테이블 기반 이메일/비밀번호 로그인. status='approved'만 허용."""
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip()
    password = data.get("password") or ""
    fail = jsonify({"ok": False, "message": "이메일 또는 비밀번호가 올바르지 않습니다."}), 401
    if not email or not password:
        return fail

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, password_hash, status FROM loan_consultants WHERE LOWER(email) = LOWER(%s) "
            "ORDER BY approved_at DESC NULLS LAST, id DESC LIMIT 1",
            (email,),
        )
        row = cur.fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], password):
            return fail
        if row["status"] != "approved":
            return jsonify({"ok": False, "message": "승인된 대출상담사 계정이 아닙니다."}), 403
    finally:
        cur.close()
        conn.close()

    session["loan_consultant_id"] = row["id"]
    session.permanent = True
    return jsonify({"ok": True})


@app.route("/api/loan-consultant/logout", methods=["POST"])
def loan_consultant_logout():
    session.pop("loan_consultant_id", None)
    return jsonify({"ok": True})


@app.route("/api/loan-consultant/password", methods=["PUT"])
@require_loan_consultant
@limiter.limit("5 per minute; 20 per hour")
def loan_consultant_change_password():
    """대출상담사 비밀번호 변경 — 현재 비밀번호 확인 후 교체 (agent와 같은 패턴)."""
    lc_id = session.get("loan_consultant_id")
    data = request.get_json(force=True, silent=True) or {}
    current_pw = data.get("current_password") or ""
    new_pw = data.get("new_password") or ""
    if len(new_pw) < 8:
        return jsonify({"ok": False, "message": "새 비밀번호는 8자 이상이어야 합니다."}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT password_hash FROM loan_consultants WHERE id = %s", (lc_id,))
        row = cur.fetchone()
        if not row or not row["password_hash"] or not check_password_hash(row["password_hash"], current_pw):
            return jsonify({"ok": False, "message": "현재 비밀번호가 올바르지 않습니다."}), 401
        cur.execute(
            "UPDATE loan_consultants SET password_hash = %s WHERE id = %s",
            (generate_password_hash(new_pw), lc_id),
        )
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/loan-consultant/dashboard")
@require_loan_consultant
def loan_consultant_dashboard_page():
    return _serve_static_html("loan_consultant_dashboard.html")


@app.route("/api/loan-consultant/me")
@require_loan_consultant
def loan_consultant_me():
    lc_id = session["loan_consultant_id"]
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT office_name, owner_name, phone, intro_text, subdomain_slug,
                   consultant_products, kakao_chat_url,
                   COALESCE(is_visible, TRUE) AS is_visible
            FROM loan_consultants WHERE id = %s
        """, [lc_id])
        me = cur.fetchone()
        if not me:
            return jsonify({"ok": False, "message": "계정을 찾을 수 없습니다."}), 404
    finally:
        cur.close()
        conn.close()
    return jsonify(dict(me))


@app.route("/api/loan-consultant/me", methods=["PUT"])
@require_loan_consultant
def loan_consultant_me_update():
    """phone / intro_text / consultant_products / kakao_chat_url 부분 업데이트 — 전달된 키만 수정."""
    lc_id = session["loan_consultant_id"]
    data = request.get_json(force=True, silent=True) or {}
    sets, params = [], []
    # 상담 가능 상품 — 허용 목록 내 다중 선택, 콤마구분 텍스트로 저장
    if "consultant_products" in data:
        v = data.get("consultant_products")
        if v is None:
            v = []
        if not isinstance(v, list) or any(not isinstance(x, str) for x in v):
            return jsonify({"ok": False, "message": "상담 가능 상품 값이 올바르지 않습니다."}), 400
        invalid = [x for x in v if x not in LOAN_CONSULTANT_PRODUCTS]
        if invalid:
            return jsonify({"ok": False, "message": "허용되지 않은 상담 상품이 포함되어 있습니다."}), 400
        # 허용 목록 순서대로 정렬해 저장 (생활숙박시설 담보대출이 항상 앞)
        ordered = [p for p in LOAN_CONSULTANT_PRODUCTS if p in v]
        sets.append("consultant_products = %s")
        params.append(",".join(ordered) or None)
    for k in ["phone", "intro_text", "kakao_chat_url"]:
        if k in data:
            v = data.get(k)
            if v is not None and not isinstance(v, str):
                return jsonify({"ok": False, "message": f"{k} 값이 올바르지 않습니다."}), 400
            v = (v or "").strip() or None
            if k == "phone":
                # 하이픈 유무와 무관하게 숫자만 저장 + 자릿수 검증 (표시할 때 재포맷)
                v = _digits_only(v) or None
                if v and not _validate_phone_digits(v):
                    return jsonify({"ok": False, "message": "전화번호 형식이 올바르지 않습니다. (숫자 10~11자리)"}), 400
            if k == "kakao_chat_url" and v and not (v.startswith("http://") or v.startswith("https://")):
                return jsonify({"ok": False, "message": "카카오톡 상담 링크는 http(s)://로 시작하는 URL이어야 합니다."}), 400
            if k == "intro_text" and v and len(v) > 100:
                return jsonify({"ok": False, "message": "한줄소개는 100자 이내로 입력해주세요."}), 400
            sets.append(f"{k} = %s")
            params.append(v)
    if not sets:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다."}), 400
    params.append(lc_id)
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"UPDATE loan_consultants SET {', '.join(sets)} WHERE id = %s", params)
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/loan-consultant/visibility", methods=["PUT"])
@require_loan_consultant
def loan_consultant_visibility_update():
    """노출 여부 토글 — 본인 세션 기준. is_visible만 갱신 (agent와 동일 패턴)."""
    data = request.get_json(force=True, silent=True) or {}
    v = data.get("is_visible")
    if not isinstance(v, bool):
        return jsonify({"ok": False, "message": "is_visible 값은 true/false여야 합니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE loan_consultants SET is_visible = %s WHERE id = %s", (v, session["loan_consultant_id"]))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "is_visible": v})


@app.route("/api/operator/buildings", methods=["POST"])
@require_operator
def operator_building_add():
    operator_id = session["operator_id"]
    data = request.get_json(force=True, silent=True) or {}
    try:
        mbid = int(data.get("master_building_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "message": "건물 ID가 올바르지 않습니다."}), 400
    note = data.get("note")
    if note is not None and not isinstance(note, str):
        return jsonify({"ok": False, "message": "메모 값이 올바르지 않습니다."}), 400
    note = (note or "").strip() or None
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT 1 FROM master_buildings WHERE id = %s", [mbid])
        if not cur.fetchone():
            return jsonify({"ok": False, "message": "존재하지 않는 건물입니다."}), 404
        # 무료 캡: 담당 건물 수가 MAX_FREE_BUILDINGS 이상이면 추가 등록 차단(안내만, 결제 미도입)
        cur.execute("SELECT COUNT(*) c FROM operator_buildings WHERE operator_id = %s", [operator_id])
        if cur.fetchone()["c"] >= MAX_FREE_BUILDINGS:
            return jsonify({
                "ok": False,
                "message": f"무료 등록 가능 건물 수({MAX_FREE_BUILDINGS}개)를 초과했습니다. 추가 등록은 준비 중입니다.",
            }), 400
        try:
            cur.execute(
                "INSERT INTO operator_buildings (operator_id, master_building_id, note) VALUES (%s, %s, %s)",
                [operator_id, mbid, note],
            )
            conn.commit()
        except psycopg2_errors.UniqueViolation:
            conn.rollback()
            return jsonify({"ok": False, "message": "이미 등록된 단지입니다."}), 400
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/operator/buildings/<int:mbid>", methods=["DELETE"])
@require_operator
def operator_building_delete(mbid):
    operator_id = session["operator_id"]
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "DELETE FROM operator_buildings WHERE operator_id = %s AND master_building_id = %s",
            [operator_id, mbid],
        )
        deleted = cur.rowcount
        conn.commit()
    finally:
        cur.close()
        conn.close()
    if not deleted:
        return jsonify({"ok": False, "message": "등록되지 않은 단지입니다."}), 404
    return jsonify({"ok": True})


@app.route("/api/operator/buildings/<int:mbid>/note", methods=["PUT"])
@require_operator
def operator_building_note(mbid):
    operator_id = session["operator_id"]
    data = request.get_json(force=True, silent=True) or {}
    note = data.get("note")
    if note is not None and not isinstance(note, str):
        return jsonify({"ok": False, "message": "메모 값이 올바르지 않습니다."}), 400
    note = (note or "").strip() or None
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE operator_buildings
            SET note = %s, updated_at = NOW()
            WHERE operator_id = %s AND master_building_id = %s
        """, [note, operator_id, mbid])
        updated = cur.rowcount
        conn.commit()
    finally:
        cur.close()
        conn.close()
    if not updated:
        return jsonify({"ok": False, "message": "등록되지 않은 단지입니다."}), 404
    return jsonify({"ok": True})


@app.route("/api/operator/buildings/search")
@require_operator
def operator_building_search():
    """단지관리 모달용 건물명 검색 — 이미 등록된 건물은 already_added 표시 (agent와 동일)."""
    operator_id = session["operator_id"]
    q = (request.args.get("q") or "").strip()
    if len(q) < 2:
        return jsonify({"items": []})
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT mb.id, mb.building_name, mb.lodging_type, mb.sgg_text, mb.umd_nm,
                   (ob.id IS NOT NULL) AS already_added
            FROM master_buildings mb
            LEFT JOIN operator_buildings ob
              ON ob.master_building_id = mb.id AND ob.operator_id = %s
            WHERE mb.building_name ILIKE %s AND mb.building_name <> '-'
            ORDER BY mb.building_name
            LIMIT 20
        """, [operator_id, f"%{q}%"])
        items = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return jsonify({"items": items})


# ---- 지도 좌표 채우기 (배포 후 운영 DB에 좌표 주입용) ----
# 에이전트/개발자는 운영(production) DB에 직접 쓸 수 없으므로, 개발에서 확보한
# 좌표를 data/building_coords.json 으로 배포에 포함시키고, 관리자가 이 API를
# 호출하면 배포된 앱(=운영 DB 연결)이 id 기준으로 lat/lng 를 채운다.
# id 기준 UPDATE라 여러 번 눌러도 결과가 동일한 idempotent 작업이다.
_COORDS_JSON_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "building_coords.json"
)


def _load_coords_seed():
    """data/building_coords.json → [(id, lat, lng), ...] (형식 오류 행은 건너뜀)."""
    with open(_COORDS_JSON_PATH, encoding="utf-8") as fp:
        rows = json.load(fp)
    out = []
    for r in rows:
        try:
            out.append((int(r["id"]), float(r["lat"]), float(r["lng"])))
        except (KeyError, TypeError, ValueError):
            continue
    return out


@app.route("/api/admin/geocode/status")
@require_admin
def admin_geocode_status():
    """좌표 확보 현황 + 마지막 실행 기록 반환 (관리자 화면 표시용)."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT COUNT(*) AS total,
                   COUNT(*) FILTER (WHERE lat IS NOT NULL AND lng IS NOT NULL) AS with_geo
            FROM master_buildings
        """)
        s = cur.fetchone()
        cur.execute("SELECT value, updated_at FROM app_meta WHERE key = 'geocode_last_run'")
        meta = cur.fetchone()
    finally:
        cur.close()
        conn.close()
    try:
        seed_count = len(_load_coords_seed())
    except Exception:
        seed_count = None
    return jsonify({
        "ok": True,
        "total": s["total"],
        "with_geo": s["with_geo"],
        "seed_count": seed_count,
        "last_run": (meta["value"] if meta else None),
        "last_run_at": (
            meta["updated_at"].strftime("%Y-%m-%d %H:%M")
            if meta and meta["updated_at"] else None
        ),
    })


@app.route("/api/admin/geocode", methods=["POST"])
@require_admin
@limiter.limit("6 per hour")
def admin_geocode_run():
    """data/building_coords.json 의 좌표를 master_buildings 에 id 기준으로 주입.
    값이 실제로 바뀐 행만 세어 반환(재실행 시 0건 → 이미 최신)."""
    try:
        seed = _load_coords_seed()
    except FileNotFoundError:
        return jsonify({"ok": False, "message": "좌표 데이터 파일(data/building_coords.json)을 찾을 수 없습니다."}), 500
    except (json.JSONDecodeError, ValueError):
        return jsonify({"ok": False, "message": "좌표 데이터 파일 형식이 올바르지 않습니다(JSON 파싱 실패)."}), 500
    if not seed:
        return jsonify({"ok": False, "message": "좌표 데이터가 비어 있습니다."}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        # 값이 실제로 달라지는 행만 UPDATE → 재실행 시 0건이 나와 idempotent 확인 가능
        execute_values(cur, """
            UPDATE master_buildings AS m
            SET lat = v.lat, lng = v.lng
            FROM (VALUES %s) AS v(id, lat, lng)
            WHERE m.id = v.id
              AND (m.lat IS DISTINCT FROM v.lat OR m.lng IS DISTINCT FROM v.lng)
        """, seed, template="(%s, %s::double precision, %s::double precision)")
        updated = cur.rowcount
        cur.execute("""
            SELECT COUNT(*) AS total,
                   COUNT(*) FILTER (WHERE lat IS NOT NULL AND lng IS NOT NULL) AS with_geo
            FROM master_buildings
        """)
        s = cur.fetchone()
        summary = f"{updated}건 갱신 (좌표 확보 {s['with_geo']}/{s['total']}건)"
        cur.execute("""
            INSERT INTO app_meta (key, value, updated_at)
            VALUES ('geocode_last_run', %s, NOW())
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
        """, (summary,))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({
        "ok": True,
        "updated": updated,
        "with_geo": s["with_geo"],
        "total": s["total"],
        "message": summary,
    })


# ---- 건축정보(표제부) 채우기 (배포 후 운영 DB에 표제부 주입용) ----
# 좌표 채우기와 완전히 같은 패턴: 개발에서 백필한 표제부 값을
# data/building_title_info.json 으로 배포에 포함시키고, 관리자가 이 API를
# 호출하면 배포된 앱(=운영 DB 연결)이 id 기준으로 컬럼을 채운다.
_TITLE_INFO_JSON_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "data", "building_title_info.json"
)


def _title_info_int(v):
    return None if v is None else int(v)


def _title_info_float(v):
    return None if v is None else float(v)


def _title_info_text(v):
    return None if v is None else str(v)


def _load_title_info_seed():
    """data/building_title_info.json → [(id, use_apr_day, ...), ...] (형식 오류 행은 건너뜀)."""
    with open(_TITLE_INFO_JSON_PATH, encoding="utf-8") as fp:
        rows = json.load(fp)
    out = []
    for r in rows:
        try:
            out.append((
                int(r["id"]),
                _title_info_text(r.get("use_apr_day")),
                _title_info_int(r.get("tot_pkng_cnt")),
                _title_info_int(r.get("grnd_flr_cnt")),
                _title_info_int(r.get("ugrnd_flr_cnt")),
                _title_info_float(r.get("tot_area")),
                _title_info_float(r.get("plat_area")),
                _title_info_int(r.get("hhld_cnt")),
                _title_info_text(r.get("strct_nm")),
                _title_info_text(r.get("mgm_bldrgst_pk")),
            ))
        except (KeyError, TypeError, ValueError):
            continue
    return out


def _title_info_counts(cur):
    cur.execute("""
        SELECT COUNT(*) AS total,
               COUNT(*) FILTER (WHERE use_apr_day IS NOT NULL
                                   OR grnd_flr_cnt IS NOT NULL
                                   OR tot_area IS NOT NULL) AS with_title
        FROM master_buildings
    """)
    return cur.fetchone()


@app.route("/api/admin/backfill-title-info/status")
@require_admin
def admin_title_info_status():
    """표제부(건축정보) 확보 현황 + 마지막 실행 기록 반환 (관리자 화면 표시용)."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        s = _title_info_counts(cur)
        cur.execute("SELECT value, updated_at FROM app_meta WHERE key = 'title_info_last_run'")
        meta = cur.fetchone()
    finally:
        cur.close()
        conn.close()
    try:
        seed_count = len(_load_title_info_seed())
    except Exception:
        seed_count = None
    return jsonify({
        "ok": True,
        "total": s["total"],
        "with_title": s["with_title"],
        "seed_count": seed_count,
        "last_run": (meta["value"] if meta else None),
        "last_run_at": (
            meta["updated_at"].strftime("%Y-%m-%d %H:%M")
            if meta and meta["updated_at"] else None
        ),
    })


@app.route("/api/admin/backfill-title-info", methods=["POST"])
@require_admin
@limiter.limit("6 per hour")
def admin_title_info_run():
    """data/building_title_info.json 의 표제부 값을 master_buildings 에 id 기준으로 주입.
    값이 실제로 바뀐 행만 세어 반환(재실행 시 0건 → 이미 최신)."""
    try:
        seed = _load_title_info_seed()
    except FileNotFoundError:
        return jsonify({"ok": False, "message": "건축정보 데이터 파일(data/building_title_info.json)을 찾을 수 없습니다."}), 500
    except (json.JSONDecodeError, ValueError):
        return jsonify({"ok": False, "message": "건축정보 데이터 파일 형식이 올바르지 않습니다(JSON 파싱 실패)."}), 500
    if not seed:
        return jsonify({"ok": False, "message": "건축정보 데이터가 비어 있습니다."}), 400

    conn = get_conn()
    cur = conn.cursor()
    try:
        # 값이 실제로 달라지는 행만 UPDATE → 재실행 시 0건이 나와 idempotent 확인 가능
        execute_values(cur, """
            UPDATE master_buildings AS m
            SET use_apr_day = v.use_apr_day,
                tot_pkng_cnt = v.tot_pkng_cnt,
                grnd_flr_cnt = v.grnd_flr_cnt,
                ugrnd_flr_cnt = v.ugrnd_flr_cnt,
                tot_area = v.tot_area,
                plat_area = v.plat_area,
                hhld_cnt = v.hhld_cnt,
                strct_nm = v.strct_nm,
                mgm_bldrgst_pk = v.mgm_bldrgst_pk
            FROM (VALUES %s) AS v(id, use_apr_day, tot_pkng_cnt, grnd_flr_cnt,
                                  ugrnd_flr_cnt, tot_area, plat_area, hhld_cnt,
                                  strct_nm, mgm_bldrgst_pk)
            WHERE m.id = v.id
              AND (m.use_apr_day IS DISTINCT FROM v.use_apr_day
                   OR m.tot_pkng_cnt IS DISTINCT FROM v.tot_pkng_cnt
                   OR m.grnd_flr_cnt IS DISTINCT FROM v.grnd_flr_cnt
                   OR m.ugrnd_flr_cnt IS DISTINCT FROM v.ugrnd_flr_cnt
                   OR m.tot_area IS DISTINCT FROM v.tot_area
                   OR m.plat_area IS DISTINCT FROM v.plat_area
                   OR m.hhld_cnt IS DISTINCT FROM v.hhld_cnt
                   OR m.strct_nm IS DISTINCT FROM v.strct_nm
                   OR m.mgm_bldrgst_pk IS DISTINCT FROM v.mgm_bldrgst_pk)
        """, seed, template=(
            "(%s, %s::text, %s::integer, %s::integer, %s::integer, "
            "%s::double precision, %s::double precision, %s::integer, %s::text, %s::text)"
        ))
        updated = cur.rowcount
        s = _title_info_counts(cur)
        summary = f"{updated}건 갱신 (표제부 확보 {s['with_title']}/{s['total']}건)"
        cur.execute("""
            INSERT INTO app_meta (key, value, updated_at)
            VALUES ('title_info_last_run', %s, NOW())
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
        """, (summary,))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({
        "ok": True,
        "updated": updated,
        "with_title": s["with_title"],
        "total": s["total"],
        "message": summary,
    })


# ---- 실거래 동기화 (관리자 버튼) ----
# 배포된 앱(=운영 DB 연결)에서 sync_runner.py 를 '독립 프로세스'로 띄운다.
# - 요청은 즉시 202 반환(웹 타임아웃 방지), 진행상황은 app_meta('tx_sync_status')에 기록.
# - app_meta UPSERT의 WHERE 조건으로 중복 실행을 원자적으로 차단(멀티 워커에서도 안전).
# - 러너가 30초마다 하트비트(updated_at 갱신) → _SYNC_STALE_MIN 이상 끊기면
#   비정상 종료로 간주하고 재실행을 허용한다.
# - start_new_session=True 라 웹 워커가 재시작/강제종료돼도 러너는 계속 실행되고
#   완료/실패를 스스로 기록한다.
_SYNC_META_KEY = "tx_sync_status"
_SYNC_STALE_MIN = 5  # 하트비트(30초 주기)가 이 시간 이상 끊기면 stale


@app.route("/api/admin/sync-transactions", methods=["POST"])
@require_admin
@limiter.limit("2 per hour")
def admin_sync_run():
    """실거래 동기화 시작. 이미 실행 중이면 409. 시작되면 즉시 202 반환."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) AS c FROM transactions")
        tx_before = cur.fetchone()["c"]
        status = {
            "run_id": _secrets.token_hex(8),  # 실행 식별자 — 이전 러너가 새 실행 상태를 덮어쓰지 못하게 함
            "state": "running",
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": None,
            "tx_before": tx_before,
            "inserted": None,
            "error": None,
        }
        # 원자적 잠금(DB 기준이라 멀티 워커/인스턴스에서도 전역 일관):
        # - running 인데 하트비트가 살아있으면 차단(중복 실행 방지)
        # - 직전 실행이 성공(done)했으면 30분 재실행 금지(국토부 API 쿼터 보호 —
        #   메모리 기반 rate limit 과 달리 워커/인스턴스가 여러 개여도 전역 강제됨)
        # - 실패(failed)했거나 하트비트가 끊겼으면(비정상 종료) 즉시 재실행 허용
        cur.execute(f"""
            INSERT INTO app_meta (key, value, updated_at)
            VALUES (%s, %s, NOW())
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            WHERE ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'running'
                   OR app_meta.updated_at < NOW() - INTERVAL '{int(_SYNC_STALE_MIN)} minutes')
              AND ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'done'
                   OR app_meta.updated_at < NOW() - INTERVAL '30 minutes')
        """, (_SYNC_META_KEY, json.dumps(status, ensure_ascii=False)))
        acquired = cur.rowcount > 0
        prev_state = None
        if not acquired:
            cur.execute("SELECT value FROM app_meta WHERE key = %s", (_SYNC_META_KEY,))
            prev = cur.fetchone()
            try:
                prev_state = json.loads(prev["value"]).get("state") if prev and prev["value"] else None
            except (TypeError, ValueError):
                prev_state = None
        conn.commit()
    finally:
        cur.close()
        conn.close()

    if not acquired:
        if prev_state == "done":
            return jsonify({"ok": False, "message": "직전 동기화가 완료된 지 30분이 지나지 않았습니다. 잠시 후 다시 시도해 주세요."}), 429
        return jsonify({"ok": False, "message": "이미 동기화가 실행 중입니다. 완료 후 다시 시도해 주세요."}), 409

    base_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        proc = subprocess.Popen(
            [sys.executable, "-u", os.path.join(base_dir, "sync_runner.py")],
            cwd=base_dir, start_new_session=True,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        # 러너 종료 시 좀비 프로세스가 남지 않도록 백그라운드에서 회수
        threading.Thread(target=proc.wait, daemon=True).start()
    except Exception as e:
        status.update({"state": "failed", "error": f"러너 실행 실패: {e}"[:300],
                       "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO app_meta (key, value, updated_at) VALUES (%s, %s, NOW())
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            """, (_SYNC_META_KEY, json.dumps(status, ensure_ascii=False)))
            conn.commit()
        finally:
            cur.close()
            conn.close()
        return jsonify({"ok": False, "message": "동기화 프로세스를 시작하지 못했습니다."}), 500

    return jsonify({"ok": True, "message": "동기화를 시작했습니다.", "started_at": status["started_at"]}), 202


@app.route("/api/admin/sync-status")
@require_admin
def admin_sync_status():
    """실거래 동기화 진행상황 + 거래 데이터 현황(총 건수·최근 계약일)."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) AS c, MAX(deal_date) AS md FROM transactions")
        row = cur.fetchone()
        cur.execute("SELECT value, updated_at FROM app_meta WHERE key = %s", (_SYNC_META_KEY,))
        meta = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    status = None
    if meta and meta["value"]:
        try:
            status = json.loads(meta["value"])
        except (TypeError, ValueError):
            status = None

    running = bool(status and status.get("state") == "running")
    stale = False
    if running and meta["updated_at"]:
        age = (datetime.now() - meta["updated_at"]).total_seconds()
        if age > _SYNC_STALE_MIN * 60:
            running, stale = False, True

    inserted = status.get("inserted") if status else None
    if running and status.get("tx_before") is not None:
        # 실행 중에는 시작 시점 대비 증가분을 실시간 표시
        inserted = max(0, row["c"] - status["tx_before"])

    return jsonify({
        "ok": True,
        "running": running,
        "state": ("stale" if stale else (status.get("state") if status else None)),
        "started_at": (status.get("started_at") if status else None),
        "finished_at": (status.get("finished_at") if status else None),
        "inserted": inserted,
        "error": ((status.get("error") if status else None)
                  or ("이전 실행이 비정상 종료된 것으로 보입니다(장시간 응답 없음). 다시 실행할 수 있습니다." if stale else None)),
        "tx_total": row["c"],
        "max_deal_date": (row["md"].strftime("%Y-%m-%d") if hasattr(row["md"], "strftime") else row["md"]) if row["md"] else None,
    })


# ---- 과거 데이터 백필 (관리자 버튼) ----
# 2020-01 부터 현재까지 개월 수를 계산해 sync_batch --months 로 전체 기간을 재수집.
# 상태는 별도 키(tx_backfill_status)에 기록 — 일반 동기화와 독립적으로 진행/표시.
# 국토부 API 일일 쿼터 보호를 위해 성공(done) 후 24시간 재실행 금지(DB 기준 전역 강제).
_BACKFILL_META_KEY = "tx_backfill_status"
_BACKFILL_FROM = "2020-01"


def _warn_if_jobs_running_at_boot():
    """[재배포 추적용] 서버 부팅 시점에 실행 중(running)인 백필/동기화가 있으면 경고 로그.
    재배포 자체를 막을 수는 없지만, '재배포 때 백필이 돌고 있었다'는 기록을 남겨
    이후 실패/중단 원인 추적을 돕는다. 부팅을 절대 막지 않도록 예외는 전부 무시."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute("SELECT key, value, updated_at FROM app_meta WHERE key IN (%s, %s)",
                        (_SYNC_META_KEY, _BACKFILL_META_KEY))
            rows = cur.fetchall()
        finally:
            cur.close()
            conn.close()
        for r in rows:
            try:
                st = json.loads(r["value"]) if r["value"] else None
            except (TypeError, ValueError):
                st = None
            if st and st.get("state") == "running":
                label = "백필" if r["key"] == _BACKFILL_META_KEY else "실거래 동기화"
                print(f"[boot-warning] 서버 부팅 시점에 {label}이(가) 실행 중 상태였습니다 "
                      f"(run_id={st.get('run_id')}, 시작 {st.get('started_at')}, "
                      f"마지막 하트비트 {r['updated_at']}). 재배포로 프로세스가 중단됐을 수 있습니다.",
                      flush=True)
    except Exception as e:
        print(f"[boot-warning] 실행중 작업 확인 실패(무시): {e}", flush=True)


_warn_if_jobs_running_at_boot()  # gunicorn 부팅(모듈 임포트) 시점에 1회 실행


def _backfill_months():
    """2020-01 부터 이번 달까지의 개월 수 (예: 2026-07 → 79)."""
    now = datetime.now()
    return (now.year - 2020) * 12 + now.month


@app.route("/api/admin/sync-backfill", methods=["POST"])
@require_admin
@limiter.limit("6 per hour")
def admin_backfill_run():
    """과거 데이터 백필 시작. 실행 중이면 409, 24시간 내 완료 이력 있으면 429.
    '1일 1회' 강제는 DB 잠금(성공 done 후 24시간)으로만 한다 — 실패(failed)로
    끝난 경우엔 당일 재시도를 허용해야 하므로 요청 자체를 하루 1회로 막지 않는다."""
    months = _backfill_months()
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) AS c FROM transactions")
        tx_before = cur.fetchone()["c"]
        run_id = _secrets.token_hex(8)
        status = {
            "run_id": run_id,
            "state": "running",
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": None,
            "tx_before": tx_before,
            "inserted": None,
            "error": None,
            "months": months,
            "target_from": _BACKFILL_FROM,
            "log_file": f"logs/backfill_{run_id}.log",
        }
        # 원자적 잠금 — sync-transactions 와 동일 패턴, 단 done 차단은 24시간
        cur.execute(f"""
            INSERT INTO app_meta (key, value, updated_at)
            VALUES (%s, %s, NOW())
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            WHERE ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'running'
                   OR app_meta.updated_at < NOW() - INTERVAL '{int(_SYNC_STALE_MIN)} minutes')
              AND ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'done'
                   OR app_meta.updated_at < NOW() - INTERVAL '24 hours')
        """, (_BACKFILL_META_KEY, json.dumps(status, ensure_ascii=False)))
        acquired = cur.rowcount > 0
        prev_state = None
        if not acquired:
            cur.execute("SELECT value FROM app_meta WHERE key = %s", (_BACKFILL_META_KEY,))
            prev = cur.fetchone()
            try:
                prev_state = json.loads(prev["value"]).get("state") if prev and prev["value"] else None
            except (TypeError, ValueError):
                prev_state = None
        conn.commit()
    finally:
        cur.close()
        conn.close()

    if not acquired:
        if prev_state == "done":
            return jsonify({"ok": False, "message": "백필이 완료된 지 24시간이 지나지 않았습니다. 내일 다시 시도해 주세요."}), 429
        return jsonify({"ok": False, "message": "이미 백필이 실행 중입니다. 완료 후 다시 시도해 주세요."}), 409

    base_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        # 러너 출력(하위 sync_batch 출력 포함)을 파일로 남겨 실패 원인을 추적한다.
        log_dir = os.path.join(base_dir, "logs")
        os.makedirs(log_dir, exist_ok=True)
        log_fh = open(os.path.join(log_dir, f"backfill_{run_id}.log"), "a", encoding="utf-8")
        try:
            proc = subprocess.Popen(
                [sys.executable, "-u", os.path.join(base_dir, "sync_runner.py"),
                 "--meta-key", _BACKFILL_META_KEY, "--months", str(months),
                 "--progress-key", "tx_backfill_progress"],
                cwd=base_dir, start_new_session=True,
                stdout=log_fh, stderr=subprocess.STDOUT,
            )
        finally:
            log_fh.close()  # 자식이 fd 를 상속했으므로 부모 쪽 핸들은 닫아도 된다
        threading.Thread(target=proc.wait, daemon=True).start()
    except Exception as e:
        status.update({"state": "failed", "error": f"러너 실행 실패: {e}"[:300],
                       "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO app_meta (key, value, updated_at) VALUES (%s, %s, NOW())
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            """, (_BACKFILL_META_KEY, json.dumps(status, ensure_ascii=False)))
            conn.commit()
        finally:
            cur.close()
            conn.close()
        return jsonify({"ok": False, "message": "백필 프로세스를 시작하지 못했습니다."}), 500

    return jsonify({"ok": True, "message": f"과거 데이터 백필을 시작했습니다 ({_BACKFILL_FROM}~현재, {months}개월).",
                    "months": months, "started_at": status["started_at"]}), 202


@app.route("/api/admin/backfill-status")
@require_admin
def admin_backfill_status():
    """과거 데이터 백필 진행상황 (sync-status 와 동일한 응답 형태)."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) AS c, MIN(deal_date) AS mind FROM transactions")
        row = cur.fetchone()
        cur.execute("SELECT value, updated_at FROM app_meta WHERE key = %s", (_BACKFILL_META_KEY,))
        meta = cur.fetchone()
        # 오늘 사용한 RTMS API 호출 수 (sync_batch가 app_meta에 날짜별로 기록)
        from sync_batch import _daily_calls_today, MAX_DAILY_BACKFILL_CALLS
        api_calls_today = _daily_calls_today(cur)
    finally:
        cur.close()
        conn.close()

    status = None
    if meta and meta["value"]:
        try:
            status = json.loads(meta["value"])
        except (TypeError, ValueError):
            status = None

    running = bool(status and status.get("state") == "running")
    stale = False
    if running and meta["updated_at"]:
        age = (datetime.now() - meta["updated_at"]).total_seconds()
        if age > _SYNC_STALE_MIN * 60:
            running, stale = False, True

    inserted = status.get("inserted") if status else None
    if running and status.get("tx_before") is not None:
        inserted = max(0, row["c"] - status["tx_before"])

    return jsonify({
        "ok": True,
        "running": running,
        "state": ("stale" if stale else (status.get("state") if status else None)),
        "started_at": (status.get("started_at") if status else None),
        "finished_at": (status.get("finished_at") if status else None),
        "inserted": inserted,
        "months": (status.get("months") if status else None),
        "target_from": (status.get("target_from") if status else None) or _BACKFILL_FROM,
        "error": ((status.get("error") if status else None)
                  or ("이전 실행이 비정상 종료된 것으로 보입니다(장시간 응답 없음). 다시 실행할 수 있습니다." if stale else None)),
        "tx_total": row["c"],
        "api_calls_today": api_calls_today,
        "api_calls_limit": MAX_DAILY_BACKFILL_CALLS,
        "min_deal_date": (row["mind"].strftime("%Y-%m-%d") if hasattr(row["mind"], "strftime") else row["mind"]) if row["mind"] else None,
        "has_log": bool(status and status.get("log_file")
                        and os.path.exists(os.path.join(os.path.dirname(os.path.abspath(__file__)), status["log_file"]))),
    })


@app.route("/api/admin/backfill-log")
@require_admin
def admin_backfill_log():
    """마지막 백필 실행 로그의 마지막 50줄 — 실패 원인 확인용.
    로그 파일 경로는 서버가 기록한 상태(log_file)에서만 가져온다(경로 조작 불가)."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT value FROM app_meta WHERE key = %s", (_BACKFILL_META_KEY,))
        meta = cur.fetchone()
    finally:
        cur.close()
        conn.close()
    status = None
    if meta and meta["value"]:
        try:
            status = json.loads(meta["value"])
        except (TypeError, ValueError):
            status = None
    log_rel = (status or {}).get("log_file")
    if not log_rel:
        return jsonify({"ok": False, "message": "이 실행에는 로그 파일 정보가 없습니다(로그 기능 도입 전 실행)."}), 404
    base_dir = os.path.dirname(os.path.abspath(__file__))
    logs_dir = os.path.join(base_dir, "logs")
    path = os.path.normpath(os.path.join(base_dir, log_rel))
    if os.path.commonpath([logs_dir, path]) != logs_dir:
        return jsonify({"ok": False, "message": "잘못된 로그 경로입니다."}), 400
    if not os.path.exists(path):
        return jsonify({"ok": False, "message": "로그 파일이 없습니다(재배포로 파일이 사라졌을 수 있습니다)."}), 404
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
    except OSError as e:
        return jsonify({"ok": False, "message": f"로그를 읽지 못했습니다: {e}"}), 500
    tail = [ln.rstrip("\n") for ln in lines[-50:]]
    # 혹시 남았을 수 있는 API 키 가림 (러너도 가리지만 이중 안전장치)
    text = "\n".join(tail)
    for key_name in ("RTMS_SERVICE_KEY", "BLD_SERVICE_KEY", "STORE_INFO_SERVICE_KEY"):
        val = os.environ.get(key_name, "")
        if val:
            text = text.replace(val, "***")
    return jsonify({"ok": True, "log_file": log_rel, "total_lines": len(lines), "tail": text})


@app.route("/admin")
@app.route("/admin/")
@require_admin
def admin_page():
    return _serve_static_html("admin.html")


@app.route("/admin/ad-products")
@require_admin
def admin_ad_products_page():
    """광고상품 안내 (관리자 전용 참고용 정보 페이지 — 판매 기능/파트너 노출 없음)."""
    return _serve_static_html("admin_ad_products.html")


# ---- 건물마스터 CRUD (모두 require_admin) ----
# 정렬 허용 컬럼 화이트리스트 (SQL 인젝션 방지 — 목록에 없으면 id로 폴백)
ADMIN_BLD_SORT = {"id", "building_name", "road_address", "units", "biz_units", "lodging_type"}
# 생성/수정 가능한 컬럼 화이트리스트 (이 목록의 키만 반영)
ADMIN_BLD_EDITABLE = [
    "building_name", "road_address", "jibun_address", "sgg_text", "sgg_cd",
    "umd_nm", "jibun", "units", "biz_units", "lodging_type", "lodging_type_detail",
]
ADMIN_BLD_INT_COLS = {"units", "biz_units"}


def _parse_int_or_none(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _clean_bld_value(col, v):
    """편집 컬럼 값을 DB 저장형으로 정규화 (숫자 컬럼은 int/None, 그 외 문자열/None)."""
    if col in ADMIN_BLD_INT_COLS:
        return _parse_int_or_none(v)
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _admin_bld_filters():
    """목록/엑셀 공용: q, sort, order, WHERE절, 파라미터를 계산."""
    q = (request.args.get("q") or "").strip()
    sort = (request.args.get("sort") or "id").strip()
    if sort not in ADMIN_BLD_SORT:
        sort = "id"
    order = "DESC" if (request.args.get("order") or "asc").strip().lower() == "desc" else "ASC"
    where = "1=1"
    params = []
    if q:
        where = "(building_name ILIKE %s OR road_address ILIKE %s)"
        params = [f"%{q}%", f"%{q}%"]
    # 명칭 미확정 건물만 필터 (관리자 주기 점검용)
    if (request.args.get("name_pending") or "").strip() == "1":
        where += " AND name_pending IS TRUE"
    return q, sort, order, where, params


@app.route("/api/admin/buildings")
@require_admin
def admin_buildings_list():
    q, sort, order, where_sql, params = _admin_bld_filters()
    try:
        page = max(int(request.args.get("page", 1)), 1)
    except (ValueError, TypeError):
        page = 1
    try:
        size = min(max(int(request.args.get("size", 50)), 1), 200)
    except (ValueError, TypeError):
        size = 50
    offset = (page - 1) * size
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) c FROM master_buildings WHERE {where_sql}", params)
    total = cur.fetchone()["c"]
    # sort/order는 화이트리스트로만 정해지므로 f-string 삽입이 안전하다.
    cur.execute(f"""
        SELECT id, building_name, name_pending, road_address, jibun_address, sgg_text,
               sgg_cd, umd_nm, jibun, units, biz_units, lodging_type, lodging_type_detail
        FROM master_buildings
        WHERE {where_sql}
        ORDER BY {sort} {order}, id ASC
        LIMIT %s OFFSET %s
    """, params + [size, offset])
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"total": total, "page": page, "size": size, "items": items})


@app.route("/api/admin/buildings", methods=["POST"])
@require_admin
def admin_buildings_create():
    data = request.get_json(force=True, silent=True) or {}
    building_name = (data.get("building_name") or "").strip()
    road_address = (data.get("road_address") or "").strip()
    if not building_name or not road_address:
        return jsonify({"ok": False, "message": "건물명과 도로명주소는 필수입니다."}), 400
    cols, vals = [], []
    for c in ADMIN_BLD_EDITABLE:
        if c in data or c in ("building_name", "road_address"):
            cols.append(c)
            vals.append(_clean_bld_value(c, data.get(c)))
    collist = ", ".join(cols)
    placeholders = ", ".join(["%s"] * len(vals))
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        f"INSERT INTO master_buildings ({collist}) VALUES ({placeholders}) RETURNING id",
        vals,
    )
    new_id = cur.fetchone()["id"]
    # 신규 건물의 좌표를 도로명주소로 즉시 채운다(실패해도 등록은 계속).
    _fill_master_coords(cur, new_id, road_address)
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/admin/buildings/<int:building_id>", methods=["PUT"])
@require_admin
def admin_buildings_update(building_id):
    data = request.get_json(force=True, silent=True) or {}
    sets, vals = [], []
    for c in ADMIN_BLD_EDITABLE:
        if c in data:
            # 필수 컬럼을 빈값으로 지우려는 시도는 막는다.
            if c in ("building_name", "road_address") and not (str(data.get(c) or "").strip()):
                return jsonify({"ok": False, "message": "건물명과 도로명주소는 비울 수 없습니다."}), 400
            sets.append(f"{c} = %s")
            vals.append(_clean_bld_value(c, data.get(c)))
    if not sets:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM master_buildings WHERE id=%s", [building_id])
    if not cur.fetchone():
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "존재하지 않는 건물입니다."}), 404
    vals.append(building_id)
    cur.execute(f"UPDATE master_buildings SET {', '.join(sets)} WHERE id = %s", vals)
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/buildings/<int:building_id>", methods=["DELETE"])
@require_admin
def admin_buildings_delete(building_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT sgg_cd, umd_nm, jibun, building_name FROM master_buildings WHERE id=%s",
        [building_id],
    )
    b = cur.fetchone()
    if not b:
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "존재하지 않는 건물입니다."}), 404
    # 참조 매물(listings) / 슬롯(slots) — master_building_id FK
    cur.execute("SELECT COUNT(*) c FROM listings WHERE master_building_id=%s", [building_id])
    listing_cnt = cur.fetchone()["c"]
    cur.execute("SELECT COUNT(*) c FROM slots WHERE master_building_id=%s", [building_id])
    slot_cnt = cur.fetchone()["c"]
    # 참조 실거래(transactions) — 지번키(sgg_cd+umd_nm+jibun) 매칭, 없으면 건물명.
    # umd_nm은 마스터/실거래 간 띄어쓰기 표기가 다를 수 있어(예: '강동면 정동진리' vs
    # '강동면정동진리') REPLACE로 공백을 제거해 정규화 매칭한다. 삭제 가드는 참조를
    # 놓쳐 고아 데이터를 만드는 것보다 넉넉히 잡는 편이 안전하다.
    if b["sgg_cd"] and b["umd_nm"] and b["jibun"]:
        cur.execute(
            "SELECT COUNT(*) c FROM transactions "
            "WHERE sgg_cd=%s AND REPLACE(umd_nm,' ','')=REPLACE(%s,' ','') AND jibun=%s",
            [b["sgg_cd"], b["umd_nm"], b["jibun"]],
        )
        tx_cnt = cur.fetchone()["c"]
    elif b["building_name"] and b["building_name"] != "-":
        cur.execute(
            "SELECT COUNT(*) c FROM transactions WHERE building_name=%s",
            [b["building_name"]],
        )
        tx_cnt = cur.fetchone()["c"]
    else:
        tx_cnt = 0
    if listing_cnt or slot_cnt or tx_cnt:
        parts = []
        if listing_cnt:
            parts.append(f"매물 {listing_cnt}건")
        if slot_cnt:
            parts.append(f"슬롯 {slot_cnt}건")
        if tx_cnt:
            parts.append(f"실거래 {tx_cnt}건")
        cur.close()
        conn.close()
        return jsonify({
            "ok": False,
            "message": "연결된 " + ", ".join(parts) + "이(가) 있어 삭제할 수 없습니다.",
        }), 400
    cur.execute("DELETE FROM master_buildings WHERE id=%s", [building_id])
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/buildings/export.xlsx")
@require_admin
def admin_buildings_export():
    q, sort, order, where_sql, params = _admin_bld_filters()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT id, building_name, road_address, sgg_text, umd_nm, jibun,
               units, biz_units, lodging_type, lodging_type_detail
        FROM master_buildings
        WHERE {where_sql}
        ORDER BY {sort} {order}, id ASC
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "건물마스터"
    headers = ["ID", "건물명", "도로명주소", "시군구", "읍면동", "지번",
               "호실수", "영업신고호수", "용도", "용도상세"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r["id"], r["building_name"], r["road_address"], r["sgg_text"],
            r["umd_nm"], r["jibun"], r["units"], r["biz_units"],
            r["lodging_type"], r["lodging_type_detail"],
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = "attachment; filename=buildings.xlsx"
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


# ---- 중개업소 데이터 동기화 + 인근 중개업소 후보 (모두 require_admin) ----
# 공공데이터포털 '전국공인중개사사무소표준데이터' (일일 쿼터 1,000건 — sync_brokers.py가
# 소프트 캡 900에서 멈추고 체크포인트로 다음날 이어서 수집). 후보 리스트는 '생성'만 하며
# 자동 이메일·SMS 발송은 하지 않는다(사람이 검토 후 수동 진행).
_BROKER_SYNC_META_KEY = "broker_sync_status"
_BROKER_DAILY_CAP = 900   # sync_brokers.MAX_DAILY_CALLS 와 동일 값 유지


@app.route("/api/admin/sync-brokers", methods=["POST"])
@require_admin
@limiter.limit("4 per hour")
def admin_broker_sync_run():
    """중개업소 데이터 동기화 시작 — 실거래 동기화와 동일한 잠금/러너 패턴."""
    if not os.environ.get("DATA_GO_KR_BROKER_API_KEY"):
        return jsonify({"ok": False, "message": "DATA_GO_KR_BROKER_API_KEY 시크릿이 등록되어 있지 않습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        status = {
            "run_id": _secrets.token_hex(8),
            "state": "running",
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": None,
            "processed": None,
            "completed": None,
            "calls_today": None,
            "error": None,
        }
        # 실행 중(하트비트 생존) 차단 + 성공 후 30분 재실행 금지 (실거래 동기화와 동일)
        cur.execute(f"""
            INSERT INTO app_meta (key, value, updated_at)
            VALUES (%s, %s, NOW())
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            WHERE ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'running'
                   OR app_meta.updated_at < NOW() - INTERVAL '{int(_SYNC_STALE_MIN)} minutes')
              AND ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'done'
                   OR app_meta.updated_at < NOW() - INTERVAL '30 minutes')
        """, (_BROKER_SYNC_META_KEY, json.dumps(status, ensure_ascii=False)))
        acquired = cur.rowcount > 0
        prev_state = None
        if not acquired:
            cur.execute("SELECT value FROM app_meta WHERE key = %s", (_BROKER_SYNC_META_KEY,))
            prev = cur.fetchone()
            try:
                prev_state = json.loads(prev["value"]).get("state") if prev and prev["value"] else None
            except (TypeError, ValueError):
                prev_state = None
        conn.commit()
    finally:
        cur.close()
        conn.close()

    if not acquired:
        if prev_state == "done":
            return jsonify({"ok": False, "message": "직전 동기화가 완료된 지 30분이 지나지 않았습니다. 잠시 후 다시 시도해 주세요."}), 429
        return jsonify({"ok": False, "message": "이미 동기화가 실행 중입니다. 완료 후 다시 시도해 주세요."}), 409

    base_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        proc = subprocess.Popen(
            [sys.executable, "-u", os.path.join(base_dir, "sync_brokers.py"),
             "--status-key", _BROKER_SYNC_META_KEY],
            cwd=base_dir, start_new_session=True,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        threading.Thread(target=proc.wait, daemon=True).start()
    except Exception as e:
        status.update({"state": "failed", "error": f"러너 실행 실패: {e}"[:300],
                       "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO app_meta (key, value, updated_at) VALUES (%s, %s, NOW())
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            """, (_BROKER_SYNC_META_KEY, json.dumps(status, ensure_ascii=False)))
            conn.commit()
        finally:
            cur.close()
            conn.close()
        return jsonify({"ok": False, "message": "동기화 프로세스를 시작하지 못했습니다."}), 500

    return jsonify({"ok": True, "message": "중개업소 데이터 동기화를 시작했습니다.", "started_at": status["started_at"]}), 202


@app.route("/api/admin/broker-sync-status")
@require_admin
def admin_broker_sync_status():
    """중개업소 동기화 진행상황 + 수집 현황 + 오늘 남은 호출 수."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) AS c FROM broker_registry")
        total = cur.fetchone()["c"]
        cur.execute("SELECT value, updated_at FROM app_meta WHERE key = %s", (_BROKER_SYNC_META_KEY,))
        meta = cur.fetchone()
        cur.execute("SELECT value FROM app_meta WHERE key = 'broker_daily_calls'")
        calls_row = cur.fetchone()
        cur.execute("SELECT value FROM app_meta WHERE key = 'broker_sync_progress'")
        prog_row = cur.fetchone()
        cur.execute("SELECT value FROM app_meta WHERE key = 'broker_last_sync'")
        last_row = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    calls_today = 0
    if calls_row and calls_row["value"]:
        try:
            d = json.loads(calls_row["value"])
            if d.get("date") == datetime.now().strftime("%Y-%m-%d"):
                calls_today = int(d.get("count", 0))
        except (TypeError, ValueError):
            pass

    progress = None
    if prog_row and prog_row["value"]:
        try:
            progress = json.loads(prog_row["value"])
        except (TypeError, ValueError):
            progress = None

    last_sync = None
    if last_row and last_row["value"]:
        try:
            last_sync = json.loads(last_row["value"])
        except (TypeError, ValueError):
            last_sync = None

    status = None
    if meta and meta["value"]:
        try:
            status = json.loads(meta["value"])
        except (TypeError, ValueError):
            status = None

    running = bool(status and status.get("state") == "running")
    stale = False
    if running and meta["updated_at"]:
        age = (datetime.now() - meta["updated_at"]).total_seconds()
        if age > _SYNC_STALE_MIN * 60:
            running, stale = False, True

    return jsonify({
        "ok": True,
        "running": running,
        "state": ("stale" if stale else (status.get("state") if status else None)),
        "started_at": (status.get("started_at") if status else None),
        "finished_at": (status.get("finished_at") if status else None),
        "processed": (status.get("processed") if status else None),
        "completed": (status.get("completed") if status else None),
        "error": ((status.get("error") if status else None)
                  or ("이전 실행이 비정상 종료된 것으로 보입니다(장시간 응답 없음). 다시 실행할 수 있습니다." if stale else None)),
        "broker_total": total,
        "calls_today": calls_today,
        "calls_remaining": max(0, _BROKER_DAILY_CAP - calls_today),
        "daily_cap": _BROKER_DAILY_CAP,
        "progress": progress,       # {"next_page":N,"total_count":M} — 미완이면 존재
        "last_sync": last_sync,     # {"finished_at":...,"total":...}
    })


# 하버사인 거리(km) SQL — master_buildings 좌표(%s lat, %s lng) 기준.
_HAVERSINE_KM = """
    6371 * acos(LEAST(1.0,
        cos(radians(%s)) * cos(radians(br.lat)) * cos(radians(br.lng) - radians(%s))
        + sin(radians(%s)) * sin(radians(br.lat))
    ))
"""


def _broker_candidates_query(building_id, radius_km):
    """건물 좌표 기준 반경 내 중개업소 목록 조회. (building dict, rows) 반환."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT mb.id, mb.building_name, mb.road_address, mb.lat, mb.lng,
                   EXISTS (SELECT 1 FROM agent_buildings ab WHERE ab.master_building_id = mb.id) AS has_agent
            FROM master_buildings mb WHERE mb.id = %s
        """, (building_id,))
        bld = cur.fetchone()
        if not bld:
            return None, None
        if bld["lat"] is None or bld["lng"] is None:
            return bld, []
        lat, lng = float(bld["lat"]), float(bld["lng"])
        # 1차 바운딩박스(인덱스 활용) → 2차 하버사인 정밀 필터
        deg = radius_km / 111.0
        cur.execute(f"""
            SELECT br.office_name, br.reg_number, br.road_address, br.jibun_address,
                   br.phone, br.reg_date, br.owner_name, br.homepage_url,
                   ROUND(({_HAVERSINE_KM})::numeric, 2) AS distance_km
            FROM broker_registry br
            WHERE br.lat IS NOT NULL AND br.lng IS NOT NULL
              AND br.lat BETWEEN %s AND %s
              AND br.lng BETWEEN %s AND %s
              AND ({_HAVERSINE_KM}) <= %s
            ORDER BY distance_km ASC, br.office_name ASC
            LIMIT 300
        """, (lat, lng, lat,
              lat - deg, lat + deg, lng - deg * 1.3, lng + deg * 1.3,
              lat, lng, lat, radius_km))
        return bld, cur.fetchall()
    finally:
        cur.close()
        conn.close()


def _parse_radius(raw):
    try:
        r = float(raw or 2)
    except (TypeError, ValueError):
        r = 2.0
    return min(5.0, max(0.5, r))


@app.route("/api/admin/broker-candidates")
@require_admin
def admin_broker_candidates():
    """건물 좌표 기준 반경 내 중개업소 후보(거리순). 후보 '생성'만 — 자동 발송 없음."""
    building_id = request.args.get("building_id", type=int)
    if not building_id:
        return jsonify({"ok": False, "message": "building_id가 필요합니다."}), 400
    radius_km = _parse_radius(request.args.get("radius_km"))
    bld, rows = _broker_candidates_query(building_id, radius_km)
    if bld is None:
        return jsonify({"ok": False, "message": "건물을 찾을 수 없습니다."}), 404
    return jsonify({
        "ok": True,
        "building": {"id": bld["id"], "building_name": bld["building_name"],
                     "road_address": bld["road_address"],
                     "has_coords": bld["lat"] is not None and bld["lng"] is not None,
                     "has_agent": bld["has_agent"]},
        "radius_km": radius_km,
        "items": rows,
    })


@app.route("/api/admin/buildings-without-agent")
@require_admin
def admin_buildings_without_agent():
    """담당중개사(agent_buildings)가 없는 건물 목록 — 후보 매칭 대상 선택용."""
    q = (request.args.get("q") or "").strip()
    conn = get_conn()
    cur = conn.cursor()
    try:
        params = []
        where = """NOT EXISTS (SELECT 1 FROM agent_buildings ab
                               WHERE ab.master_building_id = mb.id)"""
        if q:
            where += " AND (mb.building_name ILIKE %s OR mb.road_address ILIKE %s)"
            params += [f"%{q}%", f"%{q}%"]
        cur.execute(f"""
            SELECT mb.id, mb.building_name, mb.road_address, mb.sgg_text,
                   (mb.lat IS NOT NULL AND mb.lng IS NOT NULL) AS has_coords
            FROM master_buildings mb
            WHERE {where}
            ORDER BY mb.building_name ASC
            LIMIT 100
        """, params)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "items": rows})


@app.route("/api/admin/broker-candidates/export.xlsx")
@require_admin
def admin_broker_candidates_export():
    building_id = request.args.get("building_id", type=int)
    if not building_id:
        return jsonify({"ok": False, "message": "building_id가 필요합니다."}), 400
    radius_km = _parse_radius(request.args.get("radius_km"))
    bld, rows = _broker_candidates_query(building_id, radius_km)
    if bld is None:
        return jsonify({"ok": False, "message": "건물을 찾을 수 없습니다."}), 404

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "인근 중개업소 후보"
    ws.append([f"건물: {bld['building_name']} ({bld['road_address'] or '-'}) · 반경 {radius_km}km"])
    ws.append(["업소명", "거리(km)", "전화번호", "홈페이지", "등록일자",
               "대표자", "개설등록번호", "도로명주소", "지번주소"])
    for r in (rows or []):
        ws.append([
            r["office_name"], float(r["distance_km"]), r["phone"] or "", r["homepage_url"] or "",
            r["reg_date"] or "", r["owner_name"] or "", r["reg_number"],
            r["road_address"] or "", r["jibun_address"] or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = f"attachment; filename=broker_candidates_{building_id}.xlsx"
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


# ---- 숙박업 영업신고(행안부 문화_숙박업 조회서비스) 동기화 + 미등록 위탁운영 후보 ----
# 일일 쿼터 10,000건 — sync_lodgings.py가 소프트 캡 8,000에서 멈추고 체크포인트로 이어서 수집.
# 후보 리스트는 '생성'만 하며 자동 이메일·SMS 발송은 하지 않는다.
_LODGING_SYNC_META_KEY = "lodging_sync_status"
_LODGING_DAILY_CAP = 8000  # sync_lodgings.MAX_DAILY_CALLS 와 동일 값 유지

# 영업 중으로 인정하는 영업상태명 — 정확히 '영업/정상'만 (휴업/폐업/취소/말소/만료/정지/중지/제외/삭제/전출/기타 전부 제외)
_LODGING_ACTIVE_STATUS = "영업/정상"


@app.route("/api/admin/sync-lodgings", methods=["POST"])
@require_admin
@limiter.limit("4 per hour")
def admin_lodging_sync_run():
    """영업신고 데이터 동기화 시작 — 중개업소 동기화와 동일한 잠금/러너 패턴."""
    if not os.environ.get("DATA_GO_KR_BROKER_API_KEY"):
        return jsonify({"ok": False, "message": "DATA_GO_KR_BROKER_API_KEY 시크릿이 등록되어 있지 않습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        status = {
            "run_id": _secrets.token_hex(8),
            "state": "running",
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": None,
            "processed": None,
            "completed": None,
            "calls_today": None,
            "error": None,
        }
        cur.execute(f"""
            INSERT INTO app_meta (key, value, updated_at)
            VALUES (%s, %s, NOW())
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            WHERE ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'running'
                   OR app_meta.updated_at < NOW() - INTERVAL '{int(_SYNC_STALE_MIN)} minutes')
              AND ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'done'
                   OR app_meta.updated_at < NOW() - INTERVAL '30 minutes')
        """, (_LODGING_SYNC_META_KEY, json.dumps(status, ensure_ascii=False)))
        acquired = cur.rowcount > 0
        prev_state = None
        if not acquired:
            cur.execute("SELECT value FROM app_meta WHERE key = %s", (_LODGING_SYNC_META_KEY,))
            prev = cur.fetchone()
            try:
                prev_state = json.loads(prev["value"]).get("state") if prev and prev["value"] else None
            except (TypeError, ValueError):
                prev_state = None
        conn.commit()
    finally:
        cur.close()
        conn.close()

    if not acquired:
        if prev_state == "done":
            return jsonify({"ok": False, "message": "직전 동기화가 완료된 지 30분이 지나지 않았습니다. 잠시 후 다시 시도해 주세요."}), 429
        return jsonify({"ok": False, "message": "이미 동기화가 실행 중입니다. 완료 후 다시 시도해 주세요."}), 409

    base_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        proc = subprocess.Popen(
            [sys.executable, "-u", os.path.join(base_dir, "sync_lodgings.py"),
             "--status-key", _LODGING_SYNC_META_KEY],
            cwd=base_dir, start_new_session=True,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        threading.Thread(target=proc.wait, daemon=True).start()
    except Exception as e:
        status.update({"state": "failed", "error": f"러너 실행 실패: {e}"[:300],
                       "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO app_meta (key, value, updated_at) VALUES (%s, %s, NOW())
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            """, (_LODGING_SYNC_META_KEY, json.dumps(status, ensure_ascii=False)))
            conn.commit()
        finally:
            cur.close()
            conn.close()
        return jsonify({"ok": False, "message": "동기화 프로세스를 시작하지 못했습니다."}), 500

    return jsonify({"ok": True, "message": "영업신고 데이터 동기화를 시작했습니다.", "started_at": status["started_at"]}), 202


@app.route("/api/admin/lodging-sync-status")
@require_admin
def admin_lodging_sync_status():
    """영업신고 동기화 진행상황 + 수집 현황 + 오늘 남은 호출 수."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) AS c FROM lodging_registry")
        total = cur.fetchone()["c"]
        cur.execute("SELECT value, updated_at FROM app_meta WHERE key = %s", (_LODGING_SYNC_META_KEY,))
        meta = cur.fetchone()
        cur.execute("SELECT value FROM app_meta WHERE key = 'lodging_daily_calls'")
        calls_row = cur.fetchone()
        cur.execute("SELECT value FROM app_meta WHERE key = 'lodging_sync_progress'")
        prog_row = cur.fetchone()
        cur.execute("SELECT value FROM app_meta WHERE key = 'lodging_last_sync'")
        last_row = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    calls_today = 0
    if calls_row and calls_row["value"]:
        try:
            d = json.loads(calls_row["value"])
            if d.get("date") == datetime.now().strftime("%Y-%m-%d"):
                calls_today = int(d.get("count", 0))
        except (TypeError, ValueError):
            pass

    progress = last_sync = status = None
    try:
        progress = json.loads(prog_row["value"]) if prog_row and prog_row["value"] else None
    except (TypeError, ValueError):
        pass
    try:
        last_sync = json.loads(last_row["value"]) if last_row and last_row["value"] else None
    except (TypeError, ValueError):
        pass
    try:
        status = json.loads(meta["value"]) if meta and meta["value"] else None
    except (TypeError, ValueError):
        pass

    running = bool(status and status.get("state") == "running")
    stale = False
    if running and meta["updated_at"]:
        age = (datetime.now() - meta["updated_at"]).total_seconds()
        if age > _SYNC_STALE_MIN * 60:
            running, stale = False, True

    return jsonify({
        "ok": True,
        "running": running,
        "state": ("stale" if stale else (status.get("state") if status else None)),
        "started_at": (status.get("started_at") if status else None),
        "finished_at": (status.get("finished_at") if status else None),
        "processed": (status.get("processed") if status else None),
        "completed": (status.get("completed") if status else None),
        "error": ((status.get("error") if status else None)
                  or ("이전 실행이 비정상 종료된 것으로 보입니다(장시간 응답 없음). 다시 실행할 수 있습니다." if stale else None)),
        "lodging_total": total,
        "calls_today": calls_today,
        "calls_remaining": max(0, _LODGING_DAILY_CAP - calls_today),
        "daily_cap": _LODGING_DAILY_CAP,
        "progress": progress,
        "last_sync": last_sync,
    })


# ---- 건축HUB 전국 건물 발견(sync_brhub.py) 관리자 실행 ----
_BRHUB_SYNC_META_KEY = "brhub_sync_status"
_BRHUB_DAILY_CAP = 8000  # sync_brhub.py --daily-cap 기본값과 동일 유지
_BRHUB_TOTAL_DONGS = None  # bjdong_codes.json 법정동 총수 캐시


def _brhub_total_dongs():
    global _BRHUB_TOTAL_DONGS
    if _BRHUB_TOTAL_DONGS is None:
        try:
            with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "bjdong_codes.json"),
                      encoding="utf-8") as f:
                _BRHUB_TOTAL_DONGS = len(json.load(f)["dongs"])
        except Exception:
            _BRHUB_TOTAL_DONGS = 0
    return _BRHUB_TOTAL_DONGS


@app.route("/api/admin/sync-brhub", methods=["POST"])
@require_admin
@limiter.limit("2 per hour")
def admin_brhub_sync_run():
    """건축HUB 전국 건물수집 시작 — 숙박업 동기화와 동일한 잠금/러너 패턴."""
    if not os.environ.get("DATA_GO_KR_BROKER_API_KEY"):
        return jsonify({"ok": False, "message": "DATA_GO_KR_BROKER_API_KEY 시크릿이 등록되어 있지 않습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        status = {
            "run_id": _secrets.token_hex(8),
            "state": "running",
            "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "finished_at": None,
            "processed": None,
            "found": None,
            "completed": None,
            "calls_today": None,
            "error": None,
        }
        cur.execute(f"""
            INSERT INTO app_meta (key, value, updated_at)
            VALUES (%s, %s, NOW())
            ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            WHERE ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'running'
                   OR app_meta.updated_at < NOW() - INTERVAL '{int(_SYNC_STALE_MIN)} minutes')
              AND ((app_meta.value::jsonb ->> 'state') IS DISTINCT FROM 'done'
                   OR app_meta.updated_at < NOW() - INTERVAL '30 minutes')
        """, (_BRHUB_SYNC_META_KEY, json.dumps(status, ensure_ascii=False)))
        acquired = cur.rowcount > 0
        prev_state = None
        if not acquired:
            cur.execute("SELECT value FROM app_meta WHERE key = %s", (_BRHUB_SYNC_META_KEY,))
            prev = cur.fetchone()
            try:
                prev_state = json.loads(prev["value"]).get("state") if prev and prev["value"] else None
            except (TypeError, ValueError):
                prev_state = None
        conn.commit()
    finally:
        cur.close()
        conn.close()

    if not acquired:
        if prev_state == "done":
            return jsonify({"ok": False, "message": "직전 동기화가 완료된 지 30분이 지나지 않았습니다. 잠시 후 다시 시도해 주세요."}), 429
        return jsonify({"ok": False, "message": "이미 동기화가 실행 중입니다. 완료 후 다시 시도해 주세요."}), 409

    base_dir = os.path.dirname(os.path.abspath(__file__))
    try:
        proc = subprocess.Popen(
            [sys.executable, "-u", os.path.join(base_dir, "sync_brhub.py"),
             "--status-key", _BRHUB_SYNC_META_KEY],
            cwd=base_dir, start_new_session=True,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        threading.Thread(target=proc.wait, daemon=True).start()
    except Exception as e:
        status.update({"state": "failed", "error": f"러너 실행 실패: {e}"[:300],
                       "finished_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO app_meta (key, value, updated_at) VALUES (%s, %s, NOW())
                ON CONFLICT (key) DO UPDATE SET value = EXCLUDED.value, updated_at = NOW()
            """, (_BRHUB_SYNC_META_KEY, json.dumps(status, ensure_ascii=False)))
            conn.commit()
        finally:
            cur.close()
            conn.close()
        return jsonify({"ok": False, "message": "동기화 프로세스를 시작하지 못했습니다."}), 500

    return jsonify({"ok": True, "message": "건물수집(전국)을 시작했습니다.", "started_at": status["started_at"]}), 202


@app.route("/api/admin/brhub-sync-status")
@require_admin
def admin_brhub_sync_status():
    """건물수집 진행상황 + brhub_bulk 수집 현황 + 체크포인트/오늘 호출량."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT COUNT(*) AS c FROM master_buildings WHERE source = 'brhub_bulk'")
        total = cur.fetchone()["c"]
        cur.execute("""
            SELECT CASE
                     WHEN lodging_type IS NULL THEN 'unclassified'
                     WHEN lodging_type LIKE '%%·%%' THEN 'mixed'
                     WHEN lodging_type = '생활' THEN 'living_stay'
                     WHEN lodging_type = '호텔' THEN 'hotel'
                     WHEN lodging_type = '콘도' THEN 'condo'
                     ELSE 'other'
                   END AS k, COUNT(*) AS c
            FROM master_buildings WHERE source = 'brhub_bulk' GROUP BY 1
        """)
        by_type = {r["k"]: r["c"] for r in cur.fetchall()}
        cur.execute("SELECT value, updated_at FROM app_meta WHERE key = %s", (_BRHUB_SYNC_META_KEY,))
        meta = cur.fetchone()
        cur.execute("SELECT value FROM app_meta WHERE key = 'brhub_progress'")
        prog_row = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    progress = status = None
    try:
        progress = json.loads(prog_row["value"]) if prog_row and prog_row["value"] else None
    except (TypeError, ValueError):
        pass
    try:
        status = json.loads(meta["value"]) if meta and meta["value"] else None
    except (TypeError, ValueError):
        pass

    calls_today = 0
    if progress and progress.get("calls_date") == datetime.now().strftime("%Y-%m-%d"):
        try:
            calls_today = int(progress.get("calls_today") or 0)
        except (TypeError, ValueError):
            pass

    running = bool(status and status.get("state") == "running")
    stale = False
    if running and meta["updated_at"]:
        age = (datetime.now() - meta["updated_at"]).total_seconds()
        if age > _SYNC_STALE_MIN * 60:
            running, stale = False, True

    return jsonify({
        "ok": True,
        "running": running,
        "state": ("stale" if stale else (status.get("state") if status else None)),
        "started_at": (status.get("started_at") if status else None),
        "finished_at": (status.get("finished_at") if status else None),
        "processed": (status.get("processed") if status else None),
        "found": (status.get("found") if status else None),
        "completed": (status.get("completed") if status else None),
        "error": ((status.get("error") if status else None)
                  or ("이전 실행이 비정상 종료된 것으로 보입니다(장시간 응답 없음). 다시 실행할 수 있습니다." if stale else None)),
        "brhub_total": total,
        "by_type": by_type,
        "checkpoint_idx": (progress.get("idx") if progress else 0) or 0,
        "total_dongs": _brhub_total_dongs(),
        "found_total": (progress.get("found_total") if progress else 0) or 0,
        "calls_today": calls_today,
        "calls_remaining": max(0, _BRHUB_DAILY_CAP - calls_today),
        "daily_cap": _BRHUB_DAILY_CAP,
    })


def _approved_operator_name_norms(cur):
    """approved 운영업체의 (정규화명 → 목록) 매핑. 미등록 후보 판정/B화면 매칭 공용."""
    cur.execute("""
        SELECT o.id, o.company_name, o.phone, o.subdomain_slug,
               COALESCE(o.priority_score, 0) AS priority_score,
               COALESCE(o.is_visible, TRUE) AS is_visible
        FROM operators o
        WHERE o.status = 'approved'
    """)
    mapping = {}
    for r in cur.fetchall():
        norm = addr_norm.normalize_name(r["company_name"])
        if norm:
            mapping.setdefault(norm, []).append(dict(r))
    return mapping


@app.route("/api/admin/unregistered-lodging-candidates")
@require_admin
def admin_unregistered_lodging_candidates():
    """operators에 등록되지 않은 '영업/정상' 생활숙박업 사업장 목록 — 위탁운영 유치 후보."""
    q = (request.args.get("q") or "").strip()
    conn = get_conn()
    cur = conn.cursor()
    try:
        op_norms = list(_approved_operator_name_norms(cur).keys())
        params = [_LODGING_ACTIVE_STATUS]
        where = "lr.biz_status_name = %s"
        if op_norms:
            where += " AND (lr.biz_name_norm IS NULL OR NOT (lr.biz_name_norm = ANY(%s)))"
            params.append(op_norms)
        if q:
            where += " AND (lr.biz_name ILIKE %s OR lr.road_address ILIKE %s)"
            params += [f"%{q}%", f"%{q}%"]
        cur.execute(f"""
            SELECT lr.biz_name, lr.permit_number, lr.road_address, lr.jibun_address,
                   lr.permit_date, lr.biz_status_name, lr.biz_status_detail,
                   lr.room_count, lr.phone
            FROM lodging_registry lr
            WHERE {where}
            ORDER BY lr.room_count DESC NULLS LAST, lr.biz_name ASC
            LIMIT 300
        """, params)
        rows = cur.fetchall()
        cur.execute(f"SELECT COUNT(*) AS c FROM lodging_registry lr WHERE {where}", params)
        total = cur.fetchone()["c"]
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "items": rows, "total": total})


@app.route("/api/admin/unregistered-lodging-candidates/export.xlsx")
@require_admin
def admin_unregistered_lodging_export():
    q = (request.args.get("q") or "").strip()
    conn = get_conn()
    cur = conn.cursor()
    try:
        op_norms = list(_approved_operator_name_norms(cur).keys())
        params = [_LODGING_ACTIVE_STATUS]
        where = "lr.biz_status_name = %s"
        if op_norms:
            where += " AND (lr.biz_name_norm IS NULL OR NOT (lr.biz_name_norm = ANY(%s)))"
            params.append(op_norms)
        if q:
            where += " AND (lr.biz_name ILIKE %s OR lr.road_address ILIKE %s)"
            params += [f"%{q}%", f"%{q}%"]
        cur.execute(f"""
            SELECT lr.biz_name, lr.permit_number, lr.road_address, lr.jibun_address,
                   lr.permit_date, lr.room_count, lr.phone
            FROM lodging_registry lr
            WHERE {where}
            ORDER BY lr.room_count DESC NULLS LAST, lr.biz_name ASC
            LIMIT 3000
        """, params)
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "미등록 위탁운영 후보"
    ws.append(["사업장명", "객실수", "전화번호", "인허가일자", "관리번호", "도로명주소", "지번주소"])
    for r in rows:
        ws.append([r["biz_name"], r["room_count"] or 0, r["phone"] or "",
                   r["permit_date"] or "", r["permit_number"],
                   r["road_address"] or "", r["jibun_address"] or ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = "attachment; filename=unregistered_lodging_candidates.xlsx"
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


# ---- 매물(listings) 관리 (모두 require_admin) ----
# 정렬 허용 컬럼 화이트리스트 → 실제 SQL 표현식 매핑 (인젝션 방지, 없으면 l.id 폴백)
ADMIN_LST_SORT = {
    "id": "l.id", "deal_type": "l.deal_type", "price": "l.price",
    "status": "l.status", "created_at": "l.created_at",
}
# 수정 가능한 컬럼 → 값 형변환 종류
ADMIN_LST_EDITABLE = {
    "deal_type": "text", "price": "int", "monthly_rent": "int",
    "floor": "text", "area": "float", "status": "text",
}


def _parse_float_or_none(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _clean_typed_value(kind, v):
    """editable 화이트리스트의 종류(int/float/text)에 맞춰 DB 저장형으로 정규화."""
    if kind == "int":
        return _parse_int_or_none(v)
    if kind == "float":
        return _parse_float_or_none(v)
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _admin_paging():
    try:
        page = max(int(request.args.get("page", 1)), 1)
    except (ValueError, TypeError):
        page = 1
    try:
        size = min(max(int(request.args.get("size", 50)), 1), 200)
    except (ValueError, TypeError):
        size = 50
    return page, size, (page - 1) * size


def _admin_lst_filters():
    """매물 목록/엑셀 공용: sort_expr, order, WHERE절, 파라미터. 검색은 건물ID로 필터."""
    q = (request.args.get("q") or "").strip()
    sort_key = (request.args.get("sort") or "id").strip()
    sort_expr = ADMIN_LST_SORT.get(sort_key, "l.id")
    order = "DESC" if (request.args.get("order") or "asc").strip().lower() == "desc" else "ASC"
    where = "1=1"
    params = []
    if q.isdigit():
        where = "l.master_building_id = %s"
        params = [int(q)]
    return sort_expr, order, where, params


@app.route("/api/admin/listings")
@require_admin
def admin_listings_list():
    sort_expr, order, where_sql, params = _admin_lst_filters()
    page, size, offset = _admin_paging()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) c FROM listings l WHERE {where_sql}", params)
    total = cur.fetchone()["c"]
    # sort_expr/order는 화이트리스트로만 정해지므로 f-string 삽입이 안전하다.
    cur.execute(f"""
        SELECT l.id, l.master_building_id, mb.building_name,
               l.deal_type, l.price, l.monthly_rent, l.floor, l.area, l.status,
               to_char(l.created_at, 'YYYY-MM-DD') AS created_at
        FROM listings l
        LEFT JOIN master_buildings mb ON mb.id = l.master_building_id
        WHERE {where_sql}
        ORDER BY {sort_expr} {order}, l.id ASC
        LIMIT %s OFFSET %s
    """, params + [size, offset])
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"total": total, "page": page, "size": size, "items": items})


@app.route("/api/admin/listings/<int:listing_id>", methods=["PUT"])
@require_admin
def admin_listings_update(listing_id):
    data = request.get_json(force=True, silent=True) or {}
    sets, vals = [], []
    for c, kind in ADMIN_LST_EDITABLE.items():
        if c in data:
            sets.append(f"{c} = %s")
            vals.append(_clean_typed_value(kind, data.get(c)))
    if not sets:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM listings WHERE id=%s", [listing_id])
    if not cur.fetchone():
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "존재하지 않는 매물입니다."}), 404
    sets.append("updated_at = NOW()")
    vals.append(listing_id)
    cur.execute(f"UPDATE listings SET {', '.join(sets)} WHERE id = %s", vals)
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/listings/<int:listing_id>", methods=["DELETE"])
@require_admin
def admin_listings_delete(listing_id):
    # 매물은 다른 테이블이 참조하지 않으므로 참조 무결성 이슈 없이 바로 삭제한다.
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM listings WHERE id=%s RETURNING id", [listing_id])
    deleted = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    if not deleted:
        return jsonify({"ok": False, "message": "존재하지 않는 매물입니다."}), 404
    return jsonify({"ok": True})


@app.route("/api/admin/listings/export.xlsx")
@require_admin
def admin_listings_export():
    sort_expr, order, where_sql, params = _admin_lst_filters()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT l.id, l.master_building_id, mb.building_name,
               l.deal_type, l.price, l.monthly_rent, l.floor, l.area, l.status,
               to_char(l.created_at, 'YYYY-MM-DD') AS created_at
        FROM listings l
        LEFT JOIN master_buildings mb ON mb.id = l.master_building_id
        WHERE {where_sql}
        ORDER BY {sort_expr} {order}, l.id ASC
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "매물"
    ws.append(["ID", "건물ID", "건물명", "거래유형", "가격(만원)", "월세(만원)",
               "층", "면적(㎡)", "상태", "등록일"])
    for r in rows:
        ws.append([
            r["id"], r["master_building_id"], r["building_name"], r["deal_type"],
            r["price"], r["monthly_rent"], r["floor"], r["area"],
            r["status"], r["created_at"],
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = "attachment; filename=listings.xlsx"
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


# ---- 매물의뢰(listing_requests) 관리 (모두 require_admin) ----
# 관리자는 조회 + 비고(admin_note) 수정만 가능. status는 중개사 API에서만 변경(여기서는 읽기전용).
_ADMIN_LREQ_SORT = {"id": "lr.id", "created_at": "lr.created_at", "status": "lr.status"}


@app.route("/api/admin/listing-requests")
@require_admin
def admin_listing_requests_list():
    q = (request.args.get("q") or "").strip()
    sort_key = (request.args.get("sort") or "id").strip()
    sort_expr = _ADMIN_LREQ_SORT.get(sort_key, "lr.id")
    order = "DESC" if (request.args.get("order") or "asc").strip().lower() == "desc" else "ASC"
    page, size, offset = _admin_paging()
    where = "1=1"
    params = []
    if q:
        where = "(mb.building_name ILIKE %s OR lr.contact_phone ILIKE %s OR a.office_name ILIKE %s)"
        like = f"%{q}%"
        params = [like, like, like]
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"""
            SELECT COUNT(*) c
            FROM listing_requests lr
            LEFT JOIN master_buildings mb ON mb.id = lr.master_building_id
            LEFT JOIN agents a ON a.id = lr.routed_agent_id
            WHERE {where}
        """, params)
        total = cur.fetchone()["c"]
        # sort_expr/order는 화이트리스트로만 정해지므로 f-string 삽입이 안전하다.
        cur.execute(f"""
            SELECT lr.id, lr.master_building_id, mb.building_name,
                   lr.deal_type, lr.desired_price, lr.contact_phone,
                   lr.routed_reason, lr.status, lr.admin_note,
                   (lr.status = 'submitted' AND lr.created_at < NOW() - INTERVAL '7 days') AS is_delayed,
                   a.office_name AS agent_office_name, a.phone AS agent_phone,
                   to_char(lr.created_at, 'YYYY-MM-DD HH24:MI') AS created_at
            FROM listing_requests lr
            LEFT JOIN master_buildings mb ON mb.id = lr.master_building_id
            LEFT JOIN agents a ON a.id = lr.routed_agent_id
            WHERE {where}
            ORDER BY {sort_expr} {order}, lr.id ASC
            LIMIT %s OFFSET %s
        """, params + [size, offset])
        items = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return jsonify({"total": total, "page": page, "size": size, "items": items})


@app.route("/api/admin/partner-building-counts")
@require_admin
def admin_partner_building_counts():
    """중개사/운영업체별 담당 건물 수 목록 (무료 캡 대비 현황 파악용)."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT a.id, a.office_name AS name, a.phone, a.status,
                   COUNT(ab.master_building_id) AS building_count
            FROM agents a
            LEFT JOIN agent_buildings ab ON ab.agent_id = a.id
            GROUP BY a.id, a.office_name, a.phone, a.status
            ORDER BY building_count DESC, a.id ASC
        """)
        agents = [dict(r) for r in cur.fetchall()]
        cur.execute("""
            SELECT o.id, o.company_name AS name, o.phone, o.status,
                   COUNT(ob.master_building_id) AS building_count
            FROM operators o
            LEFT JOIN operator_buildings ob ON ob.operator_id = o.id
            GROUP BY o.id, o.company_name, o.phone, o.status
            ORDER BY building_count DESC, o.id ASC
        """)
        operators = [dict(r) for r in cur.fetchall()]
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "max_free_buildings": MAX_FREE_BUILDINGS,
                    "agents": agents, "operators": operators})


@app.route("/api/admin/listing-requests/<int:req_id>", methods=["PUT"])
@require_admin
def admin_listing_requests_update(req_id):
    """관리자 수정은 admin_note만 허용 — status 등 다른 필드는 값이 와도 무시한다."""
    data = request.get_json(force=True, silent=True) or {}
    if "admin_note" not in data:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다. (비고만 수정 가능)"}), 400
    note = data.get("admin_note")
    note = str(note).strip() if note is not None else None
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT 1 FROM listing_requests WHERE id = %s", [req_id])
        if not cur.fetchone():
            return jsonify({"ok": False, "message": "존재하지 않는 의뢰입니다."}), 404
        cur.execute("UPDATE listing_requests SET admin_note = %s WHERE id = %s", [note or None, req_id])
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


# ---- 실거래(transactions) 관리 (모두 require_admin) ----
# 실거래는 공공데이터 원본이라 삭제는 만들지 않고, 이상치 정정용 수정만 허용한다.
# 수정 시 반드시 사유(reason)를 받아 admin_edit_log에 필드 단위로 남긴다.
ADMIN_TX_SORT = {"id": "id", "deal_date": "deal_date", "price": "price", "area": "area"}
ADMIN_TX_EDITABLE = {
    "price": "int", "area": "float", "floor": "text",
    "deal_type": "text", "deal_date": "text",
}


def _admin_tx_filters():
    """실거래 목록/엑셀 공용: sort_expr, order, WHERE절, 파라미터. 검색은 건물명·주소."""
    q = (request.args.get("q") or "").strip()
    # 기본 정렬: 계약일 최신순(deal_date DESC). sort/order 파라미터로 기존처럼 변경 가능.
    sort_key = (request.args.get("sort") or "deal_date").strip()
    sort_expr = ADMIN_TX_SORT.get(sort_key, "deal_date")
    order = "ASC" if (request.args.get("order") or "desc").strip().lower() == "asc" else "DESC"
    where = "1=1"
    params = []
    if q:
        where = "(building_name ILIKE %s OR address ILIKE %s)"
        params = [f"%{q}%", f"%{q}%"]
    return sort_expr, order, where, params


@app.route("/api/admin/transactions")
@require_admin
def admin_transactions_list():
    sort_expr, order, where_sql, params = _admin_tx_filters()
    page, size, offset = _admin_paging()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) c FROM transactions WHERE {where_sql}", params)
    total = cur.fetchone()["c"]
    cur.execute(f"""
        SELECT id, building_name, address, area, floor, price, deal_date, deal_type
        FROM transactions
        WHERE {where_sql}
        ORDER BY {sort_expr} {order}, id ASC
        LIMIT %s OFFSET %s
    """, params + [size, offset])
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"total": total, "page": page, "size": size, "items": items})


@app.route("/api/admin/transactions/<int:tx_id>", methods=["PUT"])
@require_admin
def admin_transactions_update(tx_id):
    data = request.get_json(force=True, silent=True) or {}
    reason = (data.get("reason") or "").strip()
    if not reason:
        return jsonify({"ok": False, "message": "수정 사유(reason)는 필수입니다."}), 400
    # 계약일은 관리자 오입력 방지를 위해 YYYY-MM-DD 형식만 허용한다(데이터 오염 방지).
    if "deal_date" in data:
        dd = (str(data.get("deal_date") or "")).strip()
        if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", dd):
            return jsonify({"ok": False, "message": "계약일은 YYYY-MM-DD 형식이어야 합니다."}), 400
    # 수정 요청된 편집 컬럼만 추린다.
    changes = {}
    for c, kind in ADMIN_TX_EDITABLE.items():
        if c in data:
            changes[c] = _clean_typed_value(kind, data.get(c))
    if not changes:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        f"SELECT {', '.join(changes.keys())} FROM transactions WHERE id=%s",
        [tx_id],
    )
    old = cur.fetchone()
    if not old:
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "존재하지 않는 실거래입니다."}), 404
    # 실제로 값이 달라진 필드만 로그로 남긴다.
    logged = 0
    for field, new_val in changes.items():
        old_val = old[field]
        if str(old_val) == str(new_val):
            continue
        cur.execute(
            """INSERT INTO admin_edit_log
               (table_name, record_id, field, old_value, new_value, reason, admin)
               VALUES (%s, %s, %s, %s, %s, %s, TRUE)""",
            ["transactions", tx_id, field,
             None if old_val is None else str(old_val),
             None if new_val is None else str(new_val),
             reason],
        )
        logged += 1
    sets = [f"{c} = %s" for c in changes]
    vals = list(changes.values()) + [tx_id]
    cur.execute(f"UPDATE transactions SET {', '.join(sets)} WHERE id = %s", vals)
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True, "logged": logged})


@app.route("/api/admin/transactions/export.xlsx")
@require_admin
def admin_transactions_export():
    sort_expr, order, where_sql, params = _admin_tx_filters()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT id, building_name, address, area, floor, price, deal_date, deal_type
        FROM transactions
        WHERE {where_sql}
        ORDER BY {sort_expr} {order}, id ASC
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "실거래"
    ws.append(["ID", "건물명", "주소", "면적(㎡)", "층", "거래금액(만원)",
               "계약일", "거래유형"])
    for r in rows:
        ws.append([
            r["id"], r["building_name"], r["address"], r["area"], r["floor"],
            r["price"], r["deal_date"], r["deal_type"],
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = "attachment; filename=transactions.xlsx"
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


# ---- 공지사항(notices) 관리 (모두 require_admin) ----
# 정렬 허용 컬럼 화이트리스트 (SQL 인젝션 방지 — 없으면 기본 정렬로 폴백)
ADMIN_NOTICE_SORT = {
    "id": "id", "title": "title", "is_pinned": "is_pinned", "created_at": "created_at",
}


def _admin_notice_filters():
    """공지 목록 공용: sort_expr, order, WHERE절, 파라미터. 검색은 제목·본문."""
    q = (request.args.get("q") or "").strip()
    sort_key = (request.args.get("sort") or "").strip()
    sort_expr = ADMIN_NOTICE_SORT.get(sort_key)
    order = "DESC" if (request.args.get("order") or "").strip().lower() == "desc" else "ASC"
    where = "1=1"
    params = []
    if q:
        where = "(title ILIKE %s OR body ILIKE %s)"
        params = [f"%{q}%", f"%{q}%"]
    return sort_expr, order, where, params


def _parse_bool(v):
    """폼/JSON에서 온 다양한 표기('true','1','on', True 등)를 파이썬 bool로."""
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("1", "true", "yes", "on", "y")


@app.route("/api/admin/notices")
@require_admin
def admin_notices_list():
    sort_expr, order, where_sql, params = _admin_notice_filters()
    page, size, offset = _admin_paging()
    # 명시 정렬이 있으면 그걸 우선하되, 기본은 항상 '고정 우선 → 최신순'.
    order_sql = f"{sort_expr} {order}, id DESC" if sort_expr else "is_pinned DESC, created_at DESC, id DESC"
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) c FROM notices WHERE {where_sql}", params)
    total = cur.fetchone()["c"]
    cur.execute(f"""
        SELECT id, title, body, is_pinned,
               to_char(created_at, 'YYYY-MM-DD') AS created_at,
               to_char(updated_at, 'YYYY-MM-DD') AS updated_at
        FROM notices
        WHERE {where_sql}
        ORDER BY {order_sql}
        LIMIT %s OFFSET %s
    """, params + [size, offset])
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"total": total, "page": page, "size": size, "items": items})


@app.route("/api/admin/notices", methods=["POST"])
@require_admin
def admin_notices_create():
    data = request.get_json(force=True, silent=True) or {}
    title = (data.get("title") or "").strip()
    body = (data.get("body") or "").strip()
    if not title or not body:
        return jsonify({"ok": False, "message": "제목과 본문은 필수입니다."}), 400
    is_pinned = _parse_bool(data.get("is_pinned"))
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO notices (title, body, is_pinned) VALUES (%s, %s, %s) RETURNING id",
        [title, body, is_pinned],
    )
    new_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/admin/notices/<int:notice_id>", methods=["PUT"])
@require_admin
def admin_notices_update(notice_id):
    data = request.get_json(force=True, silent=True) or {}
    sets, vals = [], []
    if "title" in data:
        title = (data.get("title") or "").strip()
        if not title:
            return jsonify({"ok": False, "message": "제목은 비울 수 없습니다."}), 400
        sets.append("title = %s")
        vals.append(title)
    if "body" in data:
        body = (data.get("body") or "").strip()
        if not body:
            return jsonify({"ok": False, "message": "본문은 비울 수 없습니다."}), 400
        sets.append("body = %s")
        vals.append(body)
    if "is_pinned" in data:
        sets.append("is_pinned = %s")
        vals.append(_parse_bool(data.get("is_pinned")))
    if not sets:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM notices WHERE id=%s", [notice_id])
    if not cur.fetchone():
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "존재하지 않는 공지입니다."}), 404
    sets.append("updated_at = NOW()")
    vals.append(notice_id)
    cur.execute(f"UPDATE notices SET {', '.join(sets)} WHERE id = %s", vals)
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/notices/<int:notice_id>", methods=["DELETE"])
@require_admin
def admin_notices_delete(notice_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM notices WHERE id=%s RETURNING id", [notice_id])
    deleted = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    if not deleted:
        return jsonify({"ok": False, "message": "존재하지 않는 공지입니다."}), 404
    return jsonify({"ok": True})


# ============================================================
# 사이트 팝업/상단배너 (site_popups)
# 관리자 CRUD + 이미지 업로드(Object Storage) + 공개 조회/이미지 프록시.
# 표시 로직은 static/js/header.js가 담당한다.
# ============================================================

_POPUP_SCOPES = {"all", "home_only"}
_POPUP_AUDIENCES = {"all", "logged_in"}
_POPUP_DISPLAY_TYPES = {"popup", "top_banner"}
_POPUP_CLOSE_MODES = {"close", "hide_today"}


def _parse_popup_ts(value, label):
    """'YYYY-MM-DD HH:MM'(또는 T 구분, 초 포함) 문자열 → datetime. 빈 값은 None."""
    s = (value or "").strip().replace("T", " ")
    if not s:
        return None, None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt), None
        except ValueError:
            continue
    return None, f"{label} 형식이 올바르지 않습니다. (예: 2026-07-16 09:00)"


def _validate_popup_payload(data):
    """POST/PUT 공통 검증 — 정상이면 (필드 dict, None), 오류면 (None, 메시지)."""
    title = (data.get("title") or "").strip()
    if not title:
        return None, "제목(관리용)은 필수입니다."
    start_at, err = _parse_popup_ts(data.get("start_at"), "게재 시작일시")
    if err:
        return None, err
    end_at, err = _parse_popup_ts(data.get("end_at"), "게재 종료일시")
    if err:
        return None, err
    if start_at and end_at and end_at <= start_at:
        return None, "게재 종료일시는 시작일시보다 뒤여야 합니다."
    scope = (data.get("scope") or "all").strip()
    audience = (data.get("audience") or "all").strip()
    display_type = (data.get("display_type") or "popup").strip()
    close_mode = (data.get("close_mode") or "close").strip()
    if scope not in _POPUP_SCOPES or audience not in _POPUP_AUDIENCES \
            or display_type not in _POPUP_DISPLAY_TYPES or close_mode not in _POPUP_CLOSE_MODES:
        return None, "선택 값이 올바르지 않습니다."
    image_ref = (data.get("image_ref") or "").strip()
    if not image_ref:
        return None, "이미지를 업로드해주세요."
    if not storage_util.is_valid_popup_ref(image_ref):
        return None, "이미지 참조가 올바르지 않습니다. 파일을 다시 업로드해주세요."
    link_url = (data.get("link_url") or "").strip()
    if link_url and not (link_url.startswith("http://") or link_url.startswith("https://")
                         or (link_url.startswith("/") and not link_url.startswith("//"))):
        return None, "링크 URL은 http(s):// 또는 /로 시작해야 합니다."
    try:
        width_px = int(data.get("width_px") or 400)
    except (ValueError, TypeError):
        return None, "너비(px)는 숫자여야 합니다."
    width_px = min(max(width_px, 200), 1200)
    return {
        "title": title,
        "start_at": start_at,
        "end_at": end_at,
        "show_desktop": _parse_bool(data.get("show_desktop", True)),
        "show_mobile": _parse_bool(data.get("show_mobile", True)),
        "scope": scope,
        "audience": audience,
        "display_type": display_type,
        "image_ref": image_ref,
        "link_url": link_url or None,
        "open_new_tab": _parse_bool(data.get("open_new_tab", True)),
        "width_px": width_px,
        "close_mode": close_mode,
        "is_active": _parse_bool(data.get("is_active", True)),
    }, None


def _popup_row_to_dict(r):
    d = dict(r)
    for k in ("start_at", "end_at", "created_at", "updated_at"):
        if d.get(k):
            d[k] = d[k].strftime("%Y-%m-%d %H:%M")
    return d


@app.route("/api/admin/popups")
@require_admin
def admin_popups_list():
    q = (request.args.get("q") or "").strip()
    page, size, offset = _admin_paging()
    where = "TRUE"
    params = []
    if q:
        where = "title ILIKE %s"
        params.append(f"%{q}%")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) c FROM site_popups WHERE {where}", params)
    total = cur.fetchone()["c"]
    cur.execute(f"""
        SELECT * FROM site_popups
        WHERE {where}
        ORDER BY id DESC
        LIMIT %s OFFSET %s
    """, params + [size, offset])
    items = [_popup_row_to_dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"total": total, "page": page, "size": size, "items": items})


@app.route("/api/admin/popups", methods=["POST"])
@require_admin
def admin_popups_create():
    data = request.get_json(force=True, silent=True) or {}
    fields, err = _validate_popup_payload(data)
    if err:
        return jsonify({"ok": False, "message": err}), 400
    conn = get_conn()
    cur = conn.cursor()
    cols = list(fields.keys())
    cur.execute(
        f"INSERT INTO site_popups ({', '.join(cols)}) VALUES ({', '.join(['%s'] * len(cols))}) RETURNING id",
        [fields[c] for c in cols],
    )
    new_id = cur.fetchone()["id"]
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/admin/popups/<int:popup_id>", methods=["PUT"])
@require_admin
def admin_popups_update(popup_id):
    data = request.get_json(force=True, silent=True) or {}
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM site_popups WHERE id=%s", [popup_id])
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "존재하지 않는 팝업입니다."}), 404
    # 부분 수정 지원 — 안 넘어온 필드는 기존 값 유지한 채 전체 재검증.
    merged = {
        "title": data.get("title", row["title"]),
        "start_at": data.get("start_at", row["start_at"].strftime("%Y-%m-%d %H:%M") if row["start_at"] else ""),
        "end_at": data.get("end_at", row["end_at"].strftime("%Y-%m-%d %H:%M") if row["end_at"] else ""),
        "show_desktop": data.get("show_desktop", row["show_desktop"]),
        "show_mobile": data.get("show_mobile", row["show_mobile"]),
        "scope": data.get("scope", row["scope"]),
        "audience": data.get("audience", row["audience"]),
        "display_type": data.get("display_type", row["display_type"]),
        "image_ref": data.get("image_ref", row["image_ref"]),
        "link_url": data.get("link_url", row["link_url"] or ""),
        "open_new_tab": data.get("open_new_tab", row["open_new_tab"]),
        "width_px": data.get("width_px", row["width_px"]),
        "close_mode": data.get("close_mode", row["close_mode"]),
        "is_active": data.get("is_active", row["is_active"]),
    }
    fields, err = _validate_popup_payload(merged)
    if err:
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": err}), 400
    sets = ", ".join(f"{c} = %s" for c in fields) + ", updated_at = NOW()"
    cur.execute(f"UPDATE site_popups SET {sets} WHERE id = %s", list(fields.values()) + [popup_id])
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/popups/<int:popup_id>", methods=["DELETE"])
@require_admin
def admin_popups_delete(popup_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM site_popups WHERE id=%s RETURNING id", [popup_id])
    deleted = cur.fetchone()
    conn.commit()
    cur.close()
    conn.close()
    if not deleted:
        return jsonify({"ok": False, "message": "존재하지 않는 팝업입니다."}), 404
    return jsonify({"ok": True})


@app.route("/api/admin/popups/upload-image", methods=["POST"])
@require_admin
def admin_popups_upload_image():
    """팝업 이미지 업로드 — C/D 서류 업로드와 동일한 검증(확장자·5MB·매직바이트)."""
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "message": "파일을 선택해주세요."}), 400
    ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
    if ext not in storage_util.POPUP_IMAGE_EXTENSIONS:
        return jsonify({"ok": False, "message": "JPG, PNG 이미지만 업로드할 수 있습니다."}), 400
    data = f.read(storage_util.MAX_FILE_BYTES + 1)
    if len(data) > storage_util.MAX_FILE_BYTES:
        return jsonify({"ok": False, "message": "파일 크기는 5MB 이하여야 합니다."}), 400
    if len(data) < 16:
        return jsonify({"ok": False, "message": "파일이 비어 있거나 손상되었습니다."}), 400
    if not storage_util.check_magic_bytes(data, ext):
        return jsonify({"ok": False, "message": "파일 내용이 확장자와 일치하지 않습니다. 실제 JPG/PNG 이미지만 업로드해주세요."}), 400
    key = storage_util.build_popup_key(ext)
    try:
        storage_util.upload_doc(key, data)
    except Exception:
        app.logger.exception("팝업 이미지 업로드 실패")
        return jsonify({"ok": False, "message": "파일 저장 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요."}), 500
    return jsonify({"ok": True, "image_ref": key, "image_url": f"/api/popups/image/{key}"})


# ---- 팝업 공개 API (로그인 불필요) ----

def _is_mobile_ua(ua):
    """User-Agent로 모바일 여부를 대략 판별한다(완벽할 필요 없음 — 노출 필터용)."""
    ua = (ua or "").lower()
    return any(t in ua for t in ("mobile", "android", "iphone", "ipad", "ipod"))


@app.route("/api/popups/active")
def get_active_popup():
    """현재 노출 대상 팝업 1건(최신 등록순)을 반환. 없으면 popup: null.

    기기(User-Agent)·기간·audience는 서버에서 거르고,
    scope(home_only)는 현재 경로를 아는 프런트(header.js)가 판단한다.
    """
    is_mobile = _is_mobile_ua(request.headers.get("User-Agent"))
    device_col = "show_mobile" if is_mobile else "show_desktop"
    logged_in = bool(session.get("user_id"))
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT id, scope, display_type, image_ref, link_url, open_new_tab,
               width_px, close_mode
        FROM site_popups
        WHERE is_active = TRUE
          AND {device_col} = TRUE
          AND (start_at IS NULL OR start_at <= NOW())
          AND (end_at IS NULL OR end_at >= NOW())
          AND (audience = 'all' OR %s)
        ORDER BY id DESC
        LIMIT 1
    """, [logged_in])
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return jsonify({"ok": True, "popup": None})
    return jsonify({"ok": True, "popup": {
        "id": row["id"],
        "scope": row["scope"],
        "display_type": row["display_type"],
        "image_url": f"/api/popups/image/{row['image_ref']}",
        "link_url": row["link_url"],
        "open_new_tab": bool(row["open_new_tab"]),
        "width_px": row["width_px"] or 400,
        "close_mode": row["close_mode"],
    }})


@app.route("/api/popups/image/<path:key>")
def get_popup_image(key):
    """팝업 이미지 공개 프록시 — popups/… 형식 키만 허용(서류 등 다른 객체 접근 차단)."""
    if not storage_util.is_valid_popup_ref(key):
        abort(404)
    try:
        data = storage_util.download_bytes(key)
    except Exception:
        abort(404)
    ext = key.rsplit(".", 1)[-1].lower()
    mime = "image/png" if ext == "png" else "image/jpeg"
    resp = Response(data, mimetype=mime)
    resp.headers["Cache-Control"] = "public, max-age=3600"
    return resp


# ---- 공지사항 공개 API (로그인 불필요) ----
@app.route("/api/notices")
def get_notices():
    """공개 공지 목록 — 고정 우선 → 최신순. {total, page, size, items} 형태."""
    try:
        page = max(int(request.args.get("page", 1)), 1)
    except (ValueError, TypeError):
        page = 1
    try:
        size = min(max(int(request.args.get("size", 10)), 1), 100)
    except (ValueError, TypeError):
        size = 10
    offset = (page - 1) * size
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) c FROM notices")
    total = cur.fetchone()["c"]
    cur.execute("""
        SELECT id, title, body, is_pinned,
               to_char(created_at, 'YYYY-MM-DD') AS created_at
        FROM notices
        ORDER BY is_pinned DESC, created_at DESC, id DESC
        LIMIT %s OFFSET %s
    """, [size, offset])
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"total": total, "page": page, "size": size, "items": items})


# ---- 신청 승인 큐 (applications, 모두 require_admin) ----
# C/D 화면에서 들어온 중개사/운영업체 신청을 관리자가 승인/반려한다.
# 데이터 정확성 관련: 승인 시 agents/operators로 실제 INSERT되므로 중복 검사 후에만 상태를 바꾼다.
ADMIN_APP_SORT = {"id": "id", "created_at": "submitted_at", "submitted_at": "submitted_at"}


def _admin_app_filters():
    """신청 목록/엑셀 공용: sort_expr, order, WHERE절, 파라미터.
    상태 기본값은 submitted(대기중), status=all이면 전체. applicant_type로 유형 필터.
    검색어는 업체명/대표자/이메일 부분일치."""
    q = (request.args.get("q") or "").strip()
    sort_key = (request.args.get("sort") or "id").strip()
    sort_expr = ADMIN_APP_SORT.get(sort_key, "id")
    order = "DESC" if (request.args.get("order") or "asc").strip().lower() == "desc" else "ASC"
    conds, params = [], []
    status = (request.args.get("status") or "submitted").strip()
    if status and status != "all":
        conds.append("status = %s")
        params.append(status)
    atype = (request.args.get("applicant_type") or "all").strip()
    if atype in ("agent", "operator", "loan_consultant"):
        conds.append("applicant_type = %s")
        params.append(atype)
    if q:
        conds.append("(office_or_company_name ILIKE %s OR owner_name ILIKE %s OR email ILIKE %s)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    where = " AND ".join(conds) if conds else "1=1"
    return sort_expr, order, where, params


@app.route("/api/admin/applications")
@require_admin
def admin_applications_list():
    sort_expr, order, where_sql, params = _admin_app_filters()
    page, size, offset = _admin_paging()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) c FROM applications WHERE {where_sql}", params)
    total = cur.fetchone()["c"]
    cur.execute(f"""
        SELECT id, applicant_type, office_or_company_name, owner_name, reg_number,
               biz_reg_number, category, phone, email, preferred_region, preferred_building, status, reject_reason,
               doc_license_url, doc_office_reg_url, doc_biz_reg_url,
               doc_business_card_url, doc_biz_license_url, doc_logo_url, doc_photo_url,
               linked_operator_id,
               to_char(submitted_at, 'YYYY-MM-DD HH24:MI') AS submitted_at
        FROM applications
        WHERE {where_sql}
        ORDER BY {sort_expr} {order}, id ASC
        LIMIT %s OFFSET %s
    """, params + [size, offset])
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"total": total, "page": page, "size": size, "items": items})


# ---- 회원관리(통합 뷰): users/agents/operators/applications 를 한 목록으로 ----
# 읽기 전용 조회 API. 승인/반려 처리는 기존 /api/admin/applications/<id>/approve|reject 를
# 프런트에서 그대로 재사용한다(여기서 중복 구현하지 않음).
_MEMBER_GROUPS = {"all", "general", "agent", "operator", "loan_consultant", "pending"}

# 그룹별 SELECT — UNION ALL 을 위해 컬럼 구성을 통일한다.
# (id, member_type, name, email, group_label, phone, status, applicant_type, created_at, admin_tag, points)
# agent/operator 는 inactive(일괄 비활성화)도 목록에 보여야 관리자가 상태를 확인할 수 있다.
_MEMBER_SELECTS = {
    "general": """
        SELECT id, 'general' AS member_type, COALESCE(name, '-') AS name, email,
               '' AS group_label, NULL AS phone, COALESCE(status, 'active') AS status,
               NULL AS applicant_type, created_at, NULL::timestamp AS approved_at,
               admin_tag, COALESCE(points, 0) AS points, admin_memo
        FROM users
    """,
    "agent": """
        SELECT id, 'agent' AS member_type, owner_name AS name, email,
               office_name AS group_label, phone, status,
               NULL AS applicant_type, created_at, approved_at,
               admin_tag, NULL::integer AS points, admin_memo
        FROM agents WHERE status IN ('approved', 'inactive')
    """,
    "operator": """
        SELECT id, 'operator' AS member_type, owner_name AS name, email,
               (category || ' · ' || company_name) AS group_label, phone, status,
               NULL AS applicant_type, created_at, approved_at,
               admin_tag, NULL::integer AS points, admin_memo
        FROM operators WHERE status IN ('approved', 'inactive')
    """,
    "loan_consultant": """
        SELECT id, 'loan_consultant' AS member_type, owner_name AS name, email,
               office_name AS group_label, phone, status,
               NULL AS applicant_type, created_at, approved_at,
               admin_tag, NULL::integer AS points, admin_memo
        FROM loan_consultants WHERE status IN ('approved', 'inactive')
    """,
    "pending": """
        SELECT id, 'pending' AS member_type, owner_name AS name, email,
               CASE
                   WHEN applicant_type = 'operator' AND category IS NOT NULL AND category <> ''
                       THEN (category || ' · ' || office_or_company_name)
                   ELSE office_or_company_name
               END AS group_label, phone, status,
               applicant_type, submitted_at AS created_at, NULL::timestamp AS approved_at,
               NULL AS admin_tag, NULL::integer AS points, NULL AS admin_memo
        FROM applications WHERE status = 'submitted'
    """,
}


@app.route("/api/admin/members")
@require_admin
def admin_members_list():
    group = (request.args.get("group") or "all").strip()
    if group not in _MEMBER_GROUPS:
        return jsonify({"ok": False, "message": "group은 all|general|agent|operator|loan_consultant|pending 중 하나여야 합니다."}), 400
    q = (request.args.get("q") or "").strip()
    page, size, offset = _admin_paging()

    union_sql = " UNION ALL ".join(_MEMBER_SELECTS.values())
    where = []
    params = []
    if q:
        where.append("(m.name ILIKE %s OR m.email ILIKE %s)")
        like = f"%{q}%"
        params += [like, like]
    where_sql = (" WHERE " + " AND ".join(where)) if where else ""

    conn = get_conn()
    cur = conn.cursor()

    # 그룹별 인원수 (검색어 적용 — 목록 숫자와 일치하도록)
    cur.execute(f"""
        SELECT m.member_type, COUNT(*) AS c
        FROM ({union_sql}) m
        {where_sql}
        GROUP BY m.member_type
    """, params)
    counts = {r["member_type"]: r["c"] for r in cur.fetchall()}
    for k in ("general", "agent", "operator", "loan_consultant", "pending"):
        counts.setdefault(k, 0)
    counts["all"] = sum(counts.values())

    group_filter = where_sql
    group_params = list(params)
    if group != "all":
        group_filter = (group_filter + (" AND " if group_filter else " WHERE ")) + "m.member_type = %s"
        group_params.append(group)

    cur.execute(f"SELECT COUNT(*) AS c FROM ({union_sql}) m {group_filter}", group_params)
    total = cur.fetchone()["c"]

    cur.execute(f"""
        SELECT m.id, m.member_type, m.name, m.email, m.group_label, m.phone,
               m.status, m.applicant_type, m.admin_tag, m.points, m.admin_memo,
               to_char(m.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
               to_char(m.approved_at, 'YYYY-MM-DD HH24:MI') AS approved_at
        FROM ({union_sql}) m
        {group_filter}
        ORDER BY m.created_at DESC NULLS LAST, m.member_type, m.id DESC
        LIMIT %s OFFSET %s
    """, group_params + [size, offset])
    items = [dict(r) for r in cur.fetchall()]

    # 광고(revenue_records) 요약 — 현재 페이지의 파트너 행에만 배치 조회로 붙인다.
    partner_keys = [(it["member_type"], it["id"]) for it in items
                    if it["member_type"] in ("agent", "operator", "loan_consultant")]
    ads_map = {}
    if partner_keys:
        types = [t for t, _ in partner_keys]
        ids_ = [i for _, i in partner_keys]
        cur.execute("""
            SELECT partner_type, partner_id, product_type, payment_status, amount,
                   to_char(start_date, 'YY.MM.DD') AS start_date,
                   to_char(end_date, 'YY.MM.DD') AS end_date
            FROM revenue_records
            WHERE (partner_type, partner_id) IN (
                SELECT unnest(%s::text[]), unnest(%s::int[])
            )
            ORDER BY start_date DESC, id DESC
        """, [types, ids_])
        for r in cur.fetchall():
            k = (r["partner_type"], r["partner_id"])
            entry = ads_map.setdefault(k, {"count": 0, "latest": None})
            entry["count"] += 1
            if entry["latest"] is None:
                entry["latest"] = {
                    "product_type": r["product_type"], "payment_status": r["payment_status"],
                    "amount": int(r["amount"] or 0), "start_date": r["start_date"], "end_date": r["end_date"],
                }
    for it in items:
        it["ads"] = ads_map.get((it["member_type"], it["id"]))

    cur.close()
    conn.close()
    return jsonify({"total": total, "page": page, "size": size, "counts": counts, "items": items})


# ---- 회원 첨부서류 목록 (신청 시 올린 서류 → 승인 후에도 회원관리에서 열람) ----
# 승인된 회원은 applications.linked_*_id 로 원 신청서를 역추적한다.
_MEMBER_LINK_COLUMNS = {
    "agent": "linked_agent_id",
    "operator": "linked_operator_id",
    "loan_consultant": "linked_loan_consultant_id",
}
_DOC_LABELS = {
    "license": "자격증",
    "office_reg": "중개사무소 등록증",
    "biz_reg": "사업자등록증",
    "business_card": "명함",
    "biz_license": "영업허가증",
    "photo": "여권용 사진",
}


@app.route("/api/admin/members/<member_type>/<int:member_id>/docs")
@require_admin
def admin_member_docs(member_type, member_id):
    """회원의 신청 첨부서류 목록. 다운로드 URL은 기존 doc-url API(5분 서명)를 재사용한다.

    member_type='pending'이면 member_id가 곧 applications.id.
    승인된 회원(agent/operator/loan_consultant)은 linked_*_id로 최신 신청서를 찾는다.
    """
    conn = get_conn()
    cur = conn.cursor()
    if member_type == "pending":
        cur.execute("SELECT * FROM applications WHERE id=%s", [member_id])
    elif member_type in _MEMBER_LINK_COLUMNS:
        col = _MEMBER_LINK_COLUMNS[member_type]
        cur.execute(f"SELECT * FROM applications WHERE {col}=%s ORDER BY id DESC LIMIT 1", [member_id])
    else:
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "서류가 있는 회원유형이 아닙니다."}), 400
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return jsonify({"ok": True, "app_id": None, "docs": [],
                        "message": "연결된 신청서가 없습니다. (구버전 가입 등)"})
    docs = []
    for doc_key, col in _APP_DOC_COLUMNS.items():
        if row.get(col):
            docs.append({"doc": doc_key, "label": _DOC_LABELS.get(doc_key, doc_key)})
    return jsonify({
        "ok": True, "app_id": row["id"], "docs": docs,
        "applicant_name": row.get("office_or_company_name") or row.get("owner_name"),
        "submitted_at": row["submitted_at"].strftime("%Y-%m-%d %H:%M") if row.get("submitted_at") else None,
    })


# ---- 매출/광고 장부 (revenue_records) — 결제 연동 전 수동 기록 ----
_REVENUE_PARTNER_TYPES = {"agent", "operator", "loan_consultant"}
_REVENUE_PRODUCT_TYPES = {"building_slot", "priority_exposure"}
_REVENUE_STATUSES = {"대기", "완료", "만료"}
_REVENUE_PARTNER_TABLES = {"agent": "agents", "operator": "operators", "loan_consultant": "loan_consultants"}


def _revenue_validate(data, partial=False):
    """생성/수정 공통 검증. partial=True면 보낸 필드만 검사. (필드명, 값) dict 또는 오류 문자열."""
    out = {}
    def has(k):
        return k in data
    if not partial or has("product_type"):
        pt = (data.get("product_type") or "").strip()
        if pt not in _REVENUE_PRODUCT_TYPES:
            return None, "product_type은 building_slot|priority_exposure 중 하나여야 합니다."
        out["product_type"] = pt
    if not partial or has("start_date"):
        sd = (data.get("start_date") or "").strip()
        try:
            datetime.strptime(sd, "%Y-%m-%d")
        except ValueError:
            return None, "start_date는 YYYY-MM-DD 형식이어야 합니다."
        out["start_date"] = sd
    if has("end_date"):
        ed = (data.get("end_date") or "").strip()
        if ed:
            try:
                datetime.strptime(ed, "%Y-%m-%d")
            except ValueError:
                return None, "end_date는 YYYY-MM-DD 형식이어야 합니다."
            out["end_date"] = ed
        else:
            out["end_date"] = None
    if not partial or has("amount"):
        amt = data.get("amount")
        if not isinstance(amt, int) or isinstance(amt, bool) or amt < 0 or amt > 1_000_000_000:
            return None, "amount는 0 이상 10억 이하의 정수(원)여야 합니다."
        out["amount"] = amt
    if not partial or has("payment_status"):
        st = (data.get("payment_status") or "대기").strip()
        if st not in _REVENUE_STATUSES:
            return None, "payment_status는 대기|완료|만료 중 하나여야 합니다."
        out["payment_status"] = st
    if has("memo"):
        memo = (data.get("memo") or "").strip()
        if len(memo) > 500:
            return None, "메모는 500자 이내로 입력해주세요."
        out["memo"] = memo or None
    return out, None


@app.route("/api/admin/revenue-records")
@require_admin
def admin_revenue_records_list():
    """특정 파트너의 매출 기록 이력 (최신순)."""
    partner_type = (request.args.get("partner_type") or "").strip()
    if partner_type not in _REVENUE_PARTNER_TYPES:
        return jsonify({"ok": False, "message": "partner_type이 잘못되었습니다."}), 400
    try:
        partner_id = int(request.args.get("partner_id") or "")
    except ValueError:
        return jsonify({"ok": False, "message": "partner_id가 잘못되었습니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, partner_type, partner_id, product_type,
               to_char(start_date, 'YYYY-MM-DD') AS start_date,
               to_char(end_date, 'YYYY-MM-DD') AS end_date,
               amount, payment_status, memo,
               to_char(created_at, 'YYYY-MM-DD HH24:MI') AS created_at
        FROM revenue_records
        WHERE partner_type=%s AND partner_id=%s
        ORDER BY start_date DESC, id DESC
    """, [partner_type, partner_id])
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"ok": True, "items": items})


@app.route("/api/admin/revenue-records", methods=["POST"])
@require_admin
def admin_revenue_records_create():
    data = request.get_json(force=True, silent=True) or {}
    partner_type = (data.get("partner_type") or "").strip()
    if partner_type not in _REVENUE_PARTNER_TYPES:
        return jsonify({"ok": False, "message": "partner_type이 잘못되었습니다."}), 400
    try:
        partner_id = int(data.get("partner_id"))
    except (TypeError, ValueError):
        return jsonify({"ok": False, "message": "partner_id가 잘못되었습니다."}), 400
    fields, err = _revenue_validate(data, partial=False)
    if err:
        return jsonify({"ok": False, "message": err}), 400
    conn = get_conn()
    cur = conn.cursor()
    try:
        # 파트너 존재 검증 (다형 참조라 FK가 없으므로 애플리케이션에서 확인)
        cur.execute(f"SELECT id FROM {_REVENUE_PARTNER_TABLES[partner_type]} WHERE id=%s", [partner_id])
        if not cur.fetchone():
            return jsonify({"ok": False, "message": "존재하지 않는 파트너입니다."}), 404
        cur.execute("""
            INSERT INTO revenue_records
                (partner_type, partner_id, product_type, start_date, end_date, amount, payment_status, memo, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, [partner_type, partner_id, fields["product_type"], fields["start_date"],
              fields.get("end_date"), fields["amount"], fields.get("payment_status", "대기"),
              fields.get("memo"), session.get("admin_user_id")])
        new_id = cur.fetchone()["id"]
        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("매출 기록 생성 실패")
        return jsonify({"ok": False, "message": "매출 기록 저장 중 오류가 발생했습니다."}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/admin/revenue-records/<int:rec_id>", methods=["PUT"])
@require_admin
def admin_revenue_records_update(rec_id):
    data = request.get_json(force=True, silent=True) or {}
    fields, err = _revenue_validate(data, partial=True)
    if err:
        return jsonify({"ok": False, "message": err}), 400
    if not fields:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다."}), 400
    sets = [f"{k}=%s" for k in fields]
    params = list(fields.values()) + [rec_id]
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"UPDATE revenue_records SET {', '.join(sets)} WHERE id=%s", params)
        if cur.rowcount == 0:
            conn.rollback()
            return jsonify({"ok": False, "message": "기록을 찾을 수 없습니다."}), 404
        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("매출 기록 수정 실패 (id=%s)", rec_id)
        return jsonify({"ok": False, "message": "매출 기록 수정 중 오류가 발생했습니다."}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/revenue-records/<int:rec_id>", methods=["DELETE"])
@require_admin
def admin_revenue_records_delete(rec_id):
    """오입력 정정용 삭제 (장부이므로 신중히 — 프런트에서 확인창 필수)."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM revenue_records WHERE id=%s", [rec_id])
        if cur.rowcount == 0:
            conn.rollback()
            return jsonify({"ok": False, "message": "기록을 찾을 수 없습니다."}), 404
        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("매출 기록 삭제 실패 (id=%s)", rec_id)
        return jsonify({"ok": False, "message": "매출 기록 삭제 중 오류가 발생했습니다."}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/revenue-summary")
@require_admin
def admin_revenue_summary():
    """매출관리 화면용 집계 — 월(start_date 기준)×상품×파트너유형별 건수/금액.
    '완료'만 매출로 집계하고, '대기' 금액은 참고용으로 함께 반환한다."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT to_char(start_date, 'YYYY-MM') AS ym, partner_type, product_type,
               COUNT(*) AS cnt,
               COALESCE(SUM(amount) FILTER (WHERE payment_status='완료'), 0) AS amount_done,
               COALESCE(SUM(amount) FILTER (WHERE payment_status='대기'), 0) AS amount_pending
        FROM revenue_records
        GROUP BY ym, partner_type, product_type
        ORDER BY ym DESC, partner_type, product_type
    """)
    rows = [dict(r) for r in cur.fetchall()]
    for r in rows:
        r["amount_done"] = int(r["amount_done"])
        r["amount_pending"] = int(r["amount_pending"])
        r["cnt"] = int(r["cnt"])
    cur.close()
    conn.close()
    return jsonify({"ok": True, "rows": rows})


# ---- 회원 일괄 관리 (아임웹 스타일) ----
# 프런트는 선택된 행마다 {type: member_type, id: 원본 테이블 id}를 보낸다.
_BULK_MEMBER_TABLES = {"general": "users", "agent": "agents", "operator": "operators", "loan_consultant": "loan_consultants"}


def _bulk_parse_ids(data, allowed_types):
    """요청 본문의 ids 배열을 [(type, int id)] 로 정규화. 잘못된 항목은 무시하지 않고 오류."""
    raw = data.get("ids")
    if not isinstance(raw, list) or not raw:
        return None, "ids 배열이 비어 있습니다."
    out = []
    for item in raw:
        if not isinstance(item, dict):
            return None, "ids 항목 형식이 잘못되었습니다."
        t = item.get("type")
        try:
            i = int(item.get("id"))
        except (TypeError, ValueError):
            return None, "ids 항목의 id가 숫자가 아닙니다."
        if t not in allowed_types:
            return None, f"허용되지 않는 회원유형입니다: {t}"
        out.append((t, i))
    if len(out) > 500:
        return None, "한 번에 500명까지만 처리할 수 있습니다."
    return out, None


@app.route("/api/admin/members/bulk-approve", methods=["POST"])
@require_admin
def admin_members_bulk_approve():
    """승인대기(pending) 선택 건 일괄 승인 — 기존 단건 승인 함수를 그대로 반복 호출(로직 재사용)."""
    data = request.get_json(force=True, silent=True) or {}
    ids, err = _bulk_parse_ids(data, {"pending"})
    if err:
        return jsonify({"ok": False, "message": err}), 400
    results = []
    ok_count = 0
    for _t, app_id in ids:
        try:
            resp = admin_applications_approve(app_id)
            # view 함수는 Response 또는 (Response, status) 튜플을 반환한다.
            if isinstance(resp, tuple):
                body, status = resp[0].get_json(), resp[1]
            else:
                body, status = resp.get_json(), 200
        except Exception:
            app.logger.exception("일괄 승인 중 오류 (application id=%s)", app_id)
            body, status = {"ok": False, "message": "처리 중 오류"}, 500
        item = {"id": app_id, "ok": bool(status == 200 and body.get("ok"))}
        if item["ok"]:
            ok_count += 1
            for k in ("sms_sent", "sms_message", "temp_password"):
                if k in (body or {}):
                    item[k] = body[k]
        else:
            item["message"] = (body or {}).get("message", "승인 실패")
        results.append(item)
    return jsonify({"ok": True, "success": ok_count, "failed": len(ids) - ok_count, "results": results})


@app.route("/api/admin/members/bulk-reject", methods=["POST"])
@require_admin
def admin_members_bulk_reject():
    """승인대기 선택 건 일괄 반려 — 기존 단건 반려 함수 재사용. reason은 본문에서 공유."""
    data = request.get_json(force=True, silent=True) or {}
    reason = (data.get("reason") or "").strip()
    if not reason:
        return jsonify({"ok": False, "message": "반려 사유(reason)는 필수입니다."}), 400
    ids, err = _bulk_parse_ids(data, {"pending"})
    if err:
        return jsonify({"ok": False, "message": err}), 400
    results = []
    ok_count = 0
    for _t, app_id in ids:
        try:
            # admin_applications_reject 는 request.get_json()에서 reason을 읽는데,
            # 이 요청 본문에 이미 reason이 들어 있으므로 그대로 재사용된다.
            resp = admin_applications_reject(app_id)
            if isinstance(resp, tuple):
                body, status = resp[0].get_json(), resp[1]
            else:
                body, status = resp.get_json(), 200
        except Exception:
            app.logger.exception("일괄 반려 중 오류 (application id=%s)", app_id)
            body, status = {"ok": False, "message": "처리 중 오류"}, 500
        item = {"id": app_id, "ok": bool(status == 200 and body.get("ok"))}
        if not item["ok"]:
            item["message"] = (body or {}).get("message", "반려 실패")
        else:
            ok_count += 1
        results.append(item)
    return jsonify({"ok": True, "success": ok_count, "failed": len(ids) - ok_count, "results": results})


@app.route("/api/admin/members/bulk-tag", methods=["POST"])
@require_admin
def admin_members_bulk_tag():
    """선택 회원들의 admin_tag 일괄 지정. tag를 빈 값으로 보내면 태그 해제(NULL)."""
    data = request.get_json(force=True, silent=True) or {}
    tag = (data.get("tag") or "").strip() or None
    if tag and len(tag) > 50:
        return jsonify({"ok": False, "message": "태그는 50자 이내로 입력해주세요."}), 400
    ids, err = _bulk_parse_ids(data, set(_BULK_MEMBER_TABLES))
    if err:
        return jsonify({"ok": False, "message": err}), 400
    by_type = {}
    for t, i in ids:
        by_type.setdefault(t, []).append(i)
    conn = get_conn()
    cur = conn.cursor()
    updated = 0
    try:
        for t, id_list in by_type.items():
            table = _BULK_MEMBER_TABLES[t]
            cur.execute(f"UPDATE {table} SET admin_tag=%s WHERE id = ANY(%s)", [tag, id_list])
            updated += cur.rowcount
        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("일괄 태그 지정 실패")
        return jsonify({"ok": False, "message": "태그 지정 중 오류가 발생했습니다."}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "success": updated, "failed": len(ids) - updated})


@app.route("/api/admin/members/bulk-sms", methods=["POST"])
@require_admin
@limiter.limit("3 per hour")
def admin_members_bulk_sms():
    """중개사/운영업체 선택 대상에게 커스텀 문구 SMS 일괄 발송 (남용 방지 3회/시간)."""
    data = request.get_json(force=True, silent=True) or {}
    message = (data.get("message") or "").strip()
    if not message:
        return jsonify({"ok": False, "message": "보낼 메시지를 입력해주세요."}), 400
    if len(message) > 2000:
        return jsonify({"ok": False, "message": "메시지는 2000자 이내로 입력해주세요."}), 400
    ids, err = _bulk_parse_ids(data, {"agent", "operator", "loan_consultant"})
    if err:
        return jsonify({"ok": False, "message": err}), 400
    by_type = {}
    for t, i in ids:
        by_type.setdefault(t, []).append(i)
    conn = get_conn()
    cur = conn.cursor()
    targets = []
    for t, id_list in by_type.items():
        table = _BULK_MEMBER_TABLES[t]
        cur.execute(f"SELECT id, phone FROM {table} WHERE id = ANY(%s)", [id_list])
        for r in cur.fetchall():
            targets.append((t, r["id"], r["phone"]))
    cur.close()
    conn.close()
    ok_count = 0
    results = []
    for t, mid, phone in targets:
        if not (phone or "").strip():
            results.append({"type": t, "id": mid, "ok": False, "message": "전화번호 없음"})
            continue
        sent, msg = send_sms(phone, message)
        if sent:
            ok_count += 1
        results.append({"type": t, "id": mid, "ok": bool(sent), "message": msg})
    return jsonify({"ok": True, "success": ok_count, "failed": len(ids) - ok_count, "results": results})


@app.route("/api/admin/members/bulk-notify", methods=["POST"])
@require_admin
def admin_members_bulk_notify():
    """일반회원 선택 대상에게 인앱 알림(notifications) 일괄 생성 — 공지성이라 건물 정보는 NULL."""
    data = request.get_json(force=True, silent=True) or {}
    title = (data.get("title") or "").strip()
    body = (data.get("body") or "").strip() or None
    if not title:
        return jsonify({"ok": False, "message": "알림 제목을 입력해주세요."}), 400
    if len(title) > 200:
        return jsonify({"ok": False, "message": "제목은 200자 이내로 입력해주세요."}), 400
    ids, err = _bulk_parse_ids(data, {"general"})
    if err:
        return jsonify({"ok": False, "message": err}), 400
    user_ids = [i for _t, i in ids]
    conn = get_conn()
    cur = conn.cursor()
    try:
        # 존재하는 회원에게만 생성 (탈퇴/삭제된 id는 자동 제외)
        cur.execute("""
            INSERT INTO notifications (user_id, title, body)
            SELECT id, %s, %s FROM users WHERE id = ANY(%s)
        """, [title, body, user_ids])
        created = cur.rowcount
        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("일괄 알림 생성 실패")
        return jsonify({"ok": False, "message": "알림 생성 중 오류가 발생했습니다."}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "success": created, "failed": len(user_ids) - created})


@app.route("/api/admin/members/bulk-points", methods=["POST"])
@require_admin
def admin_members_bulk_points():
    """일반회원 포인트 일괄 지급/차감 — 단일 트랜잭션 + point_transactions 감사로그."""
    data = request.get_json(force=True, silent=True) or {}
    amount = data.get("amount")
    if not isinstance(amount, int) or isinstance(amount, bool) or amount == 0:
        return jsonify({"ok": False, "message": "포인트(amount)는 0이 아닌 정수여야 합니다."}), 400
    if abs(amount) > 10_000_000:
        return jsonify({"ok": False, "message": "한 번에 처리할 수 있는 포인트는 ±1,000만 이내입니다."}), 400
    reason = (data.get("reason") or "").strip()
    if not reason:
        return jsonify({"ok": False, "message": "사유(reason)는 필수입니다."}), 400
    ids, err = _bulk_parse_ids(data, {"general"})
    if err:
        return jsonify({"ok": False, "message": err}), 400
    user_ids = [i for _t, i in ids]
    admin_id = session.get("admin_user_id")
    conn = get_conn()
    cur = conn.cursor()
    try:
        # 잔액 가감 + 이력 기록을 한 트랜잭션으로 — 중간 실패 시 전체 롤백.
        cur.execute("""
            UPDATE users SET points = COALESCE(points, 0) + %s
            WHERE id = ANY(%s)
            RETURNING id
        """, [amount, user_ids])
        updated_ids = [r["id"] for r in cur.fetchall()]
        if updated_ids:
            cur.execute("""
                INSERT INTO point_transactions (user_id, amount, reason, admin_id)
                SELECT unnest(%s::int[]), %s, %s, %s
            """, [updated_ids, amount, reason, admin_id])
        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("일괄 포인트 처리 실패")
        return jsonify({"ok": False, "message": "포인트 처리 중 오류가 발생했습니다."}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "success": len(updated_ids), "failed": len(user_ids) - len(updated_ids)})


@app.route("/api/admin/members/bulk-deactivate", methods=["POST"])
@require_admin
def admin_members_bulk_deactivate():
    """일괄 비활성화(소프트 삭제) — DELETE 없이 상태값만 변경.
    users → 'withdrawn'(기존 회원탈퇴와 동일), agents/operators → 'inactive'(로그인·공개프로필 차단).
    reason(선택)이 오면 각 회원의 admin_memo에 '[날짜] 사유' 형태로 이어붙인다."""
    data = request.get_json(force=True, silent=True) or {}
    ids, err = _bulk_parse_ids(data, set(_BULK_MEMBER_TABLES))
    if err:
        return jsonify({"ok": False, "message": err}), 400
    reason = (data.get("reason") or "").strip()[:500]
    memo_line = None
    if reason:
        from datetime import date
        memo_line = f"[{date.today().isoformat()}] {reason}"
    by_type = {}
    for t, i in ids:
        by_type.setdefault(t, []).append(i)
    # 상태값: 일반회원은 탈퇴(withdrawn), 파트너는 비활성(inactive)
    new_status = {"general": "withdrawn", "agent": "inactive", "operator": "inactive", "loan_consultant": "inactive"}
    conn = get_conn()
    cur = conn.cursor()
    updated = 0
    try:
        for t, id_list in by_type.items():
            table = _BULK_MEMBER_TABLES[t]
            if memo_line:
                cur.execute(
                    f"""UPDATE {table}
                        SET status=%s,
                            admin_memo = CASE WHEN admin_memo IS NULL OR admin_memo = '' THEN %s
                                              ELSE admin_memo || E'\\n' || %s END
                        WHERE id = ANY(%s)""",
                    [new_status[t], memo_line, memo_line, id_list])
            else:
                cur.execute(f"UPDATE {table} SET status=%s WHERE id = ANY(%s)", [new_status[t], id_list])
            updated += cur.rowcount
        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("일괄 비활성화 실패")
        return jsonify({"ok": False, "message": "비활성화 처리 중 오류가 발생했습니다."}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "success": updated, "failed": len(ids) - updated})


@app.route("/api/admin/members/bulk-reactivate", methods=["POST"])
@require_admin
def admin_members_bulk_reactivate():
    """일괄 재활성화 — bulk-deactivate의 반대.
    users → 'active'(탈퇴 취소), agents/operators/loan_consultants → 'approved'(로그인·공개페이지 복구)."""
    data = request.get_json(force=True, silent=True) or {}
    ids, err = _bulk_parse_ids(data, set(_BULK_MEMBER_TABLES))
    if err:
        return jsonify({"ok": False, "message": err}), 400
    by_type = {}
    for t, i in ids:
        by_type.setdefault(t, []).append(i)
    new_status = {"general": "active", "agent": "approved", "operator": "approved", "loan_consultant": "approved"}
    # 비활성 상태였던 계정만 복구 — 승인대기/반려 등 다른 상태를 실수로 '승인'으로 바꾸지 않도록 방어
    prev_status = {"general": "withdrawn", "agent": "inactive", "operator": "inactive", "loan_consultant": "inactive"}
    conn = get_conn()
    cur = conn.cursor()
    updated = 0
    try:
        for t, id_list in by_type.items():
            table = _BULK_MEMBER_TABLES[t]
            cur.execute(f"UPDATE {table} SET status=%s WHERE id = ANY(%s) AND status=%s",
                        [new_status[t], id_list, prev_status[t]])
            updated += cur.rowcount
        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("일괄 재활성화 실패")
        return jsonify({"ok": False, "message": "재활성화 처리 중 오류가 발생했습니다."}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "success": updated, "failed": len(ids) - updated})


@app.route("/api/admin/members/<member_type>/<int:member_id>/memo", methods=["PUT"])
@require_admin
def admin_member_memo_put(member_type, member_id):
    """회원 메모(admin_memo) 수정 — 빈 값이면 삭제. 매물의뢰 비고와 동일 패턴."""
    table = _BULK_MEMBER_TABLES.get(member_type)
    if not table:
        return jsonify({"ok": False, "message": "잘못된 회원유형입니다."}), 400
    data = request.get_json(force=True, silent=True) or {}
    if "admin_memo" not in data:
        return jsonify({"ok": False, "message": "수정할 항목이 없습니다. (admin_memo)"}), 400
    memo = (data.get("admin_memo") or "").strip()[:2000] or None
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"UPDATE {table} SET admin_memo=%s WHERE id=%s", [memo, member_id])
        if cur.rowcount == 0:
            conn.rollback()
            return jsonify({"ok": False, "message": "회원을 찾을 수 없습니다."}), 404
        conn.commit()
    except Exception:
        conn.rollback()
        app.logger.exception("회원 메모 수정 실패")
        return jsonify({"ok": False, "message": "메모 저장 중 오류가 발생했습니다."}), 500
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True})


# 신청서 doc_type → applications 테이블 컬럼 매핑 (관리자 서류 열람용 화이트리스트)
_APP_DOC_COLUMNS = {
    "license": "doc_license_url",
    "office_reg": "doc_office_reg_url",
    "biz_reg": "doc_biz_reg_url",
    "business_card": "doc_business_card_url",
    "biz_license": "doc_biz_license_url",
    "photo": "doc_photo_url",
}


@app.route("/api/admin/applications/<int:app_id>/doc-url")
@require_admin
def admin_application_doc_url(app_id):
    """신청 서류의 서명된 임시 열람 URL(5분) 발급 — 관리자 전용.

    doc 파라미터는 화이트리스트(_APP_DOC_COLUMNS)로만 컬럼을 선택하므로
    임의 컬럼/키 접근이 불가능하다.
    """
    doc_type = (request.args.get("doc") or "").strip()
    col = _APP_DOC_COLUMNS.get(doc_type)
    if not col:
        return jsonify({"ok": False, "message": "알 수 없는 문서 종류입니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT {col} AS ref FROM applications WHERE id=%s", [app_id])
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return jsonify({"ok": False, "message": "신청 내역을 찾을 수 없습니다."}), 404
    ref = row["ref"]
    if not ref or not storage_util.is_valid_doc_ref(ref):
        return jsonify({"ok": False, "message": "첨부된 서류가 없습니다."}), 404
    try:
        url = storage_util.signed_get_url(ref, ttl_sec=300)
    except Exception:
        app.logger.exception("서명 URL 발급 실패 (application %s, %s)", app_id, doc_type)
        return jsonify({"ok": False, "message": "서류 열람 URL 발급에 실패했습니다."}), 500
    return jsonify({"ok": True, "url": url})


@app.route("/api/admin/operators/<int:operator_id>/logo", methods=["PUT"])
@require_admin
def admin_operator_logo_put(operator_id):
    """운영지원업체 로고 등록/수정/삭제 (건물마스터 수정처럼 PUT).

    - multipart/form-data + file: 이미지(JPG/PNG, 5MB 이하) 업로드 후 logo_url 갱신
    - JSON {"clear": true}: 로고 제거 (파트너 소개 섹션에서 내려감)
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id FROM operators WHERE id=%s", [operator_id])
    if not cur.fetchone():
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "존재하지 않는 업체입니다."}), 404

    f = request.files.get("file")
    if not f:
        data = request.get_json(silent=True) or {}
        if data.get("clear") is True:
            cur.execute("UPDATE operators SET logo_url=NULL WHERE id=%s", [operator_id])
            conn.commit()
            cur.close()
            conn.close()
            return jsonify({"ok": True, "logo_url": None})
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "로고 파일을 첨부하거나 {\"clear\": true}를 보내주세요."}), 400

    ext = f.filename.rsplit(".", 1)[-1].lower() if f.filename and "." in f.filename else ""
    if ext not in storage_util.LOGO_EXTENSIONS:
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "로고는 JPG 또는 PNG 이미지 파일만 업로드할 수 있습니다."}), 400
    data_bytes = f.read(storage_util.MAX_FILE_BYTES + 1)
    if len(data_bytes) > storage_util.MAX_FILE_BYTES:
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "파일 크기는 5MB 이하여야 합니다."}), 400
    if len(data_bytes) < 16 or not storage_util.check_magic_bytes(data_bytes, ext):
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "파일 내용이 확장자와 일치하지 않습니다. 실제 JPG/PNG 파일만 업로드해주세요."}), 400

    key = storage_util.build_doc_key("operator", "logo", ext)
    try:
        storage_util.upload_doc(key, data_bytes)
    except Exception:
        app.logger.exception("운영업체 로고 업로드 실패 (operator_id=%s)", operator_id)
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "파일 저장 중 오류가 발생했습니다. 잠시 후 다시 시도해주세요."}), 500

    cur.execute("UPDATE operators SET logo_url=%s WHERE id=%s", [key, operator_id])
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True, "logo_url": key})


@app.route("/api/admin/applications/<int:app_id>/approve", methods=["POST"])
@require_admin
def admin_applications_approve(app_id):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM applications WHERE id=%s", [app_id])
    ap = cur.fetchone()
    if not ap:
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "존재하지 않는 신청입니다."}), 404
    if ap["status"] != "submitted":
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "이미 처리된 신청입니다."}), 400

    atype = ap["applicant_type"]
    temp_pw = None
    sms_sent = False
    sms_msg = None
    try:
        if atype == "agent":
            # 등록번호(reg_number) 중복이면 승인 불가 — applications 상태는 그대로 둔다.
            cur.execute("SELECT id FROM agents WHERE reg_number=%s", [ap["reg_number"]])
            if cur.fetchone():
                cur.close()
                conn.close()
                return jsonify({"ok": False, "message": "이미 등록된 중개사무소입니다."}), 400
            # subdomain_slug: 전화번호 숫자만 추출, 중복이면 -2, -3 … 붙여 유니크화.
            # 동시 승인 경쟁 대비: SAVEPOINT + UNIQUE 충돌 시 새 slug로 재시도(최대 5회).
            base_slug = re.sub(r"\D", "", ap["phone"] or "") or f"agent{app_id}"
            # 임시 비밀번호(랜덤 8자리 영숫자) — 해시만 저장
            alphabet = "abcdefghjkmnpqrstuvwxyzABCDEFGHJKMNPQRSTUVWXYZ23456789"
            temp_pw = "".join(_secrets.choice(alphabet) for _ in range(8))
            pw_hash = generate_password_hash(temp_pw)
            created_id = None
            n = 2
            slug = base_slug
            for _attempt in range(5):
                # 다음 빈 slug 후보 탐색
                while True:
                    cur.execute("SELECT 1 FROM agents WHERE subdomain_slug=%s", [slug])
                    if not cur.fetchone():
                        break
                    slug = f"{base_slug}-{n}"
                    n += 1
                cur.execute("SAVEPOINT sp_agent_insert")
                try:
                    cur.execute("""
                        INSERT INTO agents
                            (office_name, owner_name, reg_number, biz_reg_number,
                             phone, email, status, subdomain_slug, password_hash,
                             photo_url, approved_at)
                        VALUES (%s, %s, %s, %s, %s, %s, 'approved', %s, %s, %s, NOW())
                        RETURNING id
                    """, [ap["office_or_company_name"], ap["owner_name"], _digits_only(ap["reg_number"]) or ap["reg_number"],
                          _digits_only(ap["biz_reg_number"]) or None, _digits_only(ap["phone"]), ap["email"], slug, pw_hash,
                          ap.get("doc_photo_url")])
                    created_id = cur.fetchone()["id"]
                    cur.execute("RELEASE SAVEPOINT sp_agent_insert")
                    break
                except psycopg2_errors.UniqueViolation as ue:
                    cur.execute("ROLLBACK TO SAVEPOINT sp_agent_insert")
                    cname = getattr(getattr(ue, "diag", None), "constraint_name", "") or ""
                    if "slug" in cname:
                        # 동시 승인으로 slug 선점됨 — 다음 후보로 재시도
                        slug = f"{base_slug}-{n}"
                        n += 1
                        continue
                    raise  # reg_number 등 다른 UNIQUE 충돌은 일반 오류로 처리
            if created_id is None:
                conn.rollback()
                cur.close()
                conn.close()
                return jsonify({"ok": False, "message": "페이지 주소(slug) 발급에 실패했습니다. 다시 시도해주세요."}), 409
            cur.execute(
                "UPDATE applications SET status='approved', reviewed_at=NOW(), linked_agent_id=%s WHERE id=%s",
                [created_id, app_id],
            )
            # 희망건물(preferred_building_id)이 있으면 담당중개사 자동 배정.
            # 이미 그 건물에 approved 담당중개사가 있으면 조용히 건너뛴다
            # (관리자는 필요 시 건물 상세에서 기존 담당중개사 카드로 확인).
            pref_bid = ap.get("preferred_building_id")
            if pref_bid:
                # 동시 승인 경쟁(TOCTOU) 방지: 건물 단위 어드바이저리 락으로
                # "기존 담당 확인 → 배정" 구간을 직렬화한다 (트랜잭션 종료 시 자동 해제).
                cur.execute("SELECT pg_advisory_xact_lock(%s, %s)", [911001, pref_bid])
                cur.execute("""
                    SELECT 1
                    FROM agent_buildings ab
                    JOIN agents a ON a.id = ab.agent_id AND a.status = 'approved'
                    WHERE ab.master_building_id = %s
                    LIMIT 1
                """, [pref_bid])
                if not cur.fetchone():
                    cur.execute("SAVEPOINT sp_agent_assign")
                    try:
                        cur.execute("""
                            INSERT INTO agent_buildings (agent_id, master_building_id)
                            VALUES (%s, %s)
                            ON CONFLICT ON CONSTRAINT agent_buildings_agent_building_unique DO NOTHING
                        """, [created_id, pref_bid])
                        cur.execute("RELEASE SAVEPOINT sp_agent_assign")
                    except Exception:
                        # 건물 삭제 등으로 실패해도 승인 자체는 유지 (자동 배정은 부가 기능)
                        cur.execute("ROLLBACK TO SAVEPOINT sp_agent_assign")
                        app.logger.exception("승인 시 담당건물 자동 배정 실패 (application=%s, building=%s)", app_id, pref_bid)
        elif atype == "operator":
            # 운영업체는 이메일 기준 중복 검사. category는 NOT NULL이라 값이 없으면 승인 불가.
            if not (ap["category"] or "").strip():
                cur.close()
                conn.close()
                return jsonify({"ok": False, "message": "업종 정보가 없어 승인할 수 없습니다."}), 400
            cur.execute("SELECT id FROM operators WHERE email=%s", [ap["email"]])
            if cur.fetchone():
                cur.close()
                conn.close()
                return jsonify({"ok": False, "message": "이미 등록된 운영지원업체입니다."}), 400
            # subdomain_slug + 임시비밀번호 — agent 승인 로직과 완전히 동일한 패턴.
            base_slug = re.sub(r"\D", "", ap["phone"] or "") or f"operator{app_id}"
            alphabet = "abcdefghjkmnpqrstuvwxyzABCDEFGHJKMNPQRSTUVWXYZ23456789"
            temp_pw = "".join(_secrets.choice(alphabet) for _ in range(8))
            pw_hash = generate_password_hash(temp_pw)
            created_id = None
            n = 2
            slug = base_slug
            for _attempt in range(5):
                # 다음 빈 slug 후보 탐색
                while True:
                    cur.execute("SELECT 1 FROM operators WHERE subdomain_slug=%s", [slug])
                    if not cur.fetchone():
                        break
                    slug = f"{base_slug}-{n}"
                    n += 1
                cur.execute("SAVEPOINT sp_operator_insert")
                try:
                    cur.execute("""
                        INSERT INTO operators
                            (company_name, owner_name, category, biz_reg_number,
                             phone, email, website_url, status, subdomain_slug,
                             password_hash, logo_url, approved_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, 'approved', %s, %s, %s, NOW())
                        RETURNING id
                    """, [ap["office_or_company_name"], ap["owner_name"], ap["category"],
                          _digits_only(ap["biz_reg_number"]) or None, _digits_only(ap["phone"]), ap["email"], ap["website_url"],
                          slug, pw_hash, ap.get("doc_logo_url")])
                    created_id = cur.fetchone()["id"]
                    cur.execute("RELEASE SAVEPOINT sp_operator_insert")
                    break
                except psycopg2_errors.UniqueViolation as ue:
                    cur.execute("ROLLBACK TO SAVEPOINT sp_operator_insert")
                    cname = getattr(getattr(ue, "diag", None), "constraint_name", "") or ""
                    if "slug" in cname:
                        # 동시 승인으로 slug 선점됨 — 다음 후보로 재시도
                        slug = f"{base_slug}-{n}"
                        n += 1
                        continue
                    raise
            if created_id is None:
                conn.rollback()
                cur.close()
                conn.close()
                return jsonify({"ok": False, "message": "페이지 주소(slug) 발급에 실패했습니다. 다시 시도해주세요."}), 409
            cur.execute(
                "UPDATE applications SET status='approved', reviewed_at=NOW(), linked_operator_id=%s WHERE id=%s",
                [created_id, app_id],
            )
        elif atype == "loan_consultant":
            # 대출상담사: 대출모집인 등록번호(reg_number 컬럼 재사용) 중복이면 승인 불가.
            # 관리자는 loanconsultant.or.kr에서 등록 여부 확인 후 승인하는 것이 전제.
            if not (ap["reg_number"] or "").strip():
                cur.close()
                conn.close()
                return jsonify({"ok": False, "message": "대출모집인 등록번호가 없어 승인할 수 없습니다."}), 400
            cur.execute("SELECT id FROM loan_consultants WHERE license_number=%s", [ap["reg_number"]])
            if cur.fetchone():
                cur.close()
                conn.close()
                return jsonify({"ok": False, "message": "이미 등록된 대출상담사입니다."}), 400
            # 이메일(로그인 ID) 중복도 승인 전에 차단 — 대소문자 무시
            cur.execute("SELECT id FROM loan_consultants WHERE LOWER(email)=LOWER(%s)", [ap["email"]])
            if cur.fetchone():
                cur.close()
                conn.close()
                return jsonify({"ok": False, "message": "이미 등록된 이메일의 대출상담사가 있습니다."}), 400
            # subdomain_slug — agent 승인 로직과 동일 패턴 (전화번호 기반, 충돌 시 -2, -3 …)
            base_slug = re.sub(r"\D", "", ap["phone"] or "") or f"loan{app_id}"
            # 임시 비밀번호(랜덤 8자리 영숫자) — agent/operator와 동일 패턴, 해시만 저장
            alphabet = "abcdefghjkmnpqrstuvwxyzABCDEFGHJKMNPQRSTUVWXYZ23456789"
            temp_pw = "".join(_secrets.choice(alphabet) for _ in range(8))
            pw_hash = generate_password_hash(temp_pw)
            created_id = None
            n = 2
            slug = base_slug
            for _attempt in range(5):
                while True:
                    cur.execute("SELECT 1 FROM loan_consultants WHERE subdomain_slug=%s", [slug])
                    if not cur.fetchone():
                        break
                    slug = f"{base_slug}-{n}"
                    n += 1
                cur.execute("SAVEPOINT sp_loan_insert")
                try:
                    cur.execute("""
                        INSERT INTO loan_consultants
                            (office_name, owner_name, license_number, biz_reg_number,
                             phone, email, status, subdomain_slug, password_hash, approved_at)
                        VALUES (%s, %s, %s, %s, %s, %s, 'approved', %s, %s, NOW())
                        RETURNING id
                    """, [ap["office_or_company_name"], ap["owner_name"], ap["reg_number"],
                          _digits_only(ap["biz_reg_number"]) or None, _digits_only(ap["phone"]), ap["email"], slug, pw_hash])
                    created_id = cur.fetchone()["id"]
                    cur.execute("RELEASE SAVEPOINT sp_loan_insert")
                    break
                except psycopg2_errors.UniqueViolation as ue:
                    cur.execute("ROLLBACK TO SAVEPOINT sp_loan_insert")
                    cname = getattr(getattr(ue, "diag", None), "constraint_name", "") or ""
                    if "slug" in cname:
                        slug = f"{base_slug}-{n}"
                        n += 1
                        continue
                    raise
            if created_id is None:
                conn.rollback()
                cur.close()
                conn.close()
                return jsonify({"ok": False, "message": "페이지 주소(slug) 발급에 실패했습니다. 다시 시도해주세요."}), 409
            cur.execute(
                "UPDATE applications SET status='approved', reviewed_at=NOW(), linked_loan_consultant_id=%s WHERE id=%s",
                [created_id, app_id],
            )
        else:
            cur.close()
            conn.close()
            return jsonify({"ok": False, "message": "알 수 없는 신청 유형입니다."}), 400
    except Exception:
        app.logger.exception("신청 승인 처리 실패 (application id=%s)", app_id)
        conn.rollback()
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "승인 처리 중 오류가 발생했습니다."}), 400

    conn.commit()
    cur.close()
    conn.close()

    if atype == "loan_consultant":
        # 승인 시 로그인 계정(이메일 ID + 임시비밀번호)도 함께 안내 — agent/operator와 동일 패턴
        # 링크 도메인은 관리자 접속 주소(request.host_url)가 아닌 고정 공개 URL 사용
        domain = os.environ.get("PUBLIC_BASE_URL", "https://homenstay.com").rstrip("/")
        sms_body = (
            f"[홈앤스테이] 대출상담사 승인 완료. 로그인ID(이메일): {ap['email']} / "
            f"임시비밀번호: {temp_pw} / 로그인: {domain}/loan-consultant/login — "
            f"최초 로그인 후 반드시 비밀번호를 변경해주세요."
        )
        sms_sent, sms_msg = send_sms(ap["phone"], sms_body)
        # 승인 이메일도 함께 발송 (SMS와 동일 내용) — 실패해도 승인은 확정
        email_sent, _ = _send_approval_email("loan_consultant", ap["email"], ap["email"], temp_pw, f"{domain}/loan-consultant/login")
        resp = {"ok": True, "created_id": created_id, "sms_sent": sms_sent, "sms_message": sms_msg, "email_sent": email_sent}
        if not sms_sent:
            # 문자 실패 시 관리자 화면에서 임시비밀번호를 직접 전달할 수 있게 응답에 포함
            resp["temp_password"] = temp_pw
        return jsonify(resp)

    if atype in ("agent", "operator"):
        # 승인 완료 후 문자 발송 — 실패해도 승인은 이미 확정(예외 없음, (ok, msg) 반환)
        # 링크 도메인은 관리자 접속 주소(request.host_url)가 아닌 고정 공개 URL 사용
        domain = os.environ.get("PUBLIC_BASE_URL", "https://homenstay.com").rstrip("/")
        if atype == "agent":
            sms_body = (
                f"[홈앤스테이] 중개사 승인 완료. 로그인ID(이메일): {ap['email']} / "
                f"임시비밀번호: {temp_pw} / 로그인: {domain}/agent/login — "
                f"최초 로그인 후 반드시 비밀번호를 변경해주세요."
            )
        else:
            sms_body = (
                f"[홈앤스테이] 운영지원업체 승인 완료. 로그인ID(이메일): {ap['email']} / "
                f"임시비밀번호: {temp_pw} / 로그인: {domain}/operator/login — "
                f"최초 로그인 후 반드시 비밀번호를 변경해주세요."
            )
        sms_sent, sms_msg = send_sms(ap["phone"], sms_body)
        # 승인 이메일도 함께 발송 (SMS와 동일 내용) — 실패해도 승인은 확정
        login_path = "/agent/login" if atype == "agent" else "/operator/login"
        email_sent, _ = _send_approval_email(atype, ap["email"], ap["email"], temp_pw, f"{domain}{login_path}")
        resp = {"ok": True, "created_id": created_id, "sms_sent": sms_sent, "sms_message": sms_msg, "email_sent": email_sent}
        if not sms_sent:
            # 평문 임시비번은 문자 발송 실패 시에만 반환(관리자가 수동 전달하도록)
            resp["temp_password"] = temp_pw
        return jsonify(resp)

    return jsonify({"ok": True, "created_id": created_id})


@app.route("/api/admin/applications/<int:app_id>/reject", methods=["POST"])
@require_admin
def admin_applications_reject(app_id):
    data = request.get_json(force=True, silent=True) or {}
    reason = (data.get("reason") or "").strip()
    if not reason:
        return jsonify({"ok": False, "message": "반려 사유(reason)는 필수입니다."}), 400
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT status FROM applications WHERE id=%s", [app_id])
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "존재하지 않는 신청입니다."}), 404
    if row["status"] != "submitted":
        cur.close()
        conn.close()
        return jsonify({"ok": False, "message": "이미 처리된 신청입니다."}), 400
    cur.execute(
        "UPDATE applications SET status='rejected', reject_reason=%s, reviewed_at=NOW() WHERE id=%s",
        [reason, app_id],
    )
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/applications/export.xlsx")
@require_admin
def admin_applications_export():
    sort_expr, order, where_sql, params = _admin_app_filters()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT id, applicant_type, office_or_company_name, owner_name, reg_number,
               biz_reg_number, category, phone, email, preferred_region, preferred_building, status, reject_reason,
               to_char(submitted_at, 'YYYY-MM-DD HH24:MI') AS submitted_at
        FROM applications
        WHERE {where_sql}
        ORDER BY {sort_expr} {order}, id ASC
    """, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    type_kr = {"agent": "중개사", "operator": "지원업체"}
    status_kr = {"submitted": "대기중", "approved": "승인됨", "rejected": "반려됨"}
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "신청"
    ws.append(["ID", "신청유형", "이름/업체명", "대표자", "등록번호", "사업자등록번호", "업종",
               "연락처", "이메일", "희망지역", "희망건물", "상태", "반려사유", "신청일"])
    for r in rows:
        # 번호류는 하이픈 포함 문자열로 넣어 엑셀이 숫자로 오인해 앞 0이 사라지는 것을 방지
        ws.append([
            r["id"], type_kr.get(r["applicant_type"], r["applicant_type"]),
            r["office_or_company_name"], r["owner_name"],
            str(r["reg_number"]) if r["reg_number"] else None,
            format_biz_reg_number(r["biz_reg_number"]) if r["biz_reg_number"] else None,
            r["category"], format_phone(r["phone"]), r["email"], r["preferred_region"],
            r["preferred_building"],
            status_kr.get(r["status"], r["status"]), r["reject_reason"], r["submitted_at"],
        ])
    # 등록번호(E)/사업자등록번호(F)/연락처(H) 컬럼은 셀 서식도 텍스트로 고정
    for row_cells in ws.iter_rows(min_row=2):
        row_cells[4].number_format = "@"
        row_cells[5].number_format = "@"
        row_cells[7].number_format = "@"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = "attachment; filename=applications.xlsx"
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return resp


# ---- 수정요청 이력 (building_requests, 읽기 전용) ----
# submit_building()에서 건축물대장 재검증으로 이미 자동 승인/거절되므로 관리자 액션은 없다.
# 모니터링용으로 이력만 정렬/검색해서 보여준다.
ADMIN_BREQ_SORT = {
    "id": "id", "created_at": "created_at", "processed_at": "processed_at", "status": "status",
}


def _admin_breq_filters():
    q = (request.args.get("q") or "").strip()
    sort_key = (request.args.get("sort") or "id").strip()
    sort_expr = ADMIN_BREQ_SORT.get(sort_key, "id")
    order = "DESC" if (request.args.get("order") or "asc").strip().lower() == "desc" else "ASC"
    where, params = "1=1", []
    if q:
        where = "(road_address ILIKE %s OR building_name_hint ILIKE %s)"
        params = [f"%{q}%", f"%{q}%"]
    # 상태 필터 (예: name_review = 명칭 확인 필요 건만) — 허용값만 통과
    st = (request.args.get("status") or "").strip()
    if st not in ("pending", "verified", "rejected", "name_review"):
        st = ""
    if st:
        where += " AND status = %s"
        params.append(st)
    return sort_expr, order, where, params


@app.route("/api/admin/building-requests")
@require_admin
def admin_building_requests_list():
    sort_expr, order, where_sql, params = _admin_breq_filters()
    page, size, offset = _admin_paging()
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"SELECT COUNT(*) c FROM building_requests WHERE {where_sql}", params)
    total = cur.fetchone()["c"]
    cur.execute(f"""
        SELECT id, request_type, road_address, building_name_hint, suggested_building_name, status,
               master_building_id, reject_reason, verified_lodging_type,
               to_char(created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
               to_char(processed_at, 'YYYY-MM-DD HH24:MI') AS processed_at
        FROM building_requests
        WHERE {where_sql}
        ORDER BY {sort_expr} {order}, id ASC
        LIMIT %s OFFSET %s
    """, params + [size, offset])
    items = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify({"total": total, "page": page, "size": size, "items": items})


@app.route("/api/admin/building-requests/<int:req_id>/approve-name", methods=["POST"])
@require_admin
def admin_building_request_approve_name(req_id):
    """'명칭 확인 필요'(name_review) 건 승인 — 사용자가 제안한 건물명을 마스터에 확정
    (building_name=제안값, name_pending=FALSE)하고 요청 상태를 verified로 바꾼다.
    두 UPDATE를 한 트랜잭션으로 처리해 큐와 마스터가 어긋나지 않게 한다."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            "SELECT id, status, suggested_building_name, master_building_id "
            "FROM building_requests WHERE id=%s FOR UPDATE", (req_id,))
        req = cur.fetchone()
        if not req:
            return jsonify({"ok": False, "message": "존재하지 않는 요청입니다."}), 404
        if req["status"] != "name_review":
            return jsonify({"ok": False, "message": "명칭 확인 필요 상태의 요청만 승인할 수 있습니다."}), 400
        name = (req["suggested_building_name"] or "").strip()
        if not name:
            return jsonify({"ok": False, "message": "제안된 건물명이 없습니다."}), 400
        if not req["master_building_id"]:
            return jsonify({"ok": False, "message": "대상 건물이 연결되어 있지 않습니다."}), 400
        cur.execute(
            "UPDATE master_buildings SET building_name=%s, name_pending=FALSE WHERE id=%s",
            (name, req["master_building_id"]))
        if cur.rowcount == 0:
            conn.rollback()
            return jsonify({"ok": False, "message": "대상 건물을 찾을 수 없습니다."}), 404
        cur.execute(
            "UPDATE building_requests SET status='verified', changed=TRUE, processed_at=NOW() WHERE id=%s",
            (req_id,))
        conn.commit()
    finally:
        cur.close()
        conn.close()
    return jsonify({"ok": True, "building_name": name})


@app.route("/api/admin/stats")
@require_admin
def admin_stats():
    """관리자 통계 대시보드용 집계(매출 제외). 기존 데이터 집계 + 방문 기록.
    모든 항목을 한 번에 반환한다. 빈 구간(0인 달/날)도 채워서 그래프가 끊기지 않게 한다."""
    conn = get_conn()
    cur = conn.cursor()

    # 1) 실거래 추이 — /api/monthly-trend(A/B화면)와 동일한 가변 기간 로직 공유:
    #    실제 MIN(계약월)(하한 TREND_FLOOR_YM)부터 현재까지, 24개월 초과 시 분기 버킷 자동 전환.
    #    빈 버킷 0 채움은 _trend_bucket_items가 처리한다.
    now = datetime.now()
    cur.execute("""
        SELECT substring(deal_date, 1, 7) AS ym,
               COUNT(*) AS cnt,
               COALESCE(SUM(price), 0) AS sum_price
        FROM transactions
        WHERE deal_date IS NOT NULL AND substring(deal_date, 1, 7) >= %s
        GROUP BY ym
    """, [TREND_FLOOR_YM])
    agg = {r["ym"]: {"cnt": r["cnt"], "sum_price": int(r["sum_price"] or 0)} for r in cur.fetchall()}
    trend_items, trend_granularity = _trend_bucket_items(agg, now)
    tx_monthly = [
        {"month": it["ym"], "count": int(it["count"]), "amount": int(it["sum_price"])}
        for it in trend_items
    ]

    # KPI "최근 12개월 거래"는 기존 의미 유지 — 추이 버킷(전체 기간/분기)과 무관하게
    # 월별 원집계(agg)에서 최근 12개월 건수만 합산한다.
    y12, m12 = now.year, now.month - 11
    while m12 <= 0:
        m12 += 12
        y12 -= 1
    last12_floor = f"{y12:04d}-{m12:02d}"
    tx_last12_count = sum(v["cnt"] for ym, v in agg.items() if ym >= last12_floor)

    # KPI "누적 거래" — 실거래관리 목록과 동일 기준(transactions 전체 행 수, 실시간 COUNT).
    # earliest_ym: 가장 오래된 계약월(YYYY.MM). 백필로 과거 데이터가 늘면 라벨/값이 자동 갱신된다.
    # deal_date는 TEXT('YYYY-MM-DD') — ISO 형식이라 MIN이 사전순=시간순으로 동작한다.
    cur.execute("SELECT COUNT(*) AS c, replace(left(MIN(deal_date), 7), '-', '.') AS earliest FROM transactions")
    _row = cur.fetchone()
    tx_total_count = int(_row["c"])
    tx_earliest_ym = _row["earliest"]

    # 그래프 토글 "최근 2년"용 — 최근 24개월 월별 버킷(빈 달 0 채움).
    y24, m24 = now.year, now.month - 23
    while m24 <= 0:
        m24 += 12
        y24 -= 1
    tx_recent24 = []
    y, m = y24, m24
    while (y, m) <= (now.year, now.month):
        ym = f"{y:04d}-{m:02d}"
        v = agg.get(ym, {"cnt": 0, "sum_price": 0})
        tx_recent24.append({"month": ym, "count": int(v["cnt"]), "amount": int(v["sum_price"])})
        m += 1
        if m > 12:
            m = 1
            y += 1

    # 2) 용도별(생활/호텔/콘도) 건물 수 분포
    cur.execute("""
        SELECT COALESCE(NULLIF(lodging_type, ''), '미분류') AS lodging_type, COUNT(*) AS count
        FROM master_buildings
        GROUP BY COALESCE(NULLIF(lodging_type, ''), '미분류')
        ORDER BY count DESC
    """)
    building_by_type = [{"lodging_type": r["lodging_type"], "count": int(r["count"])} for r in cur.fetchall()]

    # 3) 시/도별 건물 수 상위 10 (master_buildings.sgg_text의 첫 토큰 = 시/도)
    cur.execute("""
        SELECT split_part(sgg_text, ' ', 1) AS sido, COUNT(*) AS count
        FROM master_buildings
        WHERE sgg_text IS NOT NULL AND sgg_text <> ''
        GROUP BY split_part(sgg_text, ' ', 1)
        ORDER BY count DESC
        LIMIT 10
    """)
    building_by_sido = [{"sido": r["sido"], "count": int(r["count"])} for r in cur.fetchall()]

    # 4) 회원 현황 — 대기중/반려됨은 applications 파이프라인, 승인됨은 실제 등록된 agents/operators 수
    def _member_stats(applicant_type, member_table):
        cur.execute(
            "SELECT COUNT(*) c FROM applications WHERE applicant_type=%s AND status='submitted'",
            [applicant_type],
        )
        pending = int(cur.fetchone()["c"])
        cur.execute(
            "SELECT COUNT(*) c FROM applications WHERE applicant_type=%s AND status='rejected'",
            [applicant_type],
        )
        rejected = int(cur.fetchone()["c"])
        cur.execute(f"SELECT COUNT(*) c FROM {member_table}")
        approved = int(cur.fetchone()["c"])
        return {"pending": pending, "approved": approved, "rejected": rejected}

    members = {
        "agent": _member_stats("agent", "agents"),
        "operator": _member_stats("operator", "operators"),
    }

    # 운영업체 업종별(6개 카테고리) 건수
    cur.execute("""
        SELECT COALESCE(NULLIF(category, ''), '미지정') AS category, COUNT(*) AS count
        FROM operators
        GROUP BY COALESCE(NULLIF(category, ''), '미지정')
        ORDER BY count DESC
    """)
    operator_by_category = [{"category": r["category"], "count": int(r["count"])} for r in cur.fetchall()]

    # 5) 방문: 최근 14일 일별 페이지뷰 — 0인 날도 채움
    cur.execute("""
        WITH days AS (
            SELECT generate_series(CURRENT_DATE - 13, CURRENT_DATE, INTERVAL '1 day')::date AS d
        )
        SELECT to_char(days.d, 'YYYY-MM-DD') AS day, COUNT(pv.id) AS count
        FROM days
        LEFT JOIN page_views pv ON pv.viewed_at::date = days.d
        GROUP BY days.d
        ORDER BY days.d
    """)
    views_daily = [{"day": r["day"], "count": int(r["count"])} for r in cur.fetchall()]

    # 경로별 조회수 상위 5 (오늘 기준)
    cur.execute("""
        SELECT path, COUNT(*) AS count
        FROM page_views
        WHERE viewed_at::date = CURRENT_DATE
        GROUP BY path
        ORDER BY count DESC
        LIMIT 5
    """)
    views_top_paths = [{"path": r["path"], "count": int(r["count"])} for r in cur.fetchall()]

    # 방문 데이터 수집 시작일 (없으면 오늘) — "언제부터 쌓였는지" 안내문구용
    cur.execute("SELECT to_char(MIN(viewed_at), 'YYYY-MM-DD') AS d FROM page_views")
    row = cur.fetchone()
    collect_start = row["d"] if row and row["d"] else datetime.now().strftime("%Y-%m-%d")

    # 6) 이번달 매출 — revenue_records 중 결제 '완료' + start_date가 이번달인 합계(원)
    cur.execute("""
        SELECT COALESCE(SUM(amount), 0) AS total
        FROM revenue_records
        WHERE payment_status = '완료'
          AND to_char(start_date, 'YYYY-MM') = to_char(CURRENT_DATE, 'YYYY-MM')
    """)
    revenue_month_total = int(cur.fetchone()["total"])

    cur.close()
    conn.close()

    return jsonify({
        "transactions": {
            "monthly": tx_monthly,
            "granularity": trend_granularity,
            "last12_count": tx_last12_count,  # (레거시) 현재 UI 미사용 — 하위호환용 유지
            "total_count": tx_total_count,
            "earliest_ym": tx_earliest_ym,
            "recent24": tx_recent24,
        },
        "buildings": {"by_type": building_by_type, "by_sido": building_by_sido},
        "members": members,
        "operators": {"by_category": operator_by_category},
        "views": {"daily": views_daily, "top_paths": views_top_paths, "collect_start": collect_start},
        "revenue": {"month_total": revenue_month_total},
    })


# ---- 메인화면 좌측 패널용 공개 집계 API (인증 불필요, 읽기 전용) ----

@app.route("/api/stats/registration-rate")
def stats_registration_rate():
    """전국 숙박업 영업신고율 집계 — master_buildings 전체 기준.

    rate = SUM(biz_units) / SUM(units) * 100 (소수 1자리)
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) AS buildings,
               COALESCE(SUM(units), 0) AS total_units,
               COALESCE(SUM(biz_units), 0) AS biz_units
        FROM master_buildings
    """)
    row = cur.fetchone()
    cur.close()
    conn.close()
    total_units = int(row["total_units"])
    biz_units = int(row["biz_units"])
    rate = round(biz_units / total_units * 100, 1) if total_units > 0 else None
    return jsonify({
        "ok": True,
        "buildings": row["buildings"],
        "total_units": total_units,
        "biz_units": biz_units,
        "rate": rate,
    })


@app.route("/api/stats/agent-count")
def stats_agent_count():
    """승인(approved)된 전속중개사 수 — 메인 좌측 패널 카드용 (하우스 계정 제외)."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) AS c
        FROM agents
        WHERE status = 'approved'
          AND office_name <> '홈스퀘어부동산중개법인'
    """)
    n = cur.fetchone()["c"]
    cur.close()
    conn.close()
    return jsonify({"ok": True, "count": n})


@app.route("/api/stats/operator-counts")
def stats_operator_counts():
    """승인(approved)된 운영업체 수 — 메인 좌측 패널 카드용 그룹 집계.

    - consign(위탁정보): 위탁운영
    - housekeeping(운영지원): 청소 + 세탁 + 용품
    - finance(금융): loan_consultants 테이블(별도 엔티티) 기준
    """
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT category, COUNT(*) AS c
        FROM operators
        WHERE status = 'approved'
        GROUP BY category
    """)
    by_cat = {r["category"]: r["c"] for r in cur.fetchall()}
    cur.execute("SELECT COUNT(*) AS c FROM loan_consultants WHERE status = 'approved'")
    loan_cnt = cur.fetchone()["c"]
    cur.close()
    conn.close()
    return jsonify({
        "ok": True,
        "consign": by_cat.get("위탁운영", 0),
        "housekeeping": by_cat.get("청소", 0) + by_cat.get("세탁", 0) + by_cat.get("용품", 0),
        "finance": loan_cnt,
    })


@app.route("/api/health")
def health():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM sync_log ORDER BY id DESC LIMIT 1")
    last = cur.fetchone()
    cur.execute("SELECT COUNT(*) c FROM transactions")
    total_tx = cur.fetchone()["c"]
    cur.close()
    conn.close()
    if not last:
        return jsonify({"status": "no sync yet", "total_transactions": total_tx})
    data = dict(last)
    # datetime은 그대로 jsonify하면 RFC 형식(Tue, 07 Jul...)이 되어 프론트 파싱과 어긋남 → ISO로 통일
    for k in ("started_at", "finished_at"):
        if data.get(k) is not None:
            data[k] = data[k].isoformat(timespec="minutes")
    data["total_transactions"] = total_tx
    return jsonify(data)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
