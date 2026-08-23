#!/usr/bin/env python3
"""전국 도시철도역사정보 표준데이터를 subway_stations에 1회 적재한다.

공공데이터포털의 전국도시철도역사정보 표준데이터(XLSX)를 내려받은 뒤:

    python import_subway_stations.py ./전국도시철도역사정보.xlsx

CSV도 지원하며, 직접 다운로드 URL을 넘길 수도 있다. 데이터포털 페이지:
https://www.data.go.kr/data/15013205/standard.do

역명·노선명·위도·경도는 한국어/영문 표준 컬럼명 변형을 모두 허용한다.
기본값은 검증을 마친 원본 전체를 한 트랜잭션에서 교체하므로, 원본 좌표가
정정돼도 오래된 좌표가 남지 않는다.
"""

import argparse
import csv
import io
import os
import sys
from urllib.parse import urlparse
from urllib.request import Request, urlopen

import psycopg2.extras

from db import get_conn, init_db


SOURCE_PAGE = "https://www.data.go.kr/data/15013205/standard.do"

_ALIASES = {
    "station_name": (
        "station_name", "stationname", "역사명", "역명", "역사 명",
    ),
    "line_name": (
        "line_name", "linename", "노선명", "노선 명",
    ),
    "lat": (
        "lat", "latitude", "위도", "역위도", "역사위도", "역 위도", "역사 위도",
    ),
    "lng": (
        "lng", "lon", "longitude", "경도", "역경도", "역사경도", "역 경도", "역사 경도",
    ),
}


def _normal_header(value):
    return "".join(str(value or "").replace("\ufeff", "").split()).lower()


def _find_columns(headers):
    normalized = {_normal_header(value): index for index, value in enumerate(headers)}
    columns = {}
    for field, aliases in _ALIASES.items():
        for alias in aliases:
            index = normalized.get(_normal_header(alias))
            if index is not None:
                columns[field] = index
                break
    missing = [field for field in ("station_name", "lat", "lng") if field not in columns]
    if missing:
        raise ValueError(
            "필수 컬럼을 찾지 못했습니다: {} (원본 헤더: {})".format(
                ", ".join(missing), list(headers)
            )
        )
    return columns


def _number(value):
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _read_csv(raw):
    text = None
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV 인코딩을 확인할 수 없습니다.")
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return []
    columns = _find_columns(rows[0])
    return _normalise_rows(rows[1:], columns)


def _read_xlsx(raw):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX 적재에는 openpyxl이 필요합니다.") from exc
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if not headers:
            return []
        columns = _find_columns(headers)
        return _normalise_rows(iterator, columns)
    finally:
        workbook.close()


def _normalise_rows(rows, columns):
    result = []
    seen = set()
    for row in rows:
        values = {
            field: (row[index] if index < len(row) else None)
            for field, index in columns.items()
        }
        station_name = str(values.get("station_name") or "").strip()
        line_name = str(values.get("line_name") or "").strip() or None
        lat = _number(values.get("lat"))
        lng = _number(values.get("lng"))
        if not station_name or lat is None or lng is None:
            continue
        if not (-90 <= lat <= 90 and -180 <= lng <= 180):
            continue
        key = (station_name, line_name or "", round(lat, 8), round(lng, 8))
        if key in seen:
            continue
        seen.add(key)
        result.append((station_name, line_name, lat, lng))
    return result


def _download(source):
    request = Request(source, headers={"User-Agent": "Livingstay subway station importer"})
    with urlopen(request, timeout=60) as response:
        return response.read(), response.headers.get("Content-Type", "")


def _load_source(source):
    if urlparse(source).scheme in ("http", "https"):
        raw, content_type = _download(source)
        suffix = os.path.splitext(urlparse(source).path)[1].lower()
        if suffix in (".csv",) or "csv" in content_type.lower():
            return _read_csv(raw)
        return _read_xlsx(raw)
    with open(source, "rb") as file_obj:
        raw = file_obj.read()
    suffix = os.path.splitext(source)[1].lower()
    if suffix == ".csv":
        return _read_csv(raw)
    return _read_xlsx(raw)


def main():
    parser = argparse.ArgumentParser(description="도시철도역 좌표를 DB에 1회 적재")
    parser.add_argument(
        "source",
        help="다운로드한 XLSX/CSV 파일 경로 또는 직접 다운로드 URL",
    )
    parser.add_argument(
        "--append", action="store_true",
        help="기존 데이터를 교체하지 않고 중복 없는 행만 추가합니다.",
    )
    parser.add_argument(
        "--allow-small-source", action="store_true",
        help="100행 미만의 테스트·부분 파일도 적재하도록 허용합니다.",
    )
    args = parser.parse_args()

    rows = _load_source(args.source)
    if not rows:
        raise RuntimeError("유효한 역 좌표 행이 없습니다.")
    if len(rows) < 100 and not args.allow_small_source:
        raise RuntimeError(
            "유효 행이 100건 미만입니다. 전체 표준데이터인지 확인하거나 "
            "--allow-small-source를 사용하세요."
        )

    # 신규 테이블이 아직 없는 개발/운영 DB에서도 단독 실행 가능하게 한다.
    init_db()
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            if not args.append:
                # DELETE와 INSERT는 같은 트랜잭션이므로 독자는 이전 전체 또는
                # 새 전체만 보며, 좌표 정정본이 과거 행과 공존하지 않는다.
                cur.execute("DELETE FROM subway_stations")
            psycopg2.extras.execute_values(
                cur,
                """
                INSERT INTO subway_stations (station_name, line_name, lat, lng)
                VALUES %s
                ON CONFLICT DO NOTHING
                """,
                rows,
                page_size=500,
            )
        conn.commit()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS cnt FROM subway_stations")
            total = int(cur.fetchone()["cnt"])
        print("역 좌표 {}건 {} 완료 — DB 전체 {}건 (출처: {})".format(
            len(rows), "추가 적재" if args.append else "전체 교체", total, args.source
        ))
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("적재 실패: {}".format(exc), file=sys.stderr)
        sys.exit(1)