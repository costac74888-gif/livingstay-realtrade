# -*- coding: utf-8 -*-
"""
api_test.py — 데이터 JSON API가 조용히 깨지는 것을 배포 전에 잡아내는 체크.

홈페이지 스모크 체크(smoke_test.py)는 정적 파일(HTML/CSS/JS)만 검증한다.
하지만 화면에 실제로 뜨는 데이터는 전부 JSON API에서 온다:
  - /api/health        (배치 상태)
  - /api/regions       (지역 트리)
  - /api/years         (연도 목록)
  - /api/transactions  (실거래 목록)
쿼리 오류/스키마 드리프트 등으로 이 중 하나라도 깨지면, 페이지는 정상적으로
뜨지만 데이터가 하나도 안 보이는 "조용한 실패"가 발생한다.

이 체크는 Flask 테스트 클라이언트로 각 엔드포인트가
  1) HTTP 200
  2) JSON content-type
  3) 기대하는 형태(shape)의 JSON
을 돌려주는지 검증한다. 하나라도 어긋나면 즉시 실패(exit 1)한다.

실행: python tests/api_test.py
"""

import os
import sys
import copy
import time
import re
import hashlib
import subprocess
from concurrent.futures import ThreadPoolExecutor
from datetime import timedelta
from unittest.mock import patch
from werkzeug.security import generate_password_hash, check_password_hash

# app.py를 import할 수 있도록 프로젝트 루트를 경로에 추가
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import app as app_module  # noqa: E402
import db as db_module  # noqa: E402
import import_vworld_brokers  # noqa: E402
import sync_brokers  # noqa: E402
from app import (  # noqa: E402
    _building_share_meta,
    _canonical_sido_name,
    _create_short_link,
    _notify_lead_agents,
    _should_store_page_view,
    app,
)
from db import get_conn  # noqa: E402
import addr_norm  # noqa: E402


def check_feature_tips_admin_api(client):
    """기능 소개 관리 API의 인증·목록·생성·수정·입력 검증을 확인한다."""
    blocked = client.get("/api/admin/feature-tips")
    if blocked.status_code != 401 or not blocked.is_json:
        return "기능 소개 목록 API가 비관리자 요청을 차단하지 않음"

    with client.session_transaction() as sess:
        sess["admin"] = True
    listed = client.get("/api/admin/feature-tips")
    if listed.status_code != 200 or not listed.is_json:
        return "관리자 기능 소개 목록 API가 정상 응답하지 않음"
    payload = listed.get_json()
    if payload.get("ok") is not True or not isinstance(payload.get("items"), list):
        return "관리자 기능 소개 목록 API의 응답 형태가 잘못됨"
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT pg_get_constraintdef(c.oid) AS definition
            FROM pg_constraint c
            JOIN pg_class t ON t.oid = c.conrelid
            WHERE t.relname = 'weekly_feature_tips'
              AND c.conname = 'weekly_feature_tips_episode_check'
        """)
        row = cur.fetchone() or {}
        definition = (row.get("definition") or "").lower()
        if "episode >= 1" not in definition or "episode <= 8" not in definition:
            return "기능 소개 테이블의 ISO 회차 범위 제약(1~8)이 없음"
    finally:
        cur.close()
        conn.close()

    malformed_create = client.post("/api/admin/feature-tips", json={"episode": 0})
    if malformed_create.status_code != 400:
        return "기능 소개 생성 API가 잘못된 회차를 거절하지 않음"
    out_of_range = client.post("/api/admin/feature-tips", json={
        "episode": 9, "title": "범위 초과", "body": "범위 초과", "cta_url": "/"
    })
    if out_of_range.status_code != 400:
        return "기능 소개 생성 API가 ISO 순환 범위를 벗어난 회차를 거절하지 않음"
    malformed_update = client.patch("/api/admin/feature-tips/999999999", json={})
    if malformed_update.status_code != 400:
        return "기능 소개 수정 API가 빈 변경을 거절하지 않음"
    first_id = (payload.get("items") or [{}])[0].get("id")
    blank_label = client.patch(f"/api/admin/feature-tips/{first_id}", json={"cta_label": ""})
    if blank_label.status_code != 400:
        return "기능 소개 수정 API가 빈 CTA 문구를 거절하지 않음"
    old_label = (payload.get("items") or [{}])[0].get("cta_label")
    try:
        changed_label = client.patch(
            f"/api/admin/feature-tips/{first_id}", json={"cta_label": "api_test CTA"}
        )
        if changed_label.status_code != 200 or not (changed_label.get_json() or {}).get("ok"):
            return "기능 소개 수정 API가 CTA 문구를 저장하지 못함"
    finally:
        if first_id and old_label:
            client.patch(f"/api/admin/feature-tips/{first_id}", json={"cta_label": old_label})
    return None


def check_user_stats_admin_api(client):
    """이용자 현황 API의 관리자 인증과 집계 응답 계약을 확인한다."""
    with client.session_transaction() as sess:
        sess.clear()
    blocked = client.get("/api/admin/user-stats")
    if blocked.status_code != 401:
        return "이용자 현황 API가 비관리자 요청을 차단하지 않음"

    with client.session_transaction() as sess:
        sess["admin"] = True
    response = client.get("/api/admin/user-stats")
    if response.status_code != 200 or not response.is_json:
        return f"이용자 현황 API가 정상 응답하지 않음 (HTTP {response.status_code})"
    data = response.get_json() or {}
    required = {
        "mau", "wau", "dau", "new_this_week", "fav_this_week", "listing_this_week",
        "mau_prev", "wau_prev", "dau_prev", "daily_active", "daily_new",
        "daily_listing", "segment_counts", "page_views",
    }
    if not required <= set(data):
        return f"이용자 현황 API 필수 필드 누락: {sorted(required - set(data))}"
    for key in ("mau", "wau", "dau", "new_this_week", "fav_this_week",
                "listing_this_week", "mau_prev", "wau_prev", "dau_prev"):
        if not isinstance(data[key], int) or data[key] < 0:
            return f"{key}가 0 이상 정수가 아님"
    for key in ("daily_active", "daily_new", "daily_listing"):
        rows = data[key]
        if not isinstance(rows, list) or len(rows) != 30:
            return f"{key}가 30일 배열이 아님"
        if any(not row.get("date") or not isinstance(row.get("count"), int) for row in rows):
            return f"{key} 행의 date/count 형태가 잘못됨"
    if set((data["segment_counts"] or {})) != {"general", "agent", "operator"}:
        return "segment_counts의 general/agent/operator 구성이 잘못됨"
    views = data["page_views"]
    if not isinstance(views, dict) or not all(
        isinstance(views.get(key), int) and views[key] >= 0
        for key in ("pv_today", "pv_this_week")
    ):
        return "page_views의 오늘/이번 주 값 형태가 잘못됨"
    top_paths = views.get("top_paths")
    if not isinstance(top_paths, list) or len(top_paths) > 5:
        return "page_views.top_paths가 5건 이하 배열이 아님"
    for row in top_paths:
        if not isinstance(row.get("path"), str) or not isinstance(row.get("count"), int):
            return "page_views.top_paths 행의 path/count 형태가 잘못됨"
    if _should_store_page_view(""):
        return "빈 User-Agent가 페이지뷰 저장 대상으로 판정됨"
    for ua in ("Googlebot/2.1", "example crawler", "Spider/1.0",
               "Slurp", "facebookexternalhit/1.1"):
        if _should_store_page_view(ua):
            return f"봇 User-Agent가 페이지뷰 저장 대상으로 판정됨: {ua}"
    if not _should_store_page_view("Mozilla/5.0 (X11; Linux x86_64)"):
        return "일반 브라우저 User-Agent가 페이지뷰 저장 대상에서 제외됨"
    return None


def _check_map_poi_api(client):
    """지도 주변정보 프록시의 입력 검증·키 은닉·응답 형태를 네트워크 없이 확인한다."""
    failures = []
    app_module._MAP_POI_CACHE.clear()
    invalid_type = client.get("/api/v1/m/6b4?type=unknown&lat=37.5&lng=127.0")
    if invalid_type.status_code != 400:
        failures.append("지도 POI API가 잘못된 type을 거절하지 않음")
    invalid_coords = client.get("/api/v1/m/6b4?type=education&lat=oops&lng=127.0")
    if invalid_coords.status_code != 400:
        failures.append("지도 POI API가 잘못된 좌표를 거절하지 않음")

    with patch.dict(os.environ, {"KAKAO_REST_API_KEY": ""}, clear=False):
        missing_key = client.get("/api/v1/m/6b4?type=education&lat=37.5&lng=127.0")
    if missing_key.status_code != 503 or (missing_key.get_json() or {}).get("ok") is not False:
        failures.append("지도 POI API가 REST 키 누락을 명시적으로 처리하지 않음")

    class FakeResponse:
        def __init__(self, payload):
            self.payload = payload

        def raise_for_status(self):
            return None

        def json(self):
            return self.payload

    requests_seen = []

    def fake_get(url, headers, params, timeout):
        requests_seen.append((url, headers, params, timeout))
        code = params["category_group_code"]
        return FakeResponse({"documents": [{
            "id": f"fixture-{code}",
            "place_name": f"{code} 장소",
            "category_name": f"교육 > {code}",
            "address_name": "서울시 테스트구",
            "road_address_name": "",
            "phone": "0212345678",
            "place_url": "https://place.map.kakao.com/fixture",
            "x": "127.0123",
            "y": "37.5012",
            "distance": "120",
        }]})

    with (
        patch.dict(os.environ, {"KAKAO_REST_API_KEY": "test-kakao-key"}, clear=False),
        patch.object(app_module.requests, "get", side_effect=fake_get),
    ):
        success = client.get("/api/v1/m/6b4?type=education&lat=37.5&lng=127.0&radius=1500")
    payload = success.get_json() or {}
    if success.status_code != 200 or payload.get("ok") is not True:
        failures.append("지도 POI API가 정상 카카오 응답을 반환하지 않음")
    elif (
        payload.get("type") != "education"
        or payload.get("radius") != 1500
        or len(payload.get("items") or []) != 2
        or any(not {"name", "category", "lat", "lng"} <= set(item) for item in payload["items"])
    ):
        failures.append("지도 POI API의 성공 응답 형태가 잘못됨")
    if len(requests_seen) != 2 or any(
        call[0] != app_module._KAKAO_LOCAL_CATEGORY_URL
        or call[1].get("Authorization") != "KakaoAK test-kakao-key"
        or call[3] != 5
        for call in requests_seen
    ):
        failures.append("지도 POI API가 카카오 Local 카테고리 호출 계약을 지키지 않음")
    with patch.object(
        app_module.requests,
        "get",
        side_effect=AssertionError("캐시된 지도 POI 요청이 외부 호출을 다시 시도함"),
    ):
        cached = client.get("/api/v1/m/6b4?type=education&lat=37.5&lng=127.0&radius=1500")
    if cached.status_code != 200 or (cached.get_json() or {}).get("items") != payload.get("items"):
        failures.append("지도 POI API가 짧은 TTL 캐시를 재사용하지 않음")

    with (
        patch.dict(os.environ, {"KAKAO_REST_API_KEY": "test-kakao-key"}, clear=False),
        patch.object(
            app_module.requests,
            "get",
            side_effect=app_module.requests.RequestException("fixture"),
        ),
    ):
        upstream_error = client.get("/api/v1/m/6b4?type=convenience&lat=37.5&lng=127.0")
    if upstream_error.status_code != 502:
        failures.append("지도 POI API가 카카오 Local 장애를 502로 처리하지 않음")

    app_module._MAP_POI_CACHE.clear()

    def slow_get(*_args, **_kwargs):
        time.sleep(3)
        return FakeResponse({"documents": []})

    started = time.monotonic()
    with (
        patch.dict(os.environ, {"KAKAO_REST_API_KEY": "test-kakao-key"}, clear=False),
        patch.object(app_module.requests, "get", side_effect=slow_get),
    ):
        timed_out = client.get("/api/v1/m/6b4?type=convenience&lat=37.6&lng=127.1")
    elapsed = time.monotonic() - started
    if timed_out.status_code != 502 or elapsed > 2.8:
        failures.append("지도 POI API가 지연된 카카오 응답을 전체 제한 시간 안에 중단하지 않음")

    return failures


def _daily_count(payload, series_key, day):
    row = next((item for item in payload[series_key] if item["date"] == day.isoformat()), None)
    return None if row is None else row["count"]


def check_user_stats_aggregate_windows_and_view_writers(client):
    """날짜 경계·철회 매물·두 page_views 기록 경로를 실제 행으로 검증한다."""
    response = client.get("/api/admin/user-stats")
    if response.status_code != 200:
        return "이용자 현황 집계 기준값을 불러오지 못함"
    before = response.get_json() or {}
    tag = f"user-stats-test-{time.time_ns()}"
    user_ids, listing_ids = [], []
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT CURRENT_DATE AS today")
        today = cur.fetchone()["today"]
        cur.execute("SELECT id FROM master_buildings ORDER BY id LIMIT 1")
        building = cur.fetchone()
        if not building:
            return "이용자 현황 테스트용 마스터 건물이 없어 집계 경계를 검증할 수 없음"
        building_id = building["id"]

        def add_user(label, created_sql, login_sql, status="active"):
            cur.execute(f"""
                INSERT INTO users (email, name, user_type, status, created_at, last_login_at)
                VALUES (%s, %s, 'general', %s, {created_sql}, {login_sql})
                RETURNING id
            """, (f"{tag}-{label}@example.test", tag, status))
            user_id = cur.fetchone()["id"]
            user_ids.append(user_id)
            return user_id

        # 오늘·8일 전·31일 전: 현재/이전 MAU·WAU 및 일별 시계열의 경계점.
        today_user = add_user("today", "CURRENT_DATE + INTERVAL '1 hour'", "CURRENT_DATE + INTERVAL '1 hour'")
        add_user("wau-prev", "CURRENT_DATE - INTERVAL '8 days' + INTERVAL '1 hour'",
                 "CURRENT_DATE - INTERVAL '8 days' + INTERVAL '1 hour'")
        add_user("mau-prev", "CURRENT_DATE - INTERVAL '31 days' + INTERVAL '1 hour'",
                 "CURRENT_DATE - INTERVAL '31 days' + INTERVAL '1 hour'")
        add_user("withdrawn", "CURRENT_DATE - INTERVAL '40 days'", "CURRENT_DATE + INTERVAL '1 hour'",
                 status="withdrawn")
        cur.execute("""
            INSERT INTO user_favorites (user_id, building_name, address, master_building_id, created_at)
            VALUES (%s, %s, %s, %s, CURRENT_DATE + INTERVAL '1 hour')
        """, (today_user, tag, tag, building_id))
        # 영어·한글 철회 행은 모두 주간/일별 매물등록에서 빠져야 한다.
        for status, created_sql, target in (
            ("submitted", "CURRENT_DATE + INTERVAL '1 hour'", "unit"),
            ("withdrawn", "CURRENT_DATE + INTERVAL '1 hour'", "unit"),
            ("철회됨", "CURRENT_DATE + INTERVAL '1 hour'", "unit"),
            ("submitted", "CURRENT_DATE - INTERVAL '40 days'", "whole"),
        ):
            cur.execute(f"""
                INSERT INTO listing_requests
                    (user_id, master_building_id, deal_type, contact_phone, deal_mode,
                     transaction_target, status, created_at)
                VALUES (%s, %s, '매매', '01000000000', 'direct', %s, %s, {created_sql})
                RETURNING id
            """, (today_user, building_id, target, status))
            listing_ids.append(cur.fetchone()["id"])
        view_listing_id = listing_ids[-1]
        conn.commit()

        after_response = client.get("/api/admin/user-stats")
        if after_response.status_code != 200:
            return "테스트 행 추가 뒤 이용자 현황 API가 응답하지 않음"
        after = after_response.get_json() or {}
        expected_deltas = {
            "mau": 2, "wau": 1, "dau": 1, "new_this_week": 1,
            "fav_this_week": 1, "listing_this_week": 1,
            "mau_prev": 1, "wau_prev": 1, "dau_prev": 0,
        }
        for key, delta in expected_deltas.items():
            if after.get(key) != before.get(key, 0) + delta:
                return f"이용자 현황 {key}의 기간 또는 철회 상태 집계가 잘못됨"
        if after["segment_counts"]["general"] != before["segment_counts"]["general"] + 3:
            return "탈퇴한 일반회원이 사용자 세그먼트에 포함되거나 활성 회원이 누락됨"
        for series_key, day, delta in (
            ("daily_active", today, 1),
            ("daily_active", today - timedelta(days=8), 1),
            ("daily_new", today, 1),
            ("daily_listing", today, 1),
        ):
            old_count = _daily_count(before, series_key, day)
            new_count = _daily_count(after, series_key, day)
            if old_count is None or new_count != old_count + delta:
                return f"{series_key}의 {day.isoformat()} 일별 경계 집계가 잘못됨"

        bot_ua = f"Googlebot/2.1 {tag}"
        browser_ua = f"StatsTestBrowser/1.0 {tag}"
        cur.execute("SELECT COUNT(*) AS count FROM page_views WHERE user_agent IN (%s, %s)", (bot_ua, browser_ua))
        before_views = cur.fetchone()["count"]
        client.get("/", headers={"User-Agent": bot_ua})
        client.get("/", headers={"User-Agent": browser_ua})
        bot_view = client.post("/api/listings/views", json={"listing_ids": [view_listing_id]},
                               headers={"User-Agent": bot_ua})
        browser_view = client.post("/api/listings/views", json={"listing_ids": [view_listing_id]},
                                   headers={"User-Agent": browser_ua})
        if bot_view.status_code != 200 or browser_view.status_code != 200:
            return "매물 열람 페이지뷰 경로가 테스트 매물에 응답하지 않음"
        cur.execute("SELECT COUNT(*) AS count FROM page_views WHERE user_agent IN (%s, %s)", (bot_ua, browser_ua))
        after_views = cur.fetchone()["count"]
        # 일반 페이지 1건 + 매물열람 1건만 추가되며, 봇 UA는 어느 INSERT 경로에서도 저장되지 않아야 한다.
        if after_views != before_views + 2:
            return "빈/봇 UA 제외 또는 일반 UA의 두 페이지뷰 INSERT 경로가 잘못됨"
        return None
    except Exception as exc:
        conn.rollback()
        return f"이용자 현황 집계 경계 테스트 실행 오류: {exc}"
    finally:
        try:
            cur.execute("DELETE FROM page_views WHERE user_agent LIKE %s", (f"%{tag}%",))
            if listing_ids:
                cur.execute("DELETE FROM listing_requests WHERE id = ANY(%s)", (listing_ids,))
            if user_ids:
                cur.execute("DELETE FROM user_favorites WHERE user_id = ANY(%s)", (user_ids,))
                cur.execute("DELETE FROM users WHERE id = ANY(%s)", (user_ids,))
            conn.commit()
        except Exception:
            conn.rollback()
        cur.close()
        conn.close()


def check_health(payload):
    """/api/health: 항상 total_transactions(정수)를 포함하는 객체여야 한다."""
    if not isinstance(payload, dict):
        return "응답이 JSON 객체가 아님"
    if "total_transactions" not in payload:
        return "'total_transactions' 키 없음"
    if not isinstance(payload["total_transactions"], int):
        return "'total_transactions'가 정수가 아님"
    return None


def check_regions(payload):
    """/api/regions: 시도>시군구>읍면동 계층 트리(객체). 비어 있을 수 있음."""
    if not isinstance(payload, dict):
        return "응답이 JSON 객체(트리)가 아님"
    # 값이 있으면 각 시도 노드는 count와 sgg를 가진 객체여야 한다.
    for sido, node in payload.items():
        if not isinstance(node, dict) or "count" not in node or "sgg" not in node:
            return f"'{sido}' 노드 형태가 잘못됨 (count/sgg 필요)"
        break
    return None


def check_years(payload):
    """/api/years: {"years": [...], "current_year": "YYYY"}"""
    if not isinstance(payload, dict):
        return "응답이 JSON 객체가 아님"
    if not isinstance(payload.get("years"), list):
        return "'years'가 배열이 아님"
    if not payload.get("current_year"):
        return "'current_year' 없음"
    return None


def check_transactions(payload):
    """/api/transactions: {"total", "page", "size", "items": [...]}"""
    if not isinstance(payload, dict):
        return "응답이 JSON 객체가 아님"
    for key in ("total", "page", "size"):
        if not isinstance(payload.get(key), int):
            return f"'{key}'가 정수가 아님"
    if not isinstance(payload.get("items"), list):
        return "'items'가 배열이 아님"
    return None


def check_buildings_geo(payload):
    """/api/buildings-geo: {"total": int, "items": [...]}"""
    if not isinstance(payload, dict):
        return "응답이 JSON 객체가 아님"
    if not isinstance(payload.get("total"), int):
        return "'total'이 정수가 아님"
    if not isinstance(payload.get("items"), list):
        return "'items'가 배열이 아님"
    # 좌표가 있는 항목은 lat/lng가 숫자여야 한다
    for item in payload["items"][:10]:
        if item.get("lat") is not None and not isinstance(item["lat"], (int, float)):
            return "items[].lat이 숫자가 아님"
        if item.get("lng") is not None and not isinstance(item["lng"], (int, float)):
            return "items[].lng이 숫자가 아님"
        for key in ("txn_count", "listing_count", "total_count"):
            if not isinstance(item.get(key), int) or item[key] < 0:
                return f"items[].{key}가 0 이상의 정수가 아님"
        if item["total_count"] != item["txn_count"] + item["listing_count"]:
            return "items[].total_count가 txn_count + listing_count와 다름"
    return None


def check_platform_summary(payload):
    """/api/stats/platform-summary: 홈 신뢰지표 4개는 모두 0 이상 정수여야 한다."""
    if not isinstance(payload, dict) or payload.get("ok") is not True:
        return "응답이 성공 객체가 아님"
    for key in ("building_count", "biz_count", "transaction_count", "listing_count"):
        if not isinstance(payload.get(key), int) or payload[key] < 0:
            return f"'{key}'가 0 이상 정수가 아님"
    return None


def check_datalab_lodging_table(payload):
    """공개 데이터랩 ①의 최소 컬럼."""
    if not isinstance(payload, dict) or payload.get("ok") is not True:
        return "응답이 성공 객체가 아님"
    rows = payload.get("rows")
    if not isinstance(rows, list) or len(rows) != 11:
        return "'rows'가 개정 법정분류 11개 상위행 배열이 아님"
    expected_types = [
        "전체", "생활", "관광", "일반", "에어비앤비",
        "농어촌민박", "캠핑", "한옥", "복합", "준공전", "미분류",
    ]
    if [row.get("type") for row in rows] != expected_types:
        return "공개 용도별 행 순서가 잘못됨"
    allowed = {
        "type", "building_count", "units", "biz_count", "room_count",
        "report_rate", "sub_rows", "camping_facility_count",
        "camping_site_count", "camping_general_site_count",
        "camping_auto_site_count", "camping_glamping_site_count",
        "camping_caravan_site_count", "camping_classification_breakdown",
    }
    for row in rows:
        if set(row) - allowed or not {"type", "building_count", "units", "biz_count", "room_count", "report_rate"} <= set(row):
            return "공개 응답에 최소 컬럼 외 필드가 있거나 필수 필드가 없음"
        if row["type"] == "일반":
            sub_rows = row.get("sub_rows")
            if [sub.get("type") for sub in sub_rows or []] != ["일반호텔", "여관업", "여인숙업"]:
                return "일반숙박 세부 3종이 공개 응답에 없음"
            for sub in sub_rows:
                if set(sub) != {"type", "building_count", "units", "biz_count", "room_count", "report_rate"}:
                    return "일반숙박 세부행에 공개 최소 컬럼 외 필드가 있음"
    return None


def check_datalab_items(payload):
    """데이터랩 TOP/지역 비교 API: 성공 객체와 items 배열."""
    if not isinstance(payload, dict) or payload.get("ok") is not True:
        return "응답이 성공 객체가 아님"
    if not isinstance(payload.get("items"), list):
        return "'items'가 배열이 아님"
    if len(payload["items"]) > 5:
        return "TOP API가 5건을 초과함"
    return None


def check_datalab_consign(payload):
    """/api/stats/consign-by-sido: 시도별 영업신고현황과 전국 합계."""
    if not isinstance(payload, dict) or payload.get("ok") is not True:
        return "응답이 성공 객체가 아님"
    if not isinstance(payload.get("items"), list):
        return "'items'가 배열이 아님"
    if not isinstance(payload.get("total"), dict):
        return "'total'이 객체가 아님"
    if payload.get("is_partial") is not True:
        return "수집중 상태가 표시되지 않음"
    required = {
        "building_cnt", "total_units", "active_biz_cnt",
        "active_room_cnt", "report_rate",
    }
    for scope, rows in (("items", payload["items"]), ("total", [payload["total"]])):
        for row in rows:
            if not required <= set(row):
                return f"{scope}에 영업신고현황 필수 컬럼이 없음"
            for key in required - {"report_rate"}:
                if not isinstance(row[key], int) or row[key] < 0:
                    return f"{scope}.{key}가 0 이상 정수가 아님"
            rate = row["report_rate"]
            if rate is not None and (
                not isinstance(rate, (int, float)) or rate < 0
            ):
                return f"{scope}.report_rate가 0 이상 숫자가 아님"
            if rate is not None and rate > 100:
                return f"{scope}.report_rate가 호실수 캡을 넘어 100%를 초과함"
    return None


def expected_consign_by_sido():
    """생활숙박시설과 영업신고 주소 매칭을 사용한 독립 집계값."""
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, sgg_text, COALESCE(units, 0) AS units,
                   road_address, jibun_address
            FROM master_buildings
            WHERE lodging_type = '생활'
            ORDER BY id
        """)
        buildings = [dict(row) for row in cur.fetchall()]
        road_keys, jibun_keys = set(), set()
        for building in buildings:
            road_key = addr_norm.normalize_road_prefix(building.get("road_address"))
            jibun_key = addr_norm.get_building_jibun_key(building)
            building["_road_key"] = road_key
            building["_jibun_key"] = jibun_key
            if road_key:
                road_keys.add(road_key)
            if jibun_key:
                jibun_keys.add(jibun_key)

        cur.execute("""
            SELECT permit_number, room_count, biz_status_name, road_norm, jibun_norm
            FROM lodging_registry
            WHERE road_norm = ANY(%s) OR jibun_norm = ANY(%s)
            ORDER BY source_updated_at DESC NULLS LAST, id DESC
        """, [list(road_keys) or ["__none__"], list(jibun_keys) or ["__none__"]])
        road_permits, jibun_permits = {}, {}
        for row in cur.fetchall():
            permit = dict(row)
            if permit.get("road_norm"):
                road_permits.setdefault(permit["road_norm"], {})[
                    permit["permit_number"]
                ] = permit
            if permit.get("jibun_norm"):
                jibun_permits.setdefault(permit["jibun_norm"], {})[
                    permit["permit_number"]
                ] = permit

        regions = {}
        assigned_active_permits = set()
        for building in buildings:
            sido = _canonical_sido_name(building.get("sgg_text"))
            if not sido:
                continue
            region = regions.setdefault(sido, {
                "building_cnt": 0,
                "total_units": 0,
                "active_permits": {},
                "building_ids": [],
            })
            region["building_cnt"] += 1
            region["total_units"] += max(0, int(building.get("units") or 0))
            region["building_ids"].append(building["id"])
            road_matches = road_permits.get(building.get("_road_key"), {})
            matched_permits = road_matches or jibun_permits.get(
                building.get("_jibun_key"), {}
            )
            for permit_number, permit in matched_permits.items():
                if (
                    not permit_number
                    or permit.get("biz_status_name") != "영업/정상"
                    or permit_number in assigned_active_permits
                ):
                    continue
                assigned_active_permits.add(permit_number)
                region["active_permits"][permit_number] = permit

        candidates_by_permit = {}
        building_by_id = {building["id"]: building for building in buildings}
        for building in buildings:
            matches = road_permits.get(building.get("_road_key"), {})
            matches = matches or jibun_permits.get(building.get("_jibun_key"), {})
            for permit_number, permit in matches.items():
                if permit.get("biz_status_name") != "영업/정상":
                    continue
                candidates_by_permit.setdefault(permit_number, {
                    "permit": permit,
                    "buildings": [],
                })["buildings"].append(building)
        raw_rooms_by_building = {}
        for candidate in candidates_by_permit.values():
            representative = max(
                candidate["buildings"],
                key=lambda building: (int(building["units"] or 0), -int(building["id"])),
            )
            building_id = representative["id"]
            raw_rooms_by_building[building_id] = (
                raw_rooms_by_building.get(building_id, 0)
                + int(candidate["permit"].get("room_count") or 0)
            )
        capped_rooms_by_building = {
            building_id: min(
                room_count,
                max(0, int(building_by_id[building_id].get("units") or 0)),
            )
            for building_id, room_count in raw_rooms_by_building.items()
        }

        def summary(region):
            total_units = region["total_units"]
            active_permits = region["active_permits"]
            active_room_cnt = sum(
                capped_rooms_by_building.get(building_id, 0)
                for building_id in region["building_ids"]
            )
            return {
                "building_cnt": region["building_cnt"],
                "total_units": total_units,
                "active_biz_cnt": len(active_permits),
                "active_room_cnt": active_room_cnt,
                "report_rate": (
                    round(active_room_cnt / total_units * 100, 1)
                    if total_units else None
                ),
            }

        items = [
            {"sido": sido, **summary(region)}
            for sido, region in sorted(regions.items())
        ]
        total_units = sum(item["total_units"] for item in items)
        active_room_cnt = sum(item["active_room_cnt"] for item in items)
        total = {
            "building_cnt": sum(item["building_cnt"] for item in items),
            "total_units": total_units,
            "active_biz_cnt": len(assigned_active_permits),
            "active_room_cnt": active_room_cnt,
            "report_rate": (
                round(active_room_cnt / total_units * 100, 1)
                if total_units else None
            ),
        }
        return items, total
    finally:
        cur.close()
        conn.close()


# (경로, shape 검증 함수)
CHECKS = [
    ("/api/health", check_health),
    ("/api/regions", check_regions),
    ("/api/years", check_years),
    ("/api/transactions?with_total=1", check_transactions),
    ("/api/transactions?with_total=1&building_id=999999999", check_transactions),
    ("/api/buildings-geo", check_buildings_geo),
    ("/api/stats/platform-summary", check_platform_summary),
    ("/api/v1/d/3f7", check_datalab_lodging_table),
    ("/api/stats/price-change-top?direction=up", check_datalab_items),
    ("/api/stats/price-change-top?direction=down", check_datalab_items),
    ("/api/stats/highest-price-top?order=highest", check_datalab_items),
    ("/api/stats/highest-price-top?order=lowest", check_datalab_items),
    ("/api/stats/closure-rate-by-region", check_datalab_items),
    ("/api/stats/consign-by-sido", check_datalab_consign),
]


def run():
    failures = []
    client = app.test_client()
    for path, validate in CHECKS:
        resp = client.get(path)
        content_type = resp.headers.get("Content-Type", "")

        if resp.status_code != 200:
            failures.append(f"{path}: HTTP {resp.status_code} (기대: 200)")
            continue
        if "application/json" not in content_type:
            failures.append(
                f"{path}: content-type '{content_type}' 에 'application/json' 없음"
            )
            continue

        try:
            payload = resp.get_json()
        except Exception as e:
            failures.append(f"{path}: JSON 파싱 실패 ({e})")
            continue

        shape_error = validate(payload)
        if shape_error:
            failures.append(f"{path}: {shape_error}")
            continue

        print(f"OK  {path}  ({resp.status_code}, {content_type})")

    removed_rate = client.get("/api/stats/report-rate-by-sido")
    if removed_rate.status_code != 404:
        failures.append("삭제된 영업신고율 API가 404를 반환하지 않음")

    feature_tip_error = check_feature_tips_admin_api(client)
    if feature_tip_error:
        failures.append(feature_tip_error)
    else:
        print("OK  /api/admin/feature-tips (관리자 인증·입력 검증)")

    user_stats_error = check_user_stats_admin_api(client)
    if user_stats_error:
        failures.append(user_stats_error)
    else:
        print("OK  /api/admin/user-stats (집계 shape·기간·UA 필터)")

    user_stats_data_error = check_user_stats_aggregate_windows_and_view_writers(client)
    if user_stats_data_error:
        failures.append(user_stats_data_error)
    else:
        print("OK  /api/admin/user-stats (날짜 경계·철회 상태·두 페이지뷰 INSERT)")

    poi_errors = _check_map_poi_api(client)
    if poi_errors:
        failures += poi_errors
    else:
        print("OK  지도 편의정보 API (입력 검증·외부 응답·오류 처리)")

    # 단지 2명 정원·지역 단위 담당단지 한도·관리자 탭을 실제 등록으로 확인
    failures += _check_partner_badge_policy(client)
    # 관리자 건물의 브로커 표준데이터 상세·상권정보 폴백·목록 수 우선순위를 확인
    failures += _check_admin_building_broker_details(client)
    # 수집 응답 구조·정규화 키·영업상태 저장 및 실제 라군 표본을 함께 확인
    failures += _check_broker_sync_normalization_and_status(client)

    failures += _check_member_login_history(client)

    # /api/buildings-geo bounds 필터 추가 테스트
    failures += _check_buildings_geo_bounds(client)

    # 수정 요청 → 승인 → 지도 노출 end-to-end 테스트
    failures += _check_building_request_e2e(client)
    # 일부 재분류가 커밋된 뒤 후속 오류가 나도 통계 무효화 표식은 남는지 확인
    failures += _check_master_stats_partial_success_invalidation()
    # 관심저장 POST가 실제 DB에 남고, 새로고침 조회 뒤에도 유지되는지 확인
    failures += _check_favorite_save_persistence(client)
    # 급매 금·은색 등급, 관심단지 전용 토글과 회원/매물별 중복 알림을 확인
    failures += _check_urgent_listing_tiers_and_alerts(client)
    # 채팅 시작은 휴대폰 인증된 사용자만 가능한지 확인
    failures += _check_chat_phone_verification(client)
    # 방 재고의 보증금·만기일 저장·공실 초기화·소유자 권한을 확인
    failures += _check_room_inventory_contract_dates(client)
    # 매물의뢰 보류·보류해제·공개범위는 소유자만 변경하고 수정 시 접수됨으로 복원되는지 확인
    failures += _check_listing_hold_and_disclosure_controls(client)
    # 계약만기 90/60/30/7일 알림의 이메일·알림함 기록과 중복 방지를 확인
    failures += _check_room_expiry_alerts()
    # 새 등록자유형과 과거 agent 값의 저장 호환성을 확인
    failures += _check_listing_registrant_types(client)
    # 전국 도시철도역사정보 표준데이터의 실제 헤더 변형을 importer가 해석하는지 확인
    failures += _check_subway_station_import_headers()
    # 홈 신뢰지표가 현재 건물마스터·거래·매물 COUNT와 일치하는지 확인
    failures += _check_platform_summary(client)
    # 데이터랩 ②~⑥의 TOP·토글·표본 제외·신고율 기준을 확인
    failures += _check_datalab_stats(client)
    # 데이터랩 영업신고현황이 도로명 우선·지번 보조·폐업 제외·신고번호
    # 중복 제거라는 주소 매칭 계약을 실제 임시 데이터로 지키는지 확인
    failures += _check_datalab_report_source_contract()
    # 건물전체 입지정보의 경쟁시설·최단 지하철역·원거리 처리 확인
    failures += _check_whole_listing_location_context(client)
    # 건물전체 매물의 생성·수정·공개범위 계약을 확인
    failures += _check_whole_building_listing(client)
    # 사업주 매물의 사용자·건물별 영업신고번호 인증 캐시와 서버 우회 차단을 확인
    failures += _check_business_listing_verification(client)
    # 사업주 공개 장기방은 공실 장박가능 재고의 월 가격만 공개하고 호실수는 숨긴다.
    failures += _check_public_business_listing_summary(client)
    # 괄호 안 읍·면·동 표기와 신고 주소의 행정구역 표기가 같은 키가 되는지 확인
    failures += _check_lodging_address_normalization()
    # 생활만 객실수 대비 신고율, 나머지는 절대 객실수를 사용하는지 확인
    failures += _check_lodging_metric_contract(client)
    # 명칭 미확정 일반숙박은 영업신고 대표 사업장명으로 자동 표시되는지 확인
    failures += _check_lodging_auto_naming(client)
    # BRHUB 표제부 명칭이 없는 신규 일반숙박도 자동명명 대상 상태로 저장되는지 확인
    failures += _check_brhub_auto_naming_contract()
    # 건물마스터가 매칭된 영업신고의 정상/폐업 상태를 올바르게 필터링하는지 확인
    failures += _check_building_biz_status_filters(client)
    # 일일 캡으로 중간 종료되어도 당일 처리분 자동명명이 반영되는지 확인
    failures += _check_lodging_cap_auto_naming()
    # 관리자 통계표의 일반숙박 호실수 신뢰불가 표시와 비일반 회귀를 확인
    failures += _check_general_units_table_markup(client)
    # 마지막에 실행해 매물 등록 rate limit을 기존 API 회귀 테스트와 공유하지 않는다.
    failures += _check_weekly_email_auto_optin_apis(client)
    # 의뢰 알림 단축 링크의 생성·만료·안전한 리다이렉트와 SMS/이메일 동시 발송을 확인
    failures += _check_lead_short_links(client)
    # 이메일 회원 비밀번호 재설정의 계정 은닉·토큰 만료/1회 사용·메일·제한을 확인
    failures += _check_password_reset_flow(client)

    if failures:
        print("\nAPI 체크 실패:", file=sys.stderr)
        for f in failures:
            print(f"  - {f}", file=sys.stderr)
        return 1

    print(
        "\n모든 API 체크 통과 (/api/health, /api/regions, /api/years,"
        " /api/transactions, /api/buildings-geo, e2e 건물요청→지도노출)"
    )
    return 0


def _check_password_reset_flow(client):
    """이메일 재설정의 보안 경계와 로그인 회귀를 실제 DB 행으로 검증한다."""
    failures = []
    tag = f"password-reset-{time.time_ns()}"
    email = f"{tag}@example.test"
    kakao_email = f"{tag}-kakao@example.test"
    unknown_email = f"{tag}-unknown@example.test"
    rate_email = f"{tag}-rate@example.test"
    email_user_id = None
    kakao_user_id = None
    partner_ids = {}
    token = None
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(
            """
            INSERT INTO users (email, password_hash, name, provider, status)
            VALUES (%s, %s, %s, 'email', 'active')
            RETURNING id
            """,
            (email, generate_password_hash("before-reset-password"), "재설정 테스트"),
        )
        email_user_id = cur.fetchone()["id"]
        cur.execute(
            """
            INSERT INTO users (email, name, provider, status)
            VALUES (%s, %s, 'kakao', 'active')
            RETURNING id
            """,
            (kakao_email, "카카오 재설정 테스트"),
        )
        kakao_user_id = cur.fetchone()["id"]
        cur.execute(
            """
            INSERT INTO agents
                (office_name, owner_name, reg_number, email, status, password_hash)
            VALUES (%s, %s, %s, %s, 'approved', %s)
            RETURNING id
            """,
            (
                "중개사무소 테스트", "중개사 대표", f"{tag}-agent-reg",
                email, generate_password_hash("before-agent-password"),
            ),
        )
        partner_ids["agent"] = cur.fetchone()["id"]
        cur.execute(
            """
            INSERT INTO operators
                (company_name, owner_name, category, email, status, password_hash)
            VALUES (%s, %s, '위탁', %s, 'approved', %s)
            RETURNING id
            """,
            (
                "운영업체 테스트", "운영업체 대표", f"{tag}-operator@example.test",
                generate_password_hash("before-operator-password"),
            ),
        )
        partner_ids["operator"] = cur.fetchone()["id"]
        cur.execute(
            """
            INSERT INTO loan_consultants
                (office_name, owner_name, license_number, email, status, password_hash)
            VALUES (%s, %s, %s, %s, 'approved', %s)
            RETURNING id
            """,
            (
                "대출상담사무소 테스트", "대출상담사 대표", f"{tag}-loan-license",
                f"{tag}-loan@example.test", generate_password_hash("before-loan-password"),
            ),
        )
        partner_ids["loan_consultant"] = cur.fetchone()["id"]
        conn.commit()

        raw_token = f"raw-{tag}"
        partner_tokens = {
            "agent": f"raw-agent-{tag}",
            "operator": f"raw-operator-{tag}",
            "loan_consultant": f"raw-loan-{tag}",
        }
        partner_emails = {
            "operator": f"{tag}-operator@example.test",
            "loan_consultant": f"{tag}-loan@example.test",
        }
        with (
            patch.dict(os.environ, {"PUBLIC_BASE_URL": "https://homenstay.com"}, clear=False),
            patch.object(
                app_module._secrets,
                "token_urlsafe",
                side_effect=[raw_token, *partner_tokens.values()],
            ),
            patch.object(app_module, "_queue_password_reset_email") as queued_email,
            patch.object(app_module, "send_email") as email_sender,
        ):
            email_response = client.post(
                "/api/auth/request-password-reset",
                json={"email": email.upper()},
                headers={"Host": "attacker.example.invalid"},
            )
            unknown_response = client.post(
                "/api/auth/request-password-reset",
                json={"email": unknown_email},
            )
            kakao_response = client.post(
                "/api/auth/request-password-reset",
                json={"email": kakao_email},
            )
            partner_responses = {
                account_type: client.post(
                    "/api/auth/request-password-reset",
                    json={"email": partner_email},
                )
                for account_type, partner_email in partner_emails.items()
            }

            reset_responses = [
                email_response, unknown_response, kakao_response, *partner_responses.values()
            ]
            if any(response.status_code != 200 for response in reset_responses):
                failures.append("비밀번호 재설정 요청: 이메일/미존재/카카오 계정이 모두 200 응답하지 않음")
            else:
                bodies = [response.get_json() or {} for response in reset_responses]
                if not all(body == bodies[0] and body.get("ok") is True for body in bodies):
                    failures.append("비밀번호 재설정 요청: 계정 종류별 응답이 통일되지 않음")

            if email_sender.called:
                failures.append("비밀번호 재설정 요청: 외부 메일 발송이 HTTP 응답 경로에서 실행됨")
            if queued_email.call_count != 6:
                failures.append(f"비밀번호 재설정 요청: 모든 경우의 메일 작업을 큐잉하지 않음 ({queued_email.call_count}건)")
            elif queued_email.call_args_list:
                reset_html = queued_email.call_args_list[0].args[1]
                kakao_html = next(
                    (
                        call.args[1] for call in queued_email.call_args_list
                        if "카카오 로그인" in (call.args[1] or "")
                    ),
                    "",
                )
                if "https://homenstay.com/reset-password?token=" not in reset_html:
                    failures.append("비밀번호 재설정 메일: 공식 도메인 재설정 링크가 없음")
                if "attacker.example.invalid" in reset_html:
                    failures.append("비밀번호 재설정 메일: 요청 Host가 메일 링크에 반영됨")
                if "카카오 로그인" not in kakao_html:
                    failures.append("비밀번호 재설정 메일: 카카오 계정 안내가 없음")
                queued_recipients = [call.args[0] for call in queued_email.call_args_list]
                reset_links = [call.args[1] for call in queued_email.call_args_list
                               if "/reset-password?token=" in (call.args[1] or "")]
                if queued_recipients.count(email) != 2:
                    failures.append("비밀번호 재설정 요청: 같은 이메일의 일반회원·중개사 토큰을 모두 큐잉하지 않음")
                if len(reset_links) != 4:
                    failures.append("비밀번호 재설정 메일: 네 계정 유형의 재설정 링크가 모두 없음")

        cur.execute(
            """
            SELECT token,
                   expires_at > NOW() + INTERVAL '29 minutes'
                   AND expires_at < NOW() + INTERVAL '31 minutes' AS ttl_ok
                   , account_type
              FROM password_reset_tokens
             WHERE user_id = %s
             ORDER BY id DESC
             LIMIT 1
            """,
            (email_user_id,),
        )
        token_row = cur.fetchone() or {}
        token = raw_token
        if not token or not token_row.get("ttl_ok") or token_row.get("account_type") != "user":
            failures.append("비밀번호 재설정 요청: 30분 유효 토큰을 저장하지 않음")
        else:
            stored_token = token_row.get("token")
            if stored_token == token or stored_token != hashlib.sha256(token.encode("utf-8")).hexdigest():
                failures.append("비밀번호 재설정 요청: URL 원문 대신 토큰 다이제스트를 저장하지 않음")
            short_password = client.post(
                "/api/auth/reset-password",
                json={"token": token, "new_password": "short"},
            )
            if short_password.status_code != 400:
                failures.append("비밀번호 재설정: 8자 미만 비밀번호를 차단하지 않음")

            changed = client.post(
                "/api/auth/reset-password",
                json={"token": token, "new_password": "after-reset-password"},
            )
            if changed.status_code != 200 or not (changed.get_json() or {}).get("ok"):
                failures.append("비밀번호 재설정: 유효 토큰으로 비밀번호 변경 실패")

            cur.execute(
                """
                SELECT u.password_hash, t.used_at IS NOT NULL AS used
                  FROM users u
                  JOIN password_reset_tokens t ON t.user_id = u.id
                 WHERE t.token = %s
                """,
                (hashlib.sha256(token.encode("utf-8")).hexdigest(),),
            )
            changed_row = cur.fetchone() or {}
            if not changed_row.get("used") or not check_password_hash(
                changed_row.get("password_hash") or "", "after-reset-password"
            ):
                failures.append("비밀번호 재설정: 해시 갱신 또는 토큰 1회 사용 기록 실패")

            reused = client.post(
                "/api/auth/reset-password",
                json={"token": token, "new_password": "another-password"},
            )
            if reused.status_code != 400:
                failures.append("비밀번호 재설정: 이미 사용된 토큰을 차단하지 않음")

            logged_in = client.post(
                "/api/auth/login",
                json={"email": email, "password": "after-reset-password"},
            )
            if logged_in.status_code != 200 or not (logged_in.get_json() or {}).get("ok"):
                failures.append("비밀번호 재설정: 새 비밀번호 이메일 로그인이 동작하지 않음")

        partner_tables = {
            "agent": "agents",
            "operator": "operators",
            "loan_consultant": "loan_consultants",
        }
        for index, (account_type, table_name) in enumerate(partner_tables.items(), start=1):
            cur.execute(
                """
                SELECT token, account_type
                  FROM password_reset_tokens
                 WHERE user_id = %s AND account_type = %s
                 ORDER BY id DESC
                 LIMIT 1
                """,
                (partner_ids[account_type], account_type),
            )
            partner_token_row = cur.fetchone() or {}
            partner_token = partner_tokens[account_type]
            if (
                partner_token_row.get("account_type") != account_type
                or partner_token_row.get("token") != hashlib.sha256(partner_token.encode("utf-8")).hexdigest()
            ):
                failures.append(f"비밀번호 재설정 요청: {account_type} account_type 토큰 저장 실패")
                continue
            partner_client = app.test_client()
            reset_partner = partner_client.post(
                "/api/auth/reset-password",
                json={"token": partner_token, "new_password": f"after-{account_type}-password"},
                environ_overrides={"REMOTE_ADDR": f"198.51.100.{index}"},
            )
            cur.execute(
                f"SELECT password_hash FROM {table_name} WHERE id = %s",
                (partner_ids[account_type],),
            )
            partner_password_row = cur.fetchone() or {}
            if (
                reset_partner.status_code != 200
                or not check_password_hash(
                    partner_password_row.get("password_hash") or "",
                    f"after-{account_type}-password",
                )
            ):
                failures.append(f"비밀번호 재설정: {account_type} 비밀번호 변경 실패")
            reused_partner = partner_client.post(
                "/api/auth/reset-password",
                json={"token": partner_token, "new_password": "reused-partner-password"},
                environ_overrides={"REMOTE_ADDR": f"198.51.100.{index}"},
            )
            if reused_partner.status_code != 400:
                failures.append(f"비밀번호 재설정: {account_type} 토큰 재사용을 차단하지 않음")

        expired_token = f"expired-{tag}"
        cur.execute(
            """
            INSERT INTO password_reset_tokens (user_id, token, expires_at)
            VALUES (%s, %s, NOW() - INTERVAL '1 minute')
            """,
            (email_user_id, hashlib.sha256(expired_token.encode("utf-8")).hexdigest()),
        )
        conn.commit()
        expired = client.post(
            "/api/auth/reset-password",
            json={"token": expired_token, "new_password": "after-reset-password"},
        )
        if expired.status_code != 400:
            failures.append("비밀번호 재설정: 만료된 토큰을 차단하지 않음")

        race_token = f"race-{tag}"
        cur.execute(
            """
            INSERT INTO password_reset_tokens (user_id, token, expires_at)
            VALUES (%s, %s, NOW() + INTERVAL '30 minutes')
            """,
            (email_user_id, hashlib.sha256(race_token.encode("utf-8")).hexdigest()),
        )
        conn.commit()

        def redeem_race(password):
            race_client = app.test_client()
            return race_client.post(
                "/api/auth/reset-password",
                json={"token": race_token, "new_password": password},
            ).status_code

        with ThreadPoolExecutor(max_workers=2) as executor:
            race_statuses = list(executor.map(redeem_race, ("race-password-one", "race-password-two")))
        if sorted(race_statuses) != [200, 400]:
            failures.append(f"비밀번호 재설정: 동시 토큰 사용이 1회로 제한되지 않음 ({race_statuses})")

        with patch.object(app_module, "_queue_password_reset_email"):
            rate_responses = [
                client.post("/api/auth/request-password-reset", json={"email": rate_email})
                for _ in range(4)
            ]
        if [response.status_code for response in rate_responses[:3]] != [200, 200, 200]:
            failures.append("비밀번호 재설정 요청: 동일 이메일의 분당 3회 허용이 동작하지 않음")
        if rate_responses[3].status_code != 429:
            failures.append("비밀번호 재설정 요청: 동일 이메일의 분당 3회 제한이 동작하지 않음")

        reset_page = client.get("/reset-password")
        if reset_page.status_code != 200 or b"/api/auth/reset-password" not in reset_page.data:
            failures.append("비밀번호 재설정 페이지: 정적 페이지 또는 API 연결이 없음")
    except Exception as exc:
        conn.rollback()
        failures.append(f"비밀번호 재설정 테스트 오류: {exc}")
    finally:
        try:
            all_account_ids = [uid for uid in (email_user_id, kakao_user_id, *partner_ids.values()) if uid]
            if all_account_ids:
                cur.execute(
                    "DELETE FROM password_reset_tokens WHERE user_id = ANY(%s)",
                    (all_account_ids,),
                )
                cur.execute(
                    "DELETE FROM users WHERE id = ANY(%s)",
                    ([uid for uid in (email_user_id, kakao_user_id) if uid],),
                )
                cur.execute("DELETE FROM agents WHERE id = ANY(%s)", ([partner_ids.get("agent")] if partner_ids.get("agent") else [],))
                cur.execute("DELETE FROM operators WHERE id = ANY(%s)", ([partner_ids.get("operator")] if partner_ids.get("operator") else [],))
                cur.execute(
                    "DELETE FROM loan_consultants WHERE id = ANY(%s)",
                    ([partner_ids.get("loan_consultant")] if partner_ids.get("loan_consultant") else [],),
                )
                conn.commit()
        except Exception as cleanup_exc:
            conn.rollback()
            failures.append(f"비밀번호 재설정 테스트 정리 실패: {cleanup_exc}")
        finally:
            cur.close()
            conn.close()

    if not failures:
        print("OK  이메일 비밀번호 재설정 (계정 은닉·30분·1회용·메일·제한·로그인)")
    return failures


def _check_lead_short_links(client):
    """의뢰 알림용 6자리 링크와 best-effort SMS·이메일 계약을 점검한다."""
    failures = []
    code = None
    agent_id = None
    try:
        target_path = "/agent/dashboard?tab=leads-buy&request_id=123"
        # 발신 URL은 수신자가 신뢰하는 공식 주소여야 한다. 접수 요청의 Host
        # 헤더는 공격자가 제어할 수 있으므로 단축 링크에 반영되면 안 된다.
        with patch.dict(os.environ, {"PUBLIC_BASE_URL": "https://homenstay.com"}):
            with app.test_request_context("/", base_url="https://attacker.example.invalid"):
                short_url = _create_short_link(target_path)
                if not short_url or not re.fullmatch(
                    r"https://homenstay\.com/s/[23456789ABCDEFGHJKLMNPQRSTUVWXYZ]{6}",
                    short_url,
                ):
                    return ["의뢰 알림: 신뢰된 공식 도메인의 6자리 단축 URL을 생성하지 못함"]
                code = short_url.rsplit("/", 1)[-1]
                if _create_short_link("https://example.invalid/redirect") is not None:
                    failures.append("의뢰 알림: 외부 URL을 단축 링크 대상으로 허용함")

        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute(
                """
                SELECT target_path, expires_at > NOW() + INTERVAL '71 hours'
                       AND expires_at < NOW() + INTERVAL '73 hours' AS ttl_ok
                FROM short_links WHERE code=%s
                """,
                (code,),
            )
            row = cur.fetchone() or {}
            if row.get("target_path") != target_path or not row.get("ttl_ok"):
                failures.append("의뢰 알림: 단축 링크 대상 또는 72시간 만료시각이 잘못됨")

            active = client.get(f"/s/{code}", follow_redirects=False)
            if (
                active.status_code not in (301, 302, 303, 307, 308)
                or active.headers.get("Location") != target_path
            ):
                failures.append("의뢰 알림: 유효한 단축 링크가 중개사 대시보드로 이동하지 않음")

            with client.session_transaction() as sess:
                sess.pop("agent_id", None)
            login_redirect = client.get(target_path, follow_redirects=False)
            login_location = login_redirect.headers.get("Location") or ""
            if (
                login_redirect.status_code not in (301, 302, 303, 307, 308)
                or not login_location.startswith("/agent/login?next=")
                or target_path not in login_location.replace("%2F", "/").replace("%3F", "?").replace("%3D", "=").replace("%26", "&")
            ):
                failures.append("의뢰 알림: 비로그인 중개사의 대시보드 목적지가 로그인 뒤에도 보존되지 않음")

            login_page = client.get(login_location)
            login_html = login_page.get_data(as_text=True)
            if "nextTarget" not in login_html or "window.location.href = nextTarget" not in login_html:
                failures.append("의뢰 알림: 중개사 로그인 후 대시보드 복귀 처리가 누락됨")

            run_id = str(int(time.time() * 1000))
            agent_email = f"lead-link-agent-{run_id}@example.test"
            agent_password = "lead-link-test-password"
            cur.execute(
                """
                INSERT INTO agents (office_name, owner_name, reg_number, email, password_hash, status)
                VALUES (%s, %s, %s, %s, %s, 'approved')
                RETURNING id
                """,
                ["단축링크 테스트중개사", "테스트대표", f"lead-link-{run_id}",
                 agent_email, generate_password_hash(agent_password)],
            )
            agent_id = cur.fetchone()["id"]
            conn.commit()
            signed_in = client.post("/api/agent/login", json={
                "email": agent_email, "password": agent_password,
            })
            if signed_in.status_code != 200 or not (signed_in.get_json() or {}).get("ok"):
                failures.append("의뢰 알림: 임시 중개사 계정으로 로그인하지 못함")

            dashboard = client.get(target_path)
            dashboard_html = dashboard.get_data(as_text=True)
            if (
                dashboard.status_code != 200
                or "const REQUESTED_TAB" not in dashboard_html
                or "focusRequestedLead" not in dashboard_html
                or 'id="buy-lead-card-${r.id}"' not in dashboard_html
            ):
                failures.append("의뢰 알림: 로그인한 중개사의 의뢰 탭·요청 강조 딥링크가 누락됨")

            cur.execute("UPDATE short_links SET expires_at=NOW() - INTERVAL '1 second' WHERE code=%s", (code,))
            conn.commit()
            expired = client.get(f"/s/{code}", follow_redirects=False)
            unknown = client.get("/s/ABCDEFG", follow_redirects=False)
            malformed = client.get("/s/not-valid", follow_redirects=False)
            if any(response.headers.get("Location") != "/" for response in (expired, unknown, malformed)):
                failures.append("의뢰 알림: 만료·없는·잘못된 단축 링크가 홈으로 이동하지 않음")
        finally:
            if code:
                cur.execute("DELETE FROM short_links WHERE code=%s", (code,))
            if agent_id:
                cur.execute("DELETE FROM agents WHERE id=%s", (agent_id,))
            conn.commit()
            cur.close()
            conn.close()

        agents = [{
            "id": 31,
            "phone": "01012345678",
            "email": "broker@example.test",
            "office_name": "테스트중개사",
        }]
        with app.test_request_context("/", base_url="https://lead-link.example.test"), \
             patch.object(app_module, "_create_short_link", return_value="https://lead-link.example.test/s/ABCDEF"), \
             patch.object(app_module, "send_sms", return_value=(True, "발송 성공")) as sms_mock, \
             patch.object(app_module, "send_email", return_value=(True, "발송 성공")) as email_mock:
            results, sent_url = _notify_lead_agents(
                agents, 31, "[홈앤스테이] 매수의뢰 접수 — <테스트건물> / 매매",
                "buy", target_path,
                "<테스트건물>", "매매", "10,000", "01012345678",
            )
        sms_body = sms_mock.call_args.args[1] if sms_mock.call_args else ""
        email_html = email_mock.call_args.args[2] if email_mock.call_args else ""
        if (
            sent_url != "https://lead-link.example.test/s/ABCDEF"
            or not results or not results[0]["sent"]
            or "https://lead-link.example.test/s/ABCDEF" not in sms_body
            or len(sms_body.encode("utf-8")) <= 0
            or "https://lead-link.example.test/s/ABCDEF" not in email_html
            or "&lt;테스트건물&gt;" not in email_html
        ):
            failures.append("의뢰 알림: SMS 단축 URL 또는 이메일 HTML 이스케이프·동시 발송이 누락됨")
    except Exception as exc:
        failures.append(f"의뢰 알림 단축 링크 테스트 오류: {exc}")

    if not failures:
        print("OK  의뢰 알림 6자리 단축 URL·72시간 만료·SMS/이메일 동시 발송")
    return failures


def _check_partner_badge_policy(client):
    """테스트 전용 파트너를 실제 API로 등록해 새 뱃지 정책의 핵심 흐름을 확인한다."""
    failures = []
    tag = f"partner-badge-{time.time_ns()}"
    agent_ids, operator_ids, loan_ids, application_ids = [], [], [], []
    email_patcher = patch.object(app_module, "send_email", return_value=(True, "테스트 발송 성공"))
    email_patcher.start()
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT mb.id
            FROM master_buildings mb
            WHERE NOT EXISTS (
                SELECT 1 FROM agent_buildings ab
                WHERE ab.master_building_id=mb.id AND ab.has_priority_badge
                  AND (ab.premium_expires_at IS NULL OR ab.premium_expires_at > NOW())
            )
              AND NOT EXISTS (
                SELECT 1 FROM operator_buildings ob
                JOIN operators o ON o.id=ob.operator_id
                WHERE ob.master_building_id=mb.id AND o.category='청소' AND ob.has_priority_badge
                  AND (ob.premium_expires_at IS NULL OR ob.premium_expires_at > NOW())
            )
              AND NOT EXISTS (
                SELECT 1 FROM loan_consultant_buildings lcb
                WHERE lcb.master_building_id=mb.id AND lcb.has_priority_badge
                  AND (lcb.premium_expires_at IS NULL OR lcb.premium_expires_at > NOW())
            )
            ORDER BY mb.id
            LIMIT 40
        """)
        building_ids = [row["id"] for row in cur.fetchall()]
        if len(building_ids) < 37:
            return ["파트너 뱃지: 테스트용 마스터 건물이 37개 미만입니다."]
        agent_bld, operator_bld, loan_bld, agent_approval_bld, loan_approval_bld, *_ = building_ids
        cur.execute("""
            SELECT DISTINCT ON (mb.sgg_text)
                   mb.id, mb.building_name, mb.sgg_text, mb.umd_nm
            FROM master_buildings mb
            WHERE mb.sgg_text IS NOT NULL AND mb.sgg_text <> ''
              AND mb.building_name <> '-'
              AND (
                  SELECT COUNT(*) FROM master_buildings same_sgg
                  WHERE same_sgg.sgg_text=mb.sgg_text
                    AND same_sgg.building_name <> '-'
              ) >= 12
              AND NOT EXISTS (
                  SELECT 1 FROM agent_buildings ab
                  WHERE ab.master_building_id=mb.id AND ab.has_priority_badge
                    AND (ab.premium_expires_at IS NULL OR ab.premium_expires_at > NOW())
              )
              AND (
                  SELECT COUNT(*)
                  FROM agent_service_regions sr
                  WHERE sr.sgg_text=mb.sgg_text
                    AND sr.expires_at > NOW()
              ) = 0
              AND NOT EXISTS (
                  SELECT 1 FROM operator_service_regions osr
                  WHERE osr.sgg_text=mb.sgg_text AND osr.expires_at > NOW()
              )
            ORDER BY mb.sgg_text, mb.id
            LIMIT 2
        """)
        region_fixtures = cur.fetchall()
        if len(region_fixtures) < 2:
            return ["파트너 뱃지: 시군구 지역뱃지 테스트용 지역을 2개 찾지 못했습니다."]
        region_fixture, region_race_fixture = region_fixtures
        deadline = app_module.PARTNER_BADGE_FREE_EXPIRES_AT

        def create_agent(label):
            cur.execute("""
                INSERT INTO agents (office_name, owner_name, reg_number, phone, email, status)
                VALUES (%s, %s, %s, '01000000000', %s, 'approved') RETURNING id
            """, (f"{tag}-{label}", "정책테스트", f"{tag}-{label}", f"{tag}-{label}@example.test"))
            partner_id = cur.fetchone()["id"]
            agent_ids.append(partner_id)
            return partner_id

        def create_operator(label, category="청소"):
            cur.execute("""
                INSERT INTO operators (company_name, owner_name, category, phone, email, status)
                VALUES (%s, %s, %s, '01000000000', %s, 'approved') RETURNING id
            """, (f"{tag}-{label}", "정책테스트", category, f"{tag}-{label}@example.test"))
            partner_id = cur.fetchone()["id"]
            operator_ids.append(partner_id)
            return partner_id

        def create_loan(label):
            cur.execute("""
                INSERT INTO loan_consultants (office_name, owner_name, license_number, phone, email, status)
                VALUES (%s, %s, %s, '01000000000', %s, 'approved') RETURNING id
            """, (f"{tag}-{label}", "정책테스트", f"{tag}-{label}", f"{tag}-{label}@example.test"))
            partner_id = cur.fetchone()["id"]
            loan_ids.append(partner_id)
            return partner_id

        # ① 중개사 전속단지: 등록 즉시 골드뱃지와 고정 만료일.
        agent_id = create_agent("agent")
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = agent_id
        agent_add = client.post("/api/agent/buildings", json={"master_building_id": agent_bld})
        if agent_add.status_code != 200 or not (agent_add.get_json() or {}).get("ok"):
            failures.append(f"파트너 뱃지: 중개사 전속단지 등록 실패 ({agent_add.get_json()})")
        cur.execute("""
            SELECT has_priority_badge, premium_expires_at::text AS expires_at
            FROM agent_buildings WHERE agent_id=%s AND master_building_id=%s
        """, (agent_id, agent_bld))
        row = cur.fetchone() or {}
        if not row.get("has_priority_badge") or not (row.get("expires_at") or "").startswith(deadline):
            failures.append("파트너 뱃지: 중개사 전속단지에 즉시 골드뱃지/고정 만료일이 저장되지 않음")

        # 같은 단지에는 두 중개사까지 성공하고, 세 번째 활성 신청은 대기 등록된다.
        slot_second_id = create_agent("agent-slot-second")
        slot_third_id = create_agent("agent-slot-third")
        slot_fourth_id = create_agent("agent-slot-fourth")
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = slot_second_id
        slot_second = client.post("/api/agent/buildings", json={"master_building_id": agent_bld})
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = slot_third_id
        slot_third_rejected = client.post("/api/agent/buildings", json={"master_building_id": agent_bld})
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = slot_fourth_id
        slot_fourth_rejected = client.post("/api/agent/buildings", json={"master_building_id": agent_bld})
        if (
            slot_second.status_code != 200
            or slot_third_rejected.status_code != 200
            or not (slot_third_rejected.get_json() or {}).get("waitlisted")
            or slot_fourth_rejected.status_code != 200
            or not (slot_fourth_rejected.get_json() or {}).get("waitlisted")
        ):
            failures.append("파트너 뱃지: 건물당 두 번째 중개사를 허용하거나 초과 신청을 대기 등록하지 못함")
        waitlisted_dashboard = app_module._agent_me_data(slot_third_id) or {}
        waitlisted_items = [
            item for item in waitlisted_dashboard.get("buildings", [])
            if item.get("master_building_id") == agent_bld
        ]
        if (
            len(waitlisted_items) != 1
            or not waitlisted_items[0].get("waitlist_only")
            or waitlisted_items[0].get("waitlist_notified")
            or not waitlisted_items[0].get("occupied_by_other")
        ):
            failures.append("파트너 뱃지: 정원이 찬 대기 전용 단지 상태를 대시보드에 유지하지 못함")
        detail = client.get(f"/api/building/{agent_bld}")
        detail_agent_ids = {item.get("id") for item in (detail.get_json() or {}).get("agents", [])}
        if detail.status_code != 200 or not {agent_id, slot_second_id} <= detail_agent_ids:
            failures.append("파트너 뱃지: 건물 상세에 활성 단지뱃지 중개사 2명이 함께 표시되지 않음")

        # 만료된 뱃지는 정원에서 제외되어 직전에 거절된 세 번째 중개사가 입점할 수 있다.
        cur.execute("""
            UPDATE agent_buildings
            SET premium_expires_at=NOW() - INTERVAL '1 second'
            WHERE agent_id=%s AND master_building_id=%s
        """, (slot_second_id, agent_bld))
        conn.commit()
        if app_module._process_badge_waitlist_notifications() < 1:
            failures.append("파트너 뱃지: 단지 빈자리 발생 뒤 자동 대기 알림을 처리하지 못함")
        cur.execute("""
            SELECT agent_id, notified_at IS NOT NULL AS notified
            FROM premium_waitlist
            WHERE master_building_id=%s AND agent_id=ANY(%s)
        """, (agent_bld, [slot_third_id, slot_fourth_id]))
        first_cycle = {row["agent_id"]: row["notified"] for row in cur.fetchall()}
        if first_cycle != {slot_third_id: True, slot_fourth_id: False}:
            failures.append("파트너 뱃지: 한 빈자리에는 가장 먼저 등록한 대기자 한 명만 알리지 못함")
        available_dashboard = app_module._agent_me_data(slot_third_id) or {}
        available_items = [
            item for item in available_dashboard.get("buildings", [])
            if item.get("master_building_id") == agent_bld
        ]
        if (
            len(available_items) != 1
            or available_items[0].get("occupied_by_other")
            or not available_items[0].get("waitlist_notified")
        ):
            failures.append("파트너 뱃지: 빈자리 알림이 끝난 대기 전용 단지를 신청 가능 상태로 표시하지 못함")
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = slot_third_id
        slot_after_expiry = client.post("/api/agent/buildings", json={"master_building_id": agent_bld})
        if slot_after_expiry.status_code != 200:
            failures.append("파트너 뱃지: 만료된 뱃지를 활성 두 자리 정원에서 제외하지 않음")
        cur.execute(
            "SELECT COUNT(*) AS count FROM premium_waitlist WHERE agent_id=%s AND master_building_id=%s",
            (slot_third_id, agent_bld),
        )
        if cur.fetchone()["count"] != 0:
            failures.append("파트너 뱃지: 빈자리 신청 성공 뒤 단지 대기 행을 정리하지 못함")
        cur.execute(
            "SELECT notified_at FROM premium_waitlist WHERE agent_id=%s AND master_building_id=%s",
            (slot_fourth_id, agent_bld),
        )
        remaining_waiter = cur.fetchone()
        if not remaining_waiter or remaining_waiter["notified_at"] is not None:
            failures.append("파트너 뱃지: 먼저 알림받은 대기자가 입점한 뒤 다음 대기자를 재대기 상태로 유지하지 못함")
        cur.execute("""
            UPDATE agent_buildings
            SET premium_expires_at=NOW() - INTERVAL '1 second'
            WHERE agent_id=%s AND master_building_id=%s
        """, (slot_third_id, agent_bld))
        conn.commit()
        if app_module._process_badge_waitlist_notifications() < 1:
            failures.append("파트너 뱃지: 다음 빈자리에서 남은 대기자에게 다시 알리지 못함")
        cur.execute(
            "SELECT notified_at IS NOT NULL AS notified FROM premium_waitlist WHERE agent_id=%s AND master_building_id=%s",
            (slot_fourth_id, agent_bld),
        )
        remaining_waiter = cur.fetchone()
        if not remaining_waiter or not remaining_waiter["notified"]:
            failures.append("파트너 뱃지: 두 번째 빈자리 알림 상태를 저장하지 못함")
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = slot_fourth_id
        final_waiter_claim = client.post(
            "/api/agent/buildings", json={"master_building_id": agent_bld}
        )
        if final_waiter_claim.status_code != 200 or (final_waiter_claim.get_json() or {}).get("waitlisted"):
            failures.append("파트너 뱃지: 다음 빈자리 알림을 받은 대기자가 단지뱃지를 신청하지 못함")
        timeout_first_id = create_agent("agent-timeout-first")
        timeout_next_id = create_agent("agent-timeout-next")
        conn.commit()
        for waiting_id in (timeout_first_id, timeout_next_id):
            with client.session_transaction() as sess:
                sess.clear()
                sess["agent_id"] = waiting_id
            queued = client.post("/api/agent/buildings", json={"master_building_id": agent_bld})
            if not (queued.get_json() or {}).get("waitlisted"):
                failures.append("파트너 뱃지: 만료 순번 검증용 대기 등록에 실패함")
        cur.execute(
            "UPDATE agent_buildings SET premium_expires_at=NOW()-INTERVAL '1 day' "
            "WHERE agent_id=%s AND master_building_id=%s",
            (agent_id, agent_bld),
        )
        conn.commit()
        app_module._process_badge_waitlist_notifications()
        cur.execute(
            "UPDATE premium_waitlist SET notified_at=NOW()-INTERVAL '31 minutes' "
            "WHERE agent_id=%s AND master_building_id=%s",
            (timeout_first_id, agent_bld),
        )
        conn.commit()
        app_module._process_badge_waitlist_notifications()
        cur.execute(
            "SELECT agent_id, notified_at FROM premium_waitlist "
            "WHERE agent_id IN (%s,%s) AND master_building_id=%s",
            (timeout_first_id, timeout_next_id, agent_bld),
        )
        timeout_rows = {r["agent_id"]: r for r in cur.fetchall()}
        if timeout_first_id in timeout_rows or not timeout_rows.get(timeout_next_id, {}).get("notified_at"):
            failures.append("파트너 뱃지: 신청 기한이 지난 대기자 다음 순번으로 알림이 진행되지 않음")
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = timeout_next_id
        cancel_wait = client.delete(f"/api/agent/buildings/{agent_bld}")
        cur.execute(
            "SELECT 1 FROM premium_waitlist WHERE agent_id=%s AND master_building_id=%s",
            (timeout_next_id, agent_bld),
        )
        if cancel_wait.status_code != 200 or cur.fetchone():
            failures.append("파트너 뱃지: 단지뱃지 대기 취소가 대기 행을 정리하지 못함")
        cur.execute(
            "UPDATE agent_buildings SET premium_expires_at=%s WHERE agent_id=%s AND master_building_id=%s",
            (app_module.PARTNER_BADGE_FREE_EXPIRES_AT, agent_id, agent_bld),
        )
        conn.commit()

        # 세 중개사가 빈 단지에 동시에 등록해도 정확히 두 명만 성공한다.
        slot_race_bld = building_ids[36]
        slot_race_agent_ids = [create_agent(f"agent-slot-race-{index}") for index in range(3)]
        conn.commit()

        def parallel_building_slot_claim(race_agent_id):
            with app.test_client() as parallel_client:
                with parallel_client.session_transaction() as sess:
                    sess["agent_id"] = race_agent_id
                response = parallel_client.post(
                    "/api/agent/buildings", json={"master_building_id": slot_race_bld}
                )
                return response.status_code, bool((response.get_json() or {}).get("waitlisted"))

        with ThreadPoolExecutor(max_workers=3) as executor:
            slot_race_results = list(executor.map(
                parallel_building_slot_claim, slot_race_agent_ids
            ))
        cur.execute("""
            SELECT COUNT(*) AS count FROM agent_buildings
            WHERE master_building_id=%s AND has_priority_badge
              AND (premium_expires_at IS NULL OR premium_expires_at > NOW())
        """, (slot_race_bld,))
        active_slot_count = cur.fetchone()["count"]
        if (
            [item[0] for item in slot_race_results].count(200) != 3
            or [item[1] for item in slot_race_results].count(True) != 1
            or active_slot_count != 2
        ):
            failures.append(f"파트너 뱃지: 단지뱃지 동시 신청이 두 자리 정원을 우회함 ({slot_race_results})")

        # 같은 중개사가 9개를 가진 상태의 동시 등록도 10개 한도를 넘지 않아야 한다.
        capacity_agent_id = create_agent("agent-capacity")
        for building_id in building_ids[5:14]:
            cur.execute(
                "INSERT INTO agent_buildings (agent_id, master_building_id) VALUES (%s, %s)",
                (capacity_agent_id, building_id),
            )
        conn.commit()

        def parallel_agent_add(building_id):
            with app.test_client() as parallel_client:
                with parallel_client.session_transaction() as sess:
                    sess["agent_id"] = capacity_agent_id
                return parallel_client.post(
                    "/api/agent/buildings", json={"master_building_id": building_id}
                ).status_code

        with ThreadPoolExecutor(max_workers=2) as executor:
            capacity_results = list(executor.map(parallel_agent_add, building_ids[14:16]))
        cur.execute("SELECT COUNT(*) AS count FROM agent_buildings WHERE agent_id=%s", (capacity_agent_id,))
        if sorted(capacity_results) != [200, 400] or (cur.fetchone() or {}).get("count") != 10:
            failures.append("파트너 뱃지: 중개사 동시 등록이 담당단지 10개 한도를 우회함")

        # ② 운영업체 단지뱃지: 같은 단지·업종은 두 곳까지 허용하고 세 번째를 거절한다.
        operator_id = create_operator("operator-primary")
        operator_second_id = create_operator("operator-second")
        operator_third_id = create_operator("operator-third")
        expired_operator_id = create_operator("operator-expired-region")
        cur.execute("""
            INSERT INTO operator_service_regions (operator_id, sgg_text, expires_at)
            VALUES (%s, %s, NOW() - INTERVAL '1 day')
        """, (expired_operator_id, f"{tag}-expired"))
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["operator_id"] = expired_operator_id
        expired_same_reclaim = client.post(
            "/api/operator/service-regions",
            json={"sgg_text": f"{tag}-expired"},
        )
        cur.execute("""
            UPDATE operator_service_regions
            SET expires_at=NOW() - INTERVAL '1 day'
            WHERE operator_id=%s
        """, (expired_operator_id,))
        conn.commit()
        expired_other_reclaim = client.post(
            "/api/operator/service-regions",
            json={"sgg_text": f"{tag}-new-region"},
        )
        active_regions = client.get("/api/operator/service-regions")
        active_region_names = [
            item.get("sgg_text")
            for item in ((active_regions.get_json() or {}).get("regions") or [])
        ]
        if (
            expired_same_reclaim.status_code != 200
            or expired_other_reclaim.status_code != 200
            or active_region_names != [f"{tag}-new-region"]
        ):
            failures.append("파트너 뱃지: 운영업체 만료 지역의 동일지역 갱신·새 지역 선택·활성지역 조회가 일치하지 않음")
        with client.session_transaction() as sess:
            sess.clear()
            sess["operator_id"] = operator_id
        op_add = client.post("/api/operator/buildings", json={"master_building_id": operator_bld})
        op_claim = client.post(f"/api/operator/buildings/{operator_bld}/claim-premium")
        if op_add.status_code != 200 or op_claim.status_code != 200:
            failures.append("파트너 뱃지: 운영업체 단지 등록 또는 뱃지 신청에 실패함")
        cur.execute("""
            SELECT has_priority_badge, premium_expires_at::text AS expires_at
            FROM operator_buildings WHERE operator_id=%s AND master_building_id=%s
        """, (operator_id, operator_bld))
        row = cur.fetchone() or {}
        if not row.get("has_priority_badge") or not (row.get("expires_at") or "").startswith(deadline):
            failures.append("파트너 뱃지: 운영업체 단지뱃지의 고정 만료일이 저장되지 않음")
        with client.session_transaction() as sess:
            sess.clear()
            sess["operator_id"] = operator_second_id
        client.post("/api/operator/buildings", json={"master_building_id": operator_bld})
        operator_second = client.post(f"/api/operator/buildings/{operator_bld}/claim-premium")
        with client.session_transaction() as sess:
            sess.clear()
            sess["operator_id"] = operator_third_id
        client.post("/api/operator/buildings", json={"master_building_id": operator_bld})
        operator_third = client.post(f"/api/operator/buildings/{operator_bld}/claim-premium")
        if operator_second.status_code != 200 or operator_third.status_code != 400:
            failures.append("파트너 뱃지: 운영업체 업종별 두 번째 단지뱃지를 허용하거나 세 번째를 차단하지 못함")

        # 세 운영업체가 동시에 신청해도 정확히 두 곳만 단지뱃지를 받는다.
        operator_race_ids = [create_operator(f"operator-race-{index}") for index in range(3)]
        operator_race_bld = building_ids[16]
        conn.commit()
        for operator_race_id in operator_race_ids:
            with client.session_transaction() as sess:
                sess.clear()
                sess["operator_id"] = operator_race_id
            added = client.post("/api/operator/buildings", json={"master_building_id": operator_race_bld})
            if added.status_code != 200:
                failures.append("파트너 뱃지: 운영업체 독점 동시성 테스트용 단지 등록에 실패함")

        def parallel_operator_claim(operator_race_id):
            with app.test_client() as parallel_client:
                with parallel_client.session_transaction() as sess:
                    sess["operator_id"] = operator_race_id
                return parallel_client.post(
                    f"/api/operator/buildings/{operator_race_bld}/claim-premium"
                ).status_code

        with ThreadPoolExecutor(max_workers=3) as executor:
            operator_race_results = list(executor.map(
                parallel_operator_claim, operator_race_ids
            ))
        if sorted(operator_race_results) != [200, 200, 400]:
            failures.append("파트너 뱃지: 운영업체 같은 업종 동시 신청이 두 자리 정원을 우회함")

        # ③ 대출상담사 담당단지: 등록 즉시 골드뱃지와 고정 만료일.
        loan_id = create_loan("loan")
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["loan_consultant_id"] = loan_id
        loan_add = client.post("/api/loan-consultant/buildings", json={"master_building_id": loan_bld})
        if loan_add.status_code != 200 or not (loan_add.get_json() or {}).get("ok"):
            failures.append(f"파트너 뱃지: 대출상담사 담당단지 등록 실패 ({loan_add.get_json()})")
        cur.execute("""
            SELECT has_priority_badge, premium_expires_at::text AS expires_at
            FROM loan_consultant_buildings WHERE loan_consultant_id=%s AND master_building_id=%s
        """, (loan_id, loan_bld))
        row = cur.fetchone() or {}
        if not row.get("has_priority_badge") or not (row.get("expires_at") or "").startswith(deadline):
            failures.append("파트너 뱃지: 대출상담사 담당단지에 즉시 골드뱃지/고정 만료일이 저장되지 않음")
        loan_second_id = create_loan("loan-second")
        loan_third_id = create_loan("loan-third")
        conn.commit()
        loan_slot_results = []
        for loan_slot_id in (loan_second_id, loan_third_id):
            with client.session_transaction() as sess:
                sess.clear()
                sess["loan_consultant_id"] = loan_slot_id
            loan_slot_results.append(client.post(
                "/api/loan-consultant/buildings", json={"master_building_id": loan_bld}
            ).status_code)
        if loan_slot_results != [200, 400]:
            failures.append("파트너 뱃지: 대출상담사 두 번째 단지뱃지를 허용하거나 세 번째를 차단하지 못함")
        loan_region_only_id = create_loan("loan-region-only")
        cur.execute("""
            INSERT INTO loan_consultant_service_areas (loan_consultant_id, region_name)
            VALUES (%s, '전국'), (%s, '전국')
        """, (loan_id, loan_region_only_id))
        conn.commit()
        loan_detail = client.get(f"/api/building/{loan_bld}")
        loan_detail_rows = (loan_detail.get_json() or {}).get("loan_consultants", [])
        loan_detail_ids = [item.get("id") for item in loan_detail_rows]
        if (
            len(loan_detail_ids) != 3
            or set(loan_detail_ids[:2]) != {loan_id, loan_second_id}
            or len(set(loan_detail_ids)) != 3
            or not all(item.get("registered") for item in loan_detail_rows[:2])
            or loan_detail_rows[2].get("registered")
        ):
            failures.append("파트너 뱃지: 대출상담사 단지뱃지 우선 노출·지역 보충·중복 제외가 잘못됨")

        # 운영업체와 대출상담사도 10개를 넘겨 등록할 수 없다.
        for building_id in building_ids[17:26]:
            cur.execute(
                "INSERT INTO operator_buildings (operator_id, master_building_id) VALUES (%s, %s)",
                (operator_id, building_id),
            )
        for building_id in building_ids[27:36]:
            cur.execute(
                "INSERT INTO loan_consultant_buildings (loan_consultant_id, master_building_id) VALUES (%s, %s)",
                (loan_id, building_id),
            )
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["operator_id"] = operator_id
        operator_over_cap = client.post("/api/operator/buildings", json={"master_building_id": building_ids[26]})
        with client.session_transaction() as sess:
            sess.clear()
            sess["loan_consultant_id"] = loan_id
        loan_over_cap = client.post("/api/loan-consultant/buildings", json={"master_building_id": building_ids[36]})
        if operator_over_cap.status_code != 400 or loan_over_cap.status_code != 400:
            failures.append("파트너 뱃지: 운영업체 또는 대출상담사 담당단지 10개 한도가 적용되지 않음")

        # ④ 중개사 지역뱃지: 시군구별 다섯 명까지 허용하고 초과 신청은 대기 등록한다.
        region_sgg = region_fixture["sgg_text"]
        region_agent_ids = [create_agent(f"agent-region-{index}") for index in range(11)]
        agent_region_id = region_agent_ids[0]
        conn.commit()
        region_responses = []
        for region_agent_id in region_agent_ids:
            with client.session_transaction() as sess:
                sess.clear()
                sess["agent_id"] = region_agent_id
            region_responses.append(client.post("/api/agent/service-regions", json={
                "sgg_text": region_sgg,
            }))
        if (
            [response.status_code for response in region_responses] != ([200] * 11)
            or sum(bool((response.get_json() or {}).get("waitlisted")) for response in region_responses) != 6
        ):
            failures.append("파트너 뱃지: 시군구 지역뱃지 5명 정원과 대기 등록이 적용되지 않음")
        cur.execute("""
            SELECT umd_nm, expires_at::text AS expires_at FROM agent_service_regions
            WHERE agent_id=%s AND sgg_text=%s
        """, (agent_region_id, region_sgg))
        region_row = cur.fetchone() or {}
        if region_row.get("umd_nm") is not None or not (region_row.get("expires_at") or "").startswith(deadline):
            failures.append("파트너 뱃지: 중개사 지역뱃지가 시군구 행 또는 고정 만료일로 저장되지 않음")
        region_detail = client.get(f"/api/building/{region_fixture['id']}")
        region_detail_agent_ids = {
            item.get("id") for item in (region_detail.get_json() or {}).get("agents", [])
        }
        if not region_detail_agent_ids.intersection(region_agent_ids[:5]):
            failures.append("파트너 뱃지: 중개사 지역뱃지가 담당단지 선택 없이 같은 시군구 건물에 자동 노출되지 않음")
        cur.execute("""
            UPDATE agent_service_regions
            SET expires_at=NOW() - INTERVAL '1 second'
            WHERE agent_id=%s AND sgg_text=%s
        """, (region_agent_ids[4], region_sgg))
        conn.commit()
        if app_module._process_badge_waitlist_notifications() < 1:
            failures.append("파트너 뱃지: 지역 빈자리 발생 뒤 자동 대기 알림을 처리하지 못함")
        cur.execute("""
            SELECT agent_id, notified_at IS NOT NULL AS notified
            FROM region_badge_waitlist
            WHERE sgg_text=%s AND agent_id=ANY(%s)
        """, (region_sgg, region_agent_ids[5:7]))
        first_region_cycle = {row["agent_id"]: row["notified"] for row in cur.fetchall()}
        if first_region_cycle != {region_agent_ids[5]: True, region_agent_ids[6]: False}:
            failures.append("파트너 뱃지: 지역 빈자리에는 가장 먼저 등록한 대기자 한 명만 알리지 못함")
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = region_agent_ids[5]
        region_claim = client.post("/api/agent/service-regions", json={"sgg_text": region_sgg})
        if region_claim.status_code != 200 or (region_claim.get_json() or {}).get("waitlisted"):
            failures.append("파트너 뱃지: 지역 빈자리 알림을 받은 대기자가 지역뱃지를 신청하지 못함")
        cur.execute("""
            SELECT notified_at FROM region_badge_waitlist
            WHERE agent_id=%s AND sgg_text=%s
        """, (region_agent_ids[6], region_sgg))
        next_region_waiter = cur.fetchone()
        if not next_region_waiter or next_region_waiter["notified_at"] is not None:
            failures.append("파트너 뱃지: 지역 입점 뒤 다음 대기자를 재대기 상태로 유지하지 못함")
        cur.execute("""
            UPDATE agent_service_regions
            SET expires_at=NOW() - INTERVAL '1 second'
            WHERE agent_id=%s AND sgg_text=%s
        """, (region_agent_ids[5], region_sgg))
        conn.commit()
        if app_module._process_badge_waitlist_notifications() < 1:
            failures.append("파트너 뱃지: 다음 지역 빈자리에서 남은 대기자에게 다시 알리지 못함")
        cur.execute("""
            SELECT notified_at IS NOT NULL AS notified
            FROM region_badge_waitlist
            WHERE agent_id=%s AND sgg_text=%s
        """, (region_agent_ids[6], region_sgg))
        next_region_waiter = cur.fetchone()
        if not next_region_waiter or not next_region_waiter["notified"]:
            failures.append("파트너 뱃지: 두 번째 지역 빈자리 알림 상태를 저장하지 못함")

        # 임의/XSS 형태 지역 문자열은 마스터 조합 검증으로 거절한다.
        xss_region = f"{region_sgg}'><img src=x onerror=alert(1)>"
        xss_agent_id = create_agent("agent-xss-region")
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = xss_agent_id
        xss_region_add = client.post("/api/agent/service-regions", json={
            "sgg_text": xss_region,
        })
        if xss_region_add.status_code != 400:
            failures.append("파트너 뱃지: 마스터에 없는 임의/XSS 시군구 문자열을 허용함")

        # 담당단지 선택도 신청한 시군구 안의 건물만 허용한다.
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = agent_region_id
        region_building_add = client.post("/api/agent/region-buildings", json={
            "master_building_id": region_fixture["id"],
        })
        cur.execute("""
            SELECT id FROM master_buildings
            WHERE sgg_text <> %s
            ORDER BY id LIMIT 1
        """, (region_sgg,))
        outside_region = cur.fetchone()
        outside_add = client.post("/api/agent/region-buildings", json={
            "master_building_id": outside_region["id"],
        }) if outside_region else None
        if region_building_add.status_code != 200 or not outside_add or outside_add.status_code != 400:
            failures.append("파트너 뱃지: 담당단지를 신청한 시군구 안으로 제한하지 못함")

        # 서로 다른 담당단지를 동시에 추가해도 중개사별 10개 제한을 넘지 않는다.
        cur.execute("DELETE FROM agent_region_buildings WHERE agent_id=%s", (agent_region_id,))
        cur.execute("""
            SELECT id FROM master_buildings
            WHERE sgg_text=%s AND building_name <> '-'
            ORDER BY id LIMIT 11
        """, (region_sgg,))
        region_capacity_buildings = [row["id"] for row in cur.fetchall()]
        conn.commit()

        def parallel_region_building_add(building_id):
            with app.test_client() as parallel_client:
                with parallel_client.session_transaction() as sess:
                    sess["agent_id"] = agent_region_id
                return parallel_client.post("/api/agent/region-buildings", json={
                    "master_building_id": building_id,
                }).status_code

        with ThreadPoolExecutor(max_workers=11) as executor:
            region_building_results = list(executor.map(
                parallel_region_building_add, region_capacity_buildings
            ))
        cur.execute(
            "SELECT COUNT(*) AS count FROM agent_region_buildings WHERE agent_id=%s",
            (agent_region_id,),
        )
        if (
            len(region_capacity_buildings) != 11
            or region_building_results.count(200) != 10
            or region_building_results.count(400) != 1
            or cur.fetchone()["count"] != 10
        ):
            failures.append(
                f"파트너 뱃지: 지역 담당단지 동시 추가가 10개 제한을 우회함 ({region_building_results})"
            )

        # 열한 명이 같은 시군구에 동시에 신청해도 정확히 다섯 명만 활성화되고 나머지는 대기한다.
        region_race_agent_ids = [create_agent(f"agent-region-race-{index}") for index in range(11)]
        conn.commit()

        def parallel_region_claim(race_agent_id):
            with app.test_client() as parallel_client:
                with parallel_client.session_transaction() as sess:
                    sess["agent_id"] = race_agent_id
                response = parallel_client.post("/api/agent/service-regions", json={
                    "sgg_text": region_race_fixture["sgg_text"],
                })
                return response.status_code, bool((response.get_json() or {}).get("waitlisted"))

        with ThreadPoolExecutor(max_workers=11) as executor:
            region_race_results = list(executor.map(
                parallel_region_claim, region_race_agent_ids
            ))
        cur.execute("""
            SELECT COUNT(*) AS count FROM agent_service_regions
            WHERE sgg_text=%s AND expires_at > NOW()
        """, (region_race_fixture["sgg_text"],))
        if (
            [item[0] for item in region_race_results].count(200) != 11
            or [item[1] for item in region_race_results].count(True) != 6
            or cur.fetchone()["count"] != 5
        ):
            failures.append(f"파트너 뱃지: 중개사 지역뱃지 동시 신청이 5명 정원을 우회함 ({region_race_results})")

        # 동 단위 레거시 행은 시군구 한 행으로 병합되고 담당단지는 보존되며 재실행은 무해하다.
        legacy_agent_id = create_agent("agent-region-legacy")
        legacy_sgg = f"{tag}-legacy"
        cur.execute("ALTER TABLE agent_service_regions DROP CONSTRAINT agent_service_regions_unique")
        cur.execute("""
            INSERT INTO agent_service_regions (agent_id, sgg_text, umd_nm, expires_at)
            VALUES (%s, %s, '이충동', NOW() + INTERVAL '30 days'),
                   (%s, %s, '고덕동', NOW() + INTERVAL '20 days')
        """, (legacy_agent_id, legacy_sgg, legacy_agent_id, legacy_sgg))
        cur.execute("""
            INSERT INTO agent_region_buildings (agent_id, master_building_id)
            VALUES (%s, %s)
        """, (legacy_agent_id, region_fixture["id"]))
        cur.execute("DELETE FROM app_meta WHERE key=%s", (db_module._SGG_REGION_MIGRATION_KEY,))
        merged_count, normalized_count = db_module._migrate_agent_regions_to_sgg(cur)
        cur.execute("""
            SELECT COUNT(*) AS count, BOOL_AND(umd_nm IS NULL) AS all_null
            FROM agent_service_regions WHERE agent_id=%s AND sgg_text=%s
        """, (legacy_agent_id, legacy_sgg))
        legacy_row = cur.fetchone() or {}
        cur.execute("SELECT COUNT(*) AS count FROM agent_region_buildings WHERE agent_id=%s",
                    (legacy_agent_id,))
        first_remaining = cur.fetchone()["count"]
        rerun_counts = db_module._migrate_agent_regions_to_sgg(cur)
        cur.execute("SELECT COUNT(*) AS count FROM agent_region_buildings WHERE agent_id=%s",
                    (legacy_agent_id,))
        second_remaining = cur.fetchone()["count"]
        cur.execute("""
            ALTER TABLE agent_service_regions
            ADD CONSTRAINT agent_service_regions_unique UNIQUE (agent_id, sgg_text)
        """)
        conn.commit()
        if (
            merged_count != 1 or normalized_count != 1
            or legacy_row.get("count") != 1 or not legacy_row.get("all_null")
            or first_remaining != 1 or rerun_counts != (0, 0) or second_remaining != 1
        ):
            failures.append("파트너 뱃지: 동 단위 지역 병합·담당단지 보존·재실행이 안전하지 않음")

        # ⑤·⑥ 운영업체 지역뱃지: 시군구+업종별 두 곳까지 허용하고 건물에 자동 노출한다.
        operator_region_ids = [
            create_operator(f"operator-region-{index}", category="세탁") for index in range(3)
        ]
        operator_region_id = operator_region_ids[0]
        conn.commit()
        operator_region_results = []
        for region_operator_id in operator_region_ids:
            with client.session_transaction() as sess:
                sess.clear()
                sess["operator_id"] = region_operator_id
            operator_region_results.append(client.post(
                "/api/operator/service-regions", json={"sgg_text": region_sgg}
            ))
        if [response.status_code for response in operator_region_results] != [200, 200, 400]:
            failures.append("파트너 뱃지: 운영업체 시군구+업종별 두 곳 정원이 적용되지 않음")
        cur.execute("""
            SELECT expires_at::text AS expires_at FROM operator_service_regions
            WHERE operator_id=%s AND sgg_text=%s
        """, (operator_region_id, region_sgg))
        if not ((cur.fetchone() or {}).get("expires_at") or "").startswith(deadline):
            failures.append("파트너 뱃지: 운영업체 지역뱃지의 고정 만료일이 저장되지 않음")
        operator_region_detail = client.get(f"/api/building/{region_fixture['id']}")
        operator_card = next((
            item for item in (operator_region_detail.get_json() or {}).get("operator_by_category", [])
            if item.get("category") == "세탁"
        ), {})
        if (
            operator_card.get("tier") != "region"
            or operator_card.get("company_name")
            not in {f"{tag}-operator-region-0", f"{tag}-operator-region-1"}
        ):
            failures.append("파트너 뱃지: 운영업체 지역뱃지가 같은 시군구 건물에 자동 노출되지 않음")
        with client.session_transaction() as sess:
            sess.clear()
            sess["operator_id"] = operator_region_id
        operator_region_mine = client.get("/api/operator/service-regions")
        mine_regions = (operator_region_mine.get_json() or {}).get("regions", [])
        if (
            operator_region_mine.status_code != 200
            or not mine_regions
            or mine_regions[0].get("sgg_text") != region_sgg
        ):
            failures.append("파트너 뱃지: 운영업체 대시보드가 현재 시군구 지역뱃지를 조회하지 못함")
        routed_operator_id, routed_operator_reason = app_module._route_operator_lead(
            cur, region_fixture["id"], "세탁"
        )
        if (
            routed_operator_id not in set(operator_region_ids[:2])
            or routed_operator_reason != "region"
        ):
            failures.append("파트너 뱃지: 시군구 지역뱃지 운영업체가 같은 지역 상담에 배정되지 않음")
        with open("static/js/main.js", encoding="utf-8") as main_js_file:
            main_js_source = main_js_file.read()
        with open("static/operator_dashboard.html", encoding="utf-8") as operator_dashboard_file:
            operator_dashboard_source = operator_dashboard_file.read()
        if (
            'const items = all.filter(it => it && it.company_name);' not in main_js_source
            or 'data-operator-category="${escapeHtml(categoryLabel)}"' not in main_js_source
            or 'it.category === "위탁운영" ? "위탁"' not in main_js_source
            or 'all.filter(it => it.category === "위탁운영"' in main_js_source
        ):
            failures.append("파트너 뱃지: 건물 화면이 비위탁 운영업체 지역뱃지 또는 위탁 표준 라벨을 렌더링하지 않음")
        if (
            'fetch("/api/operator/service-regions")' not in operator_dashboard_source
            or 'body: JSON.stringify({ sgg_text: sgg })' not in operator_dashboard_source
            or 'fetch("/api/regions")' not in operator_dashboard_source
            or "/api/operator/service-areas" in operator_dashboard_source
        ):
            failures.append("파트너 뱃지: 운영업체 대시보드의 시군구 선택·등록·조회 흐름이 새 지역뱃지 API와 연결되지 않음")

        # 신규 승인에서는 세 파트너 유형 모두 slug를 비워 두고, 중개사/대출 희망단지는 자동 뱃지 처리한다.
        def create_application(applicant_type, label, preferred_building_id=None, category=None):
            cur.execute("""
                INSERT INTO applications
                    (applicant_type, office_or_company_name, owner_name, reg_number, biz_reg_number,
                     category, phone, email, preferred_building_id, password_hash)
                VALUES (%s, %s, '승인테스트', %s, %s, %s, '01000000000', %s, %s, %s)
                RETURNING id
            """, (applicant_type, f"{tag}-{label}", f"{tag}-{label}-license", f"{tag}-{label}-biz",
                  category, f"{tag}-{label}-approval@example.test", preferred_building_id,
                  generate_password_hash("partner-badge-test-password")))
            application_id = cur.fetchone()["id"]
            application_ids.append(application_id)
            return application_id

        approval_agent = create_application("agent", "approval-agent", agent_approval_bld)
        approval_operator = create_application("operator", "approval-operator", category="청소")
        approval_loan = create_application("loan_consultant", "approval-loan", loan_approval_bld)
        approval_loan_full = create_application("loan_consultant", "approval-loan-full", loan_bld)
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["admin"] = True
        with (
            patch.object(app_module, "send_sms", return_value=(True, "test")),
            patch.object(app_module, "_send_approval_email", return_value=(True, "test")),
        ):
            approval_responses = [
                client.post(f"/api/admin/applications/{application_id}/approve")
                for application_id in (
                    approval_agent, approval_operator, approval_loan, approval_loan_full
                )
            ]
        if any(response.status_code != 200 or not (response.get_json() or {}).get("ok")
               for response in approval_responses):
            failures.append("파트너 뱃지: 테스트 파트너 승인에 실패함")
        cur.execute("""
            SELECT a.subdomain_slug, ab.has_priority_badge, ab.premium_expires_at::text AS expires_at
            FROM applications ap
            JOIN agents a ON a.id=ap.linked_agent_id
            LEFT JOIN agent_buildings ab ON ab.agent_id=a.id AND ab.master_building_id=%s
            WHERE ap.id=%s
        """, (agent_approval_bld, approval_agent))
        row = cur.fetchone() or {}
        if row.get("subdomain_slug") is not None or not row.get("has_priority_badge") or not (row.get("expires_at") or "").startswith(deadline):
            failures.append("파트너 뱃지: 신규 중개사 승인 slug 또는 희망단지 자동 뱃지 정책이 잘못됨")
        cur.execute("""
            SELECT o.subdomain_slug FROM applications ap
            JOIN operators o ON o.id=ap.linked_operator_id WHERE ap.id=%s
        """, (approval_operator,))
        if (cur.fetchone() or {}).get("subdomain_slug") is not None:
            failures.append("파트너 뱃지: 신규 운영업체 승인에서 slug가 자동 발급됨")
        cur.execute("""
            SELECT lc.subdomain_slug, lcb.has_priority_badge, lcb.premium_expires_at::text AS expires_at
            FROM applications ap
            JOIN loan_consultants lc ON lc.id=ap.linked_loan_consultant_id
            LEFT JOIN loan_consultant_buildings lcb
              ON lcb.loan_consultant_id=lc.id AND lcb.master_building_id=%s
            WHERE ap.id=%s
        """, (loan_approval_bld, approval_loan))
        row = cur.fetchone() or {}
        if row.get("subdomain_slug") is not None or not row.get("has_priority_badge") or not (row.get("expires_at") or "").startswith(deadline):
            failures.append("파트너 뱃지: 신규 대출상담사 승인 slug 또는 희망단지 자동 뱃지 정책이 잘못됨")
        cur.execute("""
            SELECT lc.id AS loan_consultant_id, lcb.id AS building_link_id
            FROM applications ap
            JOIN loan_consultants lc ON lc.id=ap.linked_loan_consultant_id
            LEFT JOIN loan_consultant_buildings lcb
              ON lcb.loan_consultant_id=lc.id AND lcb.master_building_id=%s
            WHERE ap.id=%s
        """, (loan_bld, approval_loan_full))
        full_approval_row = cur.fetchone() or {}
        if not full_approval_row.get("loan_consultant_id") or full_approval_row.get("building_link_id"):
            failures.append("파트너 뱃지: 대출상담사 승인이 찬 단지의 세 번째 뱃지를 자동 배정함")

        # 관리자 목록은 세 유형의 단지·지역 뱃지를 모두 반환하며, 화면은 만기연장만 노출한다.
        premium_status = client.get("/api/admin/premium-status")
        items = (premium_status.get_json() or {}).get("items") or []
        expected = {
            ("agent", "building", agent_id), ("operator", "building", operator_id),
            ("loan_consultant", "building", loan_id), ("agent", "region", agent_region_id),
            ("operator", "region", operator_region_id),
        }
        actual = {(item.get("partner_type"), item.get("kind"), item.get("partner_id")) for item in items}
        if premium_status.status_code != 200 or not expected <= actual:
            failures.append("파트너 뱃지: 관리자 현황 목록에 세 파트너 유형의 등록 건이 모두 표시되지 않음")
        if not any(
            item.get("partner_id") == agent_region_id
            and item.get("target") == region_sgg
            and item.get("umd_nm") is None
            for item in items
        ):
            failures.append("파트너 뱃지: 관리자 현황 API의 중개사 지역뱃지가 시군구 기준이 아님")
        building_status = client.get("/api/admin/premium-status?kind=building")
        region_status = client.get("/api/admin/premium-status?kind=region")
        if (
            any(item.get("kind") != "building" for item in (building_status.get_json() or {}).get("items", []))
            or any(item.get("kind") != "region" for item in (region_status.get_json() or {}).get("items", []))
        ):
            failures.append("파트너 뱃지: 관리자 전체/단지/지역 탭용 API 필터가 잘못됨")
        legacy_email = f"{tag}-agent-region-legacy@example.test"
        member_listing = client.get("/api/admin/members", query_string={
            "group": "agent", "q": legacy_email,
        })
        member_rows = (member_listing.get_json() or {}).get("items") or []
        legacy_member = next((item for item in member_rows if item.get("id") == legacy_agent_id), None)
        member_detail = client.get(f"/api/admin/members/agent/{legacy_agent_id}/detail")
        detail_regions = ((member_detail.get_json() or {}).get("data") or {}).get("service_regions") or []
        if (
            not legacy_member
            or legacy_member.get("region_sgg") != legacy_sgg
            or legacy_member.get("region_umd") is not None
            or not detail_regions
            or detail_regions[0].get("umd_nm") is not None
        ):
            failures.append("파트너 뱃지: 관리자 회원관리에서 병합된 시군구 지역뱃지를 표시하지 않음")
        pending_listing = client.get("/api/admin/members", query_string={"group": "pending"})
        pending_payload = pending_listing.get_json() or {}
        pending_type_counts = pending_payload.get("pending_type_counts")
        if (
            pending_listing.status_code != 200
            or not isinstance(pending_type_counts, dict)
            or set(pending_type_counts) != {"agent", "operator", "loan_consultant"}
            or any(not isinstance(v, int) or v < 0 for v in pending_type_counts.values())
        ):
            failures.append("회원관리 승인대기 유형별 건수 필드가 없거나 형식이 잘못됨")
        with open("static/admin.html", encoding="utf-8") as admin_file:
            admin_source = admin_file.read()
        if (
            'onclick="extendPremiumBadge(' in admin_source
            or 'class="admin-btn premium-extend-btn" data-premium-index="${index}"' not in admin_source
            or "window.extendPremiumBadge = async function(row)" not in admin_source
            or "dgEscape(r.sgg_text || \"-\")" not in admin_source
            or "<th>읍·면·동</th>" in admin_source
            or "pending_type_counts" not in admin_source
            or "중개사 ${Number(pc.agent) || 0}" not in admin_source
        ):
            failures.append("파트너 뱃지: 관리자 만기연장 버튼이 외부 입력을 인라인 스크립트로 렌더링함")
        # 만기연장은 현재 유료 분류를 건드리지 않고, 활성 독점 뱃지가 있으면 재활성화를 거절한다.
        cur.execute(
            "UPDATE agent_buildings SET is_paid=TRUE WHERE agent_id=%s AND master_building_id=%s",
            (agent_id, agent_bld),
        )
        conn.commit()
        extension = client.post("/api/admin/premium-status/extend", json={
            "partner_type": "agent", "partner_id": agent_id, "kind": "building",
            "master_building_id": agent_bld,
        })
        cur.execute(
            "SELECT is_paid FROM agent_buildings WHERE agent_id=%s AND master_building_id=%s",
            (agent_id, agent_bld),
        )
        if extension.status_code != 200 or not (cur.fetchone() or {}).get("is_paid"):
            failures.append("파트너 뱃지: 관리자 만기연장이 기존 유료 분류를 보존하지 않음")
        expired_agent_id = create_agent("agent-expired")
        cur.execute("""
            INSERT INTO agent_buildings
                (agent_id, master_building_id, has_priority_badge, premium_granted_at, premium_expires_at)
            VALUES (%s, %s, TRUE, NOW() - INTERVAL '1 year', NOW() - INTERVAL '1 day')
        """, (expired_agent_id, agent_bld))
        cur.execute("""
            UPDATE operator_buildings
            SET has_priority_badge=TRUE, premium_granted_at=NOW() - INTERVAL '1 year',
                premium_expires_at=NOW() - INTERVAL '1 day'
            WHERE operator_id=%s AND master_building_id=%s
        """, (operator_third_id, operator_bld))
        conn.commit()
        agent_extension_conflict = client.post("/api/admin/premium-status/extend", json={
            "partner_type": "agent", "partner_id": expired_agent_id, "kind": "building",
            "master_building_id": agent_bld,
        })
        operator_extension_conflict = client.post("/api/admin/premium-status/extend", json={
            "partner_type": "operator", "partner_id": operator_third_id, "kind": "building",
            "master_building_id": operator_bld,
        })
        if agent_extension_conflict.status_code != 400 or operator_extension_conflict.status_code != 400:
            failures.append("파트너 뱃지: 관리자 만기연장이 이미 찬 단지 정원과 충돌하는 뱃지를 활성화함")
        admin_page = client.get("/admin")
        markup = admin_page.get_data(as_text=True)
        if (
            "모든 뱃지는 신청 즉시 자동 부여되며, 2026.12.31까지 무료입니다." not in markup
            or 'data-kind="all">전체<' not in markup
            or 'data-kind="building">단지뱃지<' not in markup
            or 'data-kind="region">지역뱃지<' not in markup
            or "<th>읍·면·동</th>" in markup
            or "만기연장" not in markup
            or "유료전환" in markup
        ):
            failures.append("파트너 뱃지: 관리자 안내 문구 또는 만기연장 전용 UI가 잘못됨")
        with open("static/agent_dashboard.html", encoding="utf-8") as dashboard_file:
            agent_dashboard_source = dashboard_file.read()
        with open("static/operator_dashboard.html", encoding="utf-8") as dashboard_file:
            operator_dashboard_source = dashboard_file.read()
        if (
            "시·군·구별 선착순 <b>5명</b>" not in agent_dashboard_source
            or "모든 건물에 자동으로 실버뱃지가 노출" not in agent_dashboard_source
            or 'id="regionUmdInput"' in agent_dashboard_source
            or '"umd_nm": umd' in agent_dashboard_source
            or "업종별 선착순 <b>2개</b>" not in operator_dashboard_source
            or "같은 업종 2곳까지 가능" not in operator_dashboard_source
            or "최상단에 단독 노출" in operator_dashboard_source
        ):
            failures.append("파트너 뱃지: 중개사·운영업체 대시보드 안내가 새 자동 매칭·정원 정책과 다름")
    except Exception as exc:
        conn.rollback()
        failures.append(f"파트너 뱃지 정책 테스트 오류: {exc}")
    finally:
        try:
            if application_ids:
                cur.execute("DELETE FROM applications WHERE id = ANY(%s)", (application_ids,))
            if agent_ids:
                cur.execute("DELETE FROM agents WHERE id = ANY(%s)", (agent_ids,))
            if operator_ids:
                cur.execute("DELETE FROM operators WHERE id = ANY(%s)", (operator_ids,))
            if loan_ids:
                cur.execute("DELETE FROM loan_consultants WHERE id = ANY(%s)", (loan_ids,))
            conn.commit()
            with client.session_transaction() as sess:
                sess.clear()
        except Exception:
            conn.rollback()
        finally:
            email_patcher.stop()
            cur.close()
            conn.close()
    if not failures:
        print("OK  파트너 자동 매칭·유형별 두 자리 정원·고정 만료·신규 slug 미발급")
    return failures


def _check_admin_building_broker_details(client):
    """브로커 표준데이터 상세 API·목록 우선수·상권정보 폴백을 임시 행으로 확인한다."""
    failures = []
    run_id = str(time.time_ns())
    token = run_id[-8:]
    matched_building_id = fallback_building_id = None
    broker_numbers = []
    conn = get_conn()
    cur = conn.cursor()
    try:
        with client.session_transaction() as sess:
            sess.clear()
        blocked = client.get("/api/v1/r/4c2/999999999")
        if blocked.status_code != 401:
            failures.append("건물 브로커 상세 API가 비관리자 요청을 차단하지 않음")

        matched_name = f"브로커상세매칭 {token}"
        fallback_name = f"브로커상세폴백 {token}"
        road = f"테스트특별시 브로커검증구 상세로 {token[-3:]}"
        jibun = f"테스트특별시 브로커검증동 {token[-3:]}-1번지"
        fallback_road = f"테스트특별시 브로커검증구 폴백로 {token[-3:]}"
        fallback_jibun = f"테스트특별시 브로커검증동 {token[-3:]}-2번지"
        cur.execute("""
            INSERT INTO master_buildings
                (building_name, road_address, jibun_address, sgg_text, umd_nm, jibun, source, lat, lng)
            VALUES (%s, %s, %s, '테스트특별시 브로커검증구', '브로커검증동', %s, 'api_test', 37.5, 127.0)
            RETURNING id
        """, (matched_name, road, jibun, f"{token[-3:]}-1"))
        matched_building_id = cur.fetchone()["id"]
        cur.execute("""
            INSERT INTO master_buildings
                (building_name, road_address, jibun_address, sgg_text, umd_nm, jibun, source)
            VALUES (%s, %s, %s, '테스트특별시 브로커검증구', '브로커검증동', %s, 'api_test')
            RETURNING id
        """, (fallback_name, fallback_road, fallback_jibun, f"{token[-3:]}-2"))
        fallback_building_id = cur.fetchone()["id"]

        # 동일 주소의 2건은 반환하고, 지번만 일치하는 1건은 도로명 우선 규칙상 제외한다.
        broker_specs = [
            (
                f"TEST-BROKER-{run_id}-1", '<script>alert(1)</script> 표준중개',
                road, jibun, "https://example.test/broker?a=1&b=2", "영업중",
            ),
            (
                f"TEST-BROKER-{run_id}-2", "두번째 표준중개",
                road, f"테스트특별시 다른동 {token[-3:]}-7번지", "https://example.test/second", "휴업",
            ),
            (
                f"TEST-BROKER-{run_id}-J", "지번전용 중개",
                f"테스트특별시 다른구 무관로 {token[-3:]}", jibun, "javascript:alert(1)", None,
            ),
        ]
        for index, (reg_number, office_name, broker_road, broker_jibun, homepage_url, biz_status) in enumerate(broker_specs):
            broker_numbers.append(reg_number)
            cur.execute("""
                INSERT INTO broker_registry
                    (office_name, reg_number, owner_name, phone, road_address, jibun_address,
                     phone_numbers, member_count, reg_date, homepage_url, source_updated_at,
                     lat, lng, road_norm, jibun_norm, biz_status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, '2026-01-02', %s, '2026-08-27',
                        %s, %s, %s, %s, %s)
            """, (
                office_name, reg_number, f"대표자{index + 1}", f"021234560{index}",
                broker_road, broker_jibun,
                [f"021234560{index}", f"021234990{index}"] if index == 0 else [f"021234560{index}"],
                3 if index == 0 else 1,
                homepage_url, 37.5 + index / 100, 127.0 + index / 100,
                addr_norm.normalize_road_prefix(broker_road),
                addr_norm.normalize_jibun_prefix(broker_jibun),
                biz_status,
            ))
        cur.execute("""
            INSERT INTO building_stores (master_building_id, store_name, category, floor, ho_no)
            VALUES (%s, '<img src=x onerror=alert(1)> 캐시부동산', '부동산', '3', '301')
        """, (fallback_building_id,))
        conn.commit()

        with client.session_transaction() as sess:
            sess["admin"] = True

        detail = client.get(f"/api/v1/r/4c2/{matched_building_id}")
        payload = detail.get_json() or {}
        expected_fields = {
            "office_name", "reg_number", "owner_name", "phone", "road_address", "jibun_address",
            "phone_numbers", "member_count", "reg_date", "homepage_url",
            "source_updated_at", "lat", "lng", "biz_status",
        }
        if (
            detail.status_code != 200 or payload.get("ok") is not True
            or payload.get("count") != 2 or len(payload.get("items") or []) != 2
            or any(not expected_fields <= set(item) for item in payload.get("items") or [])
        ):
            failures.append("건물 브로커 상세 API가 표준데이터 전체 필드 또는 도로명 우선 매칭을 지키지 않음")
        elif not any(item.get("office_name", "").startswith("<script>") for item in payload["items"]):
            failures.append("건물 브로커 상세 API가 원본 사무소명 데이터를 반환하지 않음")
        else:
            multi_phone = next(
                (item for item in payload["items"] if item.get("reg_number") == broker_numbers[0]),
                {},
            )
            if multi_phone.get("phone_numbers") != ["02-1234-5600", "02-1234-9900"]:
                failures.append("건물 브로커 상세 API가 여러 전화번호를 모두 포맷해 반환하지 않음")
            if multi_phone.get("member_count") != 3:
                failures.append("건물 브로커 상세 API가 중개업소 소속인원 수를 반환하지 않음")

        candidates = client.get(
            f"/api/admin/broker-candidates?building_id={matched_building_id}&radius_km=5"
        )
        candidate_items = (candidates.get_json() or {}).get("items") or []
        candidate_statuses = {
            item.get("reg_number"): item.get("biz_status")
            for item in candidate_items
            if item.get("reg_number") in set(broker_numbers[:2])
        }
        if candidate_statuses != {
            broker_numbers[0]: "영업중",
            broker_numbers[1]: "휴업",
        }:
            failures.append("인근 중개사 후보 API가 저장된 정상·휴업 영업상태를 반환하지 않음")

        fallback_detail = client.get(f"/api/v1/r/4c2/{fallback_building_id}")
        if fallback_detail.status_code != 200 or (fallback_detail.get_json() or {}).get("count") != 0:
            failures.append("브로커 표준데이터 없는 건물이 빈 상세 목록을 반환하지 않음")
        missing = client.get("/api/v1/r/4c2/999999999")
        if missing.status_code != 404:
            failures.append("건물 브로커 상세 API가 없는 건물을 404로 처리하지 않음")

        listing = client.get(f"/api/admin/buildings?q={matched_name}&size=10")
        list_items = (listing.get_json() or {}).get("items") or []
        matched_row = next((row for row in list_items if row.get("id") == matched_building_id), None)
        if (
            listing.status_code != 200 or not matched_row
            or matched_row.get("broker_realty_count") != 2
            or matched_row.get("store_realty_count") != 2
            or matched_row.get("store_realty_source") != "broker_registry"
        ):
            failures.append("건물 목록이 브로커 표준데이터 수를 입점부동산 우선값으로 표시하지 않음")
        fallback_listing = client.get(f"/api/admin/buildings?q={fallback_name}&size=10")
        fallback_items = (fallback_listing.get_json() or {}).get("items") or []
        fallback_row = next((row for row in fallback_items if row.get("id") == fallback_building_id), None)
        if (
            fallback_listing.status_code != 200 or not fallback_row
            or fallback_row.get("broker_realty_count") != 0
            or fallback_row.get("store_realty_count") != 1
            or fallback_row.get("store_realty_source") != "상권정보"
        ):
            failures.append("건물 목록이 브로커 미매칭 시 상권정보 부동산 수로 폴백하지 않음")
        total_realty = ((listing.get_json() or {}).get("totals") or {}).get("total_store_realty")
        fallback_total_realty = ((fallback_listing.get_json() or {}).get("totals") or {}).get("total_store_realty")
        if total_realty != 2 or fallback_total_realty != 1:
            failures.append("건물 목록 합계가 브로커 표준데이터 우선·상권정보 폴백 수와 일치하지 않음")
        from io import BytesIO
        from openpyxl import load_workbook
        export_response = client.get(
            f"/api/admin/buildings/export.xlsx?ids={matched_building_id},{fallback_building_id}"
        )
        if export_response.status_code != 200:
            failures.append("건물 엑셀 내보내기가 브로커 표준데이터 표본에서 실패함")
        else:
            sheet = load_workbook(BytesIO(export_response.data), data_only=True).active
            headers = [cell.value for cell in sheet[1]]
            name_col = headers.index("건물명")
            broker_name_col = headers.index("입점부동산_업체명")
            count_col = headers.index("입점부동산수")
            exported = {}
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[name_col] in {matched_name, fallback_name}:
                    exported.setdefault(row[name_col], []).append(
                        (row[broker_name_col], row[count_col])
                    )
            matched_export = exported.get(matched_name, [])
            fallback_export = exported.get(fallback_name, [])
            matched_exported_names = {name for name, _count in matched_export if name}
            if (
                not {"<script>alert(1)</script> 표준중개", "두번째 표준중개"} <= matched_exported_names
                or 2 not in {count for _name, count in matched_export}
                or fallback_export != [("<img src=x onerror=alert(1)> 캐시부동산", 1)]
            ):
                failures.append("건물 엑셀이 브로커 표준데이터 우선·상권정보 폴백 입점부동산을 일치하게 내보내지 않음")

        with open(os.path.join(os.path.dirname(__file__), "..", "static", "admin.html"), encoding="utf-8") as fh:
            markup = fh.read()
        required_markup = (
            "renderBrokerRegistryCards", "brokerStatusBadge", "주소 매칭 결과",
            "입점업소 참고", "bld-realty-expand", "bld-realty-sub-row", "bld-realty-table",
            "사무소명</th><th>등록번호</th><th>대표자", "dgEscape(value || \"-\")",
        )
        disallowed_markup = (
            "safeBrokerHomepage", "openRealtyModal", "bld-realty-modal-btn", "bld-realty-modal-card",
        )
        buildings_start = markup.index("buildings:")
        realty_col = markup.index('{ key: "store_realty_count"', buildings_start)
        lodging_col = markup.index('{ key: "lodging_expand"', buildings_start)
        report_col = markup.index('{ key: "report_rate"', buildings_start)
        if (
            any(text not in markup for text in required_markup)
            or any(text in markup for text in disallowed_markup)
            or not (realty_col < lodging_col < report_col)
        ):
            failures.append("관리자 브로커 인라인 펼치기·안전한 문자열 렌더링·컬럼 순서가 올바르지 않음")
        else:
            start = markup.index("function brokerStatusBadge")
            end = markup.index("function renderCachedRealtyTable")
            frontend_check = """
                function dgEscape(v) {
                  if (v === null || v === undefined) return "";
                  return String(v).replace(/&/g, "&amp;").replace(/</g, "&lt;")
                    .replace(/>/g, "&gt;").replace(/"/g, "&quot;").replace(/'/g, "&#39;");
                }
            """ + markup[start:end] + """
                const html = renderBrokerRegistryCards([{
                  office_name: "<img src=x onerror=alert(1)>", reg_number: "R", owner_name: "O",
                  phone: "P", phone_numbers: ["02-419-9600", "<img src=x onerror=alert(2)>"],
                  member_count: 2, road_address: "A", jibun_address: "J", reg_date: "D",
                  homepage_url: "javascript:alert(1)", source_updated_at: "U", lat: 1, lng: 2,
                  biz_status: "영업중"
                }, {
                  office_name: "비영업 중개", reg_number: "R2", biz_status: "휴업"
                }, {
                  office_name: "상태 미확인 중개", reg_number: "R3", biz_status: null
                }]);
                if (html.includes("<img") || html.includes("javascript:alert(1)")) process.exit(1);
                if (!html.includes("&lt;img")) process.exit(2);
                if (!html.includes("02-419-9600<br>&lt;img") || !html.includes("2명")) process.exit(3);
                if (!html.includes("bld-broker-status is-active") || !html.includes("bld-broker-status is-inactive")) process.exit(4);
                if (!html.includes("사무소명</th><th>등록번호</th><th>대표자")) process.exit(4);
                if ((html.match(/<th>/g) || []).length !== 8) process.exit(5);
                if (!html.includes("상태 정보 없음")) process.exit(6);
            """
            rendered = subprocess.run(
                ["node", "-e", frontend_check], capture_output=True, text=True, timeout=10
            )
            if rendered.returncode:
                failures.append("관리자 브로커 인라인 표가 악성 문자열·상태 미확인을 안전하게 표시하지 않음")
    except Exception as exc:
        conn.rollback()
        failures.append(f"관리자 건물 브로커 상세 테스트 오류: {exc}")
    finally:
        try:
            if broker_numbers:
                cur.execute("DELETE FROM broker_registry WHERE reg_number = ANY(%s)", (broker_numbers,))
            if matched_building_id or fallback_building_id:
                cur.execute(
                    "DELETE FROM master_buildings WHERE id = ANY(%s)",
                    ([building_id for building_id in (matched_building_id, fallback_building_id) if building_id],),
                )
            conn.commit()
            with client.session_transaction() as sess:
                sess.clear()
        except Exception:
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    if not failures:
        print("OK  관리자 건물 브로커 인라인 표·상태 뱃지·목록 우선수·상권정보 폴백")
    return failures


def _check_broker_sync_normalization_and_status(client):
    """실제 API 구조와 라군 주소 표기 편차, 상태 기본값/명시값 저장을 고정한다."""
    failures = []
    run_id = str(time.time_ns())
    reg_number = f"TEST-BROKER-SYNC-{run_id}"
    conn = get_conn()
    cur = conn.cursor()
    lagoon_building = "경기도 안산시 단원구 엠티브이17로 35 (성곡동)"
    lagoon_broker = "경기도 안산시 단원구 엠티브이17로 35, 성곡동 118호 (성곡동)"
    try:
        # 2026-08-27에 확인한 실 API 형식(최상위 header/body)과 레거시 wrapper를 모두 허용한다.
        current_payload = {
            "header": {"resultCode": "00", "resultMsg": "NORMAL SERVICE."},
            "body": {"items": {"item": [{"medOfficeNm": "형식검증중개"}]}, "totalCount": 1},
        }
        legacy_payload = {
            "response": {"header": {"resultCode": "00"}, "body": {"items": [{"medOfficeNm": "레거시중개"}], "totalCount": 1}},
        }
        current_items, current_total = sync_brokers._parse_page_payload(current_payload)
        legacy_items, legacy_total = sync_brokers._parse_page_payload(legacy_payload)
        if current_total != 1 or current_items[0].get("medOfficeNm") != "형식검증중개" or legacy_total != 1 or legacy_items[0].get("medOfficeNm") != "레거시중개":
            failures.append("브로커 수집기가 현재/레거시 공공데이터 응답 구조를 처리하지 못함")
        expected_phone_fields = {
            "telno", "phoneNumber", "telNo", "tel",
            "cnpsTelno", "reprsvTelno", "bsshTelno", "phoneNo",
        }
        if not expected_phone_fields.issubset(sync_brokers.FIELD_CANDIDATES["phone"]):
            failures.append("브로커 전화번호 응답 필드 후보가 누락됨")

        building_key = addr_norm.normalize_road_prefix(lagoon_building)
        broker_key = addr_norm.normalize_road_prefix(lagoon_broker)
        if building_key != "경기도안산시단원구엠티브이17로35" or broker_key != building_key:
            failures.append("라군 주소의 괄호 동·호수 표기 차이가 같은 도로명 매칭 키로 정규화되지 않음")

        cur.execute("""
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = 'broker_registry' AND column_name = 'biz_status'
        """)
        if not cur.fetchone():
            failures.append("broker_registry에 영업상태(biz_status) 컬럼이 없음")
            return failures

        source_item = {
            "medOfficeNm": "수집상태검증중개", "estblRegNo": reg_number,
            "lctnRoadNmAddr": lagoon_broker, "lctnLotnoAddr": "경기도 안산시 단원구 성곡동 999-1",
            "telno": "0212345678", "estblRegYmd": "2026-01-02", "rprsvNm": "검증대표",
            "crtrYmd": "2026-08-27",
        }
        if not sync_brokers._upsert(cur, source_item):
            failures.append("상태 없는 브로커 원본 행을 저장하지 못함")
        conn.commit()
        cur.execute("SELECT road_norm, biz_status FROM broker_registry WHERE reg_number=%s", (reg_number,))
        saved = cur.fetchone() or {}
        if saved.get("road_norm") != building_key or saved.get("biz_status") is not None:
            failures.append("상태 없는 현재 표준데이터가 상태 NULL/정규화 키로 저장되지 않음")

        source_item["bizStatus"] = "휴업"
        sync_brokers._upsert(cur, source_item)
        conn.commit()
        cur.execute("SELECT biz_status FROM broker_registry WHERE reg_number=%s", (reg_number,))
        if (cur.fetchone() or {}).get("biz_status") != "휴업":
            failures.append("원본 상태 필드가 제공될 때 영업상태를 갱신하지 못함")

        with client.session_transaction() as sess:
            sess.clear()
            sess["admin"] = True
        cur.execute("SELECT id FROM master_buildings WHERE building_name = %s", ("라군 센트럴 스테이",))
        lagoon = cur.fetchone()
        if not lagoon:
            failures.append("라군 센트럴 스테이 실매칭 검증용 건물을 찾지 못함")
        else:
            response = client.get(f"/api/v1/r/4c2/{lagoon['id']}")
            items = (response.get_json() or {}).get("items") or []
            matched = next((item for item in items if "라군부동산중개법인" in (item.get("office_name") or "")), None)
            if response.status_code != 200 or not matched:
                failures.append("라군 센트럴 스테이와 라군부동산중개법인이 실제 도로명 매칭되지 않음")
            elif matched.get("biz_status") not in (None, "영업중", "휴업", "업무정지", "휴업연장"):
                failures.append("라군부동산중개법인의 영업상태가 허용된 원본 상태값이 아님")
            listing = client.get("/api/admin/buildings", query_string={
                "q": "라군 센트럴 스테이",
                "size": 10,
            })
            lagoon_row = next(
                (
                    item for item in ((listing.get_json() or {}).get("items") or [])
                    if item.get("id") == lagoon["id"]
                ),
                None,
            )
            if (
                listing.status_code != 200
                or not lagoon_row
                or int(lagoon_row.get("broker_realty_count") or 0) < 1
                or lagoon_row.get("store_realty_source") != "broker_registry"
            ):
                failures.append("라군 센트럴 스테이 관리자 목록이 브로커 매칭 뱃지를 표시하지 않음")
    except Exception as exc:
        conn.rollback()
        failures.append(f"브로커 수집/라군 매칭 검증 오류: {exc}")
    finally:
        try:
            cur.execute("DELETE FROM broker_registry WHERE reg_number=%s", (reg_number,))
            conn.commit()
            with client.session_transaction() as sess:
                sess.clear()
        except Exception:
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    if not failures:
        print("OK  브로커 현재 API 형식·라군 주소 매칭·영업상태 저장")
    return failures


def _check_member_login_history(client):
    """이메일 로그인 이력 저장과 관리자 회원관리 조회 계약을 확인한다."""
    from werkzeug.security import generate_password_hash

    failures = []
    run_id = str(int(time.time() * 1000))
    email = f"login-history-{run_id}@example.test"
    user_id = None
    signup_user_id = None
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO users (email, password_hash, name, provider, status)
            VALUES (%s, %s, %s, 'email', 'active')
            RETURNING id
        """, [email, generate_password_hash("login-history-password"), "접속이력 테스트"])
        user_id = cur.fetchone()["id"]
        conn.commit()

        headers = {
            "User-Agent": "LivingstayLoginHistoryTest Chrome/125.0",
            "X-Forwarded-For": f"198.51.100.{(int(run_id) % 200) + 1}",
        }
        with client.session_transaction() as sess:
            sess.clear()
        blocked = client.get(f"/api/admin/members/general/{user_id}/login-history")
        if blocked.status_code != 401:
            failures.append("회원 접속이력: 비관리자 조회를 차단하지 않습니다.")
        login = client.post("/api/auth/login", json={
            "email": email,
            "password": "login-history-password",
        }, headers=headers)
        if login.status_code != 200 or not (login.get_json() or {}).get("ok"):
            failures.append(f"회원 접속이력: 이메일 로그인에 실패했습니다. ({login.get_json()})")
            return failures

        with client.session_transaction() as sess:
            sess.clear()
            sess["admin"] = True
        listing = client.get("/api/admin/members", query_string={"group": "general", "q": email})
        payload = listing.get_json() or {}
        rows = payload.get("items") or []
        member = next((row for row in rows if row.get("id") == user_id), None)
        if (
            listing.status_code != 200
            or not member
            or member.get("provider") != "email"
            or not member.get("created_at")
            or ":" not in member["created_at"]
        ):
            failures.append("회원관리: 가입일시 또는 가입경로(provider)를 반환하지 않습니다.")

        history = client.get(f"/api/admin/members/general/{user_id}/login-history")
        history_data = history.get_json() or {}
        items = history_data.get("history") or []
        if (
            history.status_code != 200
            or not history_data.get("last_login_at")
            or not items
            or "LivingstayLoginHistoryTest" not in (items[0].get("user_agent") or "")
        ):
            failures.append("회원 접속이력: 로그인 저장 또는 최근 10건 조회가 올바르지 않습니다.")
        non_general = client.get(f"/api/admin/members/agent/{user_id}/login-history")
        if non_general.status_code != 400:
            failures.append("회원 접속이력: 일반회원 외 유형 조회를 차단하지 않습니다.")

        signup_email = f"signup-history-{run_id}@example.test"
        with client.session_transaction() as sess:
            sess.clear()
        signup = client.post("/api/auth/signup", json={
            "email": signup_email,
            "password": "signup-history-password",
            "name": "가입 접속이력 테스트",
            "age14": True,
            "terms": True,
            "privacy": True,
            "marketing": False,
        }, headers={
            "User-Agent": "LivingstaySignupHistoryTest Safari/17.0",
            "X-Forwarded-For": f"198.51.101.{(int(run_id) % 200) + 1}",
        })
        if signup.status_code != 200 or not (signup.get_json() or {}).get("ok"):
            failures.append(f"회원 접속이력: 신규가입에 실패했습니다. ({signup.get_json()})")
        else:
            cur.execute("SELECT id FROM users WHERE email = %s", [signup_email])
            signup_user_id = (cur.fetchone() or {}).get("id")
            with client.session_transaction() as sess:
                sess.clear()
                sess["admin"] = True
            first_session = client.get(
                f"/api/admin/members/general/{signup_user_id}/login-history"
            )
            first_session_data = first_session.get_json() or {}
            first_items = first_session_data.get("history") or []
            if (
                first_session.status_code != 200
                or not first_items
                or "LivingstaySignupHistoryTest" not in (first_items[0].get("user_agent") or "")
            ):
                failures.append("회원 접속이력: 신규가입 직후 첫 인증 세션이 저장되지 않았습니다.")
    except Exception as exc:
        conn.rollback()
        failures.append(f"회원 접속이력 API 테스트 오류: {exc}")
    finally:
        try:
            if user_id:
                cur.execute("DELETE FROM login_history WHERE user_id = %s", [user_id])
                cur.execute("DELETE FROM users WHERE id = %s", [user_id])
            if signup_user_id:
                cur.execute("DELETE FROM login_history WHERE user_id = %s", [signup_user_id])
                cur.execute("DELETE FROM users WHERE id = %s", [signup_user_id])
                conn.commit()
            with client.session_transaction() as sess:
                sess.clear()
        except Exception:
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    if not failures:
        print("OK  회원관리 가입경로·가입일시·최근 접속이력")
    return failures


def _check_favorite_save_persistence(client):
    """아이리스모텔 관심저장 → 목록 재조회 → 삭제가 실제 DB에서 정상 동작하는지 확인."""
    import time as _time
    from db import get_conn

    failures = []
    run_id = str(int(_time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    user_id = None
    try:
        cur.execute("""
            SELECT id, building_name, road_address
            FROM master_buildings
            WHERE building_name = %s AND NULLIF(road_address, '') IS NOT NULL
            ORDER BY id
            LIMIT 1
        """, ("아이리스모텔",))
        iris = cur.fetchone()
        if not iris:
            return ["favorites persistence: 아이리스모텔 테스트 건물을 찾지 못했습니다."]

        cur.execute(
            "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
            (f"favorite-{run_id}@example.test", "관심저장 테스트"),
        )
        user_id = cur.fetchone()["id"]
        conn.commit()

        with client.session_transaction() as sess:
            sess["user_id"] = user_id

        payload = {
            "building_name": iris["building_name"],
            "address": iris["road_address"],
            "building_id": iris["id"],
        }
        saved = client.post("/api/favorites/mine", json=payload)
        saved_data = saved.get_json() or {}
        if saved.status_code != 200 or not saved_data.get("ok"):
            failures.append(
                "favorites persistence: 아이리스모텔 저장 실패 "
                f"(HTTP {saved.status_code}, {saved_data})"
            )
            return failures

        refreshed = client.get("/api/favorites/mine")
        refreshed_data = refreshed.get_json() or {}
        stored = next((
            item for item in refreshed_data.get("items", [])
            if item.get("building_name") == iris["building_name"]
            and item.get("address") == iris["road_address"]
        ), None)
        if refreshed.status_code != 200 or not refreshed_data.get("ok") or not stored:
            failures.append("favorites persistence: 새로고침 뒤 아이리스모텔 관심저장이 유지되지 않습니다.")
            return failures
        if stored.get("building_id") != iris["id"]:
            failures.append("favorites persistence: 아이리스모텔의 building_id가 저장 시점 값과 다릅니다.")
            return failures

        removed = client.delete("/api/favorites/mine", json=payload)
        removed_data = removed.get_json() or {}
        if removed.status_code != 200 or not removed_data.get("ok"):
            failures.append("favorites persistence: 아이리스모텔 관심저장 삭제에 실패했습니다.")
        else:
            print("OK  아이리스모텔 관심저장·새로고침 유지·삭제")
    except Exception as exc:
        failures.append(f"favorites persistence 테스트 오류: {exc}")
    finally:
        try:
            if user_id:
                cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
                conn.commit()
            with client.session_transaction() as sess:
                sess.pop("user_id", None)
        finally:
            cur.close()
            conn.close()
    return failures


def _check_urgent_listing_tiers_and_alerts(client):
    """급매 등급·관심단지 전용 토글·회원/매물별 중복 방지를 확인한다."""
    from app import (
        _apply_urgent_tier,
        _queue_urgent_listing_alerts,
        _send_urgent_listing_email,
        _urgent_tier_for_values,
        _whole_listing_values,
    )

    failures = []
    run_id = str(int(time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    user_id = building_id = listing_id = scope_listing_id = None
    try:
        if (
            _urgent_tier_for_values("매매", False, 900, 1000) != "urgent"
            or _urgent_tier_for_values("매매", True, 900, None) != "urgent"
            or _urgent_tier_for_values("매매", False, 900, None) is not None
            or _urgent_tier_for_values("매매", True, 1100, 1000) != "urgent"
            or _urgent_tier_for_values("통임대", True, 900, None) is not None
        ):
            failures.append("urgent listing: 사용자 체크·최신 실거래 비교 단일 급매 규칙이 잘못됨")
            return failures
        _, invalid_urgent_error = _whole_listing_values({
            "transaction_target": "whole", "deal_type": "매매",
            "disclosure_scope": "public", "is_urgent": "false",
        })
        if not invalid_urgent_error:
            failures.append("urgent listing: 문자열 false가 급매로 저장될 수 있음")
            return failures
        unit_values, unit_values_error = _whole_listing_values({
            "transaction_target": "unit", "deal_type": "매매", "is_urgent": True,
        })
        unit_public = _apply_urgent_tier({
            "deal_mode": "direct", "transaction_target": "unit",
            "deal_type": "매매", "is_urgent": True, "price_krw": 900,
            "latest_transaction_price": None,
        })
        if (
            unit_values_error
            or not unit_values.get("is_urgent")
            or unit_public.get("urgent_tier") != "urgent"
        ):
            failures.append("urgent listing: 개별호실 매매의 급매 저장·공개 판정이 누락됨")
            return failures

        cur.execute("""
            INSERT INTO master_buildings
                (building_name, road_address, sgg_text, umd_nm, lodging_type)
            VALUES (%s, %s, %s, %s, '생활')
            RETURNING id
        """, (f"급매 테스트 {run_id}", f"급매 테스트로 {run_id}", "급매테스트시", "급매동"))
        building_id = cur.fetchone()["id"]
        cur.execute("""
            INSERT INTO users (email, name, phone, phone_verified)
            VALUES (%s, '급매 알림 테스트', '01000000000', TRUE)
            RETURNING id
        """, (f"urgent-{run_id}@example.test",))
        user_id = cur.fetchone()["id"]
        cur.execute("""
            INSERT INTO user_favorites (user_id, building_name, address, master_building_id)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, [user_id, f"급매 테스트 {run_id}", f"급매 테스트로 {run_id}", building_id])
        favorite_id = cur.fetchone()["id"]
        conn.commit()

        with client.session_transaction() as sess:
            sess.clear()
            sess["user_id"] = user_id
        toggled = client.put("/api/favorites/mine/urgent-alert", json={
            "favorite_id": favorite_id, "building_id": building_id,
        })
        cur.execute(
            "SELECT urgent_alert_enabled FROM user_favorites WHERE id=%s",
            [favorite_id],
        )
        if (
            toggled.status_code != 200
            or not (toggled.get_json() or {}).get("ok")
            or not (cur.fetchone() or {}).get("urgent_alert_enabled")
        ):
            failures.append("urgent listing: 관심단지와 독립된 급매알림 토글이 저장되지 않음")
            return failures

        cur.execute("""
            INSERT INTO listing_requests
                (user_id, master_building_id, deal_type, contact_phone, deal_mode,
                 transaction_target, disclosure_scope, price_krw, is_urgent)
            VALUES (%s, %s, '매매', '01000000000', 'direct', 'whole', 'public', 900, TRUE)
            RETURNING id
        """, [user_id, building_id])
        listing_id = cur.fetchone()["id"]
        jobs = _queue_urgent_listing_alerts(
            cur, listing_id, building_id, f"급매 테스트 {run_id}",
            f"급매 테스트로 {run_id}", 900, "urgent",
        )
        duplicate_jobs = _queue_urgent_listing_alerts(
            cur, listing_id, building_id, f"급매 테스트 {run_id}",
            f"급매 테스트로 {run_id}", 900, "urgent",
        )
        conn.commit()
        urgent_only_response = client.get("/api/listings?urgent_only=1&limit=50")
        urgent_only_items = (urgent_only_response.get_json() or {}).get("items") or []
        if (
            urgent_only_response.status_code != 200
            or not any(item.get("id") == listing_id for item in urgent_only_items)
        ):
            failures.append("urgent listing: 급매 전용 목록 필터에서 급매 매물이 누락됨")
            return failures
        cur.execute(
            "SELECT COUNT(*) AS count FROM notifications WHERE user_id=%s AND listing_request_id=%s",
            [user_id, listing_id],
        )
        notification_count = int(cur.fetchone()["count"])
        if len(jobs) != 1 or duplicate_jobs or notification_count != 1:
            failures.append("urgent listing: 같은 회원·매물의 인앱 급매알림이 중복 생성됨")
            return failures

        with patch.object(app_module, "send_email", return_value=(False, "planned failure")) as email_mock:
            _send_urgent_listing_email(jobs[0])
        cur.execute("""
            SELECT email_state FROM urgent_listing_alert_logs
             WHERE user_id=%s AND listing_request_id=%s
        """, [user_id, listing_id])
        email_state = (cur.fetchone() or {}).get("email_state")
        if email_mock.call_count != 1 or email_state != "failed" or notification_count != 1:
            failures.append("urgent listing: 이메일 실패가 인앱 알림을 되돌리거나 발송 이력을 남기지 않음")

        cur.execute("""
            INSERT INTO listing_requests
                (user_id, master_building_id, deal_type, contact_phone, deal_mode,
                 transaction_target, disclosure_scope, price_krw, is_urgent)
            VALUES (%s, %s, '매매', '01000000000', 'direct', 'whole', 'limited', 800, TRUE)
            RETURNING id
        """, [user_id, building_id])
        scope_listing_id = cur.fetchone()["id"]
        conn.commit()
        with patch.object(app_module, "send_email", return_value=(True, "sent")):
            made_public = client.patch(
                f"/api/listing-requests/{scope_listing_id}/disclosure-scope",
                json={"disclosure_scope": "public"},
            )
        cur.execute(
            "SELECT COUNT(*) AS count FROM notifications WHERE user_id=%s AND listing_request_id=%s",
            [user_id, scope_listing_id],
        )
        if (
            made_public.status_code != 200
            or not (made_public.get_json() or {}).get("ok")
            or int(cur.fetchone()["count"]) != 1
        ):
            failures.append("urgent listing: 제한공개에서 전체공개 전환 시 급매알림이 생성되지 않음")

        turned_off = client.delete("/api/favorites/mine/urgent-alert", json={"favorite_id": favorite_id})
        if turned_off.status_code != 200 or not (turned_off.get_json() or {}).get("ok"):
            failures.append("urgent listing: 급매알림 끄기 API가 실패함")

        signal_columns = (
            "urgent_alert_enabled",
            "new_listing_alert_enabled",
            "permit_change_alert_enabled",
            "favorite_increase_alert_enabled",
            "nearby_change_alert_enabled",
        )
        signal_on = client.put("/api/favorites/mine/signal-alert", json={
            "favorite_id": favorite_id, "building_id": building_id,
        })
        cur.execute(
            "SELECT " + ", ".join(signal_columns) + " FROM user_favorites WHERE id=%s",
            [favorite_id],
        )
        signal_flags_on = cur.fetchone() or {}
        signal_items = (client.get("/api/favorites/mine").get_json() or {}).get("items", [])
        signal_item = next((item for item in signal_items if item.get("favorite_id") == favorite_id), {})
        if (
            signal_on.status_code != 200
            or not (signal_on.get_json() or {}).get("ok")
            or not all(signal_flags_on.get(column) for column in signal_columns)
            or not all(signal_item.get(column) for column in signal_columns)
        ):
            failures.append("숙박알리미: 통합 켜기 시 5개 알림 플래그 저장·조회가 일치하지 않음")

        signal_off = client.delete("/api/favorites/mine/signal-alert", json={
            "favorite_id": favorite_id,
        })
        cur.execute(
            "SELECT " + ", ".join(signal_columns) + " FROM user_favorites WHERE id=%s",
            [favorite_id],
        )
        signal_flags_off = cur.fetchone() or {}
        if (
            signal_off.status_code != 200
            or not (signal_off.get_json() or {}).get("ok")
            or any(signal_flags_off.get(column) for column in signal_columns)
        ):
            failures.append("숙박알리미: 통합 끄기 시 5개 알림 플래그가 모두 해제되지 않음")
    except Exception as exc:
        conn.rollback()
        failures.append(f"urgent listing 테스트 오류: {exc}")
    finally:
        try:
            if listing_id:
                cur.execute(
                    "DELETE FROM listing_request_history WHERE listing_request_id = ANY(%s)",
                    [[value for value in (listing_id, scope_listing_id) if value]],
                )
                cur.execute(
                    "DELETE FROM listing_requests WHERE id = ANY(%s)",
                    [[value for value in (listing_id, scope_listing_id) if value]],
                )
            if user_id:
                cur.execute("DELETE FROM users WHERE id=%s", [user_id])
            if building_id:
                cur.execute("DELETE FROM master_buildings WHERE id=%s", [building_id])
            conn.commit()
            with client.session_transaction() as sess:
                sess.pop("user_id", None)
        except Exception:
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    if not failures:
        print("OK  급매 등급·관심단지 토글·회원/매물 중복 알림 방지")
    return failures


def _check_weekly_email_auto_optin_apis(client):
    """네 저장 API의 자동 opt-in과 명시적 수신거부 보존을 실제 DB로 검증한다."""
    from db import get_conn

    failures = []
    run_id = str(int(time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    user_id = None
    try:
        cur.execute("""
            SELECT id, building_name, road_address
              FROM master_buildings
             WHERE NULLIF(road_address, '') IS NOT NULL
             ORDER BY id
             LIMIT 1
        """)
        building = cur.fetchone()
        if not building:
            return ["weekly auto opt-in: 테스트용 마스터 건물이 없습니다."]

        cur.execute("""
            INSERT INTO users
                (email, name, phone, phone_verified, weekly_email_enabled, updated_weekly_email_at)
            VALUES (%s, %s, %s, TRUE, FALSE, NULL)
            RETURNING id
        """, (f"weekly-optin-{run_id}@example.test", "주간자동구독 테스트", "01012345678"))
        user_id = cur.fetchone()["id"]
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["user_id"] = user_id

        def reset_auto_eligible():
            cur.execute("""
                UPDATE users
                   SET weekly_email_enabled = FALSE, updated_weekly_email_at = NULL
                 WHERE id = %s
            """, (user_id,))
            conn.commit()

        def assert_opted_in(label):
            cur.execute("""
                SELECT weekly_email_enabled, updated_weekly_email_at
                  FROM users WHERE id = %s
            """, (user_id,))
            row = cur.fetchone() or {}
            if not row.get("weekly_email_enabled") or row.get("updated_weekly_email_at") is None:
                failures.append(f"weekly auto opt-in: {label} 성공 뒤 구독이 켜지지 않았습니다.")

        favorite_payload = {
            "building_name": building["building_name"],
            "address": building["road_address"],
            "building_id": building["id"],
        }
        test_headers = {
            "X-Forwarded-For": f"198.51.100.{(int(run_id) % 200) + 1}",
        }
        reset_auto_eligible()
        favorite = client.post("/api/favorites/mine", json=favorite_payload, headers=test_headers)
        if favorite.status_code != 200 or not (favorite.get_json() or {}).get("ok"):
            failures.append("weekly auto opt-in: 관심단지 저장 API가 실패했습니다.")
        else:
            assert_opted_in("관심단지 저장")

        reset_auto_eligible()
        alert = client.post("/api/alerts/mine", json=favorite_payload, headers=test_headers)
        if alert.status_code != 200 or not (alert.get_json() or {}).get("ok"):
            failures.append("weekly auto opt-in: 실거래 알림 구독 API가 실패했습니다.")
        else:
            assert_opted_in("실거래 알림 구독")

        reset_auto_eligible()
        listing = client.post("/api/listing-requests", json={
            "master_building_id": building["id"],
            "deal_type": "매매",
            "deal_mode": "direct",
            "registrant_type": "owner",
        }, headers=test_headers)
        if listing.status_code != 200 or not (listing.get_json() or {}).get("ok"):
            failures.append(f"weekly auto opt-in: 매물의뢰 API가 실패했습니다. ({listing.get_json()})")
        else:
            assert_opted_in("매물의뢰")

        reset_auto_eligible()
        unverified_buy = client.post("/api/buy-requests", json={
            "master_building_id": building["id"],
            "deal_type": "매매",
            "contact_phone": "010-9999-9999",
        }, headers=test_headers)
        if unverified_buy.status_code != 400:
            failures.append("매수의뢰 SMS 인증: 인증된 전화번호가 아닌 요청을 차단하지 않았습니다.")
        with patch("app.send_sms", return_value=(False, "test")):
            buy = client.post("/api/buy-requests", json={
                "master_building_id": building["id"],
                "deal_type": "매매",
                "contact_phone": "010-1234-5678",
            }, headers=test_headers)
        if buy.status_code != 200 or not (buy.get_json() or {}).get("ok"):
            failures.append(f"weekly auto opt-in: 매수의뢰 API가 실패했습니다. ({buy.get_json()})")
        else:
            assert_opted_in("매수의뢰")
            buy_id = (buy.get_json() or {}).get("id")
            withdrawn = client.delete(f"/api/buy-requests/{buy_id}", headers=test_headers)
            if withdrawn.status_code != 200 or not (withdrawn.get_json() or {}).get("ok"):
                failures.append(f"매수의뢰 철회 API가 실패했습니다. ({withdrawn.get_json()})")
            else:
                cur.execute("SELECT status FROM buy_requests WHERE id = %s", (buy_id,))
                if (cur.fetchone() or {}).get("status") != "철회됨":
                    failures.append("매수의뢰 철회 API가 상태를 '철회됨'으로 저장하지 않았습니다.")
                cur.execute("SELECT id FROM agents ORDER BY id LIMIT 1")
                agent = cur.fetchone()
                if agent:
                    cur.execute("UPDATE buy_requests SET routed_agent_id = %s WHERE id = %s",
                                (agent["id"], buy_id))
                    conn.commit()
                    with client.session_transaction() as sess:
                        sess.clear()
                        sess["agent_id"] = agent["id"]
                    reactivated = client.put(
                        f"/api/agent/buy-requests/{buy_id}/status",
                        json={"status": "in_progress"},
                        headers=test_headers,
                    )
                    if reactivated.status_code != 400:
                        failures.append("매수의뢰 철회: 담당 중개사가 철회된 의뢰를 다시 처리중으로 바꿀 수 있습니다.")
                    with client.session_transaction() as sess:
                        sess.clear()
                        sess["user_id"] = user_id

        # 사용자가 이미 꺼둔 뒤에는 저장 API가 구독을 되살리지 않아야 한다.
        cur.execute("""
            UPDATE users
               SET weekly_email_enabled = FALSE, updated_weekly_email_at = NOW()
             WHERE id = %s
        """, (user_id,))
        conn.commit()
        client.post("/api/favorites/mine", json=favorite_payload, headers=test_headers)
        cur.execute("SELECT weekly_email_enabled FROM users WHERE id = %s", (user_id,))
        if (cur.fetchone() or {}).get("weekly_email_enabled") is not False:
            failures.append("weekly auto opt-in: 명시적 수신거부가 관심단지 저장으로 되살아났습니다.")
        if not failures:
            print("OK  주간 이메일 자동 opt-in 4개 API·명시적 수신거부 보존")
    except Exception as exc:
        conn.rollback()
        failures.append(f"weekly auto opt-in API 테스트 오류: {exc}")
    finally:
        try:
            if user_id:
                cur.execute("DELETE FROM listing_request_history WHERE listing_request_id IN "
                            "(SELECT id FROM listing_requests WHERE user_id = %s)", (user_id,))
                cur.execute("DELETE FROM listing_requests WHERE user_id = %s", (user_id,))
                cur.execute("DELETE FROM buy_requests WHERE user_id = %s", (user_id,))
                cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
                conn.commit()
            with client.session_transaction() as sess:
                sess.pop("user_id", None)
        except Exception:
            conn.rollback()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_datalab_report_source_contract():
    """영업신고현황 주소 매칭·폐업 제외·전국 신고번호 중복 제거를 검증."""
    from app import _report_rate_by_sido_payload

    failures = []
    conn = get_conn()
    cur = conn.cursor()
    prefix = f"API신고집계검증{int(time.time() * 1000)}"
    permit_numbers = [
        f"{prefix}-ROAD",
        f"{prefix}-JIBUN",
        f"{prefix}-CLOSED",
        f"{prefix}-DUPLICATE",
        f"{prefix}-PAUSED",
        f"{prefix}-ROAD-WINS",
    ]
    building_ids = []
    try:
        before = _report_rate_by_sido_payload()["total"]
        road_primary = f"서울특별시 신고집계구 {prefix}로 101"
        jibun_primary = f"서울특별시 신고집계동 101"
        road_jibun_fallback = f"서울특별시 신고집계구 {prefix}다른로 102"
        jibun_fallback = f"서울특별시 신고집계동 102"
        road_closed = f"부산광역시 신고집계구 {prefix}로 103"
        road_duplicate = f"서울특별시 신고집계구 {prefix}로 104"
        jibun_duplicate = f"부산광역시 신고집계동 105"
        building_specs = [
            ("서울특별시 신고집계구", road_primary, jibun_primary, 100),
            ("서울특별시 신고집계구", road_jibun_fallback, jibun_fallback, 50),
            ("부산광역시 신고집계구", road_closed, None, 30),
            ("서울특별시 신고집계구", road_duplicate, None, 10),
            ("부산광역시 신고집계구", f"부산광역시 신고집계구 {prefix}다른로 105", jibun_duplicate, 10),
        ]
        for index, (sgg_text, road_address, jibun_address, units) in enumerate(building_specs):
            cur.execute("""
                INSERT INTO master_buildings
                    (building_name, sgg_text, road_address, jibun_address,
                     lodging_type, units, source)
                VALUES (%s, %s, %s, %s, '생활', %s, 'api_test')
                RETURNING id
            """, (
                f"{prefix} 건물 {index + 1}",
                sgg_text,
                road_address,
                jibun_address,
                units,
            ))
            building_ids.append(cur.fetchone()["id"])

        registry_specs = [
            # 도로명에 신고가 있으면 같은 건물의 지번 신고보다 도로명 결과를 쓴다.
            (permit_numbers[0], road_primary, None, "영업/정상", 20),
            # road_primary의 지번도 맞지만, 도로명 결과가 있으므로 합산하지 않는다.
            (permit_numbers[5], f"서울특별시 신고집계구 {prefix}무관로 998", jibun_primary, "영업/정상", 11),
            # 도로명 결과가 없을 때만 지번 보조 매칭한다.
            (permit_numbers[1], f"서울특별시 신고집계구 {prefix}무관로 999", jibun_fallback, "영업/정상", 25),
            # 폐업은 주소가 맞아도 영업신고현황에 포함하지 않는다.
            (permit_numbers[2], road_closed, None, "폐업", 90),
            # 하나의 신고번호가 두 건물 키에 걸려도 전국 합계에는 한 번만 포함한다.
            (permit_numbers[3], road_duplicate, jibun_duplicate, "영업/정상", 7),
            # 활성 기준은 정확히 영업/정상이며 휴업은 제외한다.
            (permit_numbers[4], road_primary, None, "휴업", 9),
        ]
        for permit_number, road_address, jibun_address, status, room_count in registry_specs:
            cur.execute("""
                INSERT INTO lodging_registry
                    (biz_name, permit_number, road_address, jibun_address,
                     biz_status_name, room_count, road_norm, jibun_norm)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                f"{prefix} 신고업체",
                permit_number,
                road_address,
                jibun_address,
                status,
                room_count,
                addr_norm.normalize_road_prefix(road_address),
                addr_norm.normalize_jibun_prefix(jibun_address),
            ))
        conn.commit()

        after_payload = _report_rate_by_sido_payload()
        after = after_payload["total"]
        expected_delta = {
            "building_cnt": 5,
            "total_units": 200,
            "active_biz_cnt": 3,
            "active_room_cnt": 52,
        }
        for field, delta in expected_delta.items():
            if int(after.get(field) or 0) != int(before.get(field) or 0) + delta:
                failures.append(
                    f"영업신고현황 주소 매칭: {field} 증감이 {delta}가 아님"
                )
        items = after_payload.get("items") or []
        for field in expected_delta:
            if int(after.get(field) or 0) != sum(int(item.get(field) or 0) for item in items):
                failures.append(f"영업신고현황 주소 매칭: 전국 {field}가 시도 합계와 다름")
        if not failures:
            print("OK  영업신고현황 도로명 우선·지번 보조·폐업 제외·신고번호 중복 제거")
    except Exception as exc:
        failures.append(f"영업신고현황 주소 매칭 테스트 오류: {exc}")
    finally:
        try:
            if permit_numbers:
                cur.execute(
                    "DELETE FROM lodging_registry WHERE permit_number = ANY(%s)",
                    (permit_numbers,),
                )
            if building_ids:
                cur.execute(
                    "DELETE FROM master_buildings WHERE id = ANY(%s)",
                    (building_ids,),
                )
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_lodging_address_normalization():
    """건축물대장 괄호 표기와 영업신고 도로명주소 정규화를 검증."""
    import addr_norm

    failures = []
    matching_cases = [
        (
            "경상북도 칠곡군 팔공산로2길 8 (동명면 기성리)",
            "경상북도 칠곡군 동명면 팔공산로2길 8",
        ),
        (
            "경상북도 칠곡군 한티로 708-29 (동명면 기성리)",
            "경상북도 칠곡군 동명면 한티로 708-29",
        ),
        (
            "경상북도 칠곡군 팔공산로4길 11-12 (동명면 기성리)",
            "경상북도 칠곡군 동명면 팔공산로4길 11-12",
        ),
        # 영업신고 원본의 통합 광역명은 자치구 여부에 따라 광주·전남으로 나뉜다.
        (
            "광주광역시 북구 무등로 100",
            "전남광주통합특별시 북구 무등로 100",
        ),
        (
            "전라남도 여수시 여문로 100",
            "전남광주통합특별시 여수시 여문로 100",
        ),
    ]
    for master_address, lodging_address in matching_cases:
        master_key = addr_norm.normalize_road_prefix(master_address)
        lodging_key = addr_norm.normalize_road_prefix(lodging_address)
        if not master_key or master_key != lodging_key:
            failures.append(
                "lodging address normalization: 괄호 안 행정리 표기와 신고 도로명이 "
                f"같은 키가 되지 않음 ({master_key} != {lodging_key})"
            )

    # 한티로와 한티로1길처럼 도로명 자체가 다른 주소는 계속 분리한다.
    different_master = addr_norm.normalize_road_prefix(
        "경상북도 칠곡군 한티로1길 708-13 (동명면 기성리)"
    )
    different_lodging = addr_norm.normalize_road_prefix(
        "경상북도 칠곡군 동명면 한티로 708-13"
    )
    if different_master == different_lodging:
        failures.append(
            "lodging address normalization: 서로 다른 도로명을 같은 키로 합침"
        )
    if not failures:
        print("OK  괄호 안 읍·면·동·광주전남 통합표기 도로명 정규화 및 오매칭 방지")
    return failures

def _check_lodging_auto_naming(client):
    """미확정 일반숙박의 영업신고 대표명 자동 반영 계약을 검증한다."""
    import time as _time
    import addr_norm
    from db import get_conn
    from lodging_matching import refresh_auto_building_names

    failures = []
    run_id = str(int(_time.time() * 1000))
    road_base = f"테스트특별시 자동명명구 검증로 {run_id[-4:]}"
    inserted_buildings = []
    inserted_permits = []
    conn = get_conn()
    cur = conn.cursor()

    def add_building(label, suffix, pending=True, source="pending", with_metadata=True):
        road = f"{road_base}-{suffix}"
        if with_metadata:
            cur.execute(
                """
                INSERT INTO master_buildings
                    (building_name, road_address, jibun_address, sgg_text, sgg_cd, umd_nm, jibun,
                     source, lodging_type, name_pending, building_name_source, building_name_pending_base)
                VALUES (%s, %s, %s, '테스트특별시 자동명명구', '99999', %s, %s,
                        'api_test', '일반', %s, %s, %s)
                RETURNING id
                """,
                (label, road, f"테스트특별시 자동동 {suffix}", "자동동", suffix, pending, source, label),
            )
        else:
            cur.execute(
                """
                INSERT INTO master_buildings
                    (building_name, road_address, source, lodging_type, name_pending,
                     building_name_source, building_name_pending_base)
                VALUES (%s, %s, 'api_test', '일반', TRUE, 'pending', %s)
                RETURNING id
                """,
                (label, road, label),
            )
        building_id = cur.fetchone()["id"]
        inserted_buildings.append(building_id)
        return building_id, road, f"테스트특별시 자동동 {suffix}"

    def add_lodging(name, road, jibun, permit_suffix, rooms, date, status="영업/정상"):
        permit = f"TEST-AUTO-NAME-{run_id}-{permit_suffix}"
        cur.execute(
            """
            INSERT INTO lodging_registry
                (biz_name, permit_number, road_address, jibun_address, permit_date,
                 biz_status_name, room_count, hygiene_type, road_norm, jibun_norm)
            VALUES (%s, %s, %s, %s, %s, %s, %s, '여관업', %s, %s)
            """,
            (
                name, permit, road, jibun, date, status, rooms,
                addr_norm.normalize_road_prefix(road),
                addr_norm.normalize_jibun_prefix(jibun),
            ),
        )
        inserted_permits.append(permit)

    try:
        single_id, single_road, single_jibun = add_building("자동동 1", "1")
        add_lodging("아이리스모텔", single_road, single_jibun, "single", 8, "20220101")

        largest_id, largest_road, largest_jibun = add_building("자동동 2", "2")
        add_lodging("객실최대모텔", largest_road, largest_jibun, "largest", 31, "20200101")
        add_lodging("객실작은호텔", largest_road, largest_jibun, "small", 10, "20250101")

        tie_id, tie_road, tie_jibun = add_building("자동동 3", "3")
        add_lodging("동률기존모텔", tie_road, tie_jibun, "tie-old", 20, "20200101")
        add_lodging("동률신규호텔", tie_road, tie_jibun, "tie-new", 20, "20250101")
        add_lodging("폐업대형모텔", tie_road, tie_jibun, "closed", 999, "20260101", "폐업")

        closed_id, closed_road, closed_jibun = add_building("자동동 4", "4")
        add_lodging("폐업전용모텔", closed_road, closed_jibun, "closed-only", 99, "20260101", "폐업")

        fixed_id, fixed_road, fixed_jibun = add_building("사용자 확정명", "5", pending=False, source="user")
        add_lodging("자동으로바뀌면안됨", fixed_road, fixed_jibun, "fixed", 100, "20260101")

        unmatched_id, _, _ = add_building("자동동 6", "6")
        road_only_id, road_only_road, road_only_jibun = add_building(
            "도로명 임시명", "7", with_metadata=False
        )
        add_lodging("도로명자동모텔", road_only_road, road_only_jibun, "road-only", 7, "20260101")
        manual_id, manual_road, manual_jibun = add_building("자동동 8", "8")
        add_lodging("수동수정전모텔", manual_road, manual_jibun, "manual", 7, "20260101")
        official_pending_id, official_road, official_jibun = add_building(
            "건축HUB 정식명칭", "9", pending=True, source="official"
        )
        add_lodging("정식명칭을바꾸면안됨", official_road, official_jibun, "official-pending", 100, "20260101")
        conn.commit()
        refresh_auto_building_names(conn, inserted_buildings)

        cur.execute(
            """
            SELECT id, building_name, name_pending, building_name_source,
                   building_name_candidate_count
            FROM master_buildings WHERE id = ANY(%s)
            """,
            (inserted_buildings,),
        )
        rows = {row["id"]: row for row in cur.fetchall()}

        expected = {
            single_id: ("아이리스모텔", "lodging_report", 1),
            largest_id: ("객실최대모텔", "lodging_report", 2),
            tie_id: ("동률신규호텔", "lodging_report", 2),
            closed_id: ("자동동 4", "pending", 0),
            fixed_id: ("사용자 확정명", "user", 0),
            unmatched_id: ("자동동 6", "pending", 0),
            road_only_id: ("도로명자동모텔", "lodging_report", 1),
            manual_id: ("수동수정전모텔", "lodging_report", 1),
            official_pending_id: ("건축HUB 정식명칭", "official", 0),
        }
        for building_id, (name, source, candidate_count) in expected.items():
            row = rows.get(building_id) or {}
            if (
                row.get("building_name") != name
                or row.get("building_name_source") != source
                or int(row.get("building_name_candidate_count") or 0) != candidate_count
                or (building_id != fixed_id and row.get("name_pending") is not True)
            ):
                failures.append(
                    f"lodging auto name: id={building_id} 결과 불일치 "
                    f"({row.get('building_name')}, {row.get('building_name_source')}, "
                    f"{row.get('building_name_candidate_count')})"
                )

        # 상호 변경은 같은 신고번호 UPSERT 뒤 다음 재계산에서 즉시 반영돼야 한다.
        cur.execute(
            "UPDATE lodging_registry SET biz_name=%s WHERE permit_number=%s",
            ("아이리스모텔 리뉴얼", inserted_permits[0]),
        )
        conn.commit()
        refresh_auto_building_names(conn, [single_id])
        cur.execute(
            "SELECT building_name FROM master_buildings WHERE id=%s", (single_id,)
        )
        if (cur.fetchone() or {}).get("building_name") != "아이리스모텔 리뉴얼":
            failures.append("lodging auto name: 상호 변경이 다음 재계산에 반영되지 않음")

        # 활성 후보가 0건이 되면 자동명명 결과를 되돌리거나 덮어쓰지 않는다.
        cur.execute(
            "UPDATE lodging_registry SET biz_status_name=%s WHERE permit_number=%s",
            ("폐업", inserted_permits[0]),
        )
        conn.commit()
        refresh_auto_building_names(conn, [single_id])
        cur.execute(
            """SELECT building_name, building_name_source, building_name_candidate_count
                 FROM master_buildings WHERE id=%s""",
            (single_id,),
        )
        no_active_row = cur.fetchone() or {}
        if (
            no_active_row.get("building_name") != "아이리스모텔 리뉴얼"
            or no_active_row.get("building_name_source") != "lodging_report"
            or int(no_active_row.get("building_name_candidate_count") or 0) != 1
        ):
            failures.append("lodging auto name: 활성 후보 0건일 때 기존 자동명칭을 변경함")

        detail = client.get(f"/api/building/{tie_id}")
        payload = detail.get_json() or {}
        if (
            detail.status_code != 200
            or payload.get("building_name_source") != "lodging_report"
            or payload.get("building_name_candidate_count") != 2
            or payload.get("building_name_auto_representative") is not True
        ):
            failures.append("lodging auto name: 공개 상세 API에 자동명칭 출처/대표 정보가 없음")

        with client.session_transaction() as sess:
            sess["admin"] = True
        listing = client.get(f"/api/admin/buildings?q=동률신규호텔")
        list_row = next(
            (row for row in (listing.get_json() or {}).get("items", []) if row.get("id") == tie_id),
            None,
        )
        if (
            listing.status_code != 200
            or not list_row
            or list_row.get("building_name_source") != "lodging_report"
            or list_row.get("building_name_auto_representative") is not True
        ):
            failures.append("lodging auto name: 관리자 목록에 자동명칭 출처/대표 정보가 없음")

        # 관리자 수동 명칭 수정도 확정명으로 보호되어 다음 동기화에 덮어써지지 않아야 한다.
        manual_update = client.put(
            f"/api/admin/buildings/{manual_id}",
            json={"building_name": "관리자 확정명"},
        )
        refresh_auto_building_names(conn, [manual_id])
        cur.execute(
            """SELECT building_name, name_pending, building_name_source
                 FROM master_buildings WHERE id=%s""",
            (manual_id,),
        )
        manual_row = cur.fetchone() or {}
        if (
            manual_update.status_code != 200
            or manual_row.get("building_name") != "관리자 확정명"
            or manual_row.get("name_pending") is not False
            or manual_row.get("building_name_source") != "user"
        ):
            failures.append("lodging auto name: 관리자 수동 명칭이 자동명에 덮어써짐")

        if not failures:
            print("OK  영업신고 자동명칭 단일·복수·동률·폐업·상호변경·확정명 보호")
    except Exception as exc:
        failures.append(f"lodging auto name 테스트 오류: {exc}")
    finally:
        try:
            if inserted_permits:
                cur.execute(
                    "DELETE FROM lodging_registry WHERE permit_number = ANY(%s)",
                    (inserted_permits,),
                )
            if inserted_buildings:
                cur.execute("DELETE FROM master_buildings WHERE id = ANY(%s)", (inserted_buildings,))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_brhub_auto_naming_contract():
    """BRHUB 표제부 이름 없는 일반숙박 INSERT가 자동명명 대상이 되는지 확인."""
    import time
    import addr_norm
    import sync_brhub as brhub_module
    from db import get_conn
    from lodging_matching import refresh_auto_building_names

    failures = []
    run_id = str(int(time.time() * 1000))
    road = f"테스트특별시 BRHUB검증구 자동연동로 {run_id[-5:]}"
    permit = f"TEST-BRHUB-AUTO-NAME-{run_id}"
    building_id = None
    conn = get_conn()
    cur = conn.cursor()
    try:
        building_name, name_pending, name_source, pending_base = (
            brhub_module._building_name_metadata(
                "-", "일반", "BRHUB자동명동", "123-4", road
            )
        )
        official_metadata = brhub_module._building_name_metadata(
            "건축HUB 정식명칭", "일반", "BRHUB자동명동", "123-5", road
        )
        if (
            (building_name, name_pending, name_source, pending_base)
            != ("BRHUB자동명동 123-4", True, "pending", "BRHUB자동명동 123-4")
            or official_metadata != ("건축HUB 정식명칭", False, "official", None)
        ):
            failures.append("BRHUB 자동명명: 표제부 이름별 pending/official 메타데이터가 잘못됨")
            return failures

        # sync_brhub.py의 신규 행 INSERT 계약과 같은 명칭 메타데이터를 사용한다.
        cur.execute(
            """
            INSERT INTO master_buildings
                (building_name, road_address, jibun_address, sgg_text, umd_nm, jibun,
                 source, lodging_type, name_pending, building_name_source,
                 building_name_candidate_count, building_name_pending_base)
            VALUES (%s, %s, %s, %s, %s, %s, 'brhub_bulk', '일반', %s, %s, 0, %s)
            RETURNING id
            """,
            (
                building_name, road, f"테스트특별시 BRHUB자동명동 123-4번지",
                "테스트특별시 BRHUB검증구", "BRHUB자동명동", "123-4",
                name_pending, name_source, pending_base,
            ),
        )
        building_id = cur.fetchone()["id"]
        cur.execute(
            """
            INSERT INTO lodging_registry
                (biz_name, permit_number, road_address, jibun_address, road_norm, jibun_norm,
                 biz_status_name, hygiene_type)
            VALUES (%s, %s, %s, %s, %s, %s, '영업/정상', '여관업')
            """,
            (
                "BRHUB연동모텔", permit, road, f"테스트특별시 BRHUB자동명동 123-4번지",
                addr_norm.normalize_road_prefix(road),
                addr_norm.normalize_jibun_prefix("테스트특별시 BRHUB자동명동 123-4번지"),
            ),
        )
        conn.commit()
        refresh_auto_building_names(conn, [building_id])

        cur.execute(
            """
            SELECT building_name, name_pending, building_name_source,
                   building_name_candidate_count, building_name_pending_base
              FROM master_buildings
             WHERE id=%s
            """,
            (building_id,),
        )
        row = cur.fetchone() or {}
        if (
            row.get("building_name") != "BRHUB연동모텔"
            or row.get("name_pending") is not True
            or row.get("building_name_source") != "lodging_report"
            or int(row.get("building_name_candidate_count") or 0) != 1
            or row.get("building_name_pending_base") != "BRHUB자동명동 123-4"
        ):
            failures.append(f"BRHUB 자동명명: 신규 행 자동명칭 결과 불일치 ({row})")
        else:
            print("OK  BRHUB 신규 일반숙박 placeholder → 영업신고 자동명칭")
    except Exception as exc:
        failures.append(f"BRHUB 자동명명 테스트 오류: {exc}")
    finally:
        try:
            cur.execute("DELETE FROM lodging_registry WHERE permit_number=%s", (permit,))
            if building_id:
                cur.execute("DELETE FROM master_buildings WHERE id=%s", (building_id,))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_building_biz_status_filters(client):
    """주소 매칭 기준 정상 운영/전부 폐업/미매칭 건물 필터와 조합을 확인."""
    import time
    from urllib.parse import urlencode
    import addr_norm
    from db import get_conn
    from lodging_matching import ACTIVE_STATUS

    failures = []
    run_id = str(int(time.time() * 1000))
    token = f"STATUSFILTER{run_id}"
    specs = [
        ("활성전용", [ACTIVE_STATUS]),
        ("폐업전용", ["폐업"]),
        ("혼합", ["폐업", ACTIVE_STATUS]),
        ("미매칭", []),
    ]
    building_ids = []
    permits = []
    conn = get_conn()
    cur = conn.cursor()
    try:
        for building_index, (suffix, statuses) in enumerate(specs):
            road_address = (
                f"테스트특별시 상태검증구 상태검증로 {building_index + 100}"
            )
            cur.execute(
                """
                INSERT INTO master_buildings
                    (building_name, road_address, source, lodging_type, name_pending)
                VALUES (%s, %s, 'api_test', '일반', TRUE)
                RETURNING id
                """,
                (f"{token}-{suffix}", road_address),
            )
            building_id = cur.fetchone()["id"]
            building_ids.append(building_id)
            for index, status in enumerate(statuses):
                permit = f"TEST-BIZ-STATUS-{run_id}-{suffix}-{index}"
                permits.append(permit)
                cur.execute(
                    """
                    INSERT INTO lodging_registry
                        (biz_name, permit_number, biz_status_name, road_norm)
                    VALUES (%s, %s, %s, %s)
                    """,
                    (
                        f"{token}-{suffix}-사업장{index}",
                        permit,
                        status,
                        addr_norm.normalize_road_prefix(road_address),
                    ),
                )
        conn.commit()

        with client.session_transaction() as sess:
            sess["admin"] = True

        def filtered_ids(**filters):
            query = {"q": token, "size": "200"}
            query.update(filters)
            response = client.get(f"/api/admin/buildings?{urlencode(query)}")
            payload = response.get_json() or {}
            if response.status_code != 200:
                failures.append(
                    f"건물 영업상태 필터: HTTP {response.status_code} ({filters})"
                )
            return {row.get("building_name") for row in payload.get("items", [])}

        names = {f"{token}-{suffix}" for suffix, _ in specs}
        active_names = filtered_ids(
            lodging_type_filter="일반",
            name_pending="1",
            biz_status_filter="active",
        )
        closed_names = filtered_ids(
            lodging_type_filter="일반",
            name_pending="1",
            biz_status_filter="closed",
        )
        all_names = filtered_ids(lodging_type_filter="일반", name_pending="1")
        expected_active = {f"{token}-활성전용", f"{token}-혼합"}
        expected_closed = {f"{token}-폐업전용"}
        if active_names != expected_active:
            failures.append(
                f"건물 영업상태 필터: 정상 운영중 결과 불일치 "
                f"(기대={expected_active}, 실제={active_names})"
            )
        if closed_names != expected_closed:
            failures.append(
                f"건물 영업상태 필터: 폐업만 결과 불일치 "
                f"(기대={expected_closed}, 실제={closed_names})"
            )
        if all_names != names:
            failures.append(
                f"건물 영업상태 필터: 전체 상태 결과 불일치 "
                f"(기대={names}, 실제={all_names})"
            )
        if not failures:
            print("OK  건물마스터 정상 운영/폐업/미매칭 및 필터 조합")
    except Exception as exc:
        failures.append(f"건물 영업상태 필터 테스트 오류: {exc}")
    finally:
        try:
            if permits:
                cur.execute(
                    "DELETE FROM lodging_registry WHERE permit_number = ANY(%s)",
                    (permits,),
                )
            if building_ids:
                cur.execute(
                    "DELETE FROM master_buildings WHERE id = ANY(%s)",
                    (building_ids,),
                )
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_lodging_cap_auto_naming():
    """일일 캡 도달 시 당일 UPSERT 주소의 자동명칭을 즉시 갱신하는지 확인."""
    import os as _os
    import time as _time
    from unittest.mock import patch
    import addr_norm
    import sync_lodgings as sync_module
    from db import get_conn

    failures = []
    run_id = str(int(_time.time() * 1000))
    road = f"테스트특별시 캡검증구 자동명명대로 {run_id[-4:]}"
    permit = f"TEST-CAP-AUTO-NAME-{run_id}"
    daily_key = f"test_lodging_daily_calls_{run_id}"
    progress_key = f"test_lodging_sync_progress_{run_id}"
    inserted_buildings = []
    conn = get_conn()
    cur = conn.cursor()
    original_daily_key = sync_module.DAILY_CALLS_META_KEY
    original_progress_key = sync_module.PROGRESS_META_KEY
    try:
        for label, suffix in (("캡 처리 전 임시명", "1"), ("당일 미처리 임시명", "2")):
            cur.execute(
                """
                INSERT INTO master_buildings
                    (building_name, road_address, source, lodging_type, name_pending,
                     building_name_source, building_name_pending_base)
                VALUES (%s, %s, 'api_test', '일반', TRUE, 'pending', %s)
                RETURNING id
                """,
                (label, f"{road}-{suffix}", label),
            )
            inserted_buildings.append(cur.fetchone()["id"])
        conn.commit()

        item = {
            "BPLC_NM": "캡당일자동명모텔",
            "MNG_NO": permit,
            "ROAD_NM_ADDR": road + "-1",
            "LOTNO_ADDR": f"테스트특별시 캡동 {run_id[-3:]}-1",
            "LCPMT_YMD": "20260101",
            "SALS_STTS_NM": "영업/정상",
            "DTL_SALS_STTS_NM": "",
            "KSRM_CNT": "5",
            "WSRM_CNT": "0",
            "SNTTN_BZSTAT_NM": "여관업",
            "TELNO": "",
            "DAT_UPDT_PNT": "",
        }
        with patch.object(sync_module, "DAILY_CALLS_META_KEY", daily_key), \
             patch.object(sync_module, "PROGRESS_META_KEY", progress_key), \
             patch.object(
                 sync_module,
                 "_fetch_page_retry",
                 return_value=([item], 2, False),
             ), \
             patch.dict(_os.environ, {sync_module.SERVICE_KEY_ENV: "test-key"}):
            completed, processed, calls_today = sync_module.sync_lodgings(
                num_rows=100, sleep_sec=0, max_calls=1
            )

        cur.execute(
            "SELECT building_name FROM master_buildings WHERE id=%s",
            (inserted_buildings[0],),
        )
        target_name = (cur.fetchone() or {}).get("building_name")
        cur.execute(
            "SELECT building_name FROM master_buildings WHERE id=%s",
            (inserted_buildings[1],),
        )
        untouched_name = (cur.fetchone() or {}).get("building_name")
        if not (
            completed is False
            and processed == 1
            and calls_today == 1
            and target_name == "캡당일자동명모텔"
            and untouched_name == "당일 미처리 임시명"
        ):
            failures.append(
                "lodging cap auto name: 캡 중간 종료 후 당일 처리 건물만 즉시 자동명명되지 않음 "
                f"(completed={completed}, processed={processed}, target={target_name}, "
                f"untouched={untouched_name})"
            )
        else:
            print("OK  일일 캡 도달 시 당일 처리분 자동명칭 즉시 반영")
    except Exception as exc:
        failures.append(f"lodging cap auto name 테스트 오류: {exc}")
    finally:
        sync_module.DAILY_CALLS_META_KEY = original_daily_key
        sync_module.PROGRESS_META_KEY = original_progress_key
        try:
            cur.execute("DELETE FROM lodging_registry WHERE permit_number=%s", (permit,))
            cur.execute(
                "DELETE FROM master_buildings WHERE id = ANY(%s)",
                (inserted_buildings,),
            )
            cur.execute(
                "DELETE FROM app_meta WHERE key = ANY(%s)",
                ([daily_key, progress_key],),
            )
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_general_units_table_markup(client):
    """관리자 통계표에서 생활 외 호실수와 유형별 신고율 기준을 설명하는지 확인."""
    failures = []
    response = client.get("/admin")
    if response.status_code != 200:
        return [f"admin stats markup: 관리자 페이지 HTTP {response.status_code}"]

    html = response.get_data(as_text=True)
    required_fragments = [
        "일반숙박시설은 구분소유 호수 개념이 없어 이 값이 실제 객실수를 반영하지 않습니다.",
        "실제 객실수는 '신고율/객실수' 컬럼을 참고하세요",
        "const generalUnitsCell =",
        'row.type === "일반"',
        "? generalUnitsCell(row.units)",
        'const COVERAGE_TYPES = new Set(["관광", "에어비앤비", "농어촌민박", "캠핑", "한옥", "복합"])',
        "나머지 확정 유형은 신고 매칭 건물÷유형 건물수를 합산",
        "사이트수는 객실수와 분리",
        "원장에는 복합 업태가 없음",
        'if (c.key === "units")',
        "${n(row.units)}",
    ]
    missing = [fragment for fragment in required_fragments if fragment not in html]
    if missing:
        failures.append(
            "admin stats markup: 일반 전용 호실수 표시 또는 비일반 기존 표시가 누락됨 "
            f"({', '.join(missing)})"
        )
    else:
        print("OK  관리자 통계표 생활 외 호실수 참고·유형별 신고율 설명")
    return failures


def _check_lodging_metric_contract(client):
    """실제 표본 건물과 공개 통계가 일반숙박 지표 계약을 지키는지 확인."""
    import app as app_module
    import time as _time
    import addr_norm
    from db import get_conn
    from io import BytesIO
    from openpyxl import load_workbook

    failures = []
    conn = get_conn()
    cur = conn.cursor()
    try:
        samples = []
        for pattern in ("%라마다인천호텔%", "%THE TRINY HOTEL%"):
            cur.execute(
                "SELECT id, building_name, lodging_type FROM master_buildings "
                "WHERE building_name ILIKE %s ORDER BY id LIMIT 1",
                (pattern,),
            )
            row = cur.fetchone()
            if not row:
                failures.append(f"lodging metric: 표본 건물을 찾지 못했습니다 ({pattern})")
            else:
                samples.append(row)

        cur.execute(
            "SELECT id, building_name, lodging_type FROM master_buildings "
            "WHERE building_name ILIKE %s ORDER BY id LIMIT 1",
            ("%빌리브패러그라프해운대%",),
        )
        living_sample = cur.fetchone()
        if not living_sample:
            failures.append("lodging metric: 표본 건물을 찾지 못했습니다 (빌리브패러그라프해운대)")

        cur.execute(
            "SELECT id, building_name, lodging_type FROM master_buildings "
            "WHERE lodging_type IS DISTINCT FROM '일반' "
            "  AND lodging_type IS DISTINCT FROM 'mixed_use_excluded' "
            "ORDER BY id LIMIT 1"
        )
        non_general = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    for sample in samples:
        resp = client.get(f"/api/building/{sample['id']}")
        payload = resp.get_json() or {}
        if resp.status_code != 200:
            failures.append(
                f"lodging metric: {sample['building_name']} 상세 API HTTP {resp.status_code}"
            )
            continue
        if payload.get("lodging_type") != "일반":
            failures.append(
                f"lodging metric: {sample['building_name']} 유형이 일반이 아님 "
                f"({payload.get('lodging_type')})"
            )
        if payload.get("lodging_metric") != "room_count":
            failures.append(
                f"lodging metric: {sample['building_name']} lodging_metric이 room_count가 아님"
            )
        if payload.get("lodging_report_rate") is not None:
            failures.append(
                f"lodging metric: {sample['building_name']} 상세 응답에 신고율이 남아 있음"
            )
        if payload.get("lodging_active_business_count") != len(payload.get("lodgings") or []):
            failures.append(
                f"lodging metric: {sample['building_name']} 활성 사업장 수가 목록과 다름"
            )
        else:
            print(
                f"OK  {sample['building_name']} 일반숙박 객실수 지표 "
                f"({payload.get('lodging_room_total', 0)}실)"
            )

    if non_general:
        resp = client.get(f"/api/building/{non_general['id']}")
        payload = resp.get_json() or {}
        if resp.status_code != 200:
            failures.append(
                f"lodging metric: 비일반 표본 {non_general['building_name']} 상세 API "
                f"HTTP {resp.status_code}"
            )
        elif (
            payload.get("lodging_type") == "생활"
            and payload.get("lodging_metric") != "report_rate"
        ):
            failures.append(
                f"lodging metric: 생활 표본 {non_general['building_name']}이 신고율 지표가 아님"
            )
        elif (
            payload.get("lodging_type") != "생활"
            and payload.get("lodging_metric") != "room_count"
        ):
            failures.append(
                f"lodging metric: 비생활 표본 {non_general['building_name']}이 객실수 지표가 아님"
            )
        else:
            print(f"OK  {non_general['building_name']} 유형별 신고 지표 유지")
    else:
        failures.append("lodging metric: 비일반 표본 건물을 찾지 못했습니다.")

    if living_sample:
        resp = client.get(f"/api/building/{living_sample['id']}")
        payload = resp.get_json() or {}
        expected_rate = round(153 / 286 * 100, 1)
        if resp.status_code != 200:
            failures.append(
                f"lodging metric: {living_sample['building_name']} 상세 API "
                f"HTTP {resp.status_code}"
            )
        elif payload.get("lodging_type") != "생활":
            failures.append(
                f"lodging metric: {living_sample['building_name']} 유형이 생활이 아님 "
                f"({payload.get('lodging_type')})"
            )
        elif payload.get("lodging_metric") != "report_rate":
            failures.append(
                f"lodging metric: {living_sample['building_name']}가 신고율 지표가 아님"
            )
        elif payload.get("units") != 286 or payload.get("lodging_room_total") != 153:
            failures.append(
                f"lodging metric: {living_sample['building_name']} 객실수/호실수가 "
                f"153/286이 아님 ({payload.get('lodging_room_total')}/{payload.get('units')})"
            )
        elif payload.get("lodging_report_rate") != expected_rate:
            failures.append(
                f"lodging metric: {living_sample['building_name']} 신고율이 "
                f"{expected_rate}%가 아님 ({payload.get('lodging_report_rate')}%)"
            )
        else:
            print(
                f"OK  {living_sample['building_name']} 생활 신고율 지표 "
                f"(153실 / 286실 = {expected_rate}%)"
            )

    # 관리자 통계는 일반숙박을 업체수÷건물수, 생활만 신고객실수÷대장 호실수,
    # 나머지 확정 유형을 신고 매칭 건물÷유형 건물수로 계산한다.
    # 별도 객실수 비교 API에도 생활숙박만 포함한다.
    original_cache = app_module._bld_full_stats_cache
    try:
        app_module._bld_full_stats_cache = {"ts": 0.0, "data": None}
        fallback_response = client.get("/api/stats/registration-rate")
        fallback_stats = fallback_response.get_json() or {}

        with client.session_transaction() as sess:
            sess["admin"] = True

        full_stats_response = client.get("/api/admin/buildings/full-stats")
        full_stats = full_stats_response.get_json() or {}
        rows = {row.get("type"): row for row in full_stats.get("rows", [])}
        total_row = rows.get("전체") or {}
        general_row = rows.get("일반") or {}
        building_coverage_types = {
            "관광", "에어비앤비", "농어촌민박", "캠핑", "한옥", "복합",
        }

        if full_stats_response.status_code != 200 or not full_stats.get("ok"):
            failures.append("lodging metric: 관리자 전체 통계를 불러오지 못했습니다.")
        elif (
            general_row.get("lodging_metric") != "businesses_per_building"
            or general_row.get("report_rate")
            != (
                round(
                    int(general_row.get("permit_count") or 0)
                    / int(general_row.get("building_count") or 0) * 100,
                    1,
                )
                if int(general_row.get("building_count") or 0) else None
            )
        ):
            failures.append("lodging metric: 관리자 일반숙박 신고율이 업체수 ÷ 건물수 기준이 아님")
        elif not building_coverage_types <= set(rows):
            failures.append("lodging metric: 건물 커버리지 유형 행이 누락됨")
        elif any(
            row.get("lodging_metric") != "buildings_with_active_report"
            or row.get("report_rate_basis") != "buildings_with_active_report"
            or row.get("report_rate")
            != (
                round(
                    int(row.get("report_rate_numerator") or 0)
                    / int(row.get("report_rate_denominator") or 0) * 100,
                    1,
                )
                if int(row.get("report_rate_denominator") or 0) else None
            )
            or int(row.get("report_rate_denominator") or 0)
            != int(row.get("building_count") or 0)
            for label in building_coverage_types
            for row in [rows.get(label) or {}]
        ):
            failures.append("lodging metric: 건물 커버리지 유형의 신고율 분자·분모가 잘못됨")
        else:
            expected_sub_types = ["일반호텔", "여관업", "여인숙업"]
            sub_rows = general_row.get("sub_rows")
            if not isinstance(sub_rows, list) or [row.get("type") for row in sub_rows] != expected_sub_types:
                failures.append("lodging metric: 관리자 일반숙박 세분류 행이 3개 업태로 반환되지 않음")
            elif (
                sum(int(row.get("permit_count") or 0) for row in sub_rows) != general_row.get("permit_count")
                or sum(int(row.get("room_count") or 0) for row in sub_rows) != general_row.get("room_count")
                or any(
                    row.get("report_rate")
                    != (
                        round(
                            int(row.get("permit_count") or 0)
                            / int(row.get("building_count") or 0) * 100,
                            1,
                        )
                        if int(row.get("building_count") or 0) else None
                    )
                    for row in sub_rows
                )
            ):
                failures.append("lodging metric: 일반숙박 세분류 업체수·객실수 또는 업체÷건물 신고율이 잘못됨")
            else:
                print(
                    f"OK  관리자 일반숙박 3개 세분류·업체÷건물 신고율 "
                    f"({general_row.get('permit_count')}개 / {general_row.get('building_count')}건)"
                )

            expected_numerator = int(total_row.get("report_rate_numerator") or 0)
            expected_denominator = int(total_row.get("report_rate_denominator") or 0)
            component_types = {"생활", "일반"} | building_coverage_types
            component_numerator = sum(
                int((rows.get(label) or {}).get("report_rate_numerator") or 0)
                for label in component_types
            )
            component_denominator = sum(
                int((rows.get(label) or {}).get("report_rate_denominator") or 0)
                for label in component_types
            )
            expected_rate = round(expected_numerator * 100.0 / expected_denominator, 1) if expected_denominator else None
            if (
                expected_numerator != component_numerator
                or expected_denominator != component_denominator
                or total_row.get("report_rate") != expected_rate
            ):
                failures.append("lodging metric: 관리자 전체 신고율의 유형별 분자·분모가 불일치")
            else:
                print(
                    f"OK  관리자 전체 신고율 유형별 합산 "
                    f"({expected_numerator} / {expected_denominator} = {expected_rate}%)"
                )

        general_list_response = client.get("/api/admin/buildings?lodging_type_filter=일반")
        general_totals = (general_list_response.get_json() or {}).get("totals") or {}
        if (
            general_list_response.status_code != 200
            or general_totals.get("weighted_report_rate") is not None
            or general_totals.get("report_rate_units") != 0
        ):
            failures.append("lodging metric: 일반숙박만 필터한 관리자 합계에 신고율이 남아 있음")
        else:
            print("OK  관리자 일반숙박 필터 합계는 객실수 지표만 사용")

        cached_response = client.get("/api/stats/registration-rate")
        cached_stats = cached_response.get_json() or {}
        for label, response, stats_payload in (
            ("캐시 미스", fallback_response, fallback_stats),
            ("캐시 히트", cached_response, cached_stats),
        ):
            if (
                response.status_code != 200
                or stats_payload.get("general_excluded") is not True
                or stats_payload.get("tourism_excluded") is not True
                or stats_payload.get("non_living_excluded") is not True
                or stats_payload.get("biz_units") != total_row.get("report_rate_room_count")
                or stats_payload.get("total_units") != total_row.get("report_rate_units")
                or stats_payload.get("rate")
                != (
                    round(
                        int(total_row.get("report_rate_room_count") or 0)
                        / int(total_row.get("report_rate_units") or 0) * 100,
                        1,
                    )
                    if int(total_row.get("report_rate_units") or 0) else None
                )
            ):
                failures.append(f"lodging metric: 공개 전국 신고율 {label} 경로가 일반 제외 집계와 불일치")
            else:
                print(f"OK  공개 전국 신고율 {label} 일반 제외 집계 일치")

        # 관리자 목록과 선택 엑셀도 일반숙박을 객실수 지표로 전달하는지 확인한다.
        ramada = next((row for row in samples if row["building_name"] == "라마다인천호텔"), None)
        if ramada:
            list_response = client.get(
                f"/api/admin/buildings?lodging_type_filter=일반&q={ramada['building_name']}"
            )
            list_payload = list_response.get_json() or {}
            list_row = next(
                (row for row in list_payload.get("items", []) if row.get("id") == ramada["id"]),
                None,
            )
            if not list_row or list_row.get("lodging_metric") != "room_count":
                failures.append("lodging metric: 관리자 목록 일반숙박 지표가 객실수로 내려오지 않음")

            export_response = client.get(f"/api/admin/buildings/export.xlsx?ids={ramada['id']}")
            if export_response.status_code != 200:
                failures.append("lodging metric: 관리자 건물 엑셀 다운로드 실패")
            else:
                wb = load_workbook(BytesIO(export_response.data), data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                metric_col = headers.index("신고 지표(생활=신고율, 그 외=객실수)") + 1
                metric_value = ws.cell(2, metric_col).value
                if not isinstance(metric_value, str) or not metric_value.endswith("실"):
                    failures.append("lodging metric: 관리자 건물 엑셀에 일반숙박 객실수 표기가 없음")
                else:
                    print(f"OK  관리자 목록·엑셀 일반숙박 객실수 표기 ({metric_value})")

        # 도로명 정규화 키는 다르지만 지번 키가 같은 건물은 상세 API와
        # 관리자 목록 모두 지번 보조 매칭으로 영업신고를 보여야 한다.
        # 개발 DB의 실제 표본 유무에 의존하지 않도록 임시 행을 만들고 정리한다.
        run_id = str(int(_time.time() * 1000))
        building_id = None
        permit_number = f"TEST-ADMIN-JIBUN-{run_id}"
        building_name = f"자동 지번보조매칭 {run_id}"
        # 새 괄호 정규화 규칙이 도로명 매칭을 보강하더라도, 도로명 자체가
        # 다른 경우에는 지번 보조 매칭이 계속 작동해야 한다.
        building_road = "경상북도 칠곡군 동명면 한티로1길 8"
        building_jibun = "경상북도 칠곡군 동명면 기성리 836번지"
        lodging_road = "경상북도 칠곡군 동명면 팔공산로2길 8"
        lodging_jibun = "경상북도 칠곡군 동명면 기성리 836"
        conn = get_conn()
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO master_buildings "
                "(building_name, road_address, jibun_address, source, lodging_type) "
                "VALUES (%s, %s, %s, 'api_test', '일반') RETURNING id",
                (building_name, building_road, building_jibun),
            )
            building_id = cur.fetchone()["id"]
            cur.execute(
                "INSERT INTO lodging_registry "
                "(biz_name, permit_number, road_address, jibun_address, biz_status_name, "
                "room_count, hygiene_type, road_norm, jibun_norm) "
                "VALUES (%s, %s, %s, %s, '영업/정상', 18, '여관업', %s, %s)",
                (
                    "아이리스모텔",
                    permit_number,
                    lodging_road,
                    lodging_jibun,
                    addr_norm.normalize_road_prefix(lodging_road),
                    addr_norm.normalize_jibun_prefix(lodging_jibun),
                ),
            )
            conn.commit()

            detail_response = client.get(f"/api/building/{building_id}")
            detail_payload = detail_response.get_json() or {}
            detail_has_iris = any(
                item.get("biz_name") == "아이리스모텔"
                for item in detail_payload.get("lodgings", [])
            )
            list_response = client.get(f"/api/admin/buildings?q={building_name}")
            list_payload = list_response.get_json() or {}
            list_row = next(
                (row for row in list_payload.get("items", []) if row.get("id") == building_id),
                None,
            )
            list_has_iris = any(
                item.get("biz_name") == "아이리스모텔"
                for item in (list_row or {}).get("lodging_list", [])
            )
            if detail_response.status_code != 200 or not detail_has_iris:
                failures.append("lodging match fallback: 아이리스모텔이 상세 API 지번 보조 매칭에서 누락됨")
            elif list_response.status_code != 200 or not list_row:
                failures.append("lodging match fallback: 지번 보조 매칭 건물 행을 관리자 목록에서 찾지 못함")
            elif not list_row.get("lodging_count") or not list_has_iris:
                failures.append("lodging match fallback: 아이리스모텔이 관리자 목록에서 미매칭으로 표시됨")
            else:
                print("OK  아이리스모텔 지번 보조 매칭이 상세·관리자 목록에서 일치")
        finally:
            if building_id is not None:
                cur.execute("DELETE FROM lodging_registry WHERE permit_number = %s", (permit_number,))
                cur.execute("DELETE FROM master_buildings WHERE id = %s", (building_id,))
                conn.commit()
            cur.close()
            conn.close()
    finally:
        app_module._bld_full_stats_cache = original_cache

    return failures


def _check_chat_phone_verification(client):
    """미인증 사용자의 채팅방 생성을 서버에서 차단하고, 인증 뒤 생성되는지 확인."""
    import time as _time
    from db import get_conn

    failures = []
    run_id = str(int(_time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    seller_id = buyer_id = listing_id = room_id = None
    try:
        cur.execute("SELECT id, building_name FROM master_buildings ORDER BY id LIMIT 1")
        building = cur.fetchone()
        if not building:
            return ["chat phone verification: 테스트용 master_buildings 행이 없습니다."]
        cur.execute(
            "INSERT INTO users (email, name, phone, phone_verified) VALUES (%s, %s, %s, TRUE) RETURNING id",
            (f"chat-seller-{run_id}@example.test", "채팅 판매자", "010-0000-0000"),
        )
        seller_id = cur.fetchone()["id"]
        cur.execute(
            "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
            (f"chat-buyer-{run_id}@example.test", "채팅 구매자"),
        )
        buyer_id = cur.fetchone()["id"]
        cur.execute("""
            INSERT INTO listing_requests
                (user_id, master_building_id, deal_type, contact_phone, deal_mode, status)
            VALUES (%s, %s, '매매', %s, 'direct', 'submitted')
            RETURNING id
        """, (seller_id, building["id"], "010-0000-0000"))
        listing_id = cur.fetchone()["id"]
        conn.commit()

        with client.session_transaction() as sess:
            sess["user_id"] = buyer_id
        blocked = client.post("/api/chat/rooms", json={"listing_request_id": listing_id})
        blocked_payload = blocked.get_json() or {}
        if blocked.status_code != 403 or blocked_payload.get("code") != "PHONE_VERIFICATION_REQUIRED":
            failures.append(
                "chat phone verification: 미인증 채팅 생성 응답이 "
                f"HTTP {blocked.status_code}, code={blocked_payload.get('code')} (기대: 403, PHONE_VERIFICATION_REQUIRED)"
            )
        else:
            print("OK  /api/chat/rooms 미인증 차단 (403, PHONE_VERIFICATION_REQUIRED)")

        cur.execute(
            "UPDATE users SET phone=%s, phone_verified=TRUE WHERE id=%s",
            ("010-1111-2222", buyer_id),
        )
        conn.commit()
        allowed = client.post("/api/chat/rooms", json={"listing_request_id": listing_id})
        allowed_payload = allowed.get_json() or {}
        if allowed.status_code != 200 or not allowed_payload.get("ok") or not allowed_payload.get("room_id"):
            failures.append("chat phone verification: 인증 후 채팅방 생성에 실패했습니다.")
        else:
            room_id = allowed_payload["room_id"]
            print("OK  /api/chat/rooms 인증 후 생성")
    except Exception as exc:
        failures.append(f"chat phone verification 테스트 오류: {exc}")
    finally:
        try:
            if room_id:
                cur.execute("DELETE FROM chat_messages WHERE room_id=%s", (room_id,))
                cur.execute("DELETE FROM chat_rooms WHERE id=%s", (room_id,))
            if listing_id:
                cur.execute("DELETE FROM listing_requests WHERE id=%s", (listing_id,))
            if buyer_id:
                cur.execute("DELETE FROM users WHERE id=%s", (buyer_id,))
            if seller_id:
                cur.execute("DELETE FROM users WHERE id=%s", (seller_id,))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_room_inventory_contract_dates(client):
    """방 재고의 만기일·층·채널·벌크 생성과 소유권을 검증한다."""
    import time as _time
    from datetime import date, timedelta
    from db import get_conn

    failures = []
    run_id = str(int(_time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    owner_id = other_id = listing_id = room_id = None
    try:
        cur.execute("""
            SELECT is_nullable, column_default
              FROM information_schema.columns
             WHERE table_schema = current_schema()
               AND table_name = 'business_room_inventory'
               AND column_name = 'status'
        """)
        status_column = cur.fetchone()
        if (not status_column or status_column["is_nullable"] != "NO"
                or not status_column.get("column_default")):
            failures.append("room inventory: status가 NOT NULL 기본 공실 제약으로 마이그레이션되지 않음")
        cur.execute("""
            SELECT column_name, is_nullable, column_default
              FROM information_schema.columns
             WHERE table_schema = current_schema()
               AND table_name = 'business_room_inventory'
                AND column_name IN ('deposit_krw', 'floor', 'channel')
        """)
        inventory_columns = {row["column_name"]: row for row in cur.fetchall()}
        channel_column = inventory_columns.get("channel")
        if ("deposit_krw" not in inventory_columns or not channel_column
                or inventory_columns["deposit_krw"]["is_nullable"] != "YES"
                or "floor" not in inventory_columns or not channel_column
                or channel_column["is_nullable"] != "NO"
                or "장박가능" not in (channel_column.get("column_default") or "")):
            failures.append("room inventory: 보증금/floor/channel 마이그레이션 또는 장박가능 기본값이 없음")
        cur.execute("SELECT id FROM master_buildings ORDER BY id LIMIT 1")
        building = cur.fetchone()
        if not building:
            return ["room inventory: 테스트용 master_buildings 행이 없습니다."]
        cur.execute(
            "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
            (f"room-owner-{run_id}@example.test", "방 재고 소유자"),
        )
        owner_id = cur.fetchone()["id"]
        cur.execute(
            "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
            (f"room-other-{run_id}@example.test", "방 재고 타인"),
        )
        other_id = cur.fetchone()["id"]
        cur.execute("""
            INSERT INTO listing_requests
                (user_id, master_building_id, deal_type, contact_phone, deal_mode, status)
            VALUES (%s, %s, '월세', %s, 'direct', 'submitted')
            RETURNING id
        """, (owner_id, building["id"], "010-0000-0000"))
        listing_id = cur.fetchone()["id"]
        conn.commit()

        contract_end_date = (date.today() + timedelta(days=30)).isoformat()
        with client.session_transaction() as sess:
            sess["user_id"] = owner_id
        created = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms",
            json={
                "room_label": "201호",
                "deposit_krw": 500,
                "monthly_rent_krw": 120,
                "status": "입실",
                "contract_end_date": contract_end_date,
            },
        )
        created_payload = created.get_json() or {}
        if created.status_code != 201 or not created_payload.get("ok"):
            failures.append(f"room inventory: 입실 방 추가 실패 (HTTP {created.status_code})")
            return failures
        room = created_payload.get("item") or {}
        room_id = room.get("id")
        if (room.get("status") != "입실" or room.get("contract_end_date") != contract_end_date
                or room.get("deposit_krw") != 500
                or room.get("monthly_rent_krw") != 120
                or room.get("channel") != "장박가능" or room.get("floor") is not None):
            failures.append("room inventory: 단건 방의 보증금·월세·기본 채널 또는 계약만기일이 저장되지 않음")

        manual_floor_created = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms",
            json={"room_label": "401호", "status": "공실", "floor": 4},
        )
        manual_floor_item = (manual_floor_created.get_json() or {}).get("item") or {}
        if (manual_floor_created.status_code != 201 or manual_floor_item.get("floor") != 4
                or manual_floor_item.get("channel") != "장박가능"):
            failures.append("room inventory: 수동 추가 방의 층 또는 기본 채널 저장 실패")

        invalid_create_body = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms",
            json=[],
        )
        if invalid_create_body.status_code != 400:
            failures.append("room inventory: 객체가 아닌 추가 요청을 400으로 거부하지 않음")

        invalid_create_status = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms",
            json={"room_label": "202호", "status": False},
        )
        if invalid_create_status.status_code != 400:
            failures.append("room inventory: false 상태값을 400으로 거부하지 않음")

        invalid_create_rent = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms",
            json={"room_label": "203호", "monthly_rent_krw": 0},
        )
        if invalid_create_rent.status_code != 400:
            failures.append("room inventory: 0원 월세를 400으로 거부하지 않음")
        invalid_create_deposit = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms",
            json={"room_label": "204호", "deposit_krw": 0},
        )
        if invalid_create_deposit.status_code != 400:
            failures.append("room inventory: 0원 보증금을 400으로 거부하지 않음")

        listed = client.get(f"/api/my/listing-requests/{listing_id}/rooms")
        listed_items = (listed.get_json() or {}).get("items") or []
        if listed.status_code != 200 or not any(
            item.get("id") == room_id and item.get("contract_end_date") == contract_end_date
            and item.get("deposit_krw") == 500
            and item.get("monthly_rent_krw") == 120
            and item.get("channel") == "장박가능" and item.get("floor") is None
            for item in listed_items
        ):
            failures.append("room inventory: 소유자 조회에서 보증금·기본 채널·계약만기일을 찾지 못함")

        channel_updated = client.put(
            f"/api/my/room-inventory/{room_id}",
            json={
                "room_label": "201호 복사본",
                "deposit_krw": 750,
                "monthly_rent_krw": 135,
                "status": "입실",
                "contract_end_date": contract_end_date,
                "floor": 2,
                "channel": "OTA전용",
            },
        )
        channel_item = (channel_updated.get_json() or {}).get("item") or {}
        if (channel_updated.status_code != 200 or channel_item.get("floor") != 2
                or channel_item.get("channel") != "OTA전용"
                or channel_item.get("room_label") != "201호 복사본"
                 or channel_item.get("deposit_krw") != 750
                or channel_item.get("monthly_rent_krw") != 135):
            failures.append("room inventory: 호실·보증금·월세·층 또는 OTA전용 채널 저장 실패")

        bulk_created = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms/bulk",
            json={"floor": 3, "room_count": 10},
        )
        bulk_payload = bulk_created.get_json() or {}
        bulk_labels = {
            item.get("room_label") for item in (bulk_payload.get("items") or [])
        }
        expected_labels = {str(room_no) for room_no in range(301, 311)}
        if (bulk_created.status_code != 201 or bulk_payload.get("created_count") != 10
                or bulk_payload.get("skipped_count") != 0
                or bulk_labels != expected_labels
                or any(item.get("floor") != 3 or item.get("channel") != "장박가능"
                       for item in (bulk_payload.get("items") or []))):
            failures.append("room inventory: 3층 10실(301~310) 벌크 생성 실패")

        bulk_duplicate = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms/bulk",
            json={"floor": 3, "room_count": 10},
        )
        duplicate_payload = bulk_duplicate.get_json() or {}
        if (bulk_duplicate.status_code != 201 or duplicate_payload.get("created_count") != 0
                or duplicate_payload.get("skipped_count") != 10):
            failures.append("room inventory: 벌크 생성 시 기존 호실을 건너뛰지 않음")

        invalid_bulk_floor = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms/bulk",
            json={"floor": "3", "room_count": 10},
        )
        invalid_bulk_count = client.post(
            f"/api/my/listing-requests/{listing_id}/rooms/bulk",
            json={"floor": 3, "room_count": 100},
        )
        if invalid_bulk_floor.status_code != 400 or invalid_bulk_count.status_code != 400:
            failures.append("room inventory: 잘못된 벌크 층 또는 방 개수를 400으로 거부하지 않음")

        invalid_date = client.put(
            f"/api/my/room-inventory/{room_id}",
            json={"status": "입실", "contract_end_date": "2026-99-99"},
        )
        if invalid_date.status_code != 400:
            failures.append("room inventory: 잘못된 날짜를 400으로 거부하지 않음")
        invalid_deposit = client.put(
            f"/api/my/room-inventory/{room_id}",
            json={"status": "입실", "deposit_krw": 0},
        )
        if invalid_deposit.status_code != 400:
            failures.append("room inventory: 0원 보증금 수정을 400으로 거부하지 않음")

        vacated = client.put(
            f"/api/my/room-inventory/{room_id}",
            json={"status": "공실", "contract_end_date": contract_end_date},
        )
        vacated_item = (vacated.get_json() or {}).get("item") or {}
        if (vacated.status_code != 200 or vacated_item.get("status") != "공실"
                or vacated_item.get("contract_end_date") is not None
                or vacated_item.get("deposit_krw") != 750
                or vacated_item.get("channel") != "OTA전용"
                or vacated_item.get("floor") != 2):
            failures.append("room inventory: 공실 전환 때 만기일만 초기화하고 보증금·채널·층을 보존하지 않음")

        invalid_status = client.put(
            f"/api/my/room-inventory/{room_id}",
            json={"status": "만기임박", "contract_end_date": contract_end_date},
        )
        if invalid_status.status_code != 400:
            failures.append("room inventory: 만기임박 상태를 거부하지 않음")

        invalid_update_body = client.put(f"/api/my/room-inventory/{room_id}", json=[])
        if invalid_update_body.status_code != 400:
            failures.append("room inventory: 객체가 아닌 수정 요청을 400으로 거부하지 않음")

        with client.session_transaction() as sess:
            sess["user_id"] = other_id
        blocked = client.put(
            f"/api/my/room-inventory/{room_id}",
            json={"status": "입실", "contract_end_date": contract_end_date},
        )
        if blocked.status_code != 403:
            failures.append("room inventory: 타 사용자의 방 재고 수정을 403으로 차단하지 않음")
        if not failures:
            print("OK  방 재고 보증금·층·채널·벌크 생성·만기일·소유자 권한")
    except Exception as exc:
        failures.append(f"room inventory 테스트 오류: {exc}")
    finally:
        try:
            if listing_id:
                cur.execute(
                    "DELETE FROM business_room_inventory WHERE listing_request_id=%s",
                    (listing_id,),
                )
                cur.execute("DELETE FROM listing_requests WHERE id=%s", (listing_id,))
            if other_id:
                cur.execute("DELETE FROM users WHERE id=%s", (other_id,))
            if owner_id:
                cur.execute("DELETE FROM users WHERE id=%s", (owner_id,))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_listing_hold_and_disclosure_controls(client):
    """매물 보류·보류해제·공개범위 전환의 소유권과 상태 전이를 검증한다."""
    import time as _time
    from db import get_conn

    failures = []
    run_id = str(int(_time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    owner_id = other_id = listing_id = None
    try:
        cur.execute("SELECT id FROM master_buildings ORDER BY id LIMIT 1")
        building = cur.fetchone()
        if not building:
            return ["listing controls: 테스트용 master_buildings 행이 없습니다."]
        cur.execute(
            "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
            (f"listing-owner-{run_id}@example.test", "매물 제어 소유자"),
        )
        owner_id = cur.fetchone()["id"]
        cur.execute(
            "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
            (f"listing-other-{run_id}@example.test", "매물 제어 타인"),
        )
        other_id = cur.fetchone()["id"]
        cur.execute("""
            INSERT INTO listing_requests
                (user_id, master_building_id, deal_type, desired_price, price_krw, contact_phone,
                 deal_mode, status, disclosure_scope, transaction_target)
            VALUES (%s, %s, '매매', '테스트 조건', 10000, %s, 'direct', 'submitted', 'limited', 'whole')
            RETURNING id
        """, (owner_id, building["id"], "010-0000-0000"))
        listing_id = cur.fetchone()["id"]
        conn.commit()

        with client.session_transaction() as sess:
            sess["user_id"] = other_id
        if client.post(f"/api/listing-requests/{listing_id}/hold").status_code != 403:
            failures.append("listing controls: 타 사용자의 보류 요청을 403으로 차단하지 않음")
        if client.patch(
            f"/api/listing-requests/{listing_id}/disclosure-scope",
            json={"disclosure_scope": "public"},
        ).status_code != 403:
            failures.append("listing controls: 타 사용자의 공개범위 변경을 403으로 차단하지 않음")

        with client.session_transaction() as sess:
            sess["user_id"] = owner_id
        mine = client.get("/api/listing-requests/mine")
        mine_items = (mine.get_json() or {}).get("items") or []
        mine_item = next((item for item in mine_items if item.get("id") == listing_id), {})
        if mine.status_code != 200 or mine_item.get("disclosure_scope") != "limited":
            failures.append("listing controls: 내 매물 목록에서 공개범위를 반환하지 않음")

        held = client.post(f"/api/listing-requests/{listing_id}/hold")
        held_item = (held.get_json() or {}).get("item") or {}
        if held.status_code != 200 or held_item.get("status") != "보류":
            failures.append("listing controls: 소유자 보류 상태 변경 실패")
        if client.post(f"/api/listing-requests/{listing_id}/hold").status_code != 400:
            failures.append("listing controls: 이미 보류인 매물을 다시 보류하지 않음")
        invalid_scope = client.patch(
            f"/api/listing-requests/{listing_id}/disclosure-scope",
            json={"disclosure_scope": "private"},
        )
        if invalid_scope.status_code != 400:
            failures.append("listing controls: 잘못된 공개범위를 400으로 거부하지 않음")

        edited = client.put(f"/api/listing-requests/{listing_id}", json={
            "deal_type": "매매",
            "desired_price": "수정 조건",
            "price_krw": 10000,
            "transaction_target": "whole",
            "registrant_type": "owner",
        })
        if edited.status_code != 200:
            failures.append("listing controls: 보류 매물 수정 실패")
        cur.execute("SELECT status FROM listing_requests WHERE id=%s", (listing_id,))
        if (cur.fetchone() or {}).get("status") != "submitted":
            failures.append("listing controls: 보류 매물 수정 후 접수됨으로 복원되지 않음")

        held_again = client.post(f"/api/listing-requests/{listing_id}/hold")
        resumed = client.post(f"/api/listing-requests/{listing_id}/resume")
        resumed_item = (resumed.get_json() or {}).get("item") or {}
        if (held_again.status_code != 200 or resumed.status_code != 200
                or resumed_item.get("status") != "submitted"):
            failures.append("listing controls: 보류해제 상태 변경 실패")

        scope_changed = client.patch(
            f"/api/listing-requests/{listing_id}/disclosure-scope",
            json={"disclosure_scope": "public"},
        )
        scope_item = (scope_changed.get_json() or {}).get("item") or {}
        if scope_changed.status_code != 200 or scope_item.get("disclosure_scope") != "public":
            failures.append("listing controls: 소유자 공개범위 전체공개 전환 실패")

        public_before_hold = client.get("/api/listings?limit=50")
        public_items = (public_before_hold.get_json() or {}).get("items") or []
        if not any(item.get("id") == listing_id for item in public_items):
            failures.append("listing controls: 전체공개 건물전체 매물이 공개 목록에 나타나지 않음")

        held_for_visibility = client.post(f"/api/listing-requests/{listing_id}/hold")
        public_while_held = client.get("/api/listings?limit=50")
        held_items = (public_while_held.get_json() or {}).get("items") or []
        with app.test_request_context(f"/building/{building['id']}?listing={listing_id}"):
            held_meta = _building_share_meta(building["id"], listing_id)
            building_meta = _building_share_meta(building["id"])
        if held_for_visibility.status_code != 200 or any(item.get("id") == listing_id for item in held_items):
            failures.append("listing controls: 보류 매물이 공개 목록에 계속 노출됨")
        if (held_meta.get("title"), held_meta.get("description")) != (
                building_meta.get("title"), building_meta.get("description")):
            failures.append("listing controls: 보류 매물이 공유 메타데이터에 노출됨")
        if client.post(f"/api/listing-requests/{listing_id}/resume").status_code != 200:
            failures.append("listing controls: 공개 노출 점검 후 보류해제 실패")

        unit_changed = client.put(f"/api/listing-requests/{listing_id}", json={
            "deal_type": "단기임대",
            "desired_price": "개별호실 변경",
            "transaction_target": "unit",
            "registrant_type": "owner",
        })
        unit_scope = client.patch(
            f"/api/listing-requests/{listing_id}/disclosure-scope",
            json={"disclosure_scope": "limited"},
        )
        if unit_changed.status_code != 200 or unit_scope.status_code != 400:
            failures.append("listing controls: 개별호실 공개범위 변경을 차단하지 않음")
        cur.execute(
            "SELECT action FROM listing_request_history WHERE listing_request_id=%s ORDER BY id",
            (listing_id,),
        )
        actions = [row["action"] for row in cur.fetchall()]
        if not {"held", "resumed", "edited", "scope_changed"} <= set(actions):
            failures.append("listing controls: 상태·공개범위 변경 이력이 남지 않음")
        if not failures:
            print("OK  매물 보류·보류해제·수정 복원·공개범위·소유자 권한")
    except Exception as exc:
        failures.append(f"listing controls 테스트 오류: {exc}")
    finally:
        try:
            if listing_id:
                cur.execute("DELETE FROM listing_request_history WHERE listing_request_id=%s", (listing_id,))
                cur.execute("DELETE FROM listing_requests WHERE id=%s", (listing_id,))
            if other_id:
                cur.execute("DELETE FROM users WHERE id=%s", (other_id,))
            if owner_id:
                cur.execute("DELETE FROM users WHERE id=%s", (owner_id,))
            conn.commit()
            with client.session_transaction() as sess:
                sess.pop("user_id", None)
        finally:
            cur.close()
            conn.close()
    return failures


def _check_room_expiry_alerts():
    """계약만기 임계치별 발송·중복 방지·이메일 없는 회원의 인앱 알림을 검증한다."""
    import os
    import time as _time
    from datetime import date, timedelta
    from types import SimpleNamespace
    from unittest.mock import patch

    from db import get_conn
    import email_util
    import sync_lodgings as sync_module

    failures = []
    run_id = str(int(_time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    owner_id = no_email_id = None
    listing_ids = []
    room_ids = []
    owner_room_ids = []
    try:
        cur.execute("SELECT id FROM master_buildings ORDER BY id LIMIT 1")
        building = cur.fetchone()
        if not building:
            return ["room expiry: 테스트용 master_buildings 행이 없습니다."]

        cur.execute(
            "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
            (f"expiry-owner-{run_id}@example.test", "만기 알림 소유자"),
        )
        owner_id = cur.fetchone()["id"]
        cur.execute(
            "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
            (None, "만기 알림 이메일 없음"),
        )
        no_email_id = cur.fetchone()["id"]

        for user_id in (owner_id, no_email_id):
            cur.execute("""
                INSERT INTO listing_requests
                    (user_id, master_building_id, deal_type, contact_phone,
                     deal_mode, registrant_type, status)
                VALUES (%s, %s, '월세', '010-0000-0000',
                        'direct', 'business', 'submitted')
                RETURNING id
            """, (user_id, building["id"]))
            listing_ids.append(cur.fetchone()["id"])

        today = date.today()
        room_specs = [
            (listing_ids[0], "0901", 90),
            (listing_ids[0], "0601", 60),
            (listing_ids[0], "0701", 7),
            (listing_ids[1], "0301", 30),
        ]
        for listing_id, label, threshold in room_specs:
            cur.execute("""
                INSERT INTO business_room_inventory
                    (listing_request_id, room_label, status, contract_end_date)
                VALUES (%s, %s, '입실', %s)
                RETURNING id
            """, (listing_id, label, today + timedelta(days=threshold)))
            room_id = cur.fetchone()["id"]
            room_ids.append(room_id)
            if listing_id == listing_ids[0]:
                owner_room_ids.append(room_id)
        conn.commit()

        with patch.object(
            sync_module, "send_email", return_value=(True, "테스트 발송 성공")
        ) as email_mock:
            first = sync_module.send_room_expiry_alerts(today=today)
        if (
            first.get("target_count") != 4
            or first.get("sent_count") != 4
            or first.get("email_sent_count") != 3
            or first.get("in_app_count") != 4
            or email_mock.call_count != 3
        ):
            failures.append(
                "room expiry: 90/60/30/7일 첫 실행의 이메일·인앱 발송 건수가 다름"
            )

        cur.execute("""
            SELECT threshold, COUNT(*) AS count
              FROM room_expiry_alerts_sent
             WHERE room_id = ANY(%s)
             GROUP BY threshold
             ORDER BY threshold
        """, (room_ids,))
        thresholds = {row["threshold"]: int(row["count"]) for row in cur.fetchall()}
        if thresholds != {"7": 1, "30": 1, "60": 1, "90": 1}:
            failures.append(f"room expiry: 임계치별 발송 이력이 잘못됨 ({thresholds})")

        cur.execute(
            "SELECT COUNT(*) AS count FROM notifications WHERE user_id=%s",
            (no_email_id,),
        )
        if int(cur.fetchone()["count"]) != 1:
            failures.append("room expiry: 이메일 없는 회원의 인앱 알림이 1건 생성되지 않음")

        with patch.object(
            sync_module, "send_email", return_value=(True, "중복 테스트 발송")
        ) as duplicate_email_mock:
            second = sync_module.send_room_expiry_alerts(today=today)
        if (
            second.get("target_count") != 0
            or second.get("sent_count") != 0
            or duplicate_email_mock.call_count != 0
        ):
            failures.append("room expiry: 같은 배치 재실행에서 중복 발송됨")

        # 이메일 실패도 사이트 알림은 남기고, 다음 배치에서 이메일만 재시도해야 한다.
        cur.execute("""
            INSERT INTO business_room_inventory
                (listing_request_id, room_label, status, contract_end_date)
            VALUES (%s, '0302', '입실', %s)
            RETURNING id
        """, (listing_ids[0], today + timedelta(days=30)))
        failed_email_room_id = cur.fetchone()["id"]
        room_ids.append(failed_email_room_id)
        owner_room_ids.append(failed_email_room_id)
        conn.commit()
        with patch.object(
            sync_module, "send_email", return_value=(False, "계획된 이메일 실패")
        ) as failed_email_mock:
            failed_email_run = sync_module.send_room_expiry_alerts(today=today)
        cur.execute(
            "SELECT COUNT(*) AS count FROM notifications WHERE user_id=%s",
            (owner_id,),
        )
        owner_notification_count = int(cur.fetchone()["count"])
        cur.execute(
            "SELECT email_state FROM room_expiry_alerts_sent WHERE room_id=%s",
            (failed_email_room_id,),
        )
        failed_email_state = (cur.fetchone() or {}).get("email_state")
        if (
            failed_email_run.get("target_count") != 1
            or failed_email_run.get("in_app_count") != 1
            or failed_email_mock.call_count != 1
            or owner_notification_count != 4
            or failed_email_state != "failed"
        ):
            failures.append("room expiry: 이메일 실패 때 인앱 알림 또는 재시도 상태를 남기지 않음")

        with patch.object(
            sync_module, "send_email", return_value=(True, "재시도 성공")
        ) as retry_email_mock:
            cur.execute("""
                UPDATE room_expiry_alerts_sent
                   SET email_attempted_at = NOW() - INTERVAL '24 hours'
                 WHERE room_id=%s
            """, (failed_email_room_id,))
            conn.commit()
            retry_email_run = sync_module.send_room_expiry_alerts(today=today)
        if (
            retry_email_run.get("target_count") != 1
            or retry_email_run.get("in_app_count") != 0
            or retry_email_run.get("email_sent_count") != 1
            or retry_email_mock.call_count != 1
        ):
            failures.append("room expiry: 이메일 실패 재시도에서 인앱 중복 또는 이메일 재시도 실패")

        # 이메일 성공 뒤 상태 저장이 끊긴 상황(attempting)은 같은 멱등 키로 회복한다.
        cur.execute("""
            INSERT INTO business_room_inventory
                (listing_request_id, room_label, status, contract_end_date)
            VALUES (%s, '0303', '입실', %s)
            RETURNING id
        """, (listing_ids[0], today + timedelta(days=30)))
        stuck_room_id = cur.fetchone()["id"]
        room_ids.append(stuck_room_id)
        owner_room_ids.append(stuck_room_id)
        cur.execute("""
            INSERT INTO notifications (user_id, title, body, building_name)
            VALUES (%s, '기존 인앱 알림', '테스트', '테스트 건물')
            RETURNING id
        """, (owner_id,))
        stuck_notification_id = cur.fetchone()["id"]
        cur.execute("""
            INSERT INTO room_expiry_alerts_sent
                (room_id, threshold, notification_id, in_app_sent_at,
                 email_state, email_attempted_at)
            VALUES (%s, '30', %s, NOW(), 'attempting', NOW())
        """, (stuck_room_id, stuck_notification_id))
        conn.commit()
        with patch.object(
            sync_module, "send_email", return_value=(True, "중복 방지 확인")
        ) as stuck_email_mock:
            stuck_run = sync_module.send_room_expiry_alerts(today=today)
        stuck_key = (stuck_email_mock.call_args.kwargs.get("idempotency_key")
                     if stuck_email_mock.call_args else None)
        if (
            stuck_run.get("target_count") != 1
            or stuck_run.get("email_sent_count") != 1
            or stuck_email_mock.call_count != 1
            or not (stuck_key or "").startswith("room-expiry/")
        ):
            failures.append("room expiry: attempting 상태 이메일의 멱등 재시도 실패")

        with patch.object(
            sync_module, "send_email", return_value=(True, "이미 발송됨")
        ) as recovered_duplicate_mock:
            recovered_duplicate_run = sync_module.send_room_expiry_alerts(today=today)
        if (
            recovered_duplicate_run.get("target_count") != 0
            or recovered_duplicate_mock.call_count != 0
        ):
            failures.append("room expiry: 멱등 복구 완료 뒤 이메일을 다시 발송함")

        # Resend 5xx처럼 수락 여부가 불확실한 실패는 attempting으로 보존하고,
        # 24시간 멱등 보관 창을 지난 뒤에는 자동 재발송하지 않는다.
        cur.execute("""
            INSERT INTO business_room_inventory
                (listing_request_id, room_label, status, contract_end_date)
            VALUES (%s, '0304', '입실', %s)
            RETURNING id
        """, (listing_ids[0], today + timedelta(days=30)))
        stale_room_id = cur.fetchone()["id"]
        room_ids.append(stale_room_id)
        owner_room_ids.append(stale_room_id)
        conn.commit()
        with patch.object(
            sync_module,
            "send_email",
            return_value=(False, "Resend 발송 실패(503): temporary", "transport_error"),
        ) as transient_email_mock:
            transient_run = sync_module.send_room_expiry_alerts(today=today)
        cur.execute(
            "SELECT email_state FROM room_expiry_alerts_sent WHERE room_id=%s",
            (stale_room_id,),
        )
        transient_state = (cur.fetchone() or {}).get("email_state")
        if (
            transient_run.get("target_count") != 1
            or transient_run.get("in_app_count") != 1
            or transient_email_mock.call_count != 1
            or transient_state != "attempting"
        ):
            failures.append("room expiry: Resend 5xx를 불확실한 발송 상태로 보존하지 않음")

        cur.execute("""
            UPDATE room_expiry_alerts_sent
               SET email_attempted_at = NOW() - INTERVAL '24 hours'
             WHERE room_id=%s
        """, (stale_room_id,))
        conn.commit()
        with patch.object(
            sync_module, "send_email", return_value=(True, "오래된 시도")
        ) as stale_email_mock:
            stale_run = sync_module.send_room_expiry_alerts(today=today)
        if stale_run.get("target_count") != 0 or stale_email_mock.call_count != 0:
            failures.append("room expiry: 24시간 지난 불확실한 이메일을 자동 재발송함")

        # Resend REST 호출에도 alert ID 기반 멱등 키 헤더가 실제로 전달되어야 한다.
        with patch.dict(
            os.environ,
            {"RESEND_API_KEY": "test-key", "RESEND_FROM_EMAIL": "test@example.test"},
            clear=False,
        ), patch.object(
            email_util.requests,
            "post",
            return_value=SimpleNamespace(status_code=200),
        ) as resend_post:
            email_ok, _ = email_util.send_email(
                "recipient@example.test",
                "멱등 키 테스트",
                "<p>test</p>",
                idempotency_key="room-expiry/test-alert",
            )
        sent_headers = resend_post.call_args.kwargs.get("headers") if resend_post.call_args else {}
        if not email_ok or sent_headers.get("Idempotency-Key") != "room-expiry/test-alert":
            failures.append("room expiry: Resend Idempotency-Key 헤더를 전달하지 않음")

        with patch.dict(
            os.environ,
            {"RESEND_API_KEY": "test-key", "RESEND_FROM_EMAIL": "test@example.test"},
            clear=False,
        ), patch.object(
            email_util.requests,
            "post",
            return_value=SimpleNamespace(
                status_code=409,
                text="idempotency conflict",
                json=lambda: {"message": "idempotency conflict"},
            ),
        ):
            _, _, conflict_outcome = email_util.send_email(
                "recipient@example.test",
                "멱등 충돌 테스트",
                "<p>test</p>",
                idempotency_key="room-expiry/conflict-alert",
                detailed=True,
            )
        if conflict_outcome != "transport_error":
            failures.append("room expiry: Resend 409 응답을 불확실한 전송으로 처리하지 않음")

        # 알림 이력이 생긴 방이 있는 매물도 철회(삭제)할 수 있어야 한다.
        cur.execute("DELETE FROM listing_requests WHERE id=%s", (listing_ids[0],))
        conn.commit()
        cur.execute(
            "SELECT COUNT(*) AS count FROM room_expiry_alerts_sent WHERE room_id = ANY(%s)",
            (owner_room_ids,),
        )
        if int(cur.fetchone()["count"]) != 0:
            failures.append("room expiry: 알림 이력이 매물 삭제를 막거나 함께 삭제되지 않음")

        if not failures:
            print("OK  계약만기 90·60·30·7일 이메일·인앱 알림·실패 재시도·중복 방지")
    except Exception as exc:
        conn.rollback()
        failures.append(f"room expiry 테스트 오류: {exc}")
    finally:
        try:
            if room_ids:
                cur.execute(
                    "DELETE FROM room_expiry_alerts_sent WHERE room_id = ANY(%s)",
                    (room_ids,),
                )
            if listing_ids:
                cur.execute(
                    "DELETE FROM business_room_inventory WHERE listing_request_id = ANY(%s)",
                    (listing_ids,),
                )
                cur.execute(
                    "DELETE FROM listing_requests WHERE id = ANY(%s)",
                    (listing_ids,),
                )
            if owner_id:
                cur.execute("DELETE FROM users WHERE id=%s", (owner_id,))
            if no_email_id:
                cur.execute("DELETE FROM users WHERE id=%s", (no_email_id,))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_public_business_listing_summary(client):
    """사업주 공개 장기방은 공실 장박가능 재고의 가격만, 호실수 없이 노출되는지 확인."""
    import time as _time
    from db import get_conn

    failures = []
    run_id = str(int(_time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    user_id = None
    listing_ids = []
    try:
        cur.execute("SELECT id FROM master_buildings ORDER BY id LIMIT 1")
        building = cur.fetchone()
        if not building:
            return ["business public summary: 테스트용 master_buildings 행이 없습니다."]

        cur.execute(
            "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
            (f"business-public-{run_id}@example.test", "사업주 공개 요약 테스트"),
        )
        user_id = cur.fetchone()["id"]

        def add_listing(registrant_type, price_min, price_max, room_count):
            cur.execute(
                """
                INSERT INTO listing_requests
                    (user_id, master_building_id, deal_type, price_krw, price_krw_max,
                     monthly_rent_krw, room_count, contact_phone, verified_phone,
                     deal_mode, registrant_type, status)
                VALUES (%s, %s, '월세', %s, %s, %s, %s, '010-0000-0000', '01000000000',
                        'direct', %s, 'submitted')
                RETURNING id
                """,
                (
                    user_id, building["id"], price_min, price_max, price_min,
                    room_count, registrant_type,
                ),
            )
            listing_id = cur.fetchone()["id"]
            listing_ids.append(listing_id)
            return listing_id

        available_id = add_listing("business", 90, 120, 88)
        fixed_price_id = add_listing("business", 70, None, 77)
        unavailable_id = add_listing("business", 50, 60, 66)
        no_inventory_id = add_listing("business", 40, 50, 55)
        owner_id = add_listing("building_owner", 30, None, 3)
        cur.execute(
            """
            INSERT INTO business_room_inventory
                (listing_request_id, room_label, status, channel)
            VALUES
                (%s, '101', '공실', '장박가능'),
                (%s, '102', '공실', 'OTA전용'),
                (%s, '103', '공실', '장박가능'),
                (%s, '201', '입실', '장박가능'),
                (%s, '202', '공실', 'OTA전용')
            """,
            (available_id, available_id, fixed_price_id, unavailable_id, unavailable_id),
        )
        conn.commit()

        listed = client.get("/api/listings?limit=50")
        public_items = {
            item.get("id"): item for item in ((listed.get_json() or {}).get("items") or [])
        }
        expected_ids = {available_id, fixed_price_id, unavailable_id, no_inventory_id, owner_id}
        if listed.status_code != 200 or not expected_ids.issubset(public_items):
            failures.append("business public summary: 공개 목록에서 테스트 매물을 찾지 못함")
            return failures

        available = public_items[available_id]
        fixed_price = public_items[fixed_price_id]
        unavailable = public_items[unavailable_id]
        no_inventory = public_items[no_inventory_id]
        owner = public_items[owner_id]
        for label, item, expected_min, expected_max in (
            ("공실 장박가능", available, 90, 120),
            ("단일 가격 상속", fixed_price, 70, 70),
        ):
            if (
                item.get("is_business_listing") is not True
                or item.get("room_price_min") != expected_min
                or item.get("room_price_max") != expected_max
                or "room_count" in item
            ):
                failures.append(
                    f"business public summary: {label} 가격 범위 또는 호실수 비공개가 잘못됨"
                )
        for label, item in (("만실·OTA전용", unavailable), ("재고 없음", no_inventory)):
            if (
                item.get("room_price_min") is not None
                or item.get("room_price_max") is not None
                or "room_count" in item
            ):
                failures.append(
                    f"business public summary: {label} 사업주 매물에 가격 또는 호실수가 노출됨"
                )
        if (
            owner.get("room_count") != 3
            or "is_business_listing" in owner
            or "room_price_min" in owner
            or "room_price_max" in owner
        ):
            failures.append("business public summary: 건물주 매물의 기존 호실수·가격 표시가 바뀜")

        detail = client.get(f"/api/building/{building['id']}")
        detail_items = {
            item.get("id"): item
            for item in ((detail.get_json() or {}).get("direct_listings") or [])
        }
        detail_available = detail_items.get(available_id) or {}
        detail_unavailable = detail_items.get(unavailable_id) or {}
        detail_owner = detail_items.get(owner_id) or {}
        if (
            detail.status_code != 200
            or detail_available.get("room_price_min") != 90
            or detail_available.get("room_price_max") != 120
            or "room_count" in detail_available
            or detail_unavailable.get("room_price_min") is not None
            or detail_unavailable.get("room_price_max") is not None
            or "room_count" in detail_unavailable
            or detail_owner.get("room_count") != 3
        ):
            failures.append("business public summary: 건물상세 공개 요약 또는 기존 호실수 표시가 잘못됨")

        if not failures:
            print("OK  사업주 장기방 공개 가격범위·문의 안내·호실수 비노출 및 기존 유형 보존")
    except Exception as exc:
        conn.rollback()
        failures.append(f"business public summary 테스트 오류: {exc}")
    finally:
        try:
            if listing_ids:
                cur.execute(
                    "DELETE FROM business_room_inventory WHERE listing_request_id = ANY(%s)",
                    (listing_ids,),
                )
                cur.execute("DELETE FROM listing_requests WHERE id = ANY(%s)", (listing_ids,))
            if user_id:
                cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_business_listing_verification(client):
    """사업주 신고번호 인증 캐시, 사용자 분리, 미매칭 폴백과 우회 차단을 검증한다."""
    import time as _time
    from unittest.mock import patch
    from db import get_conn

    failures = []
    run_id = str(int(_time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    user_one_id = user_two_id = None
    listing_ids = []
    try:
        cur.execute("SELECT id FROM master_buildings ORDER BY id LIMIT 1")
        building = cur.fetchone()
        if not building:
            return ["business verification: 테스트용 master_buildings 행이 없습니다."]
        cur.execute("""
            INSERT INTO users (email, name, phone, phone_verified)
            VALUES (%s, %s, '01000000000', TRUE)
            RETURNING id
        """, (f"business-verify-one-{run_id}@example.test", "사업주 인증 사용자1"))
        user_one_id = cur.fetchone()["id"]
        cur.execute("""
            INSERT INTO users (email, name, phone, phone_verified)
            VALUES (%s, %s, '01000000001', TRUE)
            RETURNING id
        """, (f"business-verify-two-{run_id}@example.test", "사업주 인증 사용자2"))
        user_two_id = cur.fetchone()["id"]
        conn.commit()

        representative = {
            "biz_name": "테스트 대표 숙박업소",
            "permit_number": "2026-01-12345",
            "room_count": 12,
        }
        listing_body = {
            "master_building_id": building["id"],
            "deal_type": "단기임대",
            "deal_mode": "direct",
            "registrant_type": "business",
        }
        with patch("app.matched_lodgings", return_value=([representative], "road")):
            with client.session_transaction() as sess:
                sess.clear()
                sess["user_id"] = user_one_id

            status_before = client.get(
                f"/api/building/{building['id']}/business-verification"
            )
            before_data = status_before.get_json() or {}
            if (
                status_before.status_code != 200
                or not before_data.get("matched")
                or before_data.get("verified")
                or not before_data.get("requires_permit_verification")
            ):
                failures.append("business verification: 최초 인증 상태 조회가 잘못됨")

            bypassed = client.post("/api/listing-requests", json=listing_body)
            if bypassed.status_code != 403 or not (bypassed.get_json() or {}).get("requires_business_verification"):
                failures.append("business verification: 인증 캐시 없는 사업주 등록 우회를 차단하지 못함")

            wrong = client.post(
                f"/api/building/{building['id']}/business-verification",
                json={"permit_number": "2026-01-99999"},
            )
            if wrong.status_code != 400 or not (wrong.get_json() or {}).get("retryable"):
                failures.append("business verification: 틀린 신고번호를 재시도 가능 오류로 처리하지 못함")

            verified = client.post(
                f"/api/building/{building['id']}/business-verification",
                json={"permit_number": "2026 - 01 12345"},
            )
            if verified.status_code != 200 or not (verified.get_json() or {}).get("verified"):
                failures.append("business verification: 하이픈·공백을 무시한 신고번호 인증 실패")

            created = client.post("/api/listing-requests", json=listing_body)
            created_data = created.get_json() or {}
            if created.status_code != 200 or not created_data.get("id"):
                failures.append("business verification: 인증 후 사업주 등록 실패")
            elif created_data.get("id"):
                verified_listing_id = created_data["id"]
                listing_ids.append(verified_listing_id)
                cur.execute(
                    "SELECT matched_permit_number FROM listing_requests WHERE id=%s",
                    (verified_listing_id,),
                )
                stored_permit = (cur.fetchone() or {}).get("matched_permit_number")
                public_photo_key = f"listing_photos/{verified_listing_id}/{'a' * 32}.jpg"
                private_photo_key = f"listing_photos/{verified_listing_id}/{'b' * 32}.jpg"
                cur.execute("""
                    INSERT INTO listing_photos (listing_request_id, image_key, sort_order, is_public)
                    VALUES (%s, %s, 0, TRUE), (%s, %s, 1, FALSE)
                """, (
                    verified_listing_id, public_photo_key,
                    verified_listing_id, private_photo_key,
                ))
                conn.commit()
                cur.execute(
                    "SELECT id, image_key FROM listing_photos WHERE listing_request_id=%s",
                    (verified_listing_id,),
                )
                photo_ids_by_key = {row["image_key"]: row["id"] for row in cur.fetchall()}
                with client.session_transaction() as sess:
                    sess.clear()
                public_items = (client.get("/api/listings?disclosure_scope=public").get_json() or {}).get("items") or []
                public_listing = next((item for item in public_items if item.get("id") == verified_listing_id), {})
                if (
                    stored_permit != "20260112345"
                    or public_listing.get("permit_number_masked") != "••••12345"
                    or "matched_permit_number" in public_listing
                    or len(public_listing.get("photos") or []) != 1
                    or public_photo_key not in str(public_listing.get("photo_url") or "")
                ):
                    failures.append("business verification: 신고번호 복사·마스킹 또는 사진별 공개 필터가 잘못됨")
                public_path = f"/api/listing-photos/img/{public_photo_key}"
                private_path = f"/api/listing-photos/img/{private_photo_key}"
                with patch("app.storage_util.download_bytes", return_value=b"test-image"):
                    public_image = client.get(public_path)
                    private_unauthenticated = client.get(private_path)
                    with client.session_transaction() as sess:
                        sess["user_id"] = user_one_id
                    visibility_changed = client.put(
                        f"/api/listing-requests/{verified_listing_id}/photos/order",
                        json={
                            "photo_ids": [
                                photo_ids_by_key[public_photo_key],
                                photo_ids_by_key[private_photo_key],
                            ],
                            "photo_public": {
                                str(photo_ids_by_key[public_photo_key]): False,
                                str(photo_ids_by_key[private_photo_key]): False,
                            },
                        },
                    )
                    with client.session_transaction() as sess:
                        sess.clear()
                    public_after_private = client.get(public_path)
                    with client.session_transaction() as sess:
                        sess["user_id"] = user_one_id
                    private_owner = client.get(private_path)
                if (
                    public_image.status_code != 200
                    or public_image.headers.get("Cache-Control") != "private, no-store"
                    or private_unauthenticated.status_code != 404
                    or visibility_changed.status_code != 200
                    or public_after_private.status_code != 404
                    or private_owner.status_code != 200
                    or private_owner.headers.get("Cache-Control") != "private, no-store"
                ):
                    failures.append("business verification: 비공개 전환 후 사진 접근 권한 또는 공유 캐시 금지가 잘못됨")

            with client.session_transaction() as sess:
                sess.clear()
                sess["user_id"] = user_two_id
            other_user_status = client.get(
                f"/api/building/{building['id']}/business-verification"
            )
            other_data = other_user_status.get_json() or {}
            if other_user_status.status_code != 200 or other_data.get("verified"):
                failures.append("business verification: 다른 사용자에게 인증 캐시가 공유됨")

            with client.session_transaction() as sess:
                sess.clear()
                sess["user_id"] = user_one_id
            with patch("app.matched_lodgings", return_value=([], None)):
                unmatched_status = client.get(
                    f"/api/building/{building['id']}/business-verification"
                )
                unmatched_data = unmatched_status.get_json() or {}
                unmatched_listing = client.post("/api/listing-requests", json=listing_body)
            if (
                unmatched_status.status_code != 200
                or unmatched_data.get("matched")
                or not unmatched_data.get("verified")
                or unmatched_listing.status_code != 200
            ):
                failures.append("business verification: 영업신고 미매칭 건물의 휴대폰 인증 폴백 실패")
            elif (unmatched_listing.get_json() or {}).get("id"):
                listing_ids.append(unmatched_listing.get_json()["id"])

        if not failures:
            print("OK  사업주 신고번호 인증·숫자 정규화·사용자 분리·미매칭 폴백·서버 우회 차단")
    except Exception as exc:
        failures.append(f"business verification 테스트 오류: {exc}")
    finally:
        try:
            if listing_ids:
                cur.execute(
                    "DELETE FROM listing_request_history WHERE listing_request_id = ANY(%s)",
                    (listing_ids,),
                )
                cur.execute("DELETE FROM listing_requests WHERE id = ANY(%s)", (listing_ids,))
            if user_one_id or user_two_id:
                cur.execute(
                    "DELETE FROM business_building_verifications WHERE user_id = ANY(%s)",
                    ([x for x in (user_one_id, user_two_id) if x],),
                )
                cur.execute("DELETE FROM users WHERE id = ANY(%s)", ([x for x in (user_one_id, user_two_id) if x],))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_listing_registrant_types(client):
    """신규 3분류 등록자유형 저장과 과거 agent 값 수정 호환성을 검증한다."""
    import time as _time
    from unittest.mock import patch
    from db import get_conn

    failures = []
    run_id = str(int(_time.time() * 1000))
    conn = get_conn()
    cur = conn.cursor()
    user_id = listing_id = None
    try:
        cur.execute("SELECT id FROM master_buildings ORDER BY id LIMIT 1")
        building = cur.fetchone()
        if not building:
            return ["listing registrant type: 테스트용 master_buildings 행이 없습니다."]
        cur.execute(
            """
            INSERT INTO users (email, name, phone, phone_verified)
            VALUES (%s, %s, '01000000000', TRUE)
            RETURNING id
            """,
            (f"registrant-type-{run_id}@example.test", "등록자유형 테스트"),
        )
        user_id = cur.fetchone()["id"]
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["user_id"] = user_id

        created = client.post("/api/listing-requests", json={
            "master_building_id": building["id"],
            "deal_type": "단기임대",
            "deal_mode": "direct",
            "registrant_type": "building_owner",
        })
        created_item = created.get_json() or {}
        listing_id = created_item.get("id")
        if created.status_code != 200 or not listing_id:
            failures.append(
                f"listing registrant type: building_owner 등록 실패 (HTTP {created.status_code})"
            )
            return failures

        def update_without_lodging_match(payload):
            # 이 테스트는 등록자유형 값 호환만 검증한다. 신고번호 인증 차단은
            # _check_business_listing_verification에서 별도로 검증한다.
            with patch("app.matched_lodgings", return_value=([], None)):
                return client.put(f"/api/listing-requests/{listing_id}", json=payload)

        business_update = update_without_lodging_match({
                "deal_type": "월세",
                "registrant_type": "business",
                "price_krw": 3000,
                "price_krw_max": 4500,
                "room_count": 37,
            })
        cur.execute(
            """SELECT registrant_type, price_krw, price_krw_max, room_count
               FROM listing_requests WHERE id=%s""",
            (listing_id,),
        )
        business_stored = cur.fetchone() or {}
        if (
            business_update.status_code != 200
            or business_stored.get("registrant_type") != "business"
            or business_stored.get("price_krw") != 3000
            or business_stored.get("price_krw_max") != 4500
            or business_stored.get("room_count") != 37
        ):
            failures.append("business listing: 가격범위 또는 총 호실수 저장 실패")

        invalid_range = update_without_lodging_match({
                "deal_type": "월세",
                "registrant_type": "business",
                "price_krw": 4500,
                "price_krw_max": 3000,
                "room_count": 37,
            })
        if invalid_range.status_code != 400:
            failures.append("business listing: 최고가가 최저가보다 작은 가격범위를 차단하지 않음")

        sale_range = update_without_lodging_match({
                "deal_type": "매매",
                "registrant_type": "business",
                "price_krw": 4500,
                "price_krw_max": 5000,
                "room_count": 37,
            })
        if sale_range.status_code != 400:
            failures.append("business listing: 매매 가격범위를 차단하지 않음")

        short_update = update_without_lodging_match({
                "deal_type": "단기임대",
                "registrant_type": "business",
                "price_krw": 30,
                "price_krw_max": 60,
                "room_count": 37,
            })
        cur.execute(
            "SELECT deal_type, price_krw, price_krw_max FROM listing_requests WHERE id=%s",
            (listing_id,),
        )
        short_stored = cur.fetchone() or {}
        if (
            short_update.status_code != 200
            or short_stored.get("deal_type") != "단기임대"
            or short_stored.get("price_krw") != 30
            or short_stored.get("price_krw_max") != 60
        ):
            failures.append("business listing: 단기임대 가격범위 저장 실패")

        with patch("app.matched_lodgings", return_value=(
            [{"room_count": 37, "biz_name": "테스트 숙박업소"}], "road"
        )), patch("app.choose_representative", return_value={
            "room_count": 37, "biz_name": "테스트 숙박업소"
        }):
            summary = client.get(f"/api/building/{building['id']}/lodging-summary")
        summary_data = summary.get_json() or {}
        if (
            summary.status_code != 200
            or summary_data.get("room_count") != 37
            or summary_data.get("business_name") != "테스트 숙박업소"
        ):
            failures.append("business listing: 대표 숙박업 객실수 자동채움 API 실패")

        for value in ("business", "agent"):
            payload = {"deal_type": "단기임대", "registrant_type": value}
            updated = update_without_lodging_match(payload) if value == "business" else client.put(
                f"/api/listing-requests/{listing_id}", json=payload
            )
            payload = updated.get_json() or {}
            cur.execute(
                "SELECT registrant_type FROM listing_requests WHERE id=%s",
                (listing_id,),
            )
            stored = cur.fetchone() or {}
            if (
                updated.status_code != 200
                or not payload.get("ok")
                or stored.get("registrant_type") != value
            ):
                failures.append(
                    f"listing registrant type: {value} 수정 호환 실패 (HTTP {updated.status_code})"
                )
                break

        mine = client.get("/api/listing-requests/mine")
        items = (mine.get_json() or {}).get("items") or []
        if mine.status_code != 200 or not any(
            item.get("id") == listing_id and item.get("registrant_type") == "agent"
            for item in items
        ):
            failures.append("listing registrant type: 과거 agent 매물의뢰를 마이페이지 조회에서 찾지 못함")
        if not failures:
            print("OK  사업주 가격범위·총 호실수·대표객실수 API 및 과거 agent 수정·조회 호환")
    except Exception as exc:
        failures.append(f"listing registrant type 테스트 오류: {exc}")
    finally:
        try:
            if listing_id:
                cur.execute(
                    "DELETE FROM listing_request_history WHERE listing_request_id=%s",
                    (listing_id,),
                )
                cur.execute("DELETE FROM listing_requests WHERE id=%s", (listing_id,))
            if user_id:
                cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_subway_station_import_headers():
    """표준데이터의 역사위도/역사경도 헤더가 importer에서 누락되지 않게 고정한다."""
    try:
        from import_subway_stations import _find_columns, _normalise_rows
        headers = ("역번호", "역사명", "노선명", "역사위도", "역사경도")
        rows = _normalise_rows(
            [("T001", "테스트역", "테스트선", "37.5001", "127.0001")],
            _find_columns(headers),
        )
        if rows != [("테스트역", "테스트선", 37.5001, 127.0001)]:
            return ["subway importer: 전국 표준데이터 역사위도·역사경도 헤더를 해석하지 못함"]
        print("OK  전국 도시철도역사정보 표준 헤더·좌표 importer")
        return []
    except Exception as exc:
        return [f"subway importer 테스트 오류: {exc}"]


def _check_platform_summary(client):
    """플랫폼 지표 4개를 기본 COUNT 쿼리로 대조한다."""
    from app import _matched_lodging_registry_count
    from db import get_conn

    failures = []
    conn = get_conn()
    cur = conn.cursor()
    try:
        response = client.get("/api/stats/platform-summary")
        payload = response.get_json() or {}
        cur.execute("SELECT COUNT(*) AS c FROM master_buildings")
        building_count = int(cur.fetchone()["c"])
        cur.execute("SELECT COUNT(*) AS c FROM transactions")
        transaction_count = int(cur.fetchone()["c"])
        cur.execute("""
            SELECT COUNT(*) AS c FROM listing_requests
            WHERE status = '접수됨' AND disclosure_scope = 'public'
        """)
        listing_count = int(cur.fetchone()["c"])
        biz_count = _matched_lodging_registry_count(cur)
        expected = {
            "building_count": building_count,
            "biz_count": biz_count,
            "transaction_count": transaction_count,
            "listing_count": listing_count,
        }
        if response.status_code != 200 or any(payload.get(k) != v for k, v in expected.items()):
            failures.append(f"platform summary: 실제 DB COUNT와 응답 불일치 ({payload}, {expected})")
        else:
            print("OK  홈 데이터 규모 4개 지표·실제 DB COUNT 대조")
    except Exception as exc:
        failures.append(f"platform summary 테스트 오류: {exc}")
    finally:
        cur.close()
        conn.close()
    return failures


def _check_datalab_stats(client):
    """데이터랩 ②~⑥의 정렬·토글·표본 제외·공통 집계 기준을 점검한다."""
    failures = []
    try:
        ranked_items = []
        for direction, sign in (("up", 1), ("down", -1)):
            response = client.get(f"/api/stats/price-change-top?direction={direction}")
            payload = response.get_json() or {}
            items = payload.get("items") or []
            ranked_items.extend(items)
            if response.status_code != 200 or payload.get("direction") != direction:
                failures.append(f"데이터랩 가격변동 {direction}: 응답 실패")
                continue
            changes = [float(item.get("change_percent") or 0) for item in items]
            if any(item.get("transaction_count", 0) < 2 for item in items):
                failures.append(f"데이터랩 가격변동 {direction}: 2건 미만 거래 건물이 포함됨")
            if any(
                item.get("area_sqm") is None or float(item.get("area_sqm") or 0) <= 0
                for item in items
            ):
                failures.append(f"데이터랩 가격변동 {direction}: 동일 전용면적(area_sqm) 정보가 없음")
            group_keys = [
                (item.get("building_name"), item.get("address"), item.get("area_sqm"))
                for item in items
            ]
            if len(group_keys) != len(set(group_keys)):
                failures.append(f"데이터랩 가격변동 {direction}: 동일 건물·주소·전용면적 그룹이 중복됨")
            if (sign > 0 and any(value <= 0 for value in changes)) or (
                sign < 0 and any(value >= 0 for value in changes)
            ):
                failures.append(f"데이터랩 가격변동 {direction}: 상승/하락 방향이 잘못됨")
            ordered = sorted(changes, reverse=(sign > 0))
            if changes != ordered:
                failures.append(f"데이터랩 가격변동 {direction}: 변동률 정렬이 잘못됨")

        for order, reverse, label in (("highest", True, "최고가"), ("lowest", False, "최저가")):
            highest = client.get(f"/api/stats/highest-price-top?order={order}").get_json() or {}
            prices = [int(item.get("price") or 0) for item in highest.get("items") or []]
            ranked_items.extend(highest.get("items") or [])
            if (
                highest.get("order") != order
                or any(price <= 0 for price in prices)
                or prices != sorted(prices, reverse=reverse)
            ):
                failures.append(f"데이터랩 {label}: 양수 가격 정렬 TOP이 아님")

        ranking = client.get("/api/ranking").get_json() or {}
        ranked_items.extend(ranking.get("price_highs") or [])
        ranked_items.extend(ranking.get("most_traded") or [])

        for item in ranked_items:
            has_lat = item.get("lat") is not None
            has_lng = item.get("lng") is not None
            if has_lat != has_lng:
                failures.append("데이터랩 랭킹: 지도 좌표가 위도·경도 쌍으로 반환되지 않음")
                break
            if has_lat and (
                not -90 <= float(item["lat"]) <= 90
                or not -180 <= float(item["lng"]) <= 180
            ):
                failures.append("데이터랩 랭킹: 지도 좌표 범위가 올바르지 않음")
                break

        closure = client.get("/api/stats/closure-rate-by-region").get_json() or {}
        for item in closure.get("items") or []:
            total = int(item.get("total_count") or 0)
            closed = int(item.get("closed_count") or 0)
            expected_rate = round(closed / total * 100, 1) if total else None
            if total < 5 or closed > total or item.get("closure_rate") != expected_rate:
                failures.append("데이터랩 폐업 지역: 표본 5건 제외 또는 폐업률 계산이 잘못됨")
                break

        consign = client.get("/api/stats/consign-by-sido").get_json() or {}
        expected_consign_items, expected_consign_total = expected_consign_by_sido()
        if (
            consign.get("items") != expected_consign_items
            or consign.get("total") != expected_consign_total
        ):
            failures.append(
                "데이터랩 영업신고현황: 생활·주소매칭·폐업제외·신고번호 중복 제거 집계가 원본과 다름"
            )
        canonical_sidos = [item.get("sido") for item in consign.get("items") or []]
        if (
            len(canonical_sidos) != len(set(canonical_sidos))
            or any(sido in {"서울", "광주", "울산", "전남"} for sido in canonical_sidos)
        ):
            failures.append("데이터랩 영업신고현황: 시도 표기가 공식 명칭으로 통합되지 않음")
        for item in consign.get("items") or []:
            units = int(item.get("total_units") or 0)
            active_room_cnt = int(item.get("active_room_cnt") or 0)
            expected_rate = round(active_room_cnt / units * 100, 1) if units else None
            if (
                item.get("report_rate") != expected_rate
                or (item.get("report_rate") is not None and item["report_rate"] < 0)
            ):
                failures.append("데이터랩 영업신고현황: 신고율 계산이 잘못됨")
                break
        total = consign.get("total") or {}
        national_units = int(total.get("total_units") or 0)
        national_active_room_cnt = int(total.get("active_room_cnt") or 0)
        expected_national_rate = round(
            national_active_room_cnt / national_units * 100, 1
        ) if national_units else None
        if (
            total.get("report_rate") != expected_national_rate
            or (total.get("report_rate") is not None and total["report_rate"] < 0)
        ):
            failures.append("데이터랩 영업신고현황: 전국 합계 신고율 계산이 잘못됨")

        public_stats = client.get("/api/v1/d/3f7").get_json() or {}

        clusters = client.get("/api/buildings-cluster?level=sido").get_json() or {}
        for item in clusters.get("items") or []:
            by_type = item.get("by_type") or {}
            if int(item.get("total") or 0) != sum(int(value or 0) for value in by_type.values()):
                failures.append("지도 시도 배지: 전체 건물 수와 숙박유형 부분합이 다름")
                break
        with client.session_transaction() as session:
            session["admin"] = True
        master_stats = client.get("/api/admin/stats/master").get_json() or {}
        master_sections = master_stats.get("sections") or []
        expected_master_keys = {
            "lodging_stats", "region_match", "consign_stats",
            "closure_stats", "transaction_stats", "collection_stats",
        }
        if (
            not master_stats.get("ok")
            or not isinstance(master_stats.get("refreshed_at"), str)
            or not isinstance(master_stats.get("expires_in_seconds"), int)
            or {section.get("key") for section in master_sections} != expected_master_keys
            or any(section.get("status") not in {"ok", "error"} for section in master_sections)
        ):
            failures.append("통계 원본 창고: 갱신시각·TTL·6개 섹션 상태 응답이 잘못됨")
        refreshed_master_stats = client.post("/api/admin/stats/master").get_json() or {}
        if (
            not refreshed_master_stats.get("ok")
            or {section.get("key") for section in refreshed_master_stats.get("sections") or []}
            != expected_master_keys
        ):
            failures.append("통계 원본 창고: 관리자 수동 새로고침이 동작하지 않음")

        import app as app_module

        # 별도 수집 프로세스가 원본을 커밋한 뒤 표식을 기록하면, 만료된 이전
        # 통계는 즉시 응답으로 유지하고 재계산만 백그라운드에 예약한다.
        from stats_cache import mark_master_stats_invalidated
        saved_master_cache = copy.deepcopy(app_module._MASTER_STATS_CACHE)
        saved_bld_full_stats_cache = copy.deepcopy(app_module._bld_full_stats_cache)
        saved_invalidation_state = copy.deepcopy(app_module._MASTER_STATS_INVALIDATION_STATE)
        saved_refresh_signal = app_module._MASTER_STATS_NEEDS_REFRESH
        try:
            app_module._MASTER_STATS_INVALIDATION_STATE["checked_at"] = 0.0
            token_before = app_module._master_stats_invalidation_token()
            external_conn = app_module.get_conn()
            external_cur = external_conn.cursor()
            external_cur.execute("""
                UPDATE master_buildings
                   SET verified_at = NOW()
                 WHERE id = (SELECT id FROM master_buildings ORDER BY id LIMIT 1)
            """)
            external_conn.commit()
            external_cur.close()
            external_conn.close()
            mark_master_stats_invalidated("api_test_external_master_writer")
            app_module._MASTER_STATS_INVALIDATION_STATE["checked_at"] = 0.0
            token_after = app_module._master_stats_invalidation_token()
            app_module._MASTER_STATS_CACHE.update({
                "ts": time.time(),
                "data": {"consign_stats": {"ok": True, "items": [], "total": {"units": -1}}},
                "sections": {"consign_stats": {"status": "ok", "error": None}},
                "invalidation_token": token_before,
            })
            from unittest.mock import patch
            from unittest.mock import Mock
            refresh_signal = Mock()
            with patch.object(
                app_module,
                "_MASTER_STATS_NEEDS_REFRESH",
                refresh_signal,
            ):
                external_stale = client.get("/api/stats/consign-by-sido").get_json() or {}
            if (
                token_after == token_before
                or (external_stale.get("total") or {}).get("units") != -1
                or not refresh_signal.set.called
            ):
                failures.append("통계 원본 창고: 외부 원본 갱신 뒤 stale 응답·백그라운드 재검증이 동작하지 않음")

            # 신고율 API도 통합 숙박 섹션을 거쳐야 한다. 외부 무효화 후에는
            # 아직 유효한 레거시 5분 캐시에 심어 둔 값이 반환되면 안 된다.
            app_module._bld_full_stats_cache = {
                "ts": time.time(),
                "data": {
                    "rows": [{
                        "type": "전체",
                        "report_rate_building_count": 999991,
                        "report_rate_room_count": 999992,
                        "report_rate_units": 999993,
                    }],
                },
            }
            app_module._MASTER_STATS_CACHE.update({
                "ts": time.time(),
                "data": {"lodging_stats": {"rows": []}},
                "sections": {"lodging_stats": {"status": "ok", "error": None}},
                "invalidation_token": token_before,
            })
            refresh_signal = Mock()
            with patch.object(
                app_module,
                "_MASTER_STATS_NEEDS_REFRESH",
                refresh_signal,
            ):
                registration_after_external_write = client.get("/api/stats/registration-rate").get_json() or {}
            if registration_after_external_write.get("total_units") == 999993:
                failures.append("통계 원본 창고: 신고율 API가 외부 무효화 뒤 레거시 캐시를 반환함")
            elif not refresh_signal.set.called:
                failures.append("통계 원본 창고: 신고율 stale 응답이 재검증을 예약하지 않음")
        finally:
            app_module._MASTER_STATS_CACHE.clear()
            app_module._MASTER_STATS_CACHE.update(saved_master_cache)
            app_module._bld_full_stats_cache = saved_bld_full_stats_cache
            app_module._MASTER_STATS_INVALIDATION_STATE.clear()
            app_module._MASTER_STATS_INVALIDATION_STATE.update(saved_invalidation_state)
            app_module._MASTER_STATS_NEEDS_REFRESH = saved_refresh_signal

        # 마스터 캐시 한 섹션이 실패하면 기존 직접 집계로 폴백한다. 단, 배포
        # 직후처럼 캐시 시각조차 없는 콜드스타트는 요청에서 전체 원본을 만들지
        # 않고 비동기 워밍업 상태를 반환한다.
        saved_master_cache = copy.deepcopy(app_module._MASTER_STATS_CACHE)
        try:
            app_module._MASTER_STATS_CACHE.update({
                "ts": time.time(),
                "data": {},
                "sections": {"consign_stats": {"status": "error", "error": "test"}},
                "invalidation_token": app_module._master_stats_invalidation_token(),
            })
            fallback_consign = client.get("/api/stats/consign-by-sido").get_json() or {}
            if fallback_consign != consign:
                failures.append("통계 원본 창고: 영업신고 섹션 오류 때 폴백 결과가 달라짐")

            app_module._MASTER_STATS_CACHE.update({
                "ts": 0.0,
                "data": {},
                "sections": {},
                "invalidation_token": None,
            })
            with patch.object(app_module, "_master_stats_schedule_revalidation") as schedule:
                rebuilt_consign = client.get("/api/stats/consign-by-sido").get_json() or {}
            if (
                rebuilt_consign.get("ok") is not False
                or rebuilt_consign.get("status") != "warming"
                or rebuilt_consign.get("items") != []
                or not schedule.called
            ):
                failures.append("통계 원본 창고: 빈 캐시가 비동기 워밍업 상태로 응답하지 않음")
        finally:
            app_module._MASTER_STATS_CACHE.clear()
            app_module._MASTER_STATS_CACHE.update(saved_master_cache)
        admin_stats = client.get("/api/admin/buildings/full-stats").get_json() or {}
        total_row = next((row for row in admin_stats.get("rows") or [] if row.get("type") == "전체"), {})
        admin_rows = admin_stats.get("rows") or []
        admin_required = {
            "type", "building_count", "units", "favorites", "listing_requests",
            "broker_badge", "store_realty", "store_total", "report_rate",
            "permit_count", "room_count", "closed_rate", "lodging_metric",
            "report_rate_numerator", "report_rate_denominator", "report_rate_basis",
        }
        if not admin_rows or any(not admin_required <= set(row) for row in admin_rows):
            failures.append("데이터랩 전국숙박업통계: 관리자 전체 통계 원본 컬럼이 사라짐")
        public_rows = public_stats.get("rows") or []
        public_allowed = {
            "type", "building_count", "units", "biz_count", "room_count",
            "report_rate", "sub_rows", "camping_facility_count",
            "camping_site_count", "camping_general_site_count",
            "camping_auto_site_count", "camping_glamping_site_count",
            "camping_caravan_site_count", "camping_classification_breakdown",
        }
        if any(set(row) - public_allowed for row in public_rows):
            failures.append("데이터랩 전국숙박업통계: 공개 응답에 내부 운영지표가 남아 있음")
        expected_public_types = [
            "전체", "생활", "관광", "일반", "에어비앤비",
            "농어촌민박", "캠핑", "한옥", "복합", "준공전", "미분류",
        ]
        if [row.get("type") for row in public_rows] != expected_public_types:
            failures.append("데이터랩 전국숙박업통계: 개정 법정분류 행 순서가 관리자 통계와 다름")
        admin_by_type = {row.get("type"): row for row in admin_rows}
        for public_row in public_rows:
            admin_row = admin_by_type.get(public_row.get("type")) or {}
            if any(
                public_row.get(public_key) != admin_row.get(admin_key)
                for public_key, admin_key in (
                    ("building_count", "building_count"),
                    ("units", "units"),
                    ("biz_count", "permit_count"),
                    ("room_count", "room_count"),
                    ("report_rate", "report_rate"),
                )
            ):
                failures.append("데이터랩 전국숙박업통계: 공개 최소 컬럼이 관리자 계산 결과와 다름")
                break
            if public_row.get("type") == "일반":
                admin_sub_rows = {
                    row.get("type"): row for row in admin_row.get("sub_rows") or []
                }
                for public_sub in public_row.get("sub_rows") or []:
                    admin_sub = admin_sub_rows.get(public_sub.get("type")) or {}
                    if (
                        public_sub.get("building_count") != admin_sub.get("building_count")
                        or public_sub.get("biz_count") != admin_sub.get("permit_count")
                        or public_sub.get("room_count") != admin_sub.get("room_count")
                        or public_sub.get("report_rate") != admin_sub.get("report_rate")
                    ):
                        failures.append("데이터랩 전국숙박업통계: 일반 세부행 공개 수치가 관리자 결과와 다름")
                        break

        # 순위 좌표는 거래의 법정동·지번 키까지 일치하는 마스터 건물에서만 가져와야 한다.
        from db import get_conn
        conn = get_conn()
        cur = conn.cursor()
        try:
            for item in ranked_items:
                building_id = item.get("building_id")
                if not building_id:
                    continue
                cur.execute(
                    "SELECT building_name, sgg_cd, umd_nm, jibun, lat, lng FROM master_buildings WHERE id = %s",
                    [building_id],
                )
                building = cur.fetchone()
                cur.execute(
                    """
                    SELECT sgg_cd, umd_nm, jibun
                    FROM transactions
                    WHERE building_name = %s AND address = %s
                    ORDER BY id
                    LIMIT 1
                    """,
                    [item.get("building_name"), item.get("address")],
                )
                transaction = cur.fetchone()
                if (
                    not building
                    or not transaction
                    or building["building_name"] != item.get("building_name")
                    or any(building[key] != transaction[key] for key in ("sgg_cd", "umd_nm", "jibun"))
                    or (building["lat"] is None) != (item.get("lat") is None)
                    or (building["lng"] is None) != (item.get("lng") is None)
                    or (
                        building["lat"] is not None
                        and (
                            item.get("lat") != float(building["lat"])
                            or item.get("lng") != float(building["lng"])
                        )
                    )
                ):
                    failures.append("데이터랩 랭킹: 법정동·지번 불일치 지도 좌표가 반환됨")
                    break
        finally:
            cur.close()
            conn.close()

        if not failures:
            print("OK  데이터랩 가격변동 토글·최고가·폐업 표본·시도별 신고율·관리자 통계 대조")
    except Exception as exc:
        failures.append(f"데이터랩 통계 테스트 오류: {exc}")
    return failures


def _check_whole_listing_location_context(client):
    """건물전체 입지 API의 최단역·500m 경쟁업소·5km 외 null 경계를 확인한다."""
    import time as _time
    from db import get_conn

    failures = []
    conn = get_conn()
    cur = conn.cursor()
    building_ids = []
    station_id = None
    try:
        run_id = str(int(_time.time() * 1000))
        base_name = f"입지테스트-{run_id}"
        cur.execute(
            """INSERT INTO master_buildings (building_name, road_address, lat, lng, lodging_type)
               VALUES (%s, %s, 30.0000, 120.0000, '생활') RETURNING id""",
            (base_name, "입지 테스트 도로 1"),
        )
        building_ids.append(cur.fetchone()["id"])
        cur.execute(
            """INSERT INTO master_buildings (building_name, road_address, lat, lng, lodging_type)
               VALUES (%s, %s, 30.0008, 120.0000, '일반') RETURNING id""",
            (base_name + " 경쟁", "입지 테스트 도로 2"),
        )
        building_ids.append(cur.fetchone()["id"])
        cur.execute(
            """INSERT INTO master_buildings (building_name, road_address, lat, lng, lodging_type)
               VALUES (%s, %s, 30.0007, 120.0000, '생활') RETURNING id""",
            (base_name + " 생활경쟁", "입지 테스트 도로 2-1"),
        )
        building_ids.append(cur.fetchone()["id"])
        cur.execute(
            """INSERT INTO master_buildings (building_name, road_address, lat, lng, lodging_type)
               VALUES (%s, %s, 20.0000, 110.0000, '생활') RETURNING id""",
            (base_name + " 원거리", "입지 테스트 도로 3"),
        )
        building_ids.append(cur.fetchone()["id"])
        cur.execute(
            """INSERT INTO subway_stations (station_name, line_name, lat, lng)
               VALUES (%s, %s, 30.0004, 120.0000) RETURNING id""",
            (base_name + "역", "테스트선"),
        )
        station_id = cur.fetchone()["id"]
        conn.commit()

        nearby = client.get(f"/api/building/{building_ids[0]}/whole-listing-context")
        nearby_data = nearby.get_json() or {}
        subway = nearby_data.get("subway") or {}
        if (
            nearby.status_code != 200
            or subway.get("station_name") != base_name + "역"
            or subway.get("line_name") != "테스트선"
            or not (0 < int(subway.get("distance_m") or 0) < 5000)
            or not (1 <= int(subway.get("walk_minutes") or 0) <= 63)
            or int((nearby_data.get("nearby_lodgings") or {}).get("일반") or 0) != 1
            or int((nearby_data.get("nearby_lodgings") or {}).get("생활") or 0) != 1
        ):
            failures.append("whole location: 최단 지하철역·도보시간 또는 500m 경쟁업소 계산이 잘못됨")

        remote = client.get(f"/api/building/{building_ids[3]}/whole-listing-context")
        if remote.status_code != 200 or (remote.get_json() or {}).get("subway") is not None:
            failures.append("whole location: 5km 밖 지하철역을 표시함")
        # 공개 목록의 최대 50건보다 많은 31개 서로 다른 건물도 한 요청으로
        # 처리돼야 카드별 입지정보가 통째로 비지 않는다.
        bulk_ids = []
        for index in range(30):
            cur.execute(
                """INSERT INTO master_buildings (building_name, road_address, lat, lng, lodging_type)
                   VALUES (%s, %s, %s, 110.0000, '생활') RETURNING id""",
                (base_name + f" 일괄{index}", "입지 테스트 일괄", 20.1 + index / 10),
            )
            bulk_id = cur.fetchone()["id"]
            building_ids.append(bulk_id)
            bulk_ids.append(bulk_id)
        conn.commit()
        batch = client.post("/api/whole-listing-contexts", json={
            "building_ids": [building_ids[0], *bulk_ids, "invalid"],
        })
        batch_items = (batch.get_json() or {}).get("items") or {}
        if (
            batch.status_code != 200
            or batch_items.get(str(building_ids[0]), {}).get("subway", {}).get("station_name") != base_name + "역"
            or len(batch_items) != 31
        ):
            failures.append("whole location: 공개 목록용 일괄 입지 컨텍스트 API가 잘못됨")
        if not failures:
            print("OK  건물전체 경쟁업소·최단 지하철역·원거리 생략·일괄 컨텍스트")
    except Exception as exc:
        failures.append(f"whole location 테스트 오류: {exc}")
    finally:
        try:
            if station_id:
                cur.execute("DELETE FROM subway_stations WHERE id=%s", (station_id,))
            if building_ids:
                cur.execute("DELETE FROM master_buildings WHERE id = ANY(%s)", (building_ids,))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_whole_building_listing(client):
    """건물전체 매물의 전용 필드 저장·수정과 제한공개 응답을 검증한다."""
    import time as _time
    from db import get_conn

    failures = []
    conn = get_conn()
    cur = conn.cursor()
    user_id = listing_id = None
    created_building_ids = []
    try:
        run_id = str(int(_time.time() * 1000))
        test_sgg = f"매물테스트시{run_id}"
        test_umd = "권리금동"
        for index in range(5):
            cur.execute(
                """INSERT INTO master_buildings
                    (building_name, road_address, sgg_text, umd_nm, lat, lng, lodging_type)
                   VALUES (%s, %s, %s, %s, %s, %s, '생활') RETURNING id""",
                (
                    f"건물전체 테스트 {run_id}-{index}", f"매물 테스트로 {index + 1}",
                    test_sgg, test_umd, 37.100 + index * 0.002, 127.100 + index * 0.002,
                ),
            )
            created_building_ids.append(cur.fetchone()["id"])
        building = {
            "id": created_building_ids[0],
            "building_name": f"건물전체 테스트 {run_id}-0",
        }
        cur.execute(
            "INSERT INTO users (email, name, phone, phone_verified) VALUES (%s, %s, '01000000000', TRUE) RETURNING id",
            (f"whole-listing-{run_id}@example.test", "건물전체 매물 테스트"),
        )
        user_id = cur.fetchone()["id"]
        conn.commit()
        with client.session_transaction() as sess:
            sess.clear()
            sess["user_id"] = user_id

        listing_description = "■ 매물 개요\n- 테스트 제한공개 매물 설명입니다."
        created = client.post("/api/listing-requests", json={
            "master_building_id": building["id"], "deal_mode": "direct",
            "registrant_type": "building_owner", "transaction_target": "whole",
            "deal_type": "매매", "price_krw": 200000, "succession_loan_krw": 120000,
             "key_money_krw": 5000, "room_count": 22,
             "monthly_revenue_krw": 3000, "annual_revenue_krw": 36000,
             "short_stay_ratio": 32.5, "ota_revenue_ratio": 71.2,
            "operation_status": "폐업", "closed_at": "2026-08-01",
            "remodeling_info": "객실 일부 리모델링", "is_urgent": True,
            "disclosure_scope": "limited", "description": listing_description,
            "building_info_overrides": {"structure": "철근콘크리트"},
        })
        body = created.get_json() or {}
        listing_id = body.get("id")
        if created.status_code != 200 or not listing_id:
            return [f"whole listing: 생성 실패 (HTTP {created.status_code}, {body})"]
        context = client.get(f"/api/building/{building['id']}/whole-listing-context")
        context_data = context.get_json() or {}
        if (context.status_code != 200 or not context_data.get("ok")
                or "building" not in context_data or "nearby_lodgings" not in context_data
                or "subway" not in context_data or "suggested_room_count" not in context_data):
            failures.append("whole listing: 건물자동채움·주변 숙박시설 컨텍스트 API 실패")
        detail = client.get(f"/api/building/{building['id']}")
        detail_item = next(
            (item for item in ((detail.get_json() or {}).get("direct_listings") or [])
             if item.get("id") == listing_id),
            None,
        )
        if detail.status_code != 200 or detail_item is not None:
            failures.append("whole listing: 제한공개 매물이 건물상세 공개 목록에 노출됨")
        limited_checklist = client.get(f"/api/listing-requests/{listing_id}/checklist")
        if limited_checklist.status_code != 404:
            failures.append("whole listing: 제한공개 매물 체크리스트가 건물정보를 노출함")

        cur.execute("SELECT sgg_text, umd_nm FROM master_buildings WHERE id=%s", (building["id"],))
        location = cur.fetchone() or {}
        limited_location = " ".join(
            value for value in (location.get("sgg_text"), location.get("umd_nm")) if value
        )
        with client.session_transaction() as sess:
            sess.clear()
        public_items = (client.get("/api/listings?disclosure_scope=public").get_json() or {}).get("items") or []
        limited_items = (client.get("/api/listings?disclosure_scope=limited").get_json() or {}).get("items") or []
        public_item = next((item for item in public_items if item.get("id") == listing_id), None)
        limited_item = next((item for item in limited_items if item.get("id") == listing_id), {})
        if (
            public_item is not None
            or limited_item.get("building_name") != limited_location
            or limited_item.get("building_id") is not None
            or limited_item.get("monthly_revenue_krw") is not None
            or limited_item.get("succession_loan_krw") is not None
            or not limited_item.get("has_monthly_revenue")
            or not limited_item.get("has_succession_loan")
            or limited_item.get("location_precision") != "approximate"
            or limited_item.get("approx_lat") is None
            or limited_item.get("approx_lng") is None
            or limited_item.get("description") != listing_description
            or limited_item.get("remodeling_info") != "객실 일부 리모델링"
            or any(key in limited_item for key in (
                "photo_url", "photos", "building_info_overrides",
                "phone_tail", "lat", "lng",
            ))
        ):
            failures.append("whole listing: 제한공개 탭의 지역 익명화·비로그인 민감정보 마스킹이 누락됨")
        geo = client.get(f"/api/buildings-geo?q={building.get('building_name') or ''}")
        geo_item = next(
            (item for item in ((geo.get_json() or {}).get("items") or [])
             if item.get("id") == building["id"]),
            {},
        )
        if geo.status_code != 200 or geo_item.get("listing_count") != 0:
            failures.append("whole listing: 제한공개 매물이 지도 직거래 배지에서 역추적됨")
        from app import _HOME_SHARE_TITLE
        private_share_html = client.get(
            f"/building/{building['id']}?listing={listing_id}"
        ).get_data(as_text=True)
        private_share_head = private_share_html.split("</head>", 1)[0]
        building_name = building.get("building_name") or ""
        if (_HOME_SHARE_TITLE not in private_share_head
                or (building_name and building_name in private_share_head)):
            failures.append("whole listing: 제한공개 공유 메타에 정확한 건물명이 노출됨")

        with client.session_transaction() as sess:
            sess["user_id"] = user_id
        logged_limited_items = (client.get("/api/listings?disclosure_scope=limited").get_json() or {}).get("items") or []
        logged_limited_item = next((item for item in logged_limited_items if item.get("id") == listing_id), {})
        if (logged_limited_item.get("monthly_revenue_krw") != 3000
                or logged_limited_item.get("succession_loan_krw") != 120000
                or logged_limited_item.get("key_money_krw") != 5000
                or logged_limited_item.get("annual_revenue_krw") != 36000
                or logged_limited_item.get("remodeling_info") != "객실 일부 리모델링"
                or logged_limited_item.get("description") != listing_description
                or any(key in logged_limited_item for key in (
                    "photo_url", "photos", "building_info_overrides",
                    "phone_tail", "lat", "lng",
                ))):
            failures.append("whole listing: 로그인 제한공개 응답의 허용 금융정보 또는 익명화가 잘못됨")
        with client.session_transaction() as sess:
            sess.clear()
            sess["agent_id"] = -999999
        partner_items = (client.get("/api/listings?disclosure_scope=limited").get_json() or {}).get("items") or []
        partner_item = next((item for item in partner_items if item.get("id") == listing_id), {})
        if (not partner_item.get("financial_details_visible")
                or partner_item.get("monthly_revenue_krw") != 3000
                or partner_item.get("succession_loan_krw") != 120000
                or partner_item.get("key_money_krw") != 5000):
            failures.append("whole listing: 파트너 로그인 세션의 금융정보가 숨겨짐")
        with client.session_transaction() as sess:
            sess.clear()
            sess["user_id"] = user_id
        first_view = client.post("/api/listings/views", json={"listing_ids": [listing_id]})
        second_view = client.post("/api/listings/views", json={"listing_ids": [listing_id]})
        live_view = client.get("/api/listings/views", query_string=[
            ("listing_ids", listing_id),
        ])
        live_view_items = (live_view.get_json() or {}).get("items") or []
        live_view_item = next((item for item in live_view_items if item.get("id") == listing_id), {})
        limited_after_view = (
            client.get("/api/listings?disclosure_scope=limited").get_json() or {}
        ).get("items") or []
        limited_view_item = next((item for item in limited_after_view if item.get("id") == listing_id), {})
        if (
            first_view.status_code != 200
            or second_view.status_code != 200
            or live_view.status_code != 200
            or live_view_item.get("viewer_count") != 1
            or limited_view_item.get("viewer_count") != 1
        ):
            failures.append("whole listing: 제한공개 카드 열람자 기록·실시간 조회 또는 고유 IP 집계가 누락됨")

        cur.execute("""SELECT transaction_target, deal_type, price_krw, succession_loan_krw, key_money_krw, room_count,
                               short_stay_ratio, ota_revenue_ratio,
                              operation_status, closed_at, is_urgent, disclosure_scope, building_info_overrides
                       FROM listing_requests WHERE id=%s""", (listing_id,))
        stored = cur.fetchone() or {}
        if not (stored.get("transaction_target") == "whole" and stored.get("deal_type") == "매매"
                and stored.get("price_krw") == 200000 and stored.get("succession_loan_krw") == 120000
                and stored.get("key_money_krw") == 5000 and stored.get("room_count") == 22
                and float(stored.get("short_stay_ratio") or -1) == 32.5
                and float(stored.get("ota_revenue_ratio") or -1) == 71.2
                and stored.get("operation_status") == "폐업" and stored.get("is_urgent")
                and stored.get("disclosure_scope") == "limited"):
            failures.append("whole listing: 매매 전용 필드 저장 실패")

        updated = client.put(f"/api/listing-requests/{listing_id}", json={
            "transaction_target": "whole", "deal_type": "통임대", "price_krw": 5000,
            "monthly_rent_krw": 400, "key_money_krw": 2500, "succession_loan_krw": 1200,
             "room_count": 33, "operation_status": "영업중",
             "short_stay_ratio": 40, "ota_revenue_ratio": 80,
             "disclosure_scope": "public", "building_info_overrides": {"zoning": "상업지역"},
        })
        cur.execute("""SELECT deal_type, price_krw, monthly_rent_krw, key_money_krw,
                               short_stay_ratio, ota_revenue_ratio,
                               succession_loan_krw, room_count, operation_status, disclosure_scope, building_info_overrides
                       FROM listing_requests WHERE id=%s""", (listing_id,))
        changed = cur.fetchone() or {}
        if (updated.status_code != 200 or changed.get("deal_type") != "통임대"
                or changed.get("monthly_rent_krw") != 400 or changed.get("key_money_krw") != 2500
                or changed.get("succession_loan_krw") != 1200 or changed.get("room_count") != 33
                or float(changed.get("short_stay_ratio") or -1) != 40
                or float(changed.get("ota_revenue_ratio") or -1) != 80
                or changed.get("operation_status") != "영업중" or changed.get("disclosure_scope") != "public"):
            failures.append("whole listing: 통임대 수정 또는 매매 필드 초기화 실패")

        for whole_deal_type, price, key_money, loan in (
            ("운영권양도", 7000, 1800, 900),
            ("위탁운영", 6000, 1400, 800),
        ):
            response = client.put(f"/api/listing-requests/{listing_id}", json={
                "transaction_target": "whole", "deal_type": whole_deal_type,
                "price_krw": price, "key_money_krw": key_money,
                "succession_loan_krw": loan, "room_count": 33,
                "short_stay_ratio": 40, "ota_revenue_ratio": 80,
                "operation_status": "영업중", "disclosure_scope": "public",
                "building_info_overrides": {"zoning": "상업지역"},
            })
            cur.execute(
                """SELECT deal_type, price_krw, key_money_krw, succession_loan_krw
                     FROM listing_requests WHERE id=%s""",
                (listing_id,),
            )
            stored_terms = cur.fetchone() or {}
            if (response.status_code != 200 or stored_terms.get("deal_type") != whole_deal_type
                    or stored_terms.get("price_krw") != price
                    or stored_terms.get("key_money_krw") != key_money
                    or stored_terms.get("succession_loan_krw") != loan):
                failures.append(f"whole listing: {whole_deal_type} 권리금·승계융자 저장 실패")

        public_items_after = (client.get("/api/listings?disclosure_scope=public").get_json() or {}).get("items") or []
        public_item_after = next((item for item in public_items_after if item.get("id") == listing_id), {})
        if (public_item_after.get("viewer_count") != 1
                or float(public_item_after.get("short_stay_ratio") or -1) != 40
                or float(public_item_after.get("ota_revenue_ratio") or -1) != 80):
            failures.append("whole listing: 최근 5분 고유 열람자 수 집계가 정확하지 않음")

        checklist = client.get(f"/api/listing-requests/{listing_id}/checklist")
        checklist_data = checklist.get_json() or {}
        checklist_keys = {item.get("key") for item in checklist_data.get("items") or []}
        if (checklist.status_code != 200 or not checklist_data.get("ok")
                or checklist_data.get("total_items") != 14
                or checklist_data.get("is_authenticated") is not True
                or "zoning" not in checklist_keys
                or not {"building_violation", "nearby_competition", "finance", "sale_reason"} <= checklist_keys):
            failures.append("whole listing: 공개 건물전체 체크리스트 항목·자동값 응답이 잘못됨")
        saved_progress = client.post(
            f"/api/listing-requests/{listing_id}/checklist/progress",
            json={"item_key": "parking", "checked": True},
        )
        checked_after_save = client.get(f"/api/listing-requests/{listing_id}/checklist").get_json() or {}
        if (saved_progress.status_code != 200
                or "parking" not in (checked_after_save.get("checked_keys") or [])):
            failures.append("whole listing: 로그인 체크리스트 진행 상태 서버 저장 실패")
        with client.session_transaction() as sess:
            sess.clear()
        anonymous_checklist = client.get(f"/api/listing-requests/{listing_id}/checklist")
        anonymous_data = anonymous_checklist.get_json() or {}
        anonymous_save = client.post(
            f"/api/listing-requests/{listing_id}/checklist/progress",
            json={"item_key": "parking", "checked": False},
        )
        if (anonymous_checklist.status_code != 200
                or anonymous_data.get("is_authenticated") is not False
                or anonymous_data.get("checked_keys") != []
                or anonymous_save.status_code != 401):
            failures.append("whole listing: 비로그인 체크리스트의 로컬 저장 분리 또는 서버 차단 실패")
        with client.session_transaction() as sess:
            sess["user_id"] = user_id

        mine = client.get("/api/listing-requests/mine")
        mine_item = next((x for x in ((mine.get_json() or {}).get("items") or []) if x.get("id") == listing_id), {})
        if mine.status_code != 200 or mine_item.get("transaction_target") != "whole" or mine_item.get("key_money_krw") != 1400:
            failures.append("whole listing: 내 매물 조회 전용 필드 누락")

        from app import _whole_listing_values
        _invalid_values, invalid_error = _whole_listing_values({
            "transaction_target": "whole", "deal_type": "전세",
        })
        if not invalid_error:
            failures.append("whole listing: 허용하지 않는 건물전체 거래방식을 차단하지 않음")
        _ratio_values, ratio_error = _whole_listing_values({
            "transaction_target": "whole", "deal_type": "통임대",
            "short_stay_ratio": 101,
        })
        if not ratio_error:
            failures.append("whole listing: 범위를 벗어난 운영 비율을 차단하지 않음")
        if not failures:
            print("OK  건물전체 매물 생성·수정·내 매물 조회·거래방식 검증")
    except Exception as exc:
        failures.append(f"whole listing 테스트 오류: {exc}")
    finally:
        try:
            if listing_id:
                cur.execute("DELETE FROM page_views WHERE listing_request_id=%s", (listing_id,))
                cur.execute("DELETE FROM listing_request_history WHERE listing_request_id=%s", (listing_id,))
                cur.execute("DELETE FROM listing_requests WHERE id=%s", (listing_id,))
            if user_id:
                cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
            if created_building_ids:
                cur.execute("DELETE FROM master_buildings WHERE id = ANY(%s)", (created_building_ids,))
            conn.commit()
        finally:
            cur.close()
            conn.close()
    return failures


def _check_building_request_e2e(client):
    """내 건물 수정 요청 → 승인 → 지도 노출 end-to-end 회귀 점검.

    /api/submit-building의 "신규 마스터 INSERT 경로"를 검증한다.
    외부 API 의존성(건축물대장, 지오코더)은 unittest.mock으로 대체해
    data.go.kr 쿼터를 소모하지 않는다:

      1. baseline: 의정부동 cluster 배지 카운트 기록
      2. /api/submit-building (jibun_address_input 경로) 호출
           classify_lodging_type → ('생활', ...) 으로 고정 반환 (data.go.kr 대체)
           resolve_api_building_name → 고유 TEST_NAME 반환
           _fill_master_coords → lat/lng 직접 UPDATE (geocoder 대체)
         → submit_building이 master_buildings에 새 행을 INSERT하고
           building_requests를 'verified'로 완결시키는 경로 전체가 실행됨
      3. building_requests 행이 생성되고 status='verified'인지 DB 직접 확인
         → 이때 이번 호출에서 생성된 req_id · mb_id를 캡처 (cleanup 전용)
      4. master_buildings 행에 lat/lng가 채워졌는지 확인
      5. 캐시 전체 초기화
      6. /api/buildings-geo?q={이름}     — 정확 이름 검색에서 해당 건물 확인
      7. /api/buildings-geo?q={붙여쓰기} — 공백제거 ILIKE 검색에서도 확인
                                           (더 그레이스 경희 버그 회귀 방지)
      8. /api/buildings-cluster?level=umd — umd 배지 카운트가 baseline+1인지 확인
      9. 롤백: 이번 호출에서 생성된 req_id · mb_id만 삭제
               (다른 행·다른 실행 결과를 건드리지 않음)

    모든 fixture 식별자(이름·지번·도로명주소)는 ms 타임스탬프 기반으로 실행마다 고유.
    BjdongMap 전제조건 (dev 환경 확인 완료):
      extract_sgg_from_address('경기도 의정부시 ...') → sgg_cd='41150'
      find_bjdong_cd('41150', '의정부동')             → bjdong_cd='10100'
    """
    import time as _time
    from unittest import mock
    import app as _app_module
    from db import get_conn

    failures = []

    # 모든 fixture 식별자를 ms 타임스탬프로 고유화 — 같은 jibun/이름 중복 방지
    _run_ms = str(int(_time.time() * 1000))
    TEST_NAME         = f"자동검증빌딩 {_run_ms[-7:]}"   # 공백 포함 → ILIKE nospace 버그 재현
    TEST_NAME_NOSPACE = TEST_NAME.replace(" ", "")
    # 9xxx-9 범위: 9000+로 실제 지번 충돌 가능성 최소, -9 suffix로 test 전용 구별
    TEST_JIBUN        = f"9{_run_ms[-3:]}-9"             # e.g., "9328-9" — 매 실행 고유
    TEST_ROAD_ADDR    = f"경기도 의정부시 테스트로 {_run_ms[-5:]}"   # 매 실행 고유

    REAL_SGG_CD   = "41150"
    REAL_UMD_NM   = "의정부동"
    REAL_SGG_TEXT = "경기도 의정부시"
    TEST_LAT      = 37.7339
    TEST_LNG      = 127.0471
    TEST_LODGING  = "생활"

    def _clear_caches():
        """in-process 테스트 클라이언트 전용 — 전체 초기화가 부작용 없이 안전."""
        _app_module._geo_cache.clear()
        _app_module._cluster_cache.clear()

    conn = get_conn()
    cur  = conn.cursor()
    # 이번 호출에서 생성된 ID만 추적 — cleanup은 이 ID만 삭제
    captured_req_id = None
    captured_mb_id  = None
    captured_txn_id = None
    captured_other_txn_id = None
    captured_listing_id = None
    captured_user_id = None
    stats_marker = None

    try:
        # ── ① Baseline: 삽입 전 의정부동 클러스터 배지 카운트 ─────────────────
        _clear_caches()
        r_base = client.get(
            f"/api/buildings-cluster?level=umd&sgg_nm={REAL_SGG_TEXT}"
        )
        base_items   = (r_base.get_json() or {}).get("items", [])
        expected_umd = f"{REAL_SGG_TEXT} {REAL_UMD_NM}".strip()
        base_badge   = next((it for it in base_items if it.get("name") == expected_umd), None)
        base_count   = base_badge["total"] if base_badge else 0

        # ── ② /api/submit-building — 외부 API mock으로 NEW INSERT 경로 실행 ──
        mock_title = {"new_plat_plc": TEST_ROAD_ADDR, "plat_plc": None, "ho_cnt": 50}

        def _mock_fill_coords(inner_cur, master_id, road_address):
            """geocode_buildings 호출 없이 lat/lng를 직접 설정.
            submit_building이 열어 둔 cursor를 그대로 받아 같은 트랜잭션 내에서 UPDATE."""
            inner_cur.execute(
                "UPDATE master_buildings SET lat=%s, lng=%s WHERE id=%s",
                (TEST_LAT, TEST_LNG, master_id),
            )

        with (
            mock.patch(
                "building_registry.classify_lodging_type",
                return_value=(TEST_LODGING, "생활숙박시설", "", mock_title, "검증완료"),
            ),
            mock.patch(
                "building_registry.resolve_api_building_name",
                return_value=TEST_NAME,   # 고유 이름 → name_pending=False
            ),
            mock.patch("app._fill_master_coords", side_effect=_mock_fill_coords),
            mock.patch("app.mark_master_stats_invalidated") as stats_marker,
        ):
            r_sub = client.post(
                "/api/submit-building",
                json={
                    "road_address":           TEST_ROAD_ADDR,
                    "jibun_address_input":    f"{REAL_UMD_NM} {TEST_JIBUN}",
                    "suggested_lodging_type": TEST_LODGING,
                },
                headers={"X-Forwarded-For": "203.0.113.42"},  # 레이트리밋 전용 테스트 IP
            )

        if stats_marker.call_args_list != [mock.call("user_verified_building")]:
            failures.append(
                "submit-building: 원본 커밋 뒤 master stats 무효화 표식을 정확히 한 번 남기지 않음"
            )

        if r_sub.status_code != 200:
            failures.append(
                f"submit-building: HTTP {r_sub.status_code} (기대: 200)"
            )
        else:
            sub_pl = r_sub.get_json() or {}
            if sub_pl.get("status") != "verified":
                failures.append(
                    f"submit-building: status='{sub_pl.get('status')}' (기대: 'verified') "
                    f"— {sub_pl.get('message')}"
                )
            else:
                print(f"OK  /api/submit-building  (status=verified, 건물={TEST_NAME})")

        # ── ③ 이번 호출이 생성한 행 ID 캡처 + DB 연결 확인 ──────────────────
        # 고유한 (TEST_NAME, TEST_JIBUN) 조합으로 이번 호출 결과만 특정함
        cur.execute("""
            SELECT br.id AS req_id, br.status,
                   mb.id AS mb_id, mb.lat, mb.lng
            FROM building_requests br
            JOIN master_buildings mb ON mb.id = br.master_building_id
            WHERE mb.building_name = %s
              AND mb.jibun         = %s
              AND br.request_type  = 'new'
            ORDER BY br.id DESC LIMIT 1
        """, (TEST_NAME, TEST_JIBUN))
        linked = cur.fetchone()

        if not linked:
            failures.append(
                f"building_requests: '{TEST_NAME}' (jibun={TEST_JIBUN}) 연결 행 없음 "
                f"— submit-building이 master_buildings INSERT를 하지 않은 것으로 의심"
            )
            # 고아 building_request 캡처 (master 생성 전 실패 대비) — 고유 road_address 기준
            cur.execute(
                "SELECT id FROM building_requests"
                " WHERE road_address=%s ORDER BY id DESC LIMIT 1",
                (TEST_ROAD_ADDR,),
            )
            br_row = cur.fetchone()
            if br_row:
                captured_req_id = br_row["id"]
        else:
            captured_req_id = linked["req_id"]
            captured_mb_id  = linked["mb_id"]
            if linked["status"] != "verified":
                failures.append(
                    f"building_requests id={captured_req_id}: "
                    f"status='{linked['status']}' (기대: 'verified')"
                )
            else:
                print(
                    f"OK  building_requests id={captured_req_id} status=verified "
                    f"→ master_building_id={captured_mb_id}"
                )
            if linked["lat"] is None or linked["lng"] is None:
                failures.append(
                    f"master_buildings id={captured_mb_id}: lat/lng NULL — 지도 노출 불가 "
                    f"(_fill_master_coords mock이 lat/lng를 설정하지 못함)"
                )
            else:
                print(
                    f"OK  master_buildings id={captured_mb_id} "
                    f"lat={float(linked['lat']):.4f} lng={float(linked['lng']):.4f}"
                )

        # ── ④ 캐시 초기화 — baseline 조회가 채운 stale 항목 제거 ──────────────
        _clear_caches()

        # ── ⑤ /api/buildings-geo — 정확 이름 검색 ───────────────────────────
        if captured_mb_id is not None:
            r = client.get(f"/api/buildings-geo?q={TEST_NAME}")
            if r.status_code != 200:
                failures.append(f"geo(정확 이름): HTTP {r.status_code} (기대: 200)")
            else:
                payload = r.get_json() or {}
                found = [it for it in payload.get("items", [])
                         if it.get("id") == captured_mb_id]
                if not found:
                    failures.append(
                        f"geo(정확 이름): id={captured_mb_id}가 검색 결과에 없음 "
                        f"(total={payload.get('total')})"
                    )
                else:
                    print(
                        f"OK  /api/buildings-geo?q={TEST_NAME}"
                        f"  (id={captured_mb_id} 확인)"
                    )

            # ── ⑥ /api/buildings-geo — 붙여쓰기 검색 (ILIKE nospace 회귀 방지) ──
            r = client.get(f"/api/buildings-geo?q={TEST_NAME_NOSPACE}")
            if r.status_code != 200:
                failures.append(f"geo(붙여쓰기): HTTP {r.status_code} (기대: 200)")
            else:
                payload = r.get_json() or {}
                found = [it for it in payload.get("items", [])
                         if it.get("id") == captured_mb_id]
                if not found:
                    failures.append(
                        f"geo(붙여쓰기): REPLACE(building_name,' ','') ILIKE 검색에서 "
                        f"id={captured_mb_id} 없음 — nospace ILIKE 조건 확인 필요"
                    )
                else:
                    print(
                        f"OK  /api/buildings-geo?q={TEST_NAME_NOSPACE}"
                        f"  (붙여쓰기 id={captured_mb_id} 확인)"
                    )

            # ── ⑥a 지도 원형 배지 — 실거래만·직거래만·동시 조합 회귀 점검 ──
            # API가 같은 건물에서 두 수를 정확히 합산하는지 실제 DB fixture로 검증한다.
            def _assert_badge_counts(label, expected):
                _clear_caches()
                response = client.get(f"/api/buildings-geo?q={TEST_NAME}")
                items = (response.get_json() or {}).get("items", []) if response.status_code == 200 else []
                item = next((it for it in items if it.get("id") == captured_mb_id), None)
                got = (
                    (item or {}).get("txn_count"),
                    (item or {}).get("listing_count"),
                    (item or {}).get("total_count"),
                )
                if response.status_code != 200 or got != expected:
                    failures.append(
                        f"지도 배지 {label}: {got} (기대: {expected}, HTTP={response.status_code})"
                    )
                else:
                    print(f"OK  지도 배지 {label}: {got}")

            cur.execute(
                "INSERT INTO users (email, name) VALUES (%s, %s) RETURNING id",
                (f"map-badge-{_run_ms}@example.test", "지도 배지 테스트"),
            )
            captured_user_id = cur.fetchone()["id"]
            cur.execute("""
                INSERT INTO transactions
                    (building_name, address, price, deal_date, deal_type,
                     sgg_cd, umd_nm, jibun, raw_key)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                TEST_NAME, f"{REAL_UMD_NM} {TEST_JIBUN}", 10000, "2026-08-21",
                "직거래", REAL_SGG_CD, REAL_UMD_NM, TEST_JIBUN,
                f"map-badge-{_run_ms}",
            ))
            captured_txn_id = cur.fetchone()["id"]
            # 같은 필지지만 건물명이 다른 거래는, 정확 이름 거래가 있으면 배지 건수에서 제외된다.
            cur.execute("""
                INSERT INTO transactions
                    (building_name, address, price, deal_date, deal_type,
                     sgg_cd, umd_nm, jibun, raw_key)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                TEST_NAME + " 별관", f"{REAL_UMD_NM} {TEST_JIBUN}", 11000, "2026-08-20",
                "직거래", REAL_SGG_CD, REAL_UMD_NM, TEST_JIBUN,
                f"map-badge-other-{_run_ms}",
            ))
            captured_other_txn_id = cur.fetchone()["id"]
            conn.commit()
            _assert_badge_counts("실거래만", (1, 0, 1))

            cur.execute("""
                INSERT INTO listing_requests
                    (user_id, master_building_id, deal_type, contact_phone, deal_mode, status)
                VALUES (%s, %s, '매매', '01000000000', 'direct', 'submitted')
                RETURNING id
            """, (captured_user_id, captured_mb_id))
            captured_listing_id = cur.fetchone()["id"]
            # 중개 매물과 두 철회 표기는 지도 공개 직거래 집계에서 제외되어야 한다.
            cur.execute("""
                INSERT INTO listing_requests
                    (user_id, master_building_id, deal_type, contact_phone, deal_mode, status)
                VALUES
                    (%s, %s, '매매', '01000000000', 'broker', 'submitted'),
                    (%s, %s, '매매', '01000000000', 'direct', 'withdrawn'),
                    (%s, %s, '매매', '01000000000', 'direct', '철회됨')
            """, (
                captured_user_id, captured_mb_id,
                captured_user_id, captured_mb_id,
                captured_user_id, captured_mb_id,
            ))
            conn.commit()
            _assert_badge_counts("실거래+직거래 매물", (1, 1, 2))

            cur.execute("DELETE FROM transactions WHERE id=%s", (captured_txn_id,))
            captured_txn_id = None
            cur.execute("DELETE FROM transactions WHERE id=%s", (captured_other_txn_id,))
            captured_other_txn_id = None
            conn.commit()
            _assert_badge_counts("직거래 매물만", (0, 1, 1))

        # ── ⑦ /api/buildings-cluster — umd 배지 baseline+1 확인 ─────────────
        _clear_caches()
        r = client.get(
            f"/api/buildings-cluster?level=umd&sgg_nm={REAL_SGG_TEXT}"
        )
        if r.status_code != 200:
            failures.append(f"cluster(umd): HTTP {r.status_code} (기대: 200)")
        else:
            payload     = r.get_json() or {}
            items       = payload.get("items", [])
            badge       = next((it for it in items if it.get("name") == expected_umd), None)
            if badge is None:
                failures.append(
                    f"cluster(umd): '{expected_umd}' 배지 없음 "
                    f"(반환 배지 수={len(items)})"
                )
            else:
                after_count = badge.get("total", 0)
                if after_count < base_count + 1:
                    failures.append(
                        f"cluster(umd): '{expected_umd}' total={after_count}, "
                        f"기대≥{base_count + 1} (baseline={base_count}+1) "
                        f"— 신규 건물이 집계에 포함되지 않음"
                    )
                else:
                    print(
                        f"OK  /api/buildings-cluster?level=umd&sgg_nm={REAL_SGG_TEXT}"
                        f"  ('{expected_umd}' {base_count}→{after_count})"
                    )

    except Exception as e:
        failures.append(f"e2e 테스트 오류: {e}")

    finally:
        # ── ⑧ 롤백: 이번 호출에서 캡처한 ID만 삭제 ─────────────────────────
        # → 다른 실행·다른 사용자 행을 절대 건드리지 않음
        try:
            if captured_txn_id:
                cur.execute("DELETE FROM transactions WHERE id=%s", (captured_txn_id,))
            if captured_other_txn_id:
                cur.execute("DELETE FROM transactions WHERE id=%s", (captured_other_txn_id,))
            if captured_listing_id:
                cur.execute("DELETE FROM listing_requests WHERE id=%s", (captured_listing_id,))
            if captured_user_id:
                cur.execute("DELETE FROM listing_requests WHERE user_id=%s", (captured_user_id,))
            if captured_req_id:
                cur.execute(
                    "DELETE FROM building_requests WHERE id=%s", (captured_req_id,)
                )
            if captured_mb_id:
                cur.execute(
                    "DELETE FROM master_buildings WHERE id=%s", (captured_mb_id,)
                )
            if captured_user_id:
                cur.execute("DELETE FROM users WHERE id=%s", (captured_user_id,))
            conn.commit()
        except Exception as cleanup_err:
            failures.append(f"롤백 실패: {cleanup_err}")
        finally:
            cur.close()
            conn.close()
        # 사후 캐시 제거 — 다음 실행이 stale 값을 보지 않게
        try:
            _clear_caches()
        except Exception:
            pass

    return failures


def _check_master_stats_partial_success_invalidation():
    """앞선 UPDATE 커밋 뒤 다음 행 오류가 나도 외부 writer 표식이 남는지 검증한다.

    DB와 건축물대장 API는 전부 작은 fake로 대체한다. 실제
    reclassify_unclassified.run의 ``finally`` 경로를 실행하므로, 단순히 테스트
    안에서 표식 함수를 호출하는 것보다 "부분 성공 후 오류" 계약을 직접 고정한다.
    """
    from types import SimpleNamespace
    from unittest import mock

    import reclassify_unclassified as reclassify

    failures = []
    events = []
    invalidation = {"token": 17}
    token_before = invalidation["token"]

    class FakeCursor:
        rowcount = 1

        def __init__(self, rows=()):
            self.rows = list(rows)

        def execute(self, _sql, _params=None):
            events.append("source_update")

        def fetchall(self):
            return self.rows

        def close(self):
            pass

    class FakeConnection:
        def __init__(self, rows=()):
            self.cur = FakeCursor(rows)

        def cursor(self):
            return self.cur

        def commit(self):
            events.append("source_commit")

        def close(self):
            pass

    rows = [
        {"id": 1, "building_name": "부분성공", "sgg_cd": "11110",
         "umd_nm": "테스트동", "jibun": "1-1"},
        {"id": 2, "building_name": "후속오류", "sgg_cd": "11110",
         "umd_nm": "테스트동", "jibun": "2-2"},
    ]
    connections = iter((FakeConnection(rows), FakeConnection()))

    class FakeBjdongMap:
        def __init__(self):
            self.calls = 0

        def find_bjdong_cd(self, _sgg_cd, _umd_nm):
            self.calls += 1
            if self.calls == 2:
                raise RuntimeError("simulated error after committed work")
            return "10100"

    def marker(source):
        invalidation["token"] += 1
        events.append(("marker", source))

    args = SimpleNamespace(daily_cap=10, limit=None, dry_run=False, sleep=0)
    progress = {"calls_date": "", "calls_today": 0, "updated_total": 0}
    try:
        with (
            mock.patch.object(reclassify, "get_conn", side_effect=lambda: next(connections)),
            mock.patch.object(reclassify, "_get_progress", return_value=progress),
            mock.patch.object(reclassify, "_get_bjdmap", return_value=FakeBjdongMap()),
            mock.patch.object(
                reclassify,
                "classify_lodging_type",
                return_value=("생활", "생활숙박시설", "", {}, "fixture"),
            ),
            mock.patch.object(reclassify, "mark_master_stats_invalidated", side_effect=marker),
            mock.patch.object(reclassify.time, "sleep"),
        ):
            try:
                reclassify.run(args)
            except RuntimeError as exc:
                if "simulated error" not in str(exc):
                    raise
            else:
                failures.append("master stats partial success: 후속 오류 fixture가 발생하지 않음")

        commit_index = events.index("source_commit")
        marker_index = events.index(("marker", "reclassify_unclassified"))
        if marker_index <= commit_index:
            failures.append("master stats partial success: 원본 커밋 전에 무효화 표식을 기록함")
        if events.count(("marker", "reclassify_unclassified")) != 1:
            failures.append("master stats partial success: 무효화 토큰을 정확히 한 번 전진시키지 않음")
        if invalidation["token"] != token_before + 1:
            failures.append("master stats partial success: 선행 커밋 뒤 무효화 토큰이 전진하지 않음")
    except (ValueError, StopIteration) as exc:
        failures.append(f"master stats partial success: 커밋/표식 순서 누락 ({exc})")
    except Exception as exc:
        failures.append(f"master stats partial success 테스트 오류: {exc}")

    if not failures:
        print("OK  통계 원본 부분성공·후속오류에도 커밋 뒤 무효화 토큰 전진")
    return failures


def _check_buildings_geo_bounds(client):
    """bounds 파라미터 동작 검증 — 범위 필터링·잘못된 값 안전 처리."""
    failures = []

    # 전국 전체 건수 확인
    r_all = client.get("/api/buildings-geo")
    if r_all.status_code != 200:
        failures.append(f"/api/buildings-geo (전체): HTTP {r_all.status_code}")
        return failures
    total_all = (r_all.get_json() or {}).get("total", -1)

    # 서울 부근 좁은 viewport 요청 — 전체보다 적어야 한다
    bounds_qs = "sw_lat=37.4&sw_lng=126.8&ne_lat=37.7&ne_lng=127.2"
    r_bounds = client.get(f"/api/buildings-geo?{bounds_qs}")
    if r_bounds.status_code != 200:
        failures.append(f"/api/buildings-geo (bounds): HTTP {r_bounds.status_code}")
        return failures
    payload_b = r_bounds.get_json() or {}
    shape_err = check_buildings_geo(payload_b)
    if shape_err:
        failures.append(f"/api/buildings-geo (bounds): {shape_err}")
        return failures
    total_bounds = payload_b.get("total", -1)
    if total_all > 0 and total_bounds >= total_all:
        failures.append(
            f"/api/buildings-geo (bounds): 범위 필터 효과 없음"
            f" (bounds={total_bounds} >= 전체={total_all})"
        )
    else:
        print(f"OK  /api/buildings-geo?{bounds_qs}  (전체 {total_all}건 → 범위내 {total_bounds}건)")

    # 잘못된 bounds — 서버가 500 없이 응답해야 한다 (bounds 무시하고 200 반환)
    r_bad = client.get("/api/buildings-geo?sw_lat=abc&sw_lng=xyz&ne_lat=!!&ne_lng=@@")
    if r_bad.status_code != 200:
        failures.append(
            f"/api/buildings-geo (잘못된 bounds): HTTP {r_bad.status_code} (기대: 200)"
        )
    else:
        print(f"OK  /api/buildings-geo (잘못된 bounds 무시)  ({r_bad.status_code})")

    return failures


if __name__ == "__main__":
    sys.exit(run())
