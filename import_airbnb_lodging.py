#!/usr/bin/env python3
"""
외국인관광도시민박업 XLSX/CSV를 lodging_registry에 적재하고
주소가 일치하는 master_buildings를 에어비앤비 유형으로 연결한다.

사용법:
    python import_airbnb_lodging.py --dry-run 문화_외국인관광도시민박업.xlsx
    python import_airbnb_lodging.py 문화_외국인관광도시민박업.xlsx

관리번호는 자치단체마다 중복되므로 permit_number에는
AIRBNB:{개방자치단체코드}:{관리번호} 복합 식별자를 저장한다.
"""

import argparse
import csv
import io
import os
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from addr_norm import (
    get_building_jibun_key,
    normalize_jibun_prefix,
    normalize_name,
    normalize_road_prefix,
)
from address_utils import BjdongMap, normalize_umd_nm
from db import get_conn
from lodging_classification import (
    ACTIVE_STATUS,
    classify_building_use,
)
from stats_cache import mark_master_stats_invalidated


BJDONG_CODE_CSV = os.environ.get("BJDONG_CODE_CSV", "법정동코드_전체자료.zip")
HYGIENE_TYPE_FIXED = "외국인관광도시민박업"
LODGING_TYPE_FIXED = "에어비앤비"
SOURCE_KEY_PREFIX = "AIRBNB"
DRY_RUN_SAMPLE_LIMIT = 20

REQUIRED_HEADERS = frozenset({
    "개방자치단체코드",
    "관리번호",
    "인허가일자",
    "영업상태명",
    "사업장명",
    "객실수",
    "건물용도명",
    "데이터갱신시점",
    "도로명주소",
    "상세영업상태명",
    "시설규모",
    "전화번호",
    "지번주소",
    "지역구분명",
})

_AMBIGUOUS = object()
_JIBUN_RE = re.compile(
    r"([가-힣0-9]+(?:동|읍|면|리|가))\s+((?:산\s*)?\d+(?:-\d+)?)"
)
_PAREN_LOCALITY_RE = re.compile(
    r"\(\s*([가-힣0-9]+(?:동|읍|면|리|가))(?:\s*[,)]|[^)]*\))"
)


def _text(value):
    if value is None:
        return None
    result = str(value).strip()
    return result if result and result.lower() not in {"none", "nan"} else None


def _identity_text(value):
    """식별자 숫자 셀의 XLSX 정수형 `.0` 표기를 제거한다."""
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else None
    if isinstance(value, Decimal):
        return str(int(value)) if value == value.to_integral_value() else None
    text = _text(value)
    if text and re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def _date_text(value):
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return _text(value)


def _integer(value):
    text = _text(value)
    if text is None:
        return None
    try:
        number = Decimal(text.replace(",", ""))
    except (InvalidOperation, AttributeError):
        return None
    if number != number.to_integral_value():
        return None
    return int(number)


def _decimal(value):
    text = _text(value)
    if text is None:
        return None
    try:
        return Decimal(text.replace(",", ""))
    except (InvalidOperation, AttributeError):
        return None


def _phone(value):
    text = _text(value)
    if text is None:
        return None
    digits = re.sub(r"\D", "", text)
    return digits or None


def _validate_headers(headers):
    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValueError(
            "필수 컬럼이 없습니다: " + ", ".join(sorted(missing))
        )


def read_rows(filepath):
    """XLSX 또는 CSV를 한글 헤더명 기준 딕셔너리 목록으로 읽는다."""
    path = Path(filepath)
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = tuple(_text(value) or "" for value in next(values))
            _validate_headers(headers)
            return [
                {headers[index]: value for index, value in enumerate(row)}
                for row in values
            ]
        finally:
            workbook.close()

    if suffix != ".csv":
        raise ValueError("지원 형식은 .xlsx, .xlsm, .csv입니다.")

    raw = path.read_bytes()
    decoded = None
    for encoding in ("utf-8-sig", "cp949"):
        try:
            decoded = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        raise ValueError("CSV 인코딩을 확인할 수 없습니다.")
    reader = csv.DictReader(io.StringIO(decoded))
    headers = tuple(_text(value) or "" for value in (reader.fieldnames or ()))
    _validate_headers(headers)
    return [dict(row) for row in reader]


def _permit_number(authority_code, source_permit_number):
    authority = _identity_text(authority_code)
    permit = _identity_text(source_permit_number)
    if not authority or not permit:
        return None
    return f"{SOURCE_KEY_PREFIX}:{authority}:{permit}"


def parse_row(row):
    """원본 행을 DB 적재 딕셔너리로 변환한다."""
    status = _text(row.get("영업상태명"))

    permit_number = _permit_number(
        row.get("개방자치단체코드"),
        row.get("관리번호"),
    )
    biz_name = _text(row.get("사업장명"))
    if not permit_number or not biz_name:
        return None

    road_address = _text(row.get("도로명주소"))
    jibun_address = _text(row.get("지번주소"))
    return {
        "permit_number": permit_number,
        "permit_date": _date_text(row.get("인허가일자")),
        "biz_name": biz_name,
        "room_count": _integer(row.get("객실수")),
        # 캠핑장은 객실과 다른 단위로 저장한다. 일반 숙박 원본에는 없음.
        "camping_site_count": None,
        "camping_general_site_count": None,
        "camping_auto_site_count": None,
        "camping_glamping_site_count": None,
        "camping_caravan_site_count": None,
        "camping_classification": None,
        "bld_use_nm": _text(row.get("건물용도명")),
        "source_updated_at": _date_text(row.get("데이터갱신시점")),
        "road_address": road_address,
        "hygiene_type": HYGIENE_TYPE_FIXED,
        "biz_status_name": status,
        "biz_status_detail": _text(row.get("상세영업상태명")),
        "facility_area": _decimal(row.get("시설규모")),
        "phone": _phone(row.get("전화번호")),
        "jibun_address": jibun_address,
        "region_name": _text(row.get("지역구분명")),
        "road_norm": normalize_road_prefix(road_address),
        "jibun_norm": normalize_jibun_prefix(jibun_address),
        "biz_name_norm": normalize_name(biz_name),
    }


def _add_unique(index, key, building_id):
    if not key:
        return
    previous = index.get(key)
    if previous is None:
        index[key] = building_id
    elif previous != building_id:
        index[key] = _AMBIGUOUS


def _load_master_indexes(cur):
    cur.execute(
        "SELECT id, road_address, jibun_address FROM master_buildings"
    )
    road_index = {}
    jibun_index = {}
    for row in cur.fetchall():
        building = dict(row)
        _add_unique(
            road_index,
            normalize_road_prefix(building.get("road_address")),
            building["id"],
        )
        _add_unique(
            jibun_index,
            get_building_jibun_key(building),
            building["id"],
        )
    return road_index, jibun_index


def _match_master(data, road_index, jibun_index):
    road_key = data.get("road_norm")
    if road_key in road_index:
        building_id = road_index[road_key]
        return (
            (None, "도로명 주소가 여러 건물과 일치")
            if building_id is _AMBIGUOUS
            else (building_id, "road")
        )

    jibun_key = data.get("jibun_norm")
    if jibun_key in jibun_index:
        building_id = jibun_index[jibun_key]
        return (
            (None, "지번 주소가 여러 건물과 일치")
            if building_id is _AMBIGUOUS
            else (building_id, "jibun")
        )
    return None, None


def _location_from_addresses(bjdong, road_address, jibun_address):
    source_address = jibun_address or road_address
    if not source_address:
        return None
    sgg_info = bjdong.extract_sgg_from_address(source_address)
    if not sgg_info:
        return None
    si_do, sgg_nm, sgg_cd = sgg_info
    sgg_text = f"{si_do} {sgg_nm}".strip()

    locality_match = _JIBUN_RE.search(jibun_address or "")
    if locality_match:
        umd_nm = normalize_umd_nm(locality_match.group(1))
        jibun = re.sub(r"\s+", "", locality_match.group(2))
    else:
        locality_match = _PAREN_LOCALITY_RE.search(road_address or "")
        umd_nm = normalize_umd_nm(locality_match.group(1)) if locality_match else None
        jibun = None
    return {
        "sgg_cd": sgg_cd,
        "sgg_text": sgg_text,
        "umd_nm": umd_nm,
        "jibun": jibun,
    }


def _upsert_registry(cur, data, *, reset_applied_building_id=True):
    # 모든 기존 importer가 이 공통 upsert를 사용한다. 캠핑 전용 열이 추가돼도
    # 이전 원본의 payload가 깨지지 않도록 명시적으로 NULL 기본값을 보장한다.
    data = {
        **data,
        "camping_site_count": data.get("camping_site_count"),
        "camping_general_site_count": data.get("camping_general_site_count"),
        "camping_auto_site_count": data.get("camping_auto_site_count"),
        "camping_glamping_site_count": data.get("camping_glamping_site_count"),
        "camping_caravan_site_count": data.get("camping_caravan_site_count"),
        "camping_classification": data.get("camping_classification"),
    }
    applied_building_update = (
        "applied_building_id = NULL"
        if reset_applied_building_id
        else "applied_building_id = lodging_registry.applied_building_id"
    )
    cur.execute(f"""
        INSERT INTO lodging_registry (
            permit_number, biz_name, road_address, jibun_address,
            permit_date, biz_status_name, biz_status_detail,
            room_count, hygiene_type, phone,
            road_norm, jibun_norm, biz_name_norm,
            source_updated_at, bld_use_nm, facility_area, region_name,
            camping_site_count, camping_general_site_count, camping_auto_site_count,
            camping_glamping_site_count, camping_caravan_site_count, camping_classification
        ) VALUES (
            %(permit_number)s, %(biz_name)s, %(road_address)s, %(jibun_address)s,
            %(permit_date)s, %(biz_status_name)s, %(biz_status_detail)s,
            %(room_count)s, %(hygiene_type)s, %(phone)s,
            %(road_norm)s, %(jibun_norm)s, %(biz_name_norm)s,
            %(source_updated_at)s, %(bld_use_nm)s, %(facility_area)s, %(region_name)s,
            %(camping_site_count)s, %(camping_general_site_count)s, %(camping_auto_site_count)s,
            %(camping_glamping_site_count)s, %(camping_caravan_site_count)s, %(camping_classification)s
        )
        ON CONFLICT (permit_number) DO UPDATE SET
            biz_name = EXCLUDED.biz_name,
            road_address = EXCLUDED.road_address,
            jibun_address = EXCLUDED.jibun_address,
            permit_date = EXCLUDED.permit_date,
            biz_status_name = EXCLUDED.biz_status_name,
            biz_status_detail = EXCLUDED.biz_status_detail,
            room_count = EXCLUDED.room_count,
            camping_site_count = EXCLUDED.camping_site_count,
            camping_general_site_count = EXCLUDED.camping_general_site_count,
            camping_auto_site_count = EXCLUDED.camping_auto_site_count,
            camping_glamping_site_count = EXCLUDED.camping_glamping_site_count,
            camping_caravan_site_count = EXCLUDED.camping_caravan_site_count,
            camping_classification = EXCLUDED.camping_classification,
            hygiene_type = EXCLUDED.hygiene_type,
            phone = EXCLUDED.phone,
            road_norm = EXCLUDED.road_norm,
            jibun_norm = EXCLUDED.jibun_norm,
            biz_name_norm = EXCLUDED.biz_name_norm,
            source_updated_at = EXCLUDED.source_updated_at,
            bld_use_nm = EXCLUDED.bld_use_nm,
            facility_area = EXCLUDED.facility_area,
            region_name = EXCLUDED.region_name,
            {applied_building_update},
            updated_at = NOW()
        RETURNING id, (xmax = 0) AS is_new
    """, data)
    return cur.fetchone()


def _assert_schema(cur):
    cur.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public'
          AND table_name = 'lodging_registry'
          AND column_name = ANY(%s)
    """, [[
        "bld_use_nm",
        "facility_area",
        "region_name",
        "applied_building_id",
        "camping_site_count",
        "camping_general_site_count",
        "camping_auto_site_count",
        "camping_glamping_site_count",
        "camping_caravan_site_count",
        "camping_classification",
    ]])
    found = {row["column_name"] for row in cur.fetchall()}
    required = {
        "bld_use_nm",
        "facility_area",
        "region_name",
        "applied_building_id",
        "camping_site_count",
        "camping_general_site_count",
        "camping_auto_site_count",
        "camping_glamping_site_count",
        "camping_caravan_site_count",
        "camping_classification",
    }
    missing = required - found
    if missing:
        raise RuntimeError(
            "운영 DB 스키마가 아직 Publish되지 않았습니다. 누락: "
            + ", ".join(sorted(missing))
        )


def _create_master(cur, data, location):
    address = data.get("road_address") or data.get("jibun_address")
    cur.execute("""
        INSERT INTO master_buildings (
            building_name, sgg_cd, sgg_text, umd_nm, jibun,
            road_address, jibun_address, lodging_type, lodging_type_detail,
            units, source, name_pending, building_use_type, building_use_detail,
            lodging_classification_source, lodging_classification_confidence
        ) VALUES (
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, 'airbnb_import', TRUE, %s, %s, 'active_permit', 'high'
        )
        RETURNING id
    """, (
        data["biz_name"],
        location["sgg_cd"],
        location["sgg_text"],
        location["umd_nm"],
        location["jibun"],
        address,
        data.get("jibun_address"),
        LODGING_TYPE_FIXED,
        HYGIENE_TYPE_FIXED,
        data.get("room_count") or 0,
        classify_building_use(data.get("bld_use_nm")),
        data.get("bld_use_nm"),
    ))
    return cur.fetchone()["id"]


def _register_new_master_in_indexes(
    building_id, data, road_index, jibun_index
):
    _add_unique(road_index, data.get("road_norm"), building_id)
    _add_unique(jibun_index, data.get("jibun_norm"), building_id)


def run(filepath, dry_run=False):
    rows = read_rows(filepath)
    parsed = [data for data in (parse_row(row) for row in rows) if data]
    skipped = len(rows) - len(parsed)
    active = [
        data for data in parsed
        if data["biz_status_name"] == ACTIVE_STATUS
    ]
    inactive_count = len(parsed) - len(active)
    print(
        f"[로드] 전체 {len(rows):,}행 / 식별 가능 {len(parsed):,}행 / "
        f"영업·정상 {len(active):,}행 / 비영업 {inactive_count:,}행 / "
        f"제외 {skipped:,}행"
    )

    if dry_run:
        for data in active[:DRY_RUN_SAMPLE_LIMIT]:
            print(
                f"  [DRY] {data['permit_number']} | {data['biz_name']} | "
                f"{data['road_address'] or data['jibun_address'] or '-'} | "
                f"{data['facility_area'] if data['facility_area'] is not None else '-'}㎡"
            )
        if len(active) > DRY_RUN_SAMPLE_LIMIT:
            print(f"  ... 나머지 {len(active) - DRY_RUN_SAMPLE_LIMIT:,}행 생략")
        return {
            "total": len(rows),
            "valid": len(parsed),
            "active": len(active),
            "inactive": inactive_count,
            "skipped": skipped,
        }

    bjdong = BjdongMap(BJDONG_CODE_CSV)
    conn = get_conn()
    cur = conn.cursor()
    counters = {
        "inserted": 0,
        "updated": 0,
        "matched": 0,
        "created": 0,
        "inactive": 0,
        "unmatched": 0,
        "failed": 0,
    }
    try:
        _assert_schema(cur)
        road_index, jibun_index = _load_master_indexes(cur)
        for data in parsed:
            try:
                registry = _upsert_registry(cur, data)
                registry_counter = (
                    "inserted" if registry["is_new"] else "updated"
                )
                outcome_counter = None
                new_building_id = None

                if data["biz_status_name"] != ACTIVE_STATUS:
                    building_id = None
                    outcome_counter = "inactive"
                elif not data.get("road_norm") and not data.get("jibun_norm"):
                    building_id = None
                    outcome_counter = "unmatched"
                    print(
                        f"  [주소불충분] {data['biz_name']} | "
                        f"{data['road_address'] or data['jibun_address'] or '-'}"
                    )
                else:
                    building_id, match_reason = _match_master(
                        data, road_index, jibun_index
                    )
                    if building_id:
                        # 영업신고 주소 일치는 건축물대장 용도 판정 근거가 아니다.
                        # 기존 건물의 lodging_type/verified_at은 유지하고 신고만 연결한다.
                        outcome_counter = "matched"
                    elif match_reason:
                        outcome_counter = "unmatched"
                        print(
                            f"  [검토] {data['biz_name']} — {match_reason}: "
                            f"{data['road_address'] or data['jibun_address'] or '-'}"
                        )
                    else:
                        location = _location_from_addresses(
                            bjdong,
                            data.get("road_address"),
                            data.get("jibun_address"),
                        )
                        if not location:
                            outcome_counter = "unmatched"
                            print(
                                f"  [주소실패] {data['biz_name']} | "
                                f"{data['road_address'] or data['jibun_address'] or '-'}"
                            )
                        else:
                            building_id = _create_master(cur, data, location)
                            new_building_id = building_id
                            outcome_counter = "created"

                if building_id:
                    cur.execute(
                        "UPDATE lodging_registry "
                        "SET applied_building_id = %s WHERE id = %s",
                        (building_id, registry["id"]),
                    )
                conn.commit()
                counters[registry_counter] += 1
                counters[outcome_counter] += 1
                if new_building_id:
                    _register_new_master_in_indexes(
                        new_building_id, data, road_index, jibun_index
                    )
            except Exception as exc:
                conn.rollback()
                counters["failed"] += 1
                print(f"  [실패] {data['biz_name']}: {exc}")
    finally:
        cur.close()
        conn.close()

    if counters["inserted"] or counters["updated"]:
        mark_master_stats_invalidated("airbnb_import")
    print(
        "\n완료 — "
        f"신규 적재 {counters['inserted']:,} / 갱신 {counters['updated']:,} / "
        f"기존 건물 연결 {counters['matched']:,} / 신규 건물 {counters['created']:,} / "
        f"비영업 상태 {counters['inactive']:,} / 미연결 {counters['unmatched']:,} / "
        f"실패 {counters['failed']:,}"
    )
    return counters


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("filepath", help="CSV 또는 XLSX 파일 경로")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="DB 연결·스키마 변경 없이 파싱 결과만 확인",
    )
    args = parser.parse_args()
    run(args.filepath, dry_run=args.dry_run)